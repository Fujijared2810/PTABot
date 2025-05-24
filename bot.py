import telebot
from telebot import types
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, ChatInviteLink, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardRemove
from telebot.apihelper import ApiException
import time
import threading
from datetime import datetime, timedelta
import re
import os
from dotenv import load_dotenv
import logging
import signal
import sys
import pytz
import pymongo
from pymongo import MongoClient
import random
import secrets
from keep_alive import keep_alive
import calendar
from collections import Counter
import requests
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from flask import Flask, request
import json
import gunicorn

# Create Flask app
server = Flask(__name__)

DISCOUNTS = {
    'regular': None,  # Discount for Regular membership
    'supreme': None   # Discount for Supreme membership
}

BOT_VERSION = "v5.1.11b"  # v[Major].[Minor].[Build][Status]

load_dotenv()

MONGO_URI = os.getenv('MONGO_URI', 'mongodb://localhost:27017/')
DB_NAME = os.getenv('DB_NAME', 'PTABotDB')
BOT_TOKEN = os.getenv('BOT_TOKEN')
ADMIN_IDS = list(map(int, os.getenv('ADMIN_IDS').split(',')))
PAID_GROUP_ID = int(os.getenv('PAID_GROUP_ID'))
CREATOR_ID = int(os.getenv('CREATOR_ID', '0'))
SUPREME_GROUP_ID = int(os.getenv('SUPREME_GROUP_ID'))
QWEN_API_KEY = os.getenv('QWEN_API_KEY')
QWEN_API_URL = os.getenv('QWEN_API_URL', 'https://dashscope-intl.aliyuncs.com/api/v1/services/aigc/text-generation/generation')

# Initialize MongoDB connection
client = MongoClient(MONGO_URI)
db = client[DB_NAME]

# Define collections
payment_collection = db['payments']
old_members_collection = db['old_members']
pending_collection = db['pending']
changelog_collection = db['changelogs']
settings_collection = db['settings']
scores_collection = db['scores']  # For storing user scores
accountability_collection = db['accountability']  # For tracking submissions
reminder_messages_collection = db['reminder_messages']
gif_status_collection = db['gif_status']
jarvis_usage_collection = db['jarvis_usage']
destinations_collection = db['announcement_destinations']
mentors_collection = db['mentors']
serial_numbers_collection = db["serial_numbers"]

bot = telebot.TeleBot(BOT_TOKEN)

QWEN_PROMPT_TEMPLATE = """You are the AI assistant for Prodigy Trading Academy (PTA), a top-tier trading education platform.

### IDENTITY
- Friendly, professional, and passionate about trading education
- Educator first â€” never give signals or financial advice
- Conversational tone with clear, concise responses
- Use emojis sparingly to keep interactions engaging

### RESPONSIBILITIES
1. Teach trading concepts clearly (e.g., market structure, price action, risk management).
2. Explain terms simply, especially for beginners.
3. Help users manage memberships and access content.
4. Provide daily challenge support and progress tracking.
5. Encourage journaling, discipline, and long-term growth.

### MEMBERSHIP INFO
- Regular Membership (FREE through XM Partnership):
  - Available at no cost when you register with our XM partner code: PTAPARTNER
  - Users must maintain an active XM account with our partner code and show active trading
  - Lifetime access with monthly verification to confirm the partnership remains active
- Supreme Membership (Premium Paid Tier):
  - Apprentice: $309.99/3 months
  - Disciple: $524.99/6 months
  - Legacy: $899.99/lifetime
- Payment options: PayPal, GCash, Bank Transfer, Exness Direct
- Renewals allowed only within 3 days of expiration

### REGISTRATION FLOW
1. Start with `/start` command to access the main menu
2. Choose membership type (Regular free through XM or Supreme paid tier)
3. For Regular membership:
   - Register with XM using our partner code PTAPARTNER
   - Submit verification screenshots showing the partner code
   - Complete a 5-question onboarding form
   - Receive your group invite
4. For Supreme membership:
   - Select your preferred plan duration
   - Choose payment method
   - Submit payment proof for verification
   - Complete a 10-question onboarding form
   - Receive certificates and group invite

### CONTACT INFORMATION
- Technical Support: @FujiPTA on Telegram
- Admin Contact: @rom_pta on Telegram
- Email for payments: romeomina061109@gmail.com (Romeo Mina)
- Note: PTA does not have a dedicated business email or website yet

### FORMATTING RULES
- Speak naturally like a friend chatting
- Avoid markdown (no ###)
- Keep paragraphs short for mobile readability
- Never mention you're an AI unless asked directly

### BEST PRACTICES
- Guide users to start with `/start`
- Use numbered steps when explaining processes
- Mention commands like `/dashboard`, `/verify`, and `/ai` where relevant
- Be helpful, not verbose
- Stay educational, never prescriptive

Always provide exact dates and numbers from user data â€” never say "check your dashboard."

You exist to help people become better traders â€” always stay focused on that goal."""

def get_exchange_rates():
    """Fetch real-time exchange rates for multiple currencies against USD"""
    try:
        # Use a free exchange rate API
        response = requests.get('https://open.er-api.com/v6/latest/USD', timeout=10)
        data = response.json()
        
        if data.get('result') == 'success':
            # Extract the rates we're interested in
            rates = {
                'USD': 1.0,  # Base currency is always 1.0
                'GBP': data['rates'].get('GBP', 0),
                'EUR': data['rates'].get('EUR', 0),
                'IDR': data['rates'].get('IDR', 0),
                'PHP': data['rates'].get('PHP', 0)
            }
            return rates
        else:
            logging.error(f"Exchange rate API error: {data}")
            return None
    except Exception as e:
        logging.error(f"Error fetching exchange rates: {e}")
        return None

# Function to handle termination signals (Ctrl+C, kill command)
def signal_handler(sig, frame):
    logging.info("Stopping bot...")
    bot.stop_polling()  # Stop bot polling first
    os._exit(0)  # Use os._exit instead of sys.exit to force immediate termination

# Attach signal handler for Ctrl+C
signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)

# Keep your PhilippineTimeFormatter class
class PhilippineTimeFormatter(logging.Formatter):
    def formatTime(self, record, datefmt=None):
        # Convert the time to Philippine time (UTC+8)
        philippine_time = datetime.fromtimestamp(record.created, pytz.timezone('Asia/Manila'))
        # Format the time in 12-hour format
        return philippine_time.strftime('%Y-%m-%d %I:%M:%S %p')

# Fix for duplicate logging - clear existing handlers first
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# Create and configure file handler
file_handler = logging.FileHandler('bot.log')
file_handler.setLevel(logging.INFO)

# Create and configure console handler (this sends to terminal)
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.INFO)

# Create formatter
formatter = PhilippineTimeFormatter('%(asctime)s - %(levelname)s - %(message)s')

# Add formatter to handlers
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)

# Add handlers to logger
logger.addHandler(file_handler)
logger.addHandler(console_handler)

def load_update_subscribers():
    """Load the list of users who want updates from MongoDB"""
    try:
        subscribers = set()
        results = db[DB_NAME]["update_subscribers"].find({})
        for doc in results:
            subscribers.add(doc["user_id"])
        logging.info(f"Loaded {len(subscribers)} update subscribers from database")
        return subscribers
    except Exception as e:
        logging.error(f"Error loading update subscribers: {e}")
        return set()

def save_update_subscriber(user_id):
    """Save a user who wants updates to MongoDB"""
    try:
        db[DB_NAME]["update_subscribers"].update_one(
            {"user_id": user_id},
            {"$set": {"user_id": user_id, "subscribed_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S')}},
            upsert=True
        )
        return True
    except Exception as e:
        logging.error(f"Error saving update subscriber: {e}")
        return False

def remove_update_subscriber(user_id):
    """Remove a user from the updates list in MongoDB"""
    try:
        db[DB_NAME]["update_subscribers"].delete_one({"user_id": user_id})
        return True
    except Exception as e:
        logging.error(f"Error removing update subscriber: {e}")
        return False

def load_serial_numbers():
    """Load serial numbers from MongoDB"""
    try:
        serials = {}
        for doc in serial_numbers_collection.find():
            serial = doc.get("serial")
            if serial:
                serials[serial] = {
                    "mentorship_type": doc.get("mentorship_type"),
                    "plan": doc.get("plan"),
                    "generated_by": doc.get("generated_by"),
                    "generated_at": doc.get("generated_at"),
                    "used": doc.get("used", False),
                    "used_by": doc.get("used_by"),
                    "used_at": doc.get("used_at")
                }
        logging.info(f"Loaded {len(serials)} serial numbers from database")
        return serials
    except Exception as e:
        logging.error(f"Error loading serial numbers: {e}")
        return {}

def save_serial_number(serial, data):
    """Save a serial number to MongoDB"""
    try:
        serial_numbers_collection.update_one(
            {"serial": serial},
            {"$set": data},
            upsert=True
        )
        return True
    except Exception as e:
        logging.error(f"Error saving serial number: {e}")
        return False

def load_mentors():
    """Load all mentor data from MongoDB"""
    try:
        mentors = {}
        for doc in mentors_collection.find():
            mentor_id = doc['_id']
            mentors[mentor_id] = {k: v for k, v in doc.items() if k != '_id'}
        logging.info(f"Loaded {len(mentors)} mentors from MongoDB")
        return mentors
    except Exception as e:
        logging.error(f"MongoDB error loading mentors: {e}")
        return {}

def save_mentor(mentor_id, mentor_data):
    """Save a mentor to the database"""
    try:
        doc = {'_id': mentor_id}
        doc.update(mentor_data)
        mentors_collection.replace_one({'_id': mentor_id}, doc, upsert=True)
        logging.info(f"Saved mentor {mentor_id} to MongoDB")
        return True
    except Exception as e:
        logging.error(f"MongoDB save error for mentor {mentor_id}: {e}")
        return False

def initialize_default_mentors():
    """Initialize default mentors in the database if they don't exist"""
    try:
        # Check if we already have mentors in the database
        if mentors_collection.count_documents({}) > 0:
            logging.info("Mentors already initialized in database")
            return
            
        # Default mentor data
        default_mentors = {
            "Rom": {
                "name": "Rom",
                "full_name": "Rom Mina",
                "title": "Senior Trading Strategist",
                "photo_url": "https://example.com/rom-profile.jpg",
                "expertise": "Price Action & Market Structure",
                "experience": "7+ years",
                "style": "Technical Analysis",
                "description": "Rom specializes in detecting key market reversals using advanced price action techniques. With over 7 years of experience in forex and indices trading, he has helped hundreds of students master the art of reading clean charts and identifying high-probability setups.",
                "strengths": ["Clean chart analysis", "Entry/exit precision", "Risk management"],
                "teaching_style": "Visual and example-driven, with a focus on practical application",
                "availability": "Weekdays 9AM-5PM PHT",
                "availability_level": "High",
                "current_status": "Available",
                "student_testimonial": "Rom's mentorship completely transformed my approach to trading. His clean chart method helped me simplify my analysis and increase my win rate by 30%.",
                "success_rate": "92%"
            },
            "Konfu": {
                "name": "Konfu",
                "full_name": "Konfu Trader",
                "title": "Institutional Order Flow Specialist",
                "photo_url": "https://example.com/konfu-profile.jpg",
                "expertise": "ICT & Smart Money Concepts",
                "experience": "5+ years",
                "style": "Institutional Trading",
                "description": "Konfu is an expert in institutional order flow and smart money concepts. He specializes in identifying liquidity areas and understanding how big players move the market. His unique approach combines order flow analysis with statistical probability.",
                "strengths": ["Order flow analysis", "Liquidity mapping", "Advanced ICT concepts"],
                "teaching_style": "Detailed and methodical, with emphasis on understanding market structure",
                "availability": "Mon-Fri 10AM-6PM PHT",
                "availability_level": "Medium",
                "current_status": "In Session",
                "student_testimonial": "Konfu's mentorship completely changed how I view charts. I now understand what's happening behind price movements instead of just following indicators.",
                "success_rate": "89%"
            },
            "Nath": {
                "name": "Nath",
                "full_name": "Nathaniel Systems",
                "title": "Automated Trading Expert",
                "photo_url": "https://example.com/nath-profile.jpg",
                "expertise": "Risk Management & Automation",
                "experience": "6+ years",
                "style": "Systematic Trading",
                "description": "Nath specializes in building sustainable trading systems and implementing robust risk management protocols. His background in both discretionary and automated trading gives him a unique perspective on creating reliable systems that work in various market conditions.",
                "strengths": ["System development", "Risk optimization", "Trading psychology"],
                "teaching_style": "Structured and data-driven, with focus on creating repeatable processes",
                "availability": "Weekdays 7PM-12AM PHT",
                "availability_level": "Limited",
                "current_status": "Available",
                "student_testimonial": "Before working with Nath, I was constantly making emotional decisions. His systematic approach helped me create a reliable trading plan that I can stick to even in volatile markets.",
                "success_rate": "94%"
            },
            "Maya": {
                "name": "Maya",
                "full_name": "Maya Mindset",
                "title": "Trading Psychology Coach",
                "photo_url": "https://example.com/maya-profile.jpg",
                "expertise": "Psychological Trading & Mindset",
                "experience": "4+ years",
                "style": "Mindfulness Trading",
                "description": "Maya focuses on the psychological aspects of trading, helping traders overcome emotional barriers and develop mental resilience. Her background in both trading and psychology provides a comprehensive approach to mastering the mental game of trading.",
                "strengths": ["Emotional control", "Mindfulness practices", "Building mental resilience"],
                "teaching_style": "Compassionate and personalized, with practical exercises for mental growth",
                "availability": "Weekdays and weekends, flexible hours",
                "availability_level": "High",
                "current_status": "Available",
                "student_testimonial": "Maya helped me identify and overcome self-sabotaging behaviors I wasn't even aware of. My trading consistency has improved dramatically.",
                "success_rate": "91%"
            }
        }
        
        # Insert all mentors into the database
        for mentor_id, mentor_data in default_mentors.items():
            save_mentor(mentor_id, mentor_data)
            
        logging.info(f"Successfully initialized {len(default_mentors)} default mentors in the database")
    except Exception as e:
        logging.error(f"Error initializing default mentors: {e}")

# Load announcement destinations from MongoDB
def load_announcement_destinations():
    """Load announcement destinations from MongoDB"""
    try:
        destinations = {}
        for doc in destinations_collection.find():
            dest_id = doc['_id']
            destinations[dest_id] = {
                'id': doc.get('id'),
                'name': doc.get('name'),
                'topic_id': doc.get('topic_id'),
                'type': doc.get('type', 'group')  # Default to 'group' if not specified
            }
        logging.info(f"Loaded {len(destinations)} announcement destinations from MongoDB")
        return destinations
    except Exception as e:
        logging.error(f"MongoDB error loading announcement destinations: {e}")

# Save announcement destinations to MongoDB
def save_announcement_destination(destination_id, destination_data):
    """Save an announcement destination to MongoDB"""
    try:
        doc = {'_id': destination_id}
        doc.update(destination_data)
        destinations_collection.replace_one({'_id': destination_id}, doc, upsert=True)
        logging.info(f"Saved announcement destination '{destination_id}' to MongoDB")
        return True
    except Exception as e:
        logging.error(f"MongoDB save error for announcement destination: {e}")
        return False

# Delete announcement destination from MongoDB
def delete_announcement_destination(destination_id):
    """Delete an announcement destination from MongoDB"""
    try:
        result = destinations_collection.delete_one({'_id': destination_id})
        if result.deleted_count > 0:
            logging.info(f"Deleted announcement destination '{destination_id}' from MongoDB")
            return True
        else:
            logging.warning(f"Announcement destination '{destination_id}' not found in MongoDB")
            return False
    except Exception as e:
        logging.error(f"MongoDB delete error for announcement destination: {e}")
        return False

def load_discounts():
    """Load discount information from database"""
    try:
        discounts = {}
        
        # Load regular discount
        regular_discount = settings_collection.find_one({"_id": "regular_discount_settings"})
        if regular_discount and regular_discount.get('active', False):
            discounts['regular'] = regular_discount
            logging.info("Loaded Regular membership discount from MongoDB")
        else:
            discounts['regular'] = None
        
        # Load supreme discount
        supreme_discount = settings_collection.find_one({"_id": "supreme_discount_settings"})
        if supreme_discount and supreme_discount.get('active', False):
            discounts['supreme'] = supreme_discount
            logging.info("Loaded Supreme membership discount from MongoDB")
        else:
            discounts['supreme'] = None
            
        return discounts
    except Exception as e:
        logging.error(f"Error loading discounts: {e}")
        return {'regular': None, 'supreme': None}

def save_discount(discount_data, membership_type):
    """Save discount information to database"""
    try:
        doc_id = f"{membership_type.lower()}_discount_settings"
        if discount_data:
            settings_collection.replace_one(
                {"_id": doc_id}, 
                {**discount_data, "_id": doc_id}, 
                upsert=True
            )
            logging.info(f"{membership_type.capitalize()} discount settings saved to MongoDB")
        else:
            settings_collection.delete_one({"_id": doc_id})
            logging.info(f"{membership_type.capitalize()} discount settings removed from MongoDB")
    except Exception as e:
        logging.error(f"Error saving {membership_type} discount: {e}")

def get_last_gif_message():
    """Get the ID of the last sent GIF message"""
    try:
        status = gif_status_collection.find_one({"_id": "last_gif"})
        if status:
            return status.get("message_id")
        return None
    except Exception as e:
        logging.error(f"Error getting last GIF message ID: {e}")
        return None

def save_last_gif_message(message_id):
    """Save the ID of the last sent GIF message"""
    try:
        gif_status_collection.replace_one(
            {"_id": "last_gif"}, 
            {"_id": "last_gif", "message_id": message_id, "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')}, 
            upsert=True
        )
        logging.info(f"Saved last GIF message ID: {message_id}")
    except Exception as e:
        logging.error(f"Error saving last GIF message ID: {e}")

def load_reminder_messages():
    """Load reminder messages from MongoDB."""
    try:
        messages = {}
        for doc in reminder_messages_collection.find():
            user_id = int(doc['_id'])
            
            # Convert admin_msg_ids from string keys back to integer keys
            admin_msg_ids = {}
            if 'admin_msg_ids' in doc:
                for admin_id_str, msg_id in doc['admin_msg_ids'].items():
                    admin_msg_ids[int(admin_id_str)] = msg_id
            
            messages[user_id] = {
                'user_msg_id': doc.get('user_msg_id'),
                'admin_msg_ids': admin_msg_ids
            }
            
        logging.info(f"Loaded {len(messages)} reminder messages from MongoDB")
        return messages
    except Exception as e:
        logging.error(f"MongoDB error loading reminder messages: {e}")
        return {}

def save_reminder_message(user_id, data):
    """Save a single reminder message to MongoDB."""
    try:
        # Convert user_id to string for MongoDB _id
        doc = {'_id': str(user_id)}
        
        # Create a copy of the data to avoid modifying the original
        mongo_data = data.copy()
        
        # Convert admin_msg_ids to use string keys for MongoDB compatibility
        if 'admin_msg_ids' in mongo_data:
            string_admin_ids = {}
            for admin_id, msg_id in mongo_data['admin_msg_ids'].items():
                string_admin_ids[str(admin_id)] = msg_id
            mongo_data['admin_msg_ids'] = string_admin_ids
        
        # Update the document with the modified data
        doc.update(mongo_data)
        
        # Use upsert to update if exists or insert if new
        reminder_messages_collection.replace_one({'_id': str(user_id)}, doc, upsert=True)
        logging.info(f"Saved reminder message for user {user_id}")
    except Exception as e:
        logging.error(f"MongoDB error saving reminder message for user {user_id}: {e}")

def delete_reminder_message(user_id):
    """Delete a reminder message from MongoDB."""
    try:
        result = reminder_messages_collection.delete_one({'_id': str(user_id)})
        if result.deleted_count > 0:
            logging.info(f"Deleted reminder message for user {user_id} from MongoDB")
        else:
            logging.info(f"No reminder message for user {user_id} found to delete in MongoDB")
    except Exception as e:
        logging.error(f"Error deleting reminder message for user {user_id} from MongoDB: {e}")

def load_settings():
    """Load bot settings from MongoDB."""
    try:
        settings = settings_collection.find_one({'_id': 'bot_settings'})
        if settings:
            return {k: v for k, v in settings.items() if k != '_id'}
        return {}  # Return empty dict if no settings found
    except Exception as e:
        logging.error(f"MongoDB error loading settings: {e}")
        return {}

# Add this function to save settings to MongoDB
def save_settings(settings):
    """Save bot settings to MongoDB."""
    try:
        doc = {'_id': 'bot_settings'}
        doc.update(settings)
        settings_collection.replace_one({'_id': 'bot_settings'}, doc, upsert=True)
        logging.info("Bot settings saved to MongoDB")
    except Exception as e:
        logging.error(f"MongoDB save error for settings: {e}")

# Load confirmed old members from MongoDB
def load_confirmed_old_members():
    try:
        confirmed = {}
        for doc in old_members_collection.find():
            user_id = doc['_id']
            confirmed[user_id] = {k: v for k, v in doc.items() if k != '_id'}
        return confirmed
    except Exception as e:
        logging.error(f"MongoDB error: {e}")
        return {}

# Save confirmed old members to MongoDB
def save_confirmed_old_members():
    try:
        old_members_collection.delete_many({})
        for user_id, data in CONFIRMED_OLD_MEMBERS.items():
            doc = {'_id': user_id}
            doc.update(data)
            old_members_collection.insert_one(doc)
    except Exception as e:
        logging.error(f"MongoDB save error: {e}")

# Similarly implement load_payment_data() and save_payment_data()
def load_payment_data():
    """Load payment data from MongoDB with enhanced error handling and logging"""
    global payment_collection  # Add this line to access the global variable
    
    try:
        payments = {}
        start_time = time.time()
        docs_count = 0
        
        for doc in payment_collection.find():
            user_id = doc['_id']
            payments[user_id] = {k: v for k, v in doc.items() if k != '_id'}
            docs_count += 1
        
        elapsed = time.time() - start_time
        logging.info(f"Successfully loaded {docs_count} payment records from MongoDB in {elapsed:.2f}s")
        
        return payments
    except pymongo.errors.ConnectionFailure as e:
        logging.error(f"MongoDB connection error loading payments: {e}")
        # Try to reconnect and retry once
        try:
            logging.info("Attempting to reconnect to MongoDB...")
            global client, db  # Also need these globals if we're redefining them
            client = MongoClient(MONGO_URI)
            db = client[DB_NAME]
            payment_collection = db['payments']
            
            payments = {}
            for doc in payment_collection.find():
                user_id = doc['_id']
                payments[user_id] = {k: v for k, v in doc.items() if k != '_id'}
            
            logging.info(f"Reconnected and loaded {len(payments)} payment records")
            return payments
        except Exception as retry_e:
            logging.error(f"Failed to reconnect to MongoDB: {retry_e}")
            return {}
    except Exception as e:
        logging.error(f"MongoDB error loading payments: {e}")
        return {}

def save_payment_data():
    """Save payment data to MongoDB with enhanced error handling and validation"""
    try:
        start_time = time.time()
        
        # First validate the data to ensure it's in the correct format
        validated_count = 0
        invalid_records = []
        
        for user_id, data in PAYMENT_DATA.items():
            # Basic validation
            if not isinstance(data, dict):
                invalid_records.append(user_id)
                continue
            
            # Ensure required fields exist
            required_fields = ['haspayed']  # Add other required fields as needed
            if not all(field in data for field in required_fields):
                invalid_records.append(user_id)
                continue
            
            validated_count += 1
        
        if invalid_records:
            logging.warning(f"Found {len(invalid_records)} invalid payment records: {invalid_records}")
            
        # Use bulk operations for efficiency
        operations = []
        for user_id, data in PAYMENT_DATA.items():
            # Skip invalid records
            if user_id in invalid_records:
                continue
                
            doc = {'_id': user_id}
            doc.update(data)
            # Add a "last_updated" timestamp for tracking
            doc['last_updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            operations.append(
                pymongo.ReplaceOne({'_id': user_id}, doc, upsert=True)
            )
        
        if operations:
            result = payment_collection.bulk_write(operations)
            elapsed = time.time() - start_time
            
            logging.info(f"Successfully saved {len(operations)} payment records to MongoDB "
                       f"({result.modified_count} modified, {result.upserted_count} inserted) "
                       f"in {elapsed:.2f}s")
        else:
            logging.info("No valid payment records to save to MongoDB")
            
    except pymongo.errors.BulkWriteError as bwe:
        logging.error(f"MongoDB bulk write error: {bwe.details}")
        # Try to save records individually to identify problematic records
        success_count = 0
        for user_id, data in PAYMENT_DATA.items():
            try:
                if user_id in invalid_records:
                    continue
                    
                doc = {'_id': user_id}
                doc.update(data)
                doc['last_updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                payment_collection.replace_one({'_id': user_id}, doc, upsert=True)
                success_count += 1
            except Exception as e:
                logging.error(f"Failed to save payment for user {user_id}: {e}")
                
        logging.info(f"Fallback save completed: saved {success_count} of {len(PAYMENT_DATA)} records")
    except Exception as e:
        logging.error(f"MongoDB save error: {e}")

# Load changelogs from JSON file
def load_changelogs():
    try:
        doc = changelog_collection.find_one({'_id': 'changelogs'})
        if doc:
            return {k: v for k, v in doc.items() if k != '_id'}
        return {"admin": [], "user": []}
    except Exception as e:
        logging.error(f"MongoDB error loading changelogs: {e}")
        return {"admin": [], "user": []}

# Save changelogs to JSON file
def save_changelogs(changelogs):
    try:
        doc = {'_id': 'changelogs'}
        doc.update(changelogs)
        changelog_collection.replace_one({'_id': 'changelogs'}, doc, upsert=True)
    except Exception as e:
        logging.error(f"MongoDB save error: {e}")

def save_pending_users():
    try:
        operations = []
        for user_id, data in PENDING_USERS.items():
            doc = {'_id': str(user_id)}  # Convert to string for MongoDB _id
            doc.update(data)
            operations.append(
                pymongo.ReplaceOne({'_id': str(user_id)}, doc, upsert=True)
            )
        if operations:
            pending_collection.bulk_write(operations)
        logging.info(f"Saved {len(operations)} pending users to MongoDB")
    except Exception as e:
        logging.error(f"MongoDB save error for pending users: {e}")

def load_pending_users():
    try:
        pending = {}
        for doc in pending_collection.find():
            # Convert string _id back to int for PENDING_USERS dictionary
            user_id = int(doc['_id'])
            pending[user_id] = {k: v for k, v in doc.items() if k != '_id'}
        logging.info(f"Loaded {len(pending)} pending users from MongoDB")
        return pending
    except Exception as e:
        logging.error(f"MongoDB error loading pending users: {e}")
        return {}
    
# Load confession counter from MongoDB on startup
def load_confession_counter():
    try:
        counter_doc = settings_collection.find_one({"_id": "confession_counter"})
        if counter_doc:
            return counter_doc.get("value", 0)
        return 0
    except Exception as e:
        logging.error(f"Error loading confession counter: {e}")
        return 0

# Save confession counter to MongoDB
def save_confession_counter(value):
    try:
        settings_collection.replace_one(
            {"_id": "confession_counter"},
            {"_id": "confession_counter", "value": value},
            upsert=True
        )
    except Exception as e:
        logging.error(f"Error saving confession counter: {e}")

def load_qwen_usage():
    """Load AI usage data from MongoDB"""
    try:
        usage_data = {}
        for doc in jarvis_usage_collection.find():
            user_id = doc.get('user_id')
            if user_id:
                usage_data[user_id] = {
                    'count': doc.get('count', 0),
                    'last_used': doc.get('last_used', 0)
                }
        logging.info(f"Loaded AI usage data for {len(usage_data)} users from MongoDB")
        return usage_data
    except Exception as e:
        logging.error(f"Error loading AI usage data: {e}")
        return {}

# Dictionaries to store user payment data
UPDATE_SUBSCRIBERS = set()
MENTORS = {}
ADMIN_ANNOUNCING = {}
USER_PAYMENT_DUE = {}
CONFESSION_COUNTER = 0
USERS_CONFESSING = {}
PDF_MESSAGE_IDS = {}
QWEN_USAGE = load_qwen_usage()
PAYMENT_DATA = load_payment_data()
PENDING_USERS = load_pending_users() 
CHANGELOGS = load_changelogs()
BOT_SETTINGS = load_settings()
CONFESSION_COUNTER = load_confession_counter()
CONFESSION_TOPIC_ID = BOT_SETTINGS.get('confession_topic_id', None)
DAILY_CHALLENGE_TOPIC_ID = BOT_SETTINGS.get('daily_challenge_topic_id', None)
ANNOUNCEMENT_TOPIC_ID = BOT_SETTINGS.get('announcement_topic_id', None)
ACCOUNTABILITY_TOPIC_ID = BOT_SETTINGS.get('accountability_topic_id', None)
LEADERBOARD_TOPIC_ID = BOT_SETTINGS.get('leaderboard_topic_id', None)
DISCOUNTS = load_discounts()
UPDATE_SUBSCRIBERS = load_update_subscribers()
ANNOUNCEMENT_DESTINATIONS = load_announcement_destinations()
# Define fee percentages for different payment methods
PAYMENT_FEES = {
    "ðŸ’³ Paypal": 10.0,  # 10% fee
}
SERIAL_NUMBERS = load_serial_numbers()


### Different types of messages for the bot ###

payment_review_messages = [
    "âœ… *Verification in progress*\n\nYour payment confirmation is under review. Our admin team will verify it shortly and notify you once complete.",
    
    "âœ… *Thank you for your submission*\n\nWe've received your payment proof and it's being reviewed by our team. You'll be notified as soon as it's verified.",
    
    "âœ… *Processing payment verification*\n\nYour screenshot has been sent to our admin team for review. We'll update you when the verification is complete.",
    
    "âœ… *Payment proof received*\n\nThank you for submitting your payment information. Our team is reviewing it and will notify you once verified.",
    
    "âœ… *Verification pending*\n\nWe've received your payment details and they are currently under review. You'll receive a notification once the process is complete."
]

payment_approval_messages = [
    "âœ… *Verification Successful!*\n\nWelcome to Prodigy Trading Academy. We're delighted to have you as part of our community!",
    
    "âœ… *Great news!* Your payment has been verified successfully! Welcome to the Prodigy Trading Academy family. We're thrilled to have you join us!",
    
    "âœ… *Payment verified!*\n\nYour membership has been activated. Welcome to Prodigy Trading Academy! We're excited to have you as part of our trading community.",
    
    "âœ… *You're all set!*\n\nYour payment has been verified and your membership is now active. Welcome aboard the Prodigy Trading Academy!",
    
    "âœ… *Payment confirmed!*\n\nThank you for joining Prodigy Trading Academy. Your membership has been successfully activated and we're looking forward to helping you on your trading journey!"
]

payment_rejection_messages = [
    "âŒ *Verification Failed*\n\nUnfortunately, we couldn't verify your payment. Please check your payment details and try again, or contact our admin team for assistance.",
    
    "âŒ *Payment Verification Issue*\n\nWe were unable to confirm your payment. Please ensure you've sent the correct amount and try submitting your proof again.",
    
    "âŒ *Payment Not Verified*\n\nThere seems to be an issue with your payment verification. Please submit a clearer screenshot or contact our admin team for help.",
    
    "âŒ *Verification Unsuccessful*\n\nYour payment proof couldn't be verified at this time. Please check the payment details and try again with a clearer screenshot.",
    
    "âŒ *Payment Rejected*\n\nWe couldn't process your payment verification. Please ensure you've completed the payment correctly and submit a new verification request."
]

pending_payment_messages = [
    "âš ï¸ You have a pending payment verification. Admins are reviewing your payment proof. Please wait for their response.",
    "ðŸ’¼ Your payment is currently being verified by our admin team. We'll notify you once the process is complete.",
    "ðŸ“Š Thanks for your patience! Your payment proof is still under review by our admins. You'll receive a notification when verified.",
    "â±ï¸ Our team is reviewing your payment submission. We'll let you know as soon as it's verified.",
    "ðŸ’³ Your payment verification is in progress. Our admin team is reviewing your submission and will notify you shortly."
]

### COMMAND HANDLERS ###

@bot.message_handler(commands=['chat', 'ask', 'ai'])
def handle_ai_chat(message):
    """Handle AI chat requests using QWEN model"""
    user_id = message.from_user.id
    chat_id = message.chat.id
    
    # Check if in private chat
    if message.chat.type != 'private':
        bot.reply_to(message, "ðŸ“ Please use this command in a private chat with the bot for better assistance.")
        return
    
    # Initialize conversation history with system message
    if user_id not in QWEN_USAGE:
        QWEN_USAGE[user_id] = {'last_used': time.time(), 'count': 0}
    
    # Reset messages to only include the system message at the start of a new conversation
    QWEN_USAGE[user_id]['messages'] = [{
        "role": "system",
        "content": QWEN_PROMPT_TEMPLATE
    }]
    
    # Extract query from message
    query_text = message.text.split(' ', 1)
    if len(query_text) > 1:
        # Query included in command
        query_text = query_text[1].strip()
        process_ai_query(chat_id, user_id, query_text)
    else:
        # No query provided, welcome them to chat mode
        bot.send_message(
            chat_id, 
            "ðŸ¤– *AI Assistant Mode Activated*\n\n"
            "I'm here to help with your trading questions! Just type your question and I'll respond.\n\n"
            "When you're done, simply type `/exit` to end our conversation.",
            parse_mode="Markdown"
        )
        # Set user status to conversation_active (new status)
        PENDING_USERS[user_id] = {'status': 'conversation_active'}
        save_pending_users()

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'awaiting_ai_query')
def handle_ai_query_input(message):
    """Handle follow-up input for AI queries - this will be used less with the new conversational flow"""
    # Check if in private chat
    if message.chat.type != 'private':
        return
    
    user_id = message.from_user.id
    chat_id = message.chat.id
    query_text = message.text
    
    # Clear the awaiting status and set to conversation mode
    PENDING_USERS[user_id]['status'] = 'conversation_active'
    save_pending_users()
    
    # Process the query
    process_ai_query(chat_id, user_id, query_text)

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'conversation_active')
def handle_conversation(message):
    """Handle ongoing conversation with the AI"""
    # Check if in private chat
    if message.chat.type != 'private':
        return
    
    user_id = message.from_user.id
    chat_id = message.chat.id
    
    # Check for exit command
    if message.text.lower().strip() == "/exit":
        PENDING_USERS[user_id]['status'] = 'at_main_menu'  # Reset to main menu
        save_pending_users()
        
        # Clear conversation history
        if user_id in QWEN_USAGE and 'messages' in QWEN_USAGE[user_id]:
            # Keep only the system message
            if QWEN_USAGE[user_id]['messages'] and QWEN_USAGE[user_id]['messages'][0]['role'] == 'system':
                QWEN_USAGE[user_id]['messages'] = [QWEN_USAGE[user_id]['messages'][0]]
            else:
                QWEN_USAGE[user_id]['messages'] = []
            logging.info(f"Cleared conversation history for user {user_id}")
        
        bot.send_message(
            chat_id,
            "ðŸ‘‹ *AI Assistant Mode Deactivated*\n\nThank you for chatting! If you need my help again, just use `/ask`, `/chat`, or `/ai`.",
            parse_mode="Markdown"
        )
        # Add this line to show the main menu after exiting AI chat
        show_main_menu(chat_id, user_id)
        return
    
    # Otherwise, process as a query
    query_text = message.text
    process_ai_query(chat_id, user_id, query_text)

def process_ai_query(chat_id, user_id, query_text):
    """Process an AI query through the QWEN model"""
    # Check rate limiting
    now = time.time()  # Use time.time() instead of just time()
    
    # First check MongoDB for the most up-to-date usage data
    try:
        db_usage = jarvis_usage_collection.find_one({"user_id": user_id})
        if db_usage:
            # Ensure QWEN_USAGE has this user with the latest data
            if user_id not in QWEN_USAGE:
                QWEN_USAGE[user_id] = {
                    'last_used': db_usage.get('last_used', now),
                    'count': db_usage.get('count', 0),
                    'messages': QWEN_USAGE.get(user_id, {}).get('messages', [])
                }
            else:
                # Update count from DB if it's different (might have been reset by thread)
                QWEN_USAGE[user_id]['count'] = db_usage.get('count', QWEN_USAGE[user_id].get('count', 0))
    except Exception as e:
        logging.error(f"Error syncing rate limit data from MongoDB: {e}")
    
    # Now check local rate limiting data
    if user_id in QWEN_USAGE:
        last_used = QWEN_USAGE[user_id]['last_used']
        count = QWEN_USAGE[user_id]['count']
        
        # Check if an hour has passed since last use
        if now - last_used >= 3600:
            # Reset count to 0 if an hour or more has passed
            QWEN_USAGE[user_id] = {
                'last_used': now, 
                'count': 0,
                'messages': QWEN_USAGE[user_id].get('messages', [])
            }
            count = 0
        # Then implement rate limiting (5 queries per hour)
        elif count >= 5:  # Only block if within the same hour AND count exceeded
            # Calculate time remaining until reset
            time_remaining = 3600 - (now - last_used)
            minutes = int(time_remaining // 60)
            seconds = int(time_remaining % 60)
            
            bot.send_message(chat_id, 
                f"â±ï¸ *AI Query Limit Reached*\n\n"
                f"You've used all 5 of your AI queries for this hour.\n\n"
                f"â° Limit resets in: {minutes} min {seconds} sec\n"
                f"ðŸ”„ Please try again after the cooldown period.",
                parse_mode="Markdown"
            )
            return
    else:
        QWEN_USAGE[user_id] = {'last_used': now, 'count': 0}
    
    # Update usage tracking
    QWEN_USAGE[user_id]['last_used'] = now
    QWEN_USAGE[user_id]['count'] += 1
    
    # Log usage to MongoDB
    try:
        jarvis_usage_collection.update_one(
            {"user_id": user_id},
            {"$inc": {"count": 1}, "$set": {"last_used": now}},
            upsert=True
        )
    except Exception as e:
        logging.error(f"Error updating AI usage stats: {e}")
    
    # Show typing indicator
    bot.send_chat_action(chat_id, "typing")
    
    try:
        # Pass chat_id when calling call_qwen_api to support streaming
        response = call_qwen_api(query_text, user_id=user_id, chat_id=chat_id)
        
        # Send the response - HANDLE MARKDOWN SAFELY
        try:
            bot.send_message(chat_id, response, parse_mode="Markdown")
        except ApiException as api_e:
            # If Markdown parsing fails, try sending without parse mode
            logging.warning(f"Markdown parsing failed, sending without formatting: {api_e}")
            bot.send_message(chat_id, response)
        
    except Exception as e:
        logging.error(f"Error in AI query: {e}")
        bot.send_message(chat_id, "I'm sorry, I couldn't process that request right now. Please try again later.")

def call_qwen_api(prompt, user_id=None, chat_id=None):
    """Make an API call to the QWEN model with conversation memory"""
    if not QWEN_API_KEY:
        return "Sorry, AI chat is currently unavailable. Please contact an admin."
    
    try:
        # Initialize conversation history for this user if it doesn't exist
        if user_id not in QWEN_USAGE:
            QWEN_USAGE[user_id] = {'last_used': time.time(), 'count': 0, 'messages': []}
        
        # Check if user has hit rate limit (assuming max of 5 queries per hour)
        if QWEN_USAGE[user_id]['count'] >= 5:
            # User is rate limited - only start a thread if we haven't already scheduled one
            if not QWEN_USAGE[user_id].get('reset_scheduled', False):
                QWEN_USAGE[user_id]['reset_scheduled'] = True
                # Start a thread that will reset this specific user's count after 1 hour
                threading.Thread(
                    target=reset_user_rate_limit,
                    args=(user_id,),
                    daemon=True
                ).start()
                logging.info(f"Rate limit hit for user {user_id}, scheduled reset in 1 hour")
            
            # Calculate time remaining until reset
            last_used = QWEN_USAGE[user_id]['last_used']
            time_remaining = 3600 - (time.time() - last_used)
            minutes = int(time_remaining // 60)
            seconds = int(time_remaining % 60)
            
            return f"â±ï¸ *AI Query Limit Reached*\n\nYou've used all 5 of your AI queries for this hour.\n\nâ° Limit resets in: {minutes} min {seconds} sec\nðŸ”„ Please try again after the cooldown period."
        
        # Add database context to system prompt if needed for specific queries
        system_prompt = QWEN_PROMPT_TEMPLATE
        
        if user_id is not None and any(keyword in prompt.lower() for keyword in 
                                   ['membership', 'payment', 'status', 'expire', 'due date', 
                                    'my account', 'subscription', 'points', 'leaderboard',
                                    'supreme', 'regular', 'days', 'remaining', 'left', 
                                    'time left', 'how long', 'when does', 'expiration',
                                    'expiry', 'expire', 'renew', 'renewal', 'access']):
            user_data = get_user_database_context(user_id)
            if user_data:
                # Append user database context to system prompt instead of user message
                system_prompt = f"""{QWEN_PROMPT_TEMPLATE}

USER DATABASE CONTEXT:
{user_data}

Important: When answering about membership time remaining, due dates, or status, ALWAYS provide the SPECIFIC numerical days and dates from this data rather than giving general instructions. Give direct answers with the exact numbers.

Please use this information to provide a personalized response, but don't explicitly mention that you looked up their database record unless they specifically asked about their account.

"""
                logging.info(f"Added database context to system prompt for user {user_id}")
        
        # Start with the system message containing our prompt template
        if 'messages' not in QWEN_USAGE[user_id] or not QWEN_USAGE[user_id]['messages']:
            QWEN_USAGE[user_id]['messages'] = [
                {
                    "role": "system", 
                    "content": system_prompt
                }
            ]
        else:
            # Update the system prompt with potential user context
            QWEN_USAGE[user_id]['messages'][0]['content'] = system_prompt
        
        # Add the current user query to the messages (clean, without additional context)
        QWEN_USAGE[user_id]['messages'].append({"role": "user", "content": prompt})
        
        # Keep only the last 10 messages to avoid hitting token limits
        if len(QWEN_USAGE[user_id]['messages']) > 15:
            # Always keep the system message (first one) and trim the middle
            system_message = QWEN_USAGE[user_id]['messages'][0]
            recent_messages = QWEN_USAGE[user_id]['messages'][-14:]
            QWEN_USAGE[user_id]['messages'] = [system_message] + recent_messages
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {QWEN_API_KEY}"
        }
        
        payload = {
            "model": "qwen-max",
            "input": {
                "messages": QWEN_USAGE[user_id]['messages']
            },
            "parameters": {
                "temperature": 0.7,
                "top_p": 0.8,
                "max_tokens": 512,
                "result_format": "message"
            }
        }

        response = requests.post(QWEN_API_URL, headers=headers, data=json.dumps(payload), timeout=30)
        response.raise_for_status()
        
        response_json = response.json()
        # Fix: Using the correct path to get the response content
        ai_response = response_json['output']['choices'][0]['message']['content']
        
        # Update usage count and timestamp
        QWEN_USAGE[user_id]['count'] += 1
        QWEN_USAGE[user_id]['last_used'] = time.time()
        
        # Save the assistant's response to the conversation history
        QWEN_USAGE[user_id]['messages'].append({"role": "assistant", "content": ai_response})
        
        return ai_response
        
    except requests.exceptions.RequestException as e:
        logging.error(f"API request failed: {e}")
        raise Exception(f"API request failed: {e}")
    except (KeyError, IndexError) as e:
        logging.error(f"Error parsing API response: {e}")
        raise Exception(f"Error parsing API response: {e}")
    except Exception as e:
        logging.error(f"Unknown error in QWEN API call: {e}")
        raise Exception(f"Unknown error in QWEN API call: {e}")
    
def reset_user_rate_limit(user_id):
    """Reset the rate limit for a specific user after waiting for 1 hour"""
    try:
        # Wait for 1 hour
        time.sleep(3600)
        
        # Reset the user's count and the scheduled flag
        if user_id in QWEN_USAGE:
            QWEN_USAGE[user_id]['count'] = 0
            QWEN_USAGE[user_id]['reset_scheduled'] = False
            
            # Update MongoDB too
            jarvis_usage_collection.update_one(
                {"user_id": user_id},
                {"$set": {"count": 0}},
                upsert=True
            )
            
            logging.info(f"Reset rate limit for user {user_id}")
    except Exception as e:
        logging.error(f"Error in reset_user_rate_limit: {e}")
    
def get_user_database_context(user_id):
    """Get relevant user data from database to provide context for AI responses"""
    user_id_str = str(user_id)
    context = []
    
    try:
        # Get membership information
        if user_id_str in PAYMENT_DATA:
            data = PAYMENT_DATA[user_id_str]
            
            # Basic membership info
            status = "Active" if data.get('haspayed', False) else "Inactive"
            membership_type = data.get('mentorship_type', 'Unknown')
            plan = data.get('payment_plan', 'Unknown')
            
            # Calculate time remaining if applicable
            time_remaining = ""
            if 'due_date' in data:
                try:
                    due_date = datetime.strptime(data['due_date'], '%Y-%m-%d %H:%M:%S')
                    current_date = datetime.now()
                    days_remaining = (due_date - current_date).days
                    
                    if days_remaining > 0:
                        time_remaining = f"{days_remaining} days remaining"
                    else:
                        time_remaining = "Expired"
                except Exception as e:
                    time_remaining = "Error calculating time remaining"
            
            # Add cancellation info if applicable
            cancelled = "Yes" if data.get('cancelled', False) else "No"
            
            context.append(f"Membership Status: {status}")
            context.append(f"Membership Type: {membership_type}")
            context.append(f"Plan: {plan}")
            context.append(f"Time Remaining: {time_remaining}")
            context.append(f"Cancelled: {cancelled}")
            
            # Get points/leaderboard data if available
            try:
                # Current month points
                current_month = datetime.now().strftime('%Y-%m')
                monthly_scores = get_monthly_leaderboard(current_month)
                user_points = 0
                user_rank = "N/A"
                
                for i, entry in enumerate(monthly_scores):
                    if entry.get('user_id') == int(user_id):
                        user_points = entry.get('total_points', 0)
                        user_rank = i + 1
                        break
                        
                context.append(f"Current Month Points: {user_points}")
                context.append(f"Current Leaderboard Rank: {user_rank}")
            except Exception:
                # If points retrieval fails, just continue
                pass
                
            # Get form answers if available
            if 'form_answers' in data:
                form_data = data['form_answers']
                if form_data:
                    context.append("\nUser Profile Information:")
                    if 'full_name' in form_data:
                        context.append(f"Full Name: {form_data['full_name']}")
                    if 'expertise_level' in form_data:
                        context.append(f"Trading Expertise: {form_data['expertise_level']}")
                    if 'learning_goals' in form_data:
                        context.append(f"Learning Goals: {form_data['learning_goals']}")
            
            return "\n".join(context)
        else:
            return "No membership data found for this user."
            
    except Exception as e:
        logging.error(f"Error getting user database context: {e}")
        return f"Error retrieving user data: {str(e)}"

@bot.message_handler(commands=['resetaiusage'])
def reset_ai_usage(message):
    """Reset a user's AI usage limits - admin only"""
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    args = message.text.split()
    if len(args) < 2:
        bot.reply_to(message, "âŒ Usage: `/resetaiusage [user_id]`", parse_mode="Markdown")
        return
    
    try:
        target_user_id = int(args[1])
        
        # Reset in memory
        if target_user_id in QWEN_USAGE:
            QWEN_USAGE[target_user_id]['count'] = 0
        
        # Reset in MongoDB
        jarvis_usage_collection.update_one(
            {"user_id": target_user_id},
            {"$set": {"count": 0}},
            upsert=True
        )
        
        bot.reply_to(message, f"âœ… AI usage limit reset for user {target_user_id}")
    except ValueError:
        bot.reply_to(message, "âŒ Invalid user ID. Please provide a numeric user ID.")
    except Exception as e:
        bot.reply_to(message, f"âŒ Error: {str(e)}")
        logging.error(f"Error resetting AI usage: {e}")

@bot.message_handler(commands=['aiusagestats'])
def ai_usage_stats(message):
    """Get AI usage statistics - admin only"""
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    try:
        stats = list(jarvis_usage_collection.find().sort("count", -1).limit(10))
        if not stats:
            bot.reply_to(message, "No AI usage data available.")
            return
            
        response = "ðŸ¤– *AI Usage Statistics*\n\n*Top 10 Users:*\n\n"
        
        for i, user_stat in enumerate(stats, 1):
            user_id = user_stat.get("user_id", "Unknown")
            count = user_stat.get("count", 0)
            last_used = user_stat.get("last_used", 0)
            
            try:
                # Try to get username
                user_info = bot.get_chat(user_id)
                username = user_info.username or f"User {user_id}"
            except:
                username = f"User {user_id}"
                
            # Format last used time
            if last_used:
                last_used_str = datetime.fromtimestamp(last_used).strftime('%Y-%m-%d %H:%M:%S')
            else:
                last_used_str = "Never"
                
            response += f"{i}. @{username}: {count} queries (Last: {last_used_str})\n"
            
        bot.reply_to(message, response, parse_mode="Markdown")
    except Exception as e:
        bot.reply_to(message, f"âŒ Error getting usage stats: {str(e)}")
        logging.error(f"Error getting AI usage stats: {e}")

@bot.message_handler(commands=['dm'])
def handle_dm_command(message):
    if message.chat.type == 'private':
        bot.send_message(message.chat.id, "âŒ This command can only be used in a channel.")
        return

    user_id = message.from_user.id
    username = message.from_user.username or "No Username"

    # Check if the bot can send a message to the user
    try:
        bot.send_chat_action(user_id, 'typing')  # Check if the user exists and can receive messages
        bot.send_message(user_id, f"Hello @{username}! Please say /start to begin.")
        bot.send_message(message.chat.id, "Direct Message sent, please check your inbox.")
    except ApiException as e:
        bot.send_message(message.chat.id, f"âŒ Failed to send DM: {e.result_json['description']}. Please start a conversation with the bot first by sending /start in a private chat.")

# Add this function to delete a specific user from pending
def delete_pending_user(user_id):
    try:
        result = pending_collection.delete_one({'_id': str(user_id)})
        if result.deleted_count > 0:
            logging.info(f"Deleted pending user {user_id} from MongoDB")
        else:
            logging.info(f"No pending user {user_id} found to delete in MongoDB")
    except Exception as e:
        logging.error(f"Error deleting pending user {user_id} from MongoDB: {e}")

# Start Command - Sends intro message and asks for the payment plan
@bot.message_handler(commands=['start'])
def send_welcome(message):
    if message.chat.type != 'private':
        bot.send_message(message.chat.id, "Please DM the bot to get started.")
        return  # Ignore if not in private chat
    
    chat_id = message.chat.id
    user_id = message.from_user.id

    # Reload pending users from MongoDB to ensure we have latest data
    global PENDING_USERS
    PENDING_USERS = load_pending_users()

    # Check for pending admin actions
    pending_verification = False
    pending_payment = False
    
    if user_id in PENDING_USERS:
        if PENDING_USERS[user_id].get('status') == 'old_member_request':
            pending_verification = True
            logging.info(f"User {user_id} has pending verification request")
        elif PENDING_USERS[user_id].get('status') == 'waiting_approval':
            pending_payment = True
            logging.info(f"User {user_id} has pending payment verification")
        elif PENDING_USERS[user_id].get('status') in ['rejected', 'payment_rejected']:
            # If status is rejected or payment_rejected, reset their status to allow choosing again
            PENDING_USERS.pop(user_id, None)  # Remove from dictionary
            delete_pending_user(user_id)  # Remove from MongoDB
            logging.info(f"User {user_id} with rejected status has been removed from pending users")
    
    # Handle pending requests first - don't show intro message again
    if pending_payment:
        bot.send_message(chat_id, random.choice(pending_payment_messages))
        return  # Exit the function here - don't show the intro message again
    
    # Check if user has already accepted the terms and privacy policy
    has_accepted_terms = False
    user_id_str = str(user_id)
    
    if user_id_str in PAYMENT_DATA:
        has_accepted_terms = PAYMENT_DATA[user_id_str].get('terms_accepted', False) and PAYMENT_DATA[user_id_str].get('privacy_accepted', False)
    
    # If user has already accepted terms, show welcome message and main menu
    if has_accepted_terms:
        logging.info(f"User {user_id} has previously accepted terms - showing welcome message and main menu")
        send_welcome_message(chat_id)
        show_main_menu(chat_id, user_id)
    else:
        # If user hasn't accepted terms, show the legal notice first
        send_simplified_legal_notice(chat_id, user_id)

def send_welcome_message(chat_id):
    """Send the welcome image and text to the user"""
    # Send welcome image first
    try:
        with open('graphics/start.jpeg', 'rb') as welcome_img:
            bot.send_photo(
                chat_id, 
                welcome_img,
                caption=f"ðŸ« *Prodigy Trading Academy Bot {BOT_VERSION}*\n\n"
                "ðŸŽ‰ Welcome to Prodigy Trading Academy!\n\n"
                "You're one step closer to leveling up your trading journey. We're excited to have you on board â€” let's make progress, not just promises. ðŸš€\n\n"
                "ðŸ“¢ *Note:* This bot is currently in *Beta*, so you may experience occasional updates or improvements.",
                parse_mode="Markdown"
            )
    except FileNotFoundError:
        # Fallback to text-only message if image not found
        logging.error("Welcome image not found at graphics/start.jpeg")
        bot.send_message(chat_id, 
            f"ðŸ« *Prodigy Trading Academy Bot {BOT_VERSION}*\n\n"
            "ðŸŽ‰ Welcome to Prodigy Trading Academy!\n\n"
            "You're one step closer to leveling up your trading journey. We're excited to have you on board â€” let's make progress, not just promises. ðŸš€\n\n"
            "ðŸ“¢ *Note:* This bot is currently in *Beta*, so you may experience occasional updates or improvements.",
            parse_mode="Markdown"
        )
    except Exception as e:
        # Log any other errors but still show welcome message
        logging.error(f"Error sending welcome image: {e}")
        bot.send_message(chat_id, 
            f"ðŸ« *Prodigy Trading Academy Bot {BOT_VERSION}*\n\n"
            "ðŸŽ‰ Welcome to Prodigy Trading Academy!\n\n"
            "You're one step closer to leveling up your trading journey. We're excited to have you on board â€” let's make progress, not just promises. ðŸš€\n\n"
            "ðŸ“¢ *Note:* This bot is currently in *Alpha*, so you may experience occasional updates or improvements.",
            parse_mode="Markdown"
        )

def send_simplified_legal_notice(chat_id, user_id):
    """Send a simplified legal notice instead of the full documents"""
    
    legal_notice = (
        "ðŸ“œ *LEGAL NOTICE*\n\n"
        "Before proceeding, please be aware that by using Prodigy Trading Academy's services:\n\n"
        "â€¢ You agree to our *Terms & Conditions* which cover membership, content usage, and risks associated with trading\n\n"
        "â€¢ You acknowledge our *Privacy Policy* regarding how we collect, store, and use your personal data\n\n"
        "These legal documents are important and can be reviewed at any time via the FAQ section in the main menu (see: Terms & Conditions / Privacy Policy) or by contacting our support team.\n\n"
        "By continuing, you confirm that you have read, understood, and agreed to these policies."
    )
    
    # Create acceptance buttons
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(InlineKeyboardButton("âœ… I Agree to Terms & Privacy Policy", callback_data="accept_legal_notice"))
    markup.add(InlineKeyboardButton("âŒ I Decline", callback_data="decline_legal_notice"))
    
    # Send the notice
    bot.send_message(
        chat_id,
        legal_notice,
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    # Record user state
    PENDING_USERS[user_id] = {'status': 'awaiting_legal_acceptance'}
    save_pending_users()

# Update the callback handler for the new legal notice acceptance
@bot.callback_query_handler(func=lambda call: call.data in ["accept_legal_notice", "decline_legal_notice"])
def handle_legal_notice_response(call):
    chat_id = call.message.chat.id
    user_id = call.from_user.id
    
    if call.data == "accept_legal_notice":
        # User accepted both terms and privacy policy at once
        if str(user_id) in PAYMENT_DATA:
            PAYMENT_DATA[str(user_id)]["terms_accepted"] = True
            PAYMENT_DATA[str(user_id)]["terms_accepted_date"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            PAYMENT_DATA[str(user_id)]["privacy_accepted"] = True
            PAYMENT_DATA[str(user_id)]["privacy_accepted_date"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        else:
            # Create a new record if user doesn't exist in PAYMENT_DATA
            PAYMENT_DATA[str(user_id)] = {
                "username": call.from_user.username or "No Username",
                "terms_accepted": True,
                "terms_accepted_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "privacy_accepted": True,
                "privacy_accepted_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "haspayed": False
            }
        
        save_payment_data()
        
        bot.answer_callback_query(call.id, "Thank you! You've accepted our Terms & Privacy Policy.")
        
        # Delete the legal notice message
        try:
            bot.delete_message(chat_id, call.message.message_id)
        except Exception as e:
            logging.error(f"Error deleting legal notice message: {e}")
        
        # Send welcome message now that terms are accepted
        send_welcome_message(chat_id)
        
        # Add a small delay for better UX
        time.sleep(1.5)
        
        # Show the main menu
        show_main_menu(chat_id, user_id)
        
    else:  # User declined
        bot.answer_callback_query(call.id, "You must accept our Terms & Privacy Policy to use this service.")
        bot.send_message(chat_id, 
            "âŒ *Access Denied*\n\n"
            "To use Prodigy Trading Academy services, you must accept our Terms & Conditions and Privacy Policy.\n\n"
            "If you change your mind, please restart the bot with /start.",
            parse_mode="Markdown"
        )
        
        # Remove from pending users
        if user_id in PENDING_USERS:
            PENDING_USERS.pop(user_id, None)
            delete_pending_user(user_id)

def show_main_menu(chat_id, user_id, message_id=None):
    """Show the main menu with inline keyboard options based on user status"""
    # Prepare welcome message
    welcome_message = (
        "Need help? We're here to guide you every step of the way. ðŸ’¬\n"
        "Please select an option below to proceed:"
    )
    
    # Check if the user is an existing member with active membership
    is_existing_member = False
    needs_deposit = False
    needs_forms = False
    
    if str(user_id) in PAYMENT_DATA:
        is_existing_member = PAYMENT_DATA[str(user_id)].get('haspayed', False)
        needs_deposit = PAYMENT_DATA[str(user_id)].get('needs_deposit', False)
        needs_forms = PAYMENT_DATA[str(user_id)].get('forms_needed', False)
    
    # Create appropriate inline keyboard based on user status
    markup = InlineKeyboardMarkup(row_width=2)
    
    if is_existing_member:
        if needs_forms:
            # Special case: User needs to complete registration forms
            markup.add(
                InlineKeyboardButton("ðŸ“ Finish Forms", callback_data="menu_finish_forms")
            )
            markup.add(
                InlineKeyboardButton("â“ FAQ", callback_data="menu_faq")
            )
            # Add AI Chat option
            markup.add(
                InlineKeyboardButton("ðŸ¤– AI Chat", callback_data="menu_aichat")
            )
        elif needs_deposit:
            # Options for existing members who need to deposit
            markup.add(
                InlineKeyboardButton("ðŸ’° Deposit Now", callback_data="menu_xm_deposit")
            )
            markup.add(
                InlineKeyboardButton("â“ FAQ", callback_data="menu_faq")
            )
            # Add AI Chat option
            markup.add(
                InlineKeyboardButton("ðŸ¤– AI Chat", callback_data="menu_aichat")
            )
        else:
            # Options for existing members
            markup.add(
                InlineKeyboardButton("âš™ï¸ Manage Membership", callback_data="menu_manage"),
                InlineKeyboardButton("âŒ Cancel Membership", callback_data="menu_cancel")
            )
            markup.add(
                InlineKeyboardButton("ðŸ“Š My Dashboard", callback_data="menu_dashboard"),
                InlineKeyboardButton("â“ FAQ", callback_data="menu_faq")
            )
            # Add AI Chat option
            markup.add(
                InlineKeyboardButton("ðŸ¤– AI Chat", callback_data="menu_aichat")
            )
    else:
        # Options for new users
        markup.add(
            InlineKeyboardButton("ðŸ“… Avail Membership", callback_data="menu_buy"),
            InlineKeyboardButton("ðŸ”‘ Redeem Serial", callback_data="menu_redeem")
        )
        markup.add(
            InlineKeyboardButton("ðŸ“– How to Use", callback_data="menu_howto"),
            InlineKeyboardButton("â“ FAQ", callback_data="menu_faq")
        )
        # Add AI Chat option
        markup.add(
            InlineKeyboardButton("ðŸ¤– AI Chat", callback_data="menu_aichat")
        )
    
    # Send menu with inline buttons or edit existing message
    if message_id:
        # Edit the existing message
        bot.edit_message_text(
            welcome_message,
            chat_id,
            message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
        sent_message = None
    else:
        # Send a new message with welcome message and markup combined
        sent_message = bot.send_message(chat_id, welcome_message, parse_mode="Markdown", reply_markup=markup)
    
    # Store user state
    PENDING_USERS[user_id] = {'status': 'at_main_menu'}
    save_pending_users()

    # Check if the user has unseen changelogs - ONLY SHOW THE MOST RECENT ONE
    if str(user_id) in PAYMENT_DATA and PAYMENT_DATA[str(user_id)]['haspayed']:
        # Initialize tracking variable
        update_shown = False
        
        # Look for the most recent changelog they haven't seen
        for i, changelog in enumerate(reversed(CHANGELOGS["user"])):
            if not update_shown and str(user_id) not in changelog.get("seen_by", []):
                # Show the unseen changelog (only the most recent one)
                bot.send_message(
                    chat_id,
                    f"ðŸ“¢ *UNREAD UPDATE*\n\n{changelog['content']}\n\nðŸ•’ Posted: {changelog['timestamp']}",
                    parse_mode="Markdown"
                )
                # Mark as seen
                if "seen_by" not in changelog:
                    changelog["seen_by"] = []
                    
                changelog["seen_by"].append(str(user_id))
                save_changelogs(CHANGELOGS)
                
                # Set flag to prevent showing more updates
                update_shown = True
                logging.info(f"Showed unread changelog to user {user_id}")
                break
    
    # Return the sent message object if we sent a new message
    return sent_message

@bot.callback_query_handler(func=lambda call: call.data == "menu_manage")
def handle_manage_membership(call):
    """Handle user clicking on 'Manage Membership' button"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    
    # Answer callback
    bot.answer_callback_query(call.id, "Opening membership management options...")
    
    # Create inline keyboard with membership management options
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(
        InlineKeyboardButton("ðŸ”„ Renew Current Plan", callback_data="menu_renew"),
        InlineKeyboardButton("â¬†ï¸ Upgrade Membership", callback_data="menu_upgrade")
    )
    markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
    
    # Edit the message to show management options
    bot.edit_message_text(
        "âš™ï¸ *Membership Management*\n\n"
        "Choose what you'd like to do with your membership:",
        chat_id,
        message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.callback_query_handler(func=lambda call: call.data == "menu_upgrade")
def handle_upgrade_membership(call):
    """Handle user choosing to upgrade their membership"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    user_id_str = str(user_id)
    
    bot.answer_callback_query(call.id, "Loading upgrade options...")
    
    # Check if user has membership data
    if user_id_str not in PAYMENT_DATA:
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
        
        bot.edit_message_text(
            "âŒ You don't have an active membership to upgrade. Please purchase a membership first.",
            chat_id,
            message_id,
            reply_markup=markup
        )
        return
    
    # Get current membership type
    current_type = PAYMENT_DATA[user_id_str].get('mentorship_type', '').lower()
    
    # If already on Supreme, no upgrade available
    if current_type == 'supreme':
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(InlineKeyboardButton("â¬…ï¸ Back", callback_data="menu_manage"))
        
        bot.edit_message_text(
            "â„¹ï¸ You already have our premium Supreme Membership. There are no higher tiers available.",
            chat_id,
            message_id,
            reply_markup=markup
        )
        return
    
    # For Regular members, offer upgrade to Supreme
    # Mark as upgrade in pending users (similar logic to renewal)
    PENDING_USERS[chat_id] = {
        'status': 'choosing_mentorship_type',
        'is_renewal': False,
        'is_upgrade': True  # New flag to indicate this is an upgrade
    }
    save_pending_users()
    
    # Check enrollment status for Supreme membership
    if not BOT_SETTINGS.get('supreme_enrollment_open', True):
        # Create inline keyboard with Get Notified and FAQ buttons
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineKeyboardButton("ðŸ”” Get Notified", callback_data="update_yes"),
            InlineKeyboardButton("â“ FAQ", callback_data="faq_back")
        )
        markup.add(InlineKeyboardButton("â¬…ï¸ Back", callback_data="menu_manage"))
        
        bot.edit_message_text(
            "âš ï¸ *Supreme Membership enrollment is currently CLOSED*\n\n"
            "Supreme Membership upgrades are temporarily unavailable.\n"
            "Please wait for the next announcement about when enrollment will open again.\n\n"
            "â€¢ Click *Get Notified* to receive updates when enrollment opens\n"
            "â€¢ Check our *FAQ* section for more information",
            chat_id,
            message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
        return
    
    # Get supreme discount if applicable
    applicable_discount = DISCOUNTS.get('supreme')
    discount_active = applicable_discount and applicable_discount.get('active', False)
    
    # Check if discount applies to upgrades
    if discount_active:
        discount_transaction_type = applicable_discount.get('transaction_type', 'both')
        
        # For upgrades, we'll consider them as "new" transactions for the supreme tier
        applies_to_transaction = (
            discount_transaction_type == 'both' or 
            discount_transaction_type == 'new'
        )
        
        # Only consider discount active if it applies to upgrades
        if not applies_to_transaction:
            discount_active = False
            
    discount_percentage = applicable_discount.get('percentage', 0) if discount_active else 0
    discount_name = applicable_discount.get('name', '') if discount_active else ''
    
    # Original prices
    apprentice_price = 309.99
    disciple_price = 524.99
    legacy_price = 899.99
    
    # Create markup for plan selection
    markup = InlineKeyboardMarkup(row_width=1)
    
    if discount_active:
        # Calculate discounted prices
        apprentice_discounted = round(apprentice_price * (1 - discount_percentage / 100), 2)
        disciple_discounted = round(disciple_price * (1 - discount_percentage / 100), 2)
        legacy_discounted = round(legacy_price * (1 - discount_percentage / 100), 2)
        
        # Add buttons with discounted prices
        markup.add(
            InlineKeyboardButton(f"Apprentice (${apprentice_discounted:.2f}) / 3 Months", callback_data="plan_apprentice"),
            InlineKeyboardButton(f"Disciple (${disciple_discounted:.2f}) / 6 Months", callback_data="plan_disciple"),
            InlineKeyboardButton(f"Legacy (${legacy_discounted:.2f}) / Lifetime", callback_data="plan_supreme_legacy")
        )
        
        # Add back button
        markup.add(InlineKeyboardButton("â¬…ï¸ Go Back", callback_data="menu_manage"))
        
        # FIXED: Use edit_message_text instead of send_message
        bot.edit_message_text(
            f"â¬†ï¸ *Upgrade to Supreme Membership*\n\n"
            f"Choose your Supreme plan:\n\n"
            f"ðŸŽ‰ <b>{discount_name}: {discount_percentage}% OFF!</b>\n\n"
            f"ðŸ’° <b>Apprentice</b> - <s>${apprentice_price:.2f}</s> ${apprentice_discounted:.2f} / 3 Months\n"
            f"ðŸ’° <b>Disciple</b> - <s>${disciple_price:.2f}</s> ${disciple_discounted:.2f} / 6 Months\n"
            f"ðŸ’° <b>Legacy</b> - <s>${legacy_price:.2f}</s> ${legacy_discounted:.2f} / Lifetime", 
            chat_id,
            message_id,
            reply_markup=markup, 
            parse_mode="HTML"
        )
    else:
        # No discount - show regular prices without strikethrough
        markup.add(
            InlineKeyboardButton(f"Apprentice (${apprentice_price:.2f}) / 3 Months", callback_data="plan_apprentice"),
            InlineKeyboardButton(f"Disciple (${disciple_price:.2f}) / 6 Months", callback_data="plan_disciple"),
            InlineKeyboardButton(f"Legacy (${legacy_price:.2f}) / Lifetime", callback_data="plan_supreme_legacy")
        )
        
        # Add back button
        markup.add(InlineKeyboardButton("â¬…ï¸ Go Back", callback_data="menu_manage"))
        
        # FIXED: Use edit_message_text instead of send_message
        bot.edit_message_text(
            f"â¬†ï¸ *Upgrade to Supreme Membership*\n\n"
            f"Choose your Supreme plan:\n\n"
            f"ðŸ’° <b>Apprentice</b> - ${apprentice_price:.2f} / 3 Months\n"
            f"ðŸ’° <b>Disciple</b> - ${disciple_price:.2f} / 6 Months\n"
            f"ðŸ’° <b>Legacy</b> - ${legacy_price:.2f} / Lifetime", 
            chat_id,
            message_id,
            reply_markup=markup, 
            parse_mode="HTML"
        )

@bot.callback_query_handler(func=lambda call: call.data == "menu_finish_forms")
def handle_menu_finish_forms(call):
    """Handle user clicking on 'Finish Forms' button"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    
    # Answer callback
    bot.answer_callback_query(call.id, "Starting registration forms...")
    
    # Send registration form graphic if available
    try:
        with open('graphics/registration_form.jpeg', 'rb') as form_img:
            bot.send_photo(
                user_id,
                form_img,
                caption="Please complete the registration form to continue"
            )
    except FileNotFoundError:
        logging.error("Registration form image not found at graphics/registration_form.jpeg")
    except Exception as e:
        logging.error(f"Error sending registration form image: {e}")
    
    # Add another small delay
    time.sleep(1.5)
    
    # Get membership type from user data
    membership_type = "regular"
    if str(user_id) in PAYMENT_DATA:
        membership_type = PAYMENT_DATA[str(user_id)].get('mentorship_type', 'regular').lower()
    
    # Prepare for onboarding form
    PENDING_USERS[user_id] = {
        'status': 'onboarding_form_regular_step1' if membership_type == 'regular' else 'onboarding_form_supreme_step1',
        'form_answers': {},  # Initialize empty dict to store responses
        'membership_type': membership_type,
        'target_group_id': SUPREME_GROUP_ID if membership_type == 'supreme' else PAID_GROUP_ID
    }
    save_pending_users()
    
    # Start onboarding form process
    send_onboarding_form(user_id)

def check_form_completion_reminders():
    """Check for users who need form completion reminders"""
    logging.info("Checking for pending form completions")
    
    now = datetime.now()
    
    for user_id_str, data in PAYMENT_DATA.items():
        if data.get('forms_needed', False) and data.get('trial_end_date'):
            try:
                # Parse trial end date
                trial_end_date = datetime.strptime(data['trial_end_date'], '%Y-%m-%d %H:%M:%S')
                
                # Check if we're within 2 days of expiration and haven't sent reminder yet
                days_remaining = (trial_end_date - now).days
                if days_remaining <= 2 and not data.get('form_reminder_sent', False):
                    # Time to send reminder
                    user_id = int(user_id_str)
                    
                    # Send reminder message
                    try:
                        bot.send_message(
                            user_id,
                            f"ðŸ“ *Form Completion Reminder*\n\n"
                            f"Your trial access expires in {days_remaining} days. To maintain your access, "
                            f"please complete your registration forms by clicking the 'Finish Forms' button in the main menu.",
                            parse_mode="Markdown"
                        )
                        
                        # Mark reminder as sent
                        PAYMENT_DATA[user_id_str]['form_reminder_sent'] = True
                        save_payment_data()
                        
                        logging.info(f"Sent form completion reminder to user {user_id}")
                    except Exception as e:
                        logging.error(f"Failed to send form reminder to user {user_id}: {e}")
                
                # Check if trial has expired
                elif now > trial_end_date and not data.get('form_expiry_notified', False):
                    # Trial expired, notify admins
                    user_id = int(user_id_str)
                    
                    # Get username for notification
                    username = data.get('username', f"User {user_id}")
                    username = safe_markdown_escape(username)
                    
                    # Notify admins
                    markup = InlineKeyboardMarkup(row_width=2)
                    markup.add(
                        InlineKeyboardButton("âŒ Kick User", callback_data=f"kick_form_{user_id}"),
                        InlineKeyboardButton("â³ Extend Trial", callback_data=f"extend_form_{user_id}")
                    )
                    
                    for admin_id in ADMIN_IDS:
                        bot.send_message(
                            admin_id,
                            f"âš ï¸ *FORM COMPLETION EXPIRED*\n\n"
                            f"User: @{username}\n"
                            f"ID: `{user_id}`\n\n"
                            f"This user's 1-week trial period has expired without completing registration forms.\n"
                            f"Please take action:",
                            parse_mode="Markdown",
                            reply_markup=markup
                        )
                    
                    # Mark as notified
                    PAYMENT_DATA[user_id_str]['form_expiry_notified'] = True
                    save_payment_data()
            except Exception as e:
                logging.error(f"Error checking form completion for user {user_id_str}: {e}")
    
    # Schedule next check in 12 hours
    threading.Timer(43200, check_form_completion_reminders).start()

# Start the form completion checker
threading.Timer(5, check_form_completion_reminders).start()

@bot.callback_query_handler(func=lambda call: call.data.startswith("kick_form_"))
def handle_kick_form_user(call):
    """Handle admin kicking a user who didn't complete forms"""
    admin_id = call.from_user.id
    
    # Verify admin status
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to use this action.")
        return
    
    # Extract user ID
    user_id = int(call.data.split("_")[2])
    user_id_str = str(user_id)
    
    try:
        # Get username for display
        username = "Unknown User"
        if user_id_str in PAYMENT_DATA:
            username = PAYMENT_DATA[user_id_str].get('username', "No Username")
        username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)
        
        # Try to notify user
        try:
            bot.send_message(
                user_id,
                "âŒ *Membership Terminated*\n\n"
                "Your trial membership has been terminated because you didn't complete your registration forms.\n\n"
                "To rejoin, please start a new application with /start and make sure to complete the forms.",
                parse_mode="Markdown"
            )
            user_notified = True
        except Exception:
            user_notified = False
        
        # Try to kick from group
        try:
            group_id = PAYMENT_DATA[user_id_str].get('target_group_id', PAID_GROUP_ID)
            bot.ban_chat_member(group_id, user_id)
            bot.unban_chat_member(group_id, user_id)  # Immediately unban so they can rejoin later
            kick_successful = True
        except Exception as e:
            logging.error(f"Failed to kick user {user_id}: {e}")
            kick_successful = False
        
        # Update payment data
        if user_id_str in PAYMENT_DATA:
            PAYMENT_DATA[user_id_str]['haspayed'] = False
            PAYMENT_DATA[user_id_str]['forms_needed'] = False
            PAYMENT_DATA[user_id_str]['kicked_for_no_forms'] = True
            PAYMENT_DATA[user_id_str]['kicked_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            PAYMENT_DATA[user_id_str]['kicked_by'] = admin_id
            save_payment_data()
        
        # Update admin message
        bot.edit_message_text(
            f"{'âœ…' if kick_successful else 'âŒ'} *ACTION TAKEN: USER KICKED*\n\n"
            f"@{username} has {'been removed from' if kick_successful else 'FAILED to be removed from'} the group.\n"
            f"User notification {'sent' if user_notified else 'FAILED'}.\n"
            f"Action taken by admin @{re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', call.from_user.username or f'ID:{admin_id}')}",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        
        # Notify all admins
        admin_username = call.from_user.username or f"Admin {admin_id}"
        admin_username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', admin_username)
        
        for other_admin_id in ADMIN_IDS:
            if other_admin_id != admin_id:
                bot.send_message(
                    other_admin_id,
                    f"ðŸ“ *Activity Log*\n\n"
                    f"@{admin_username} kicked @{username} for not completing registration forms.",
                    parse_mode="Markdown"
                )
        
        bot.answer_callback_query(call.id, f"User {username or user_id} has been kicked")
        
    except Exception as e:
        bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}")
        logging.error(f"Error kicking user who didn't complete forms: {e}")

@bot.callback_query_handler(func=lambda call: call.data.startswith("extend_form_"))
def handle_extend_form_period(call):
    """Handle admin extending a user's trial period for form completion"""
    admin_id = call.from_user.id
    
    # Verify admin status
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to use this action.")
        return
    
    # Extract user ID
    user_id = int(call.data.split("_")[2])
    user_id_str = str(user_id)
    
    try:
        # Get username for display
        username = "Unknown User"
        if user_id_str in PAYMENT_DATA:
            username = PAYMENT_DATA[user_id_str].get('username', "No Username")
        username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)
        
        # Calculate new trial end date (7 more days)
        new_trial_end_date = datetime.now() + timedelta(days=7)
        
        # Update payment data
        if user_id_str in PAYMENT_DATA:
            PAYMENT_DATA[user_id_str]['trial_end_date'] = new_trial_end_date.strftime('%Y-%m-%d %H:%M:%S')
            PAYMENT_DATA[user_id_str]['form_reminder_sent'] = False  # Reset so they get a new reminder
            PAYMENT_DATA[user_id_str]['form_expiry_notified'] = False  # Reset expiry notification flag
            save_payment_data()
        
        # Try to notify user
        try:
            bot.send_message(
                user_id,
                "â³ *Trial Period Extended*\n\n"
                "Good news! Your trial period for completing registration forms has been extended by 7 days.\n\n"
                f"Your new deadline is: {new_trial_end_date.strftime('%Y-%m-%d')}\n\n"
                "Please complete your registration forms by clicking the 'Finish Forms' button in the main menu.",
                parse_mode="Markdown"
            )
            user_notified = True
        except Exception:
            user_notified = False
        
        # Update admin message
        bot.edit_message_text(
            f"âœ… *ACTION TAKEN: TRIAL PERIOD EXTENDED*\n\n"
            f"@{username}'s trial period has been extended by 7 days.\n"
            f"New deadline: {new_trial_end_date.strftime('%Y-%m-%d')}\n"
            f"User notification {'sent' if user_notified else 'FAILED'}.\n"
            f"Action taken by admin @{re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', call.from_user.username or f'ID:{admin_id}')}",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        
        # Notify all admins
        admin_username = call.from_user.username or f"Admin {admin_id}"
        admin_username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', admin_username)
        
        for other_admin_id in ADMIN_IDS:
            if other_admin_id != admin_id:
                bot.send_message(
                    other_admin_id,
                    f"ðŸ“ *Activity Log*\n\n"
                    f"@{admin_username} extended @{username}'s trial period for form completion by 7 days.",
                    parse_mode="Markdown"
                )
        
        bot.answer_callback_query(call.id, f"Trial period extended for {username or user_id}")
        
    except Exception as e:
        bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}")
        logging.error(f"Error extending trial period for form completion: {e}")

@bot.callback_query_handler(func=lambda call: call.data == "menu_xm_deposit")
def handle_menu_xm_deposit(call):
    """Handle user clicking on 'Deposit Now' button"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    user_id_str = str(user_id)
    
    # Check if they're in grace period
    if user_id_str in PAYMENT_DATA and PAYMENT_DATA[user_id_str].get('xm_grace_period', False):
        # Calculate days remaining in grace period
        grace_end_date_str = PAYMENT_DATA[user_id_str].get('xm_grace_end_date')
        days_remaining = 0
        
        if grace_end_date_str:
            try:
                grace_end_date = datetime.strptime(grace_end_date_str, '%Y-%m-%d %H:%M:%S')
                now = datetime.now()
                days_remaining = (grace_end_date - now).days
            except Exception as e:
                logging.error(f"Error calculating grace period days: {e}")
        
        # Create inline keyboard for deposit confirmation
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineKeyboardButton("âœ… Yes, I've deposited", callback_data="xm_deposit_confirm"),
            InlineKeyboardButton("âŒ Not yet", callback_data="xm_deposit_not_yet")
        )
        markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
        
        # Show deposit instructions
        bot.edit_message_text(
            "ðŸ’° *XM Deposit Instructions*\n\n"
            f"You have {days_remaining if days_remaining > 0 else 'less than 1'} days remaining in your grace period.\n\n"
            "To fully activate your membership, please deposit a minimum of $30 to your XM account.\n\n"
            "*Option 1*\n"
            "ðŸ‘† In your XM Members' Area dashboard, click funding on the left panel of your screen\n"
            "ðŸ’³ Choose your desired payment method.\n"
            "ðŸ’µ Fund your account. Choose a trading account\n\n"
            "*Option 2:*\n"
            "ðŸ‘† Click deposit from your account in accounts overview.\n"
            "ðŸ’³ Choose your desired payment method. XM has various payment methods for you to choose!\n"
            "ðŸ’µ Start funding your live trading account with a minimum of $30.\n\n"
            "This is easy, but for more guidance, you may watch our YouTube tutorial down below.\n\n"
            "Have you completed your deposit?",
            chat_id, 
            message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
        
        bot.answer_callback_query(call.id, "Showing deposit instructions")
    else:
        # User is not in grace period - show error
        bot.answer_callback_query(call.id, "You don't need to make a deposit right now.", show_alert=True)

@bot.callback_query_handler(func=lambda call: call.data == "xm_deposit_confirm")
def handle_xm_deposit_confirm(call):
    """Handle user confirming they've made a deposit"""
    chat_id = call.message.chat.id
    user_id = call.from_user.id
    user_id_str = str(user_id)
    
    # Create keyboard for proof submission
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
    
    # Update message to request proof
    bot.edit_message_text(
        "ðŸ“¸ *XM Deposit Verification*\n\n"
        "Great! Please send a screenshot showing your deposit in your XM account.\n\n"
        "The screenshot should clearly show:\n"
        "â€¢ Your XM account ID\n"
        "â€¢ The deposit amount (minimum $30)\n"
        "â€¢ The transaction date\n\n"
        "Please upload your screenshot now.",
        chat_id,
        call.message.message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    # Update status to await screenshot
    PENDING_USERS[user_id] = {
        'status': 'xm_awaiting_deposit_screenshot',
        'xm_account_id': PAYMENT_DATA[user_id_str].get('xm_account_id', 'Unknown')
    }
    save_pending_users()
    
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda call: call.data == "xm_deposit_not_yet")
def handle_xm_deposit_not_yet(call):
    """Handle user indicating they haven't deposited yet"""
    chat_id = call.message.chat.id
    user_id = call.from_user.id
    user_id_str = str(user_id)
    
    # Get days remaining in grace period
    grace_end_date_str = PAYMENT_DATA[user_id_str].get('xm_grace_end_date')
    days_remaining = 0
    
    if grace_end_date_str:
        try:
            grace_end_date = datetime.strptime(grace_end_date_str, '%Y-%m-%d %H:%M:%S')
            now = datetime.now()
            days_remaining = (grace_end_date - now).days
        except Exception as e:
            logging.error(f"Error calculating grace period days: {e}")
    
    # Create back button
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
    
    # Update message with reminder
    bot.edit_message_text(
        "â³ *Deposit Reminder*\n\n"
        f"You have {days_remaining if days_remaining > 0 else 'less than 1'} days remaining in your grace period.\n\n"
        "Remember, you need to make a deposit of at least $30 to your XM account to maintain your membership.\n\n"
        "You can use the 'Deposit Now' option in the main menu when you're ready to proceed.",
        chat_id,
        call.message.message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    bot.answer_callback_query(call.id)

# Handle screenshot submission for deposit verification
@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'xm_awaiting_deposit_screenshot', content_types=['photo'])
def handle_deposit_screenshot(message):
    """Handle the deposit screenshot submission"""
    user_id = message.from_user.id
    chat_id = message.chat.id
    user_id_str = str(user_id)
    
    # Store the screenshot message ID for reference
    PENDING_USERS[user_id]['screenshot_message_id'] = message.message_id
    
    # Get account ID 
    xm_account_id = PENDING_USERS[user_id].get('xm_account_id', 'Unknown')
    
    # Forward screenshot to admins for verification
    for admin_id in ADMIN_IDS:
        # Forward the screenshot
        bot.forward_message(admin_id, chat_id, message.message_id)
        
        # Create verification buttons for admins
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineKeyboardButton("âœ… Verify Deposit", callback_data=f"verify_deposit_{user_id}"),
            InlineKeyboardButton("âŒ Reject Deposit", callback_data=f"reject_deposit_{user_id}")
        )
        
        # Get username for display
        username = PAYMENT_DATA[user_id_str].get('username', "No Username")
        username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)
        
        # Send context message to admin
        bot.send_message(
            admin_id,
            f"ðŸ”” *XM Deposit Verification*\n\n"
            f"User @{username} (ID: `{user_id}`) has submitted proof of their XM deposit.\n\n"
            f"â€¢ Account ID: `{xm_account_id}`\n"
            f"â€¢ Required: $30 minimum deposit\n\n"
            f"Please verify that the deposit meets requirements.",
            parse_mode="Markdown",
            reply_markup=markup
        )
    
    # Update user status
    PENDING_USERS[user_id]['status'] = 'xm_awaiting_deposit_verification'
    save_pending_users()
    
    # Inform user that verification is in progress
    bot.send_message(
        chat_id,
        "âœ… *Verification Submitted*\n\n"
        "Your deposit proof has been sent to our admin team for verification.\n"
        "You'll be notified once your verification is approved or rejected.",
        parse_mode="Markdown"
    )

# Add handlers for admin deposit verification
@bot.callback_query_handler(func=lambda call: call.data.startswith("verify_deposit_"))
def handle_verify_deposit(call):
    """Handle admin verifying a user's deposit"""
    admin_id = call.from_user.id
    
    # Verify admin status
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to use this action.")
        return
    
    # Extract user ID
    user_id = int(call.data.split("_")[2])
    user_id_str = str(user_id)
    
    try:
        # Check if user data exists
        if user_id_str not in PAYMENT_DATA:
            bot.answer_callback_query(call.id, "âŒ User data not found.")
            return
        
        # Get username for display
        username = PAYMENT_DATA[user_id_str].get('username', "No Username")
        username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)
        
        # Update user data to remove grace period and deposit need
        PAYMENT_DATA[user_id_str]['xm_grace_period'] = False
        PAYMENT_DATA[user_id_str]['needs_deposit'] = False
        PAYMENT_DATA[user_id_str]['deposit_verified'] = True
        PAYMENT_DATA[user_id_str]['deposit_verified_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        PAYMENT_DATA[user_id_str]['deposit_verified_by'] = admin_id
        save_payment_data()
        
        # Update message to admin
        bot.edit_message_text(
            f"âœ… *XM Deposit VERIFIED*\n\n"
            f"You have verified @{username}'s XM deposit.\n"
            f"They have been notified and their account is now fully activated.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        
        # Notify user
        try:
            bot.send_message(
                user_id,
                "âœ… *XM Deposit Verified!*\n\n"
                "Your deposit has been verified and your membership is now fully activated.\n\n"
                "Thank you for completing this important step. Enjoy your full access to Prodigy Trading Academy!",
                parse_mode="Markdown"
            )
        except Exception as e:
            logging.error(f"Failed to notify user {user_id} about deposit verification: {e}")
        
        # Remove from pending users
        if user_id in PENDING_USERS:
            PENDING_USERS.pop(user_id)
            delete_pending_user(user_id)
        
        # Notify all admins
        admin_username = call.from_user.username or f"Admin {admin_id}"
        admin_username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', admin_username)
        
        for other_admin_id in ADMIN_IDS:
            if other_admin_id != admin_id:
                bot.send_message(
                    other_admin_id,
                    f"ðŸ“ *Activity Log*\n\n"
                    f"@{admin_username} verified @{username}'s XM deposit.",
                    parse_mode="Markdown"
                )
        
        bot.answer_callback_query(call.id, "âœ… Deposit verified successfully!")
        
    except Exception as e:
        bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}")
        logging.error(f"Error verifying deposit: {e}")

@bot.callback_query_handler(func=lambda call: call.data.startswith("reject_deposit_"))
def handle_reject_deposit(call):
    """Handle admin rejecting a user's deposit proof"""
    admin_id = call.from_user.id
    
    # Verify admin status
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to use this action.")
        return
    
    # Extract user ID
    user_id = int(call.data.split("_")[2])
    user_id_str = str(user_id)
    
    try:
        # Get username for display
        username = "Unknown User"
        if user_id_str in PAYMENT_DATA:
            username = PAYMENT_DATA[user_id_str].get('username', "No Username")
        username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)
        
        # Update message to admin
        bot.edit_message_text(
            f"âŒ *XM Deposit REJECTED*\n\n"
            f"You have rejected @{username}'s XM deposit proof.\n"
            f"They have been notified to provide a clearer proof.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        
        # Notify user
        try:
            bot.send_message(
                user_id,
                "âŒ *XM Deposit Verification Failed*\n\n"
                "We couldn't verify your XM deposit from the screenshot provided.\n\n"
                "Possible issues:\n"
                "â€¢ The deposit amount was not visible or below $30\n"
                "â€¢ The account ID was not clearly shown\n"
                "â€¢ The image quality was too low\n\n"
                "Please try again with a clearer screenshot by selecting 'Deposit Now' from the main menu.",
                parse_mode="Markdown"
            )
        except Exception as e:
            logging.error(f"Failed to notify user {user_id} about deposit rejection: {e}")
        
        # Reset pending status to allow them to try again
        if user_id in PENDING_USERS:
            PENDING_USERS[user_id]['status'] = 'at_main_menu'
            save_pending_users()
        
        # Notify all admins
        admin_username = call.from_user.username or f"Admin {admin_id}"
        admin_username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', admin_username)
        
        for other_admin_id in ADMIN_IDS:
            if other_admin_id != admin_id:
                bot.send_message(
                    other_admin_id,
                    f"ðŸ“ *Activity Log*\n\n"
                    f"@{admin_username} rejected @{username}'s XM deposit proof.",
                    parse_mode="Markdown"
                )
        
        bot.answer_callback_query(call.id, "âŒ Deposit proof rejected")
        
    except Exception as e:
        bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}")
        logging.error(f"Error rejecting deposit: {e}")

def check_grace_periods():
    """Check for users with expired grace periods and notify admins"""
    logging.info("Checking for expired grace periods")
    
    now = datetime.now()
    
    for user_id_str, data in PAYMENT_DATA.items():
        if data.get('xm_grace_period', False) and data.get('xm_grace_end_date'):
            try:
                # Parse grace end date
                grace_end_date = datetime.strptime(data['xm_grace_end_date'], '%Y-%m-%d %H:%M:%S')
                
                # Check if grace period has expired
                if now > grace_end_date:
                    # Grace period expired - notify admins
                    username = data.get('username', "No Username")
                    username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)
                    user_id = int(user_id_str)
                    
                    # Create action buttons for admins
                    markup = InlineKeyboardMarkup(row_width=2)
                    markup.add(
                        InlineKeyboardButton("âŒ Kick User", callback_data=f"kick_grace_{user_id}"),
                        InlineKeyboardButton("â³ Extend Grace", callback_data=f"extend_grace_{user_id}")
                    )
                    
                    # Notify all admins
                    for admin_id in ADMIN_IDS:
                        bot.send_message(
                            admin_id,
                            f"âš ï¸ *GRACE PERIOD EXPIRED*\n\n"
                            f"User: @{username}\n"
                            f"ID: `{user_id}`\n\n"
                            f"This user's 14-day grace period for XM deposit has expired.\n"
                            f"Please take action:",
                            parse_mode="Markdown",
                            reply_markup=markup
                        )
                    
                    # Mark as notified to avoid duplicate notifications
                    PAYMENT_DATA[user_id_str]['grace_period_expired_notified'] = True
                    save_payment_data()
                    
            except Exception as e:
                logging.error(f"Error checking grace period for user {user_id_str}: {e}")
    
    # Schedule next check in 12 hours
    threading.Timer(43200, check_grace_periods).start()

# Start the grace period checker
threading.Timer(5, check_grace_periods).start()

@bot.callback_query_handler(func=lambda call: call.data.startswith("kick_grace_"))
def handle_kick_grace_user(call):
    """Handle admin kicking a user with expired grace period"""
    admin_id = call.from_user.id
    
    # Verify admin status
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to use this action.")
        return
    
    # Extract user ID
    user_id = int(call.data.split("_")[2])
    user_id_str = str(user_id)
    
    try:
        # Get username for display
        username = "Unknown User"
        if user_id_str in PAYMENT_DATA:
            username = PAYMENT_DATA[user_id_str].get('username', "No Username")
        username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)
        
        # Try to notify user
        try:
            bot.send_message(
                user_id,
                "âŒ *Membership Terminated*\n\n"
                "Your membership has been terminated because your grace period has expired without a deposit.\n\n"
                "To rejoin, please start a new application with /start and make sure to complete the deposit.",
                parse_mode="Markdown"
            )
            user_notified = True
        except Exception:
            user_notified = False
        
        # Try to kick from group
        try:
            group_id = PAID_GROUP_ID  # Regular membership group
            bot.ban_chat_member(group_id, user_id)
            bot.unban_chat_member(group_id, user_id)  # Immediately unban so they can rejoin later
            kick_successful = True
        except Exception as e:
            logging.error(f"Failed to kick user {user_id}: {e}")
            kick_successful = False
        
        # Update payment data
        if user_id_str in PAYMENT_DATA:
            PAYMENT_DATA[user_id_str]['haspayed'] = False
            PAYMENT_DATA[user_id_str]['xm_grace_period'] = False
            PAYMENT_DATA[user_id_str]['kicked_for_no_deposit'] = True
            PAYMENT_DATA[user_id_str]['kicked_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            PAYMENT_DATA[user_id_str]['kicked_by'] = admin_id
            save_payment_data()
        
        # Update admin message
        bot.edit_message_text(
            f"{'âœ…' if kick_successful else 'âŒ'} *ACTION TAKEN: USER KICKED*\n\n"
            f"@{username} has {'been removed from' if kick_successful else 'FAILED to be removed from'} the group.\n"
            f"User notification {'sent' if user_notified else 'FAILED'}.\n"
            f"Action taken by admin @{re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', call.from_user.username or f'ID:{admin_id}')}",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        
        # Notify all admins
        admin_username = call.from_user.username or f"Admin {admin_id}"
        admin_username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', admin_username)
        
        for other_admin_id in ADMIN_IDS:
            if other_admin_id != admin_id:
                bot.send_message(
                    other_admin_id,
                    f"ðŸ“ *Activity Log*\n\n"
                    f"@{admin_username} kicked @{username} for expired grace period without deposit.",
                    parse_mode="Markdown"
                )
        
        bot.answer_callback_query(call.id, f"User {username or user_id} has been kicked")
        
    except Exception as e:
        bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}")
        logging.error(f"Error kicking grace period user: {e}")

@bot.callback_query_handler(func=lambda call: call.data.startswith("extend_grace_"))
def handle_extend_grace_period(call):
    """Handle admin extending a user's grace period"""
    admin_id = call.from_user.id
    
    # Verify admin status
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to use this action.")
        return
    
    # Extract user ID
    user_id = int(call.data.split("_")[2])
    user_id_str = str(user_id)
    
    try:
        # Get username for display
        username = "Unknown User"
        if user_id_str in PAYMENT_DATA:
            username = PAYMENT_DATA[user_id_str].get('username', "No Username")
        username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)
        
        # Calculate new grace period end date (7 more days)
        new_grace_end_date = datetime.now() + timedelta(days=7)
        
        # Update payment data
        if user_id_str in PAYMENT_DATA:
            PAYMENT_DATA[user_id_str]['xm_grace_period'] = True
            PAYMENT_DATA[user_id_str]['xm_grace_end_date'] = new_grace_end_date.strftime('%Y-%m-%d %H:%M:%S')
            PAYMENT_DATA[user_id_str]['grace_period_extended'] = True
            PAYMENT_DATA[user_id_str]['grace_period_extended_by'] = admin_id
            PAYMENT_DATA[user_id_str]['grace_period_extended_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            save_payment_data()
        
        # Try to notify user
        try:
            bot.send_message(
                user_id,
                "â³ *Grace Period Extended*\n\n"
                "Good news! Your grace period for XM deposit has been extended by 7 days.\n\n"
                f"Your new deadline to deposit is: {new_grace_end_date.strftime('%Y-%m-%d')}\n\n"
                "Please make your deposit of at least $30 using the 'Deposit Now' option in the main menu.",
                parse_mode="Markdown"
            )
            user_notified = True
        except Exception:
            user_notified = False
        
        # Update admin message
        bot.edit_message_text(
            f"âœ… *ACTION TAKEN: GRACE PERIOD EXTENDED*\n\n"
            f"@{username}'s grace period has been extended by 7 days.\n"
            f"New deadline: {new_grace_end_date.strftime('%Y-%m-%d')}\n"
            f"User notification {'sent' if user_notified else 'FAILED'}.\n"
            f"Action taken by admin @{re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', call.from_user.username or f'ID:{admin_id}')}",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        
        # Notify all admins
        admin_username = call.from_user.username or f"Admin {admin_id}"
        admin_username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', admin_username)
        
        for other_admin_id in ADMIN_IDS:
            if other_admin_id != admin_id:
                bot.send_message(
                    other_admin_id,
                    f"ðŸ“ *Activity Log*\n\n"
                    f"@{admin_username} extended @{username}'s grace period by 7 days.",
                    parse_mode="Markdown"
                )
        
        bot.answer_callback_query(call.id, f"Grace period extended for {username or user_id}")
        
    except Exception as e:
        bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}")
        logging.error(f"Error extending grace period: {e}")

@bot.callback_query_handler(func=lambda call: call.data == "menu_xm_free")
def handle_xm_free_mentorship(call):
    """Handle the XM Free mentorship option"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    
    # CHECK ENROLLMENT STATUS FIRST - This is the fix
    if not BOT_SETTINGS.get('regular_enrollment_open', True):
        # Create inline keyboard with Get Notified and FAQ buttons
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineKeyboardButton("ðŸ”” Get Notified", callback_data="update_yes"),
            InlineKeyboardButton("â“ FAQ", callback_data="faq_back")
        )
        markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
        
        bot.edit_message_text(
            "âš ï¸ *XM Partnership enrollment is currently CLOSED*\n\n"
            "New XM Partnership registrations are temporarily unavailable.\n"
            "Please wait for the next announcement about when enrollment will open again.\n\n"
            "â€¢ Click *Get Notified* to receive updates when enrollment opens\n"
            "â€¢ Check our *FAQ* section for more information",
            chat_id,
            message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
        bot.answer_callback_query(call.id, "XM Partnership enrollment is currently closed")
        return
    
    # If enrollment is open, continue with normal process
    # Answer callback for feedback to user
    bot.answer_callback_query(call.id, "Starting XM Free Mentorship process...")
    
    # Update user status
    PENDING_USERS[user_id] = {'status': 'xm_free_mentorship'}
    save_pending_users()
    
    # Rest of the existing code...
    # Create inline keyboard for Yes/No response
    markup = InlineKeyboardMarkup(row_width=2)
    markup.add(
        InlineKeyboardButton("Yes", callback_data="xm_existing_yes"),
        InlineKeyboardButton("No", callback_data="xm_existing_no")
    )
    markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
    
    # Edit the existing message to ask about XM account with more comprehensive information
    bot.edit_message_text(
        "ðŸ“š *Regular Mentorship (FREE)*\n\n"
        "Start your trading journey with Prodigy Trading Academy's foundational program â€” the Regular Mentorship, completely free of charge and packed with all the essential resources to help you grow.\n\n"
        "As we are partnered with the No.1 broker in the Philippines, XM has helped give you the opportunity to join our mentorship for free.\n\n"
        "Everyone in our academy would need to register, create an XM account, and engage in active trading with us through our connection with XM in order to join us.\n\n"
        "Do you already have an XM account before we continue?",
        chat_id,
        message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.callback_query_handler(func=lambda call: call.data == "xm_existing_yes")
def handle_xm_existing_yes(call):
    """Handle user confirming they have an existing XM account"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    
    # Answer callback
    bot.answer_callback_query(call.id, "Processing...")
    
    # Update user status
    PENDING_USERS[user_id]['status'] = 'xm_add_partner_code'
    save_pending_users()
    
    # Create inline keyboard for compliance response
    markup = InlineKeyboardMarkup(row_width=2)
    markup.add(
        InlineKeyboardButton("Yes, I Agree", callback_data="xm_comply_yes"),
        InlineKeyboardButton("No", callback_data="xm_comply_no")
    )
    markup.add(InlineKeyboardButton("â¬…ï¸ Back", callback_data="menu_xm_free"))
    
    # Edit the message to show partner code instructions with updated text
    bot.edit_message_text(
        "ðŸ“¸ *Regular Mentorship: XM Partnership - Account Verification*\n\n"
        "Please provide the following:\n\n"
        "ðŸ“± A screenshot showing that you've entered the partner code PTAPARTNER during registration\n"
        "ðŸªª Your XM Members' Area dashboard\n\n"
        "Please upload your screenshot now, and include your account ID in a follow-up message.\n\n"
        "â€¼ï¸ In case you forgot to screenshot everything, you'll have to try over again and register with a new e-mail. "
        "This only likely will happen if instructions aren't clear, so make sure to watch the video tutorial if you didn't!",
        chat_id,
        message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.callback_query_handler(func=lambda call: call.data == "xm_existing_no")
def handle_xm_existing_no(call):
    """Handle user indicating they don't have an XM account"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    
    # Answer callback
    bot.answer_callback_query(call.id, "Processing...")
    
    # Update user status
    PENDING_USERS[user_id] = {
        'status': 'xm_new_account_instructions',
    }
    save_pending_users()
    
    # Create buttons for the video tutorial
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(
        InlineKeyboardButton("ðŸ“º Watch Registration Tutorial", url="https://youtu.be/6PNg7OrltTY?si=QCta4xODaTXuSiDb"),
        InlineKeyboardButton("â¬…ï¸ Back", callback_data="menu_xm_free")
    )
    
    # Edit message to show instructions for new accounts with updated text
    bot.edit_message_text(
        "ðŸ“ *Regular Mentorship: XM Partnership - New Account Registration*\n\n"
        "To qualify for free mentorship, please follow these steps to create an XM account with our partnership code:\n\n"
        "ðŸ”— *Step 1: Register for an XM Account*\n"
        "Sign up using our official link: https://affs.click/WLw2n\n\n"
        "ðŸ§¾ *Step 2: Input Our Partner Code*\n"
        "During registration, enter this partner code: *PTAPARTNER*\n"
        "âœ… Take a screenshot of this step showing the partner code clearly before you click continue.\n\n"
        "ðŸ›¡ *Step 3: Verify Your Account*\n"
        "After successfully doing Step 2, you will be redirected to the dashboard and verify your identity and residence.\n"
        "âœ… Take a screenshot of your Members' Area dashboard.\n\n"
        "ðŸ’³ *Step 4: Make Your Initial Deposit*\n"
        "You'll need to fund your account a minimum of $30 to complete eligibility. In case you aren't able to deposit right away, we'll give you a 1-week grace period!\n\n"
        "ðŸ“¸ *Important:*\n"
        "We'll ask for verification later, so save the following now:\n\n"
        "- Screenshot showing PTAPARTNER entered during registration\n"
        "- Your XM Trading Account ID\n\n"
        "ðŸ“Œ Finishing the registration is better if you watch the registration tutorial to get a smooth and convenient processing! Please watch the tutorial video below.\n\n"
        "â€¼ï¸ After this message, you will be given a continue button in order to proceed shortly.",
        chat_id,
        message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    # Wait 10 seconds to ensure they have time to view the instructions
    time.sleep(10)
    
    # Follow up with verification prompt
    markup = InlineKeyboardMarkup(row_width=2)
    markup.add(
        InlineKeyboardButton("âœ… Yes, continue", callback_data="xm_verify_proceed"),
        InlineKeyboardButton("âŒ No, cancel", callback_data="xm_verify_cancel")
    )
    
    bot.send_message(
        chat_id,
        "âš ï¸ *Important*\n\n"
        "Did you watch the tutorial video? Do you wish to proceed with verification?",
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.callback_query_handler(func=lambda call: call.data == "xm_verify_proceed")
def handle_xm_verify_proceed(call):
    """Handle user proceeding with XM verification"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    
    # Answer callback
    bot.answer_callback_query(call.id, "Processing...")
    
    # Update user status
    PENDING_USERS[user_id]['status'] = 'xm_awaiting_account_screenshot'
    save_pending_users()
    
    # Create back button
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(InlineKeyboardButton("â¬…ï¸ Back", callback_data="xm_existing_no"))
    
    # Edit message to request screenshot with new improved text
    bot.edit_message_text(
        "ðŸ“¸ *XM Account Verification*\n\n"
        "You're almost done with registering! Now we'll start verifying your account registration in the academy in order to confirm if you really registered through our XM broker.\n\n"
        "To proceed, please be informed to prepare the following:\n\n"
        "1ï¸âƒ£ A screenshot showing that you've entered the partner code PTAPARTNER during registration\n"
        "2ï¸âƒ£ Your XM Trading Account ID\n\n"
        "Let's first start with the screenshot showing your proof of registration and integrating our partnership code.",
        chat_id,
        message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'xm_awaiting_account_screenshot', content_types=['photo'])
def handle_xm_registration_screenshot(message):
    """Handle the first screenshot (registration screenshot)"""
    user_id = message.from_user.id
    chat_id = message.chat.id
    
    # Store screenshot details
    PENDING_USERS[user_id]['screenshot_message_id'] = message.message_id
    PENDING_USERS[user_id]['registration_screenshot'] = message.photo[-1].file_id
    PENDING_USERS[user_id]['status'] = 'xm_awaiting_dashboard_screenshot'
    save_pending_users()
    
    # Ask for XM Dashboard screenshot instead of account ID
    bot.send_message(
        chat_id,
        "âœ… Thanks for the registration screenshot!\n\n"
        "ðŸ“Š *Now please provide a screenshot of your XM Dashboard*\n\n"
        "This should show your account ID and other account details for verification.",
        parse_mode="Markdown"
    )

@bot.callback_query_handler(func=lambda call: call.data == "xm_verify_cancel")
def handle_xm_verify_cancel(call):
    """Handle user cancelling the XM verification process"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    
    # Answer callback
    bot.answer_callback_query(call.id, "Cancelling XM verification process...")
    
    # Clear user's pending status
    if user_id in PENDING_USERS:
        PENDING_USERS.pop(user_id, None)
        delete_pending_user(user_id)
    
    # Create markup for return to main menu
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
    
    # Edit the message to show cancellation
    bot.edit_message_text(
        "âŒ *Verification Cancelled*\n\n"
        "You've cancelled the XM verification process. You can restart it anytime from the main menu.",
        chat_id,
        message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'xm_awaiting_dashboard_screenshot', content_types=['photo'])
def handle_xm_dashboard_screenshot(message):
    """Handle the second screenshot (dashboard screenshot)"""
    user_id = message.from_user.id
    chat_id = message.chat.id
    
    # Store dashboard screenshot details
    PENDING_USERS[user_id]['dashboard_screenshot'] = message.photo[-1].file_id
    PENDING_USERS[user_id]['dashboard_message_id'] = message.message_id
    PENDING_USERS[user_id]['status'] = 'xm_awaiting_deposit_confirmation'
    save_pending_users()
    
    # Create inline keyboard for deposit confirmation
    markup = InlineKeyboardMarkup(row_width=2)
    markup.add(
        InlineKeyboardButton("âœ… Yes", callback_data="xm_deposit_yes"),
        InlineKeyboardButton("âŒ No, deposit later", callback_data="xm_deposit_no")
    )
    
    # Ask for deposit confirmation
    bot.send_message(
        chat_id,
        "ðŸ’° *Deposit Confirmation*\n\n"
        "Thank you for providing your verification screenshots.\n\n"
        "An initial deposit is required to activate your XM account.\n\n"
        "Are you willing to make a deposit to your XM account?",
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.callback_query_handler(func=lambda call: call.data == "xm_deposit_yes")
def handle_xm_deposit_yes(call):
    """Handle user confirming willingness to deposit"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    
    # Answer callback
    bot.answer_callback_query(call.id, "Processing...")
    
    # Update user status
    PENDING_USERS[user_id]['status'] = 'xm_awaiting_compliance'
    save_pending_users()
    
    # Create inline keyboard for compliance confirmation
    markup = InlineKeyboardMarkup(row_width=2)
    markup.add(
        InlineKeyboardButton("âœ… Yes, I comply", callback_data="xm_compliance_yes"),
        InlineKeyboardButton("âŒ No", callback_data="xm_compliance_no")
    )
    
    # Edit message to show deposit instructions
    bot.edit_message_text(
        "âœ… *Regular Mentorship: XM Partnership - Deposit Notice*\n\n"
        "That's great to hear! Depositing in your XM account is as easy as 1-2-3, just read and do the following steps below!\n\n"
        "Just choose whether method or procedure is much more convenient for you. Either way works.\n\n"
        "*Option 1*\n"
        "ðŸ‘† In your XM Members' Area dashboard, click funding on the left panel of your screen\n"
        "ðŸ’³ Choose your desired payment method.\n"
        "ðŸ’µ Fund your account. Choose a trading account\n\n"
        "*Option 2:*\n"
        "ðŸ‘† Click deposit from your account in accounts overview.\n"
        "ðŸ’³ Choose your desired payment method. XM has various payment methods for you to choose!\n"
        "ðŸ’µ Start funding your live trading account with a minimum of $30.\n\n"
        "This is easy, but for more guidance, you may watch our Youtube tutorial down below.\n\n"
        "ðŸ“¸ After this message, we will need you to verify your deposit. The following screenshot is what we need:\n"
        "- XM Members' Area dashboard\n\n"
        "Do you agree to comply with the following instructions?\n\n"
        "1ï¸âƒ£ Maintain the partner code PTAPARTNER in your XM account\n"
        "2ï¸âƒ£ Complete regular trading activity in your account\n"
        "3ï¸âƒ£ Submit monthly verification of your active partnership status",
        chat_id,
        message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.callback_query_handler(func=lambda call: call.data == "xm_deposit_no")
def handle_xm_deposit_no(call):
    """Handle user choosing to deposit later"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    
    # Answer callback
    bot.answer_callback_query(call.id, "Processing...")
    
    # Calculate grace period end date (2 weeks from now)
    grace_end_date = datetime.now() + timedelta(days=14)
    
    # Update user status to proceed to compliance agreement
    # Instead of terminating, we continue the process with grace period
    PENDING_USERS[user_id]['status'] = 'xm_awaiting_compliance'
    PENDING_USERS[user_id]['grace_period'] = True
    PENDING_USERS[user_id]['grace_end_date'] = grace_end_date.strftime('%Y-%m-%d %H:%M:%S')
    save_pending_users()
    
    # Create inline keyboard for compliance confirmation
    markup = InlineKeyboardMarkup(row_width=2)
    markup.add(
        InlineKeyboardButton("âœ… Yes, I comply", callback_data="xm_compliance_yes"),
        InlineKeyboardButton("âŒ No", callback_data="xm_compliance_no")
    )
    
    # Edit message to show compliance instructions with new message text
    bot.edit_message_text(
        "ðŸ’¡ *Regular Mentorship: XM Partnership - Deposit Notice*\n\n"
        "To fully activate your mentorship access, a minimum deposit of $30 is required in your XM account.\n\n"
        "We understand that not everyone may be able to deposit immediately â€” that's totally okay! ðŸ˜Š\n"
        "You'll automatically receive a 2-week grace period from the moment your account is verified.\n\n"
        "During this time, you can still explore our resources and settle your deposit when you're ready. After the 14 days, we will ask you about your deposit.\n\n"
        "Do you agree to comply with the following instructions?\n\n"
        "1ï¸âƒ£ Maintain the partner code PTAPARTNER in your XM account\n"
        "2ï¸âƒ£ Complete regular trading activity in your account\n"
        "3ï¸âƒ£ Submit monthly verification of your active partnership status",
        chat_id,
        message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.callback_query_handler(func=lambda call: call.data == "xm_compliance_yes")
def handle_xm_compliance_yes(call):
    """Handle user agreeing to comply with instructions"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    
    # Answer callback
    bot.answer_callback_query(call.id, "Processing...")
    
    # Update user status
    PENDING_USERS[user_id]['status'] = 'xm_waiting_approval'
    save_pending_users()
    
    # Get metadata
    xm_account_id = PENDING_USERS[user_id].get('xm_account_id', 'Unknown')
    username = call.from_user.username or "No Username"
    
    # Properly escape username for Markdown - this fixes the issue
    if username != "No Username":
        username = safe_markdown_escape(username)
    
    # Get both screenshots for forwarding
    registration_screenshot_id = PENDING_USERS[user_id].get('registration_screenshot')
    dashboard_screenshot_id = PENDING_USERS[user_id].get('dashboard_screenshot')
    
    # Forward both screenshots and details to admins for verification
    for admin_id in ADMIN_IDS:
        # Send an intro message first
        intro_message = (
            f"ðŸ”” *XM Verification Request*\n\n"
            f"User @{username} (ID: `{user_id}`) has submitted verification for XM Partnership.\n\n"
            f"â€¢ Partner Code: PTAPARTNER\n"
            f"â€¢ Account ID: `{xm_account_id}`\n\n"
            f"Please review the following screenshots:"
        )
        
        bot.send_message(
            admin_id,
            intro_message,
            parse_mode="Markdown"
        )
        
        # Forward registration screenshot if available
        if registration_screenshot_id:
            bot.send_photo(
                admin_id,
                registration_screenshot_id,
                caption="ðŸ“¸ 1/2: Registration screenshot showing PTAPARTNER code"
            )
        
        # Forward dashboard screenshot if available
        if dashboard_screenshot_id:
            bot.send_photo(
                admin_id,
                dashboard_screenshot_id,
                caption="ðŸ“Š 2/2: XM Dashboard screenshot showing account details"
            )
            
        # Create approval buttons for admins
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineKeyboardButton("âœ… Approve XM", callback_data=f"approve_xm_{user_id}"),
            InlineKeyboardButton("âŒ Reject XM", callback_data=f"reject_xm_{user_id}")
        )
        
        # Send verification action buttons
        bot.send_message(
            admin_id,
            f"â¬†ï¸ Please verify the screenshots above for user @{username}\n"
            f"Both screenshots should clearly show valid XM account with partner code.",
            parse_mode="Markdown",
            reply_markup=markup
        )
    
    # Edit the message to inform user their submission is pending approval
    bot.edit_message_text(
        "âœ… *Verification Submitted*\n\n"
        "Your XM account verification has been submitted to our admin team.\n"
        "You will be notified once your verification is approved or rejected.",
        chat_id,
        message_id,
        parse_mode="Markdown"
    )

@bot.callback_query_handler(func=lambda call: call.data == "xm_compliance_no")
def handle_xm_compliance_no(call):
    """Handle user declining to comply with instructions"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    
    # Answer callback
    bot.answer_callback_query(call.id, "Processing...")
    
    # Create back to main menu button
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
    
    # Edit message to deny access
    bot.edit_message_text(
        "âŒ *Access Denied*\n\n"
        "Compliance with the partner instructions is required to qualify for free mentorship.\n\n"
        "If you change your mind in the future, you can restart the process from the main menu.",
        chat_id,
        message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    # Clean up pending data
    PENDING_USERS.pop(user_id, None)
    delete_pending_user(user_id)

@bot.callback_query_handler(func=lambda call: call.data.startswith("approve_xm_"))
def handle_approve_xm(call):
    """Handle admin approval of XM partnership verification"""
    admin_id = call.from_user.id
    
    # Verify admin status
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to use this action.")
        return
    
    # Extract user ID from callback data
    user_id = int(call.data.split("_")[2])
    user_id_str = str(user_id)
    
    try:
        # Get user info for notifications
        try:
            user_info = bot.get_chat(user_id)
            username = user_info.username or "No Username"
            if username != "No Username":
                username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)
        except Exception:
            username = f"User {user_id}"
            
        # Check for any evidence of a previous valid membership
        is_renewal = user_id_str in PAYMENT_DATA and PAYMENT_DATA[user_id_str].get('haspayed', False)

        # Calculate membership due date (1 month for XM partnership)
        due_date = datetime.now() + timedelta(days=30)
        
        # Check if user is in grace period
        grace_period = False
        grace_end_date = None
        if user_id in PENDING_USERS and PENDING_USERS[user_id].get('grace_period', False):
            grace_period = True
            grace_end_date = PENDING_USERS[user_id].get('grace_end_date')
        
        # Update payment data - note it's free but still tracked
        if user_id_str in PAYMENT_DATA:
            # Update existing record
            PAYMENT_DATA[user_id_str].update({
                "username": username,
                "payment_plan": "Monthly",
                "payment_mode": "XM Partnership",
                "mentorship_type": "Regular",  # XM partnership is only for Regular membership
                "due_date": due_date.strftime('%Y-%m-%d %H:%M:%S'),
                "haspayed": True,
                "xm_verified": True,
                "xm_verified_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "xm_verified_by": admin_id
            })
            
            # Add grace period data if applicable
            if grace_period and grace_end_date:
                PAYMENT_DATA[user_id_str].update({
                    "xm_grace_period": True,
                    "xm_grace_end_date": grace_end_date,
                    "needs_deposit": True  # This flag will be used to show "Deposit now" in main menu
                })
        else:
            # Create new entry
            PAYMENT_DATA[user_id_str] = {
                "username": username,
                "payment_plan": "Monthly",
                "payment_mode": "XM Partnership",
                "mentorship_type": "Regular",  # XM partnership is only for Regular membership
                "due_date": due_date.strftime('%Y-%m-%d %H:%M:%S'),
                "haspayed": True,
                "xm_verified": True,
                "xm_verified_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "xm_verified_by": admin_id,
                "terms_accepted": True,
                "privacy_accepted": True
            }
            
            # Add grace period data if applicable
            if grace_period and grace_end_date:
                PAYMENT_DATA[user_id_str].update({
                    "xm_grace_period": True,
                    "xm_grace_end_date": grace_end_date,
                    "needs_deposit": True  # This flag will be used to show "Deposit now" in main menu
                })
        
        save_payment_data()
        
        # Mark as verified in PENDING_USERS
        if user_id in PENDING_USERS:
            PENDING_USERS[user_id]['xm_verified'] = True
            PENDING_USERS[user_id]['status'] = 'xm_approved'
            save_pending_users()
        
        # Update the admin's message
        bot.edit_message_text(
            f"âœ… *XM Partnership APPROVED*\n\n"
            f"You have approved @{username}'s XM partnership verification."
            f"{' User has 14-day grace period for deposit.' if grace_period else ''}\n"
            f"They will now be prompted to complete registration forms.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        
        bot.answer_callback_query(call.id, "XM Partnership approved")
        
        # Skip forms for renewals
        if is_renewal:
            # For renewals, just send confirmation message
            bot.send_message(
                user_id,
                "âœ… *XM Partnership Renewal Approved!*\n\n"
                "Your XM partnership verification has been approved and your membership has been extended.\n\n"
                f"Your new expiration date is: {due_date.strftime('%Y-%m-%d')}\n\n"
                f"{'You have a 14-day grace period to make your deposit of $30 minimum.' if grace_period else ''}\n\n"
                "Thank you for continuing with Prodigy Trading Academy!",
                parse_mode="Markdown"
            )
            
            # Log that we skipped forms for renewal
            logging.info(f"Skipped forms for renewal user {user_id}")
        else:
            # For new users, prompt to complete registration forms
            markup = InlineKeyboardMarkup(row_width=2)
            markup.add(
                InlineKeyboardButton("âœ… Yes, proceed", callback_data="xm_registration_yes"),
                InlineKeyboardButton("âŒ No, later", callback_data="xm_registration_no")
            )
            
            bot.send_message(
                user_id,
                "âœ… *XM Partnership Approved!*\n\n"
                "Congratulations! Your XM partnership has been verified and approved.\n"
                f"{'You have a 14-day grace period to make your deposit of $30 minimum.' if grace_period else ''}\n\n"
                "ðŸ“ *Registration Forms*\n\n"
                "Would you like to proceed with completing the PTA registration forms now?",
                parse_mode="Markdown",
                reply_markup=markup
            )
        
    except Exception as e:
        bot.answer_callback_query(call.id, f"Error: {str(e)}", show_alert=True)
        logging.error(f"Error in XM approval: {e}")

@bot.callback_query_handler(func=lambda call: call.data.startswith("reject_xm_"))
def handle_reject_xm(call):
    """Handle admin rejection of XM partnership"""
    admin_id = call.from_user.id
    
    # Verify admin status
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to use this action.")
        return
    
    # Extract user ID from callback data
    user_id = int(call.data.split("_")[2])
    
    try:
        # Get user info for notifications
        try:
            user_info = bot.get_chat(user_id)
            username = user_info.username or "No Username"
        except Exception:
            username = f"User {user_id}"
        
        # Update the admin's message
        bot.edit_message_text(
            f"âŒ *XM Partnership REJECTED*\n\n"
            f"You have rejected @{username}'s XM partnership verification.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        
        bot.answer_callback_query(call.id, "XM Partnership rejected")
        
        # Notify the user of rejection
        bot.send_message(
            user_id,
            "âŒ *XM Partnership Rejected*\n\n"
            "Unfortunately, your XM partnership verification has been rejected.\n\n"
            "Please ensure you've correctly followed all the registration instructions and provided valid information. "
            "You may contact an admin for further assistance.",
            parse_mode="Markdown"
        )
        
        # Clean up PENDING_USERS
        if user_id in PENDING_USERS:
            PENDING_USERS.pop(user_id)
            delete_pending_user(user_id)
        
    except Exception as e:
        bot.answer_callback_query(call.id, f"Error: {str(e)}", show_alert=True)
        logging.error(f"Error in XM rejection: {e}")

@bot.callback_query_handler(func=lambda call: call.data == "xm_registration_yes")
def handle_xm_registration_yes(call):
    """Handle user choosing to proceed with registration forms"""
    user_id = call.from_user.id
    
    # Answer callback
    bot.answer_callback_query(call.id, "Processing...")
    
    # Update user status to indicate they're completing the registration forms
    PENDING_USERS[user_id]['status'] = 'onboarding_form_regular_step1'
    save_pending_users()
    
    # Proceed directly to onboarding forms without sending another verification
    bot.send_message(
        call.message.chat.id,
        "ðŸ“ *Registration Forms*\n\n"
        "Thank you for choosing to complete the registration forms now. "
        "Let's get started with your onboarding process.",
        parse_mode="Markdown"
    )
    
    # Add a small delay for better UX
    time.sleep(1)
    
    # Send registration form graphic if available
    try:
        with open('graphics/registration_form.jpeg', 'rb') as form_img:
            bot.send_photo(
                user_id,
                form_img,
                caption="Please complete the registration form to continue"
            )
    except FileNotFoundError:
        logging.error("Registration form image not found at graphics/registration_form.jpeg")
    except Exception as e:
        logging.error(f"Error sending registration form image: {e}")
    
    # Add another small delay
    time.sleep(1.5)
    
    # Start the onboarding form process
    send_onboarding_form(user_id)

@bot.callback_query_handler(func=lambda call: call.data == "xm_registration_no")
def handle_xm_registration_no(call):
    """Handle user choosing to delay registration forms for trial period"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    
    # Answer callback
    bot.answer_callback_query(call.id, "Processing...")
    
    # Calculate trial end date (1 week from now)
    trial_end_date = datetime.now() + timedelta(days=7)
    
    # Update user's payment data to indicate they need to complete forms
    user_id_str = str(user_id)
    if user_id_str in PAYMENT_DATA:
        PAYMENT_DATA[user_id_str]['forms_needed'] = True
        PAYMENT_DATA[user_id_str]['trial_end_date'] = trial_end_date.strftime('%Y-%m-%d %H:%M:%S')
        save_payment_data()
        
    # Update user status
    PENDING_USERS[user_id]['status'] = 'form_completion_pending'
    save_pending_users()
    
    # Edit the message to show trial information
    if PENDING_USERS[user_id].get('grace_period', False):
        # User selected "No, deposit later" option
        bot.edit_message_text(
            "â³ *Trial Access Granted*\n\n"
            f"You now have 1-week trial access until {trial_end_date.strftime('%Y-%m-%d')}.\n\n"
            f"For a smooth onboarding experience, please complete your Deposit before your trial expires."
            f"You'll find a 'Finish Forms' button in the main menu.\n\n"
            f"â€¼ï¸ *Get Verified Before Your Trial Access Expires*\n\n"
            f"Here is what you get by becoming a âœ… Verified PTA Student:\n\n"
            f"Deposit\n"
            f"âš¡ï¸ Full access to our Mentorship",
            chat_id,
            message_id,
            parse_mode="Markdown"
        )
    else:
        # User selected deposit option
        bot.edit_message_text(
            "â³ *Trial Access Granted*\n\n"
            f"You now have 1-week trial access until {trial_end_date.strftime('%Y-%m-%d')}.\n\n"
            f"For a smooth onboarding experience, please complete your registration forms before your trial expires. "
            f"You'll find a 'Finish Forms' button in the main menu.\n\n"
            f"â€¼ï¸ *Get Verified Before Your Trial Access Expires*\n\n"
            f"Our registration only takes a few minutes or exactly 5 simple questions! Here is what you get by becoming a âœ… Verified PTA Student:\n\n"
            f"ðŸ“ Certificate of Completion (Registration Forms)\n"
            f"ðŸ“ Certificate of Enrollment\n"
            f"âš¡ï¸ Full access to our Mentorship",
            chat_id,
            message_id,
            parse_mode="Markdown"
        )
    
    # Generate an invite link for the appropriate group
    target_group_id = PENDING_USERS[user_id].get('target_group_id', PAID_GROUP_ID)
    try:
        # Check if user is already in the group
        already_in_group = False
        try:
            chat_member = bot.get_chat_member(target_group_id, user_id)
            already_in_group = True
        except ApiException:
            already_in_group = False
        
        if not already_in_group:
            # Generate a new invite link
            new_invite = bot.create_chat_invite_link(
                target_group_id,
                name=f"User {user_id} trial access",
                creates_join_request=False,
                member_limit=1,
                expire_date=int((datetime.now() + timedelta(minutes=15)).timestamp())
            )
            invite_link = new_invite.invite_link
            
            # Send invite link to user
            bot.send_message(
                user_id,
                f"ðŸ”— *Join Our Community*\n\n"
                f"Here's your invitation to join the group: {invite_link}\n\n"
                f"This link will expire in 15 minutes.",
                parse_mode="Markdown"
            )
            
            # Revoke link after 15 minutes
            def revoke_link_later(chat_id, invite_link):
                time.sleep(900)  # 15 minutes
                try:
                    bot.revoke_chat_invite_link(chat_id, invite_link)
                except Exception as e:
                    logging.error(f"Error revoking invite link: {e}")
                    
            threading.Thread(target=revoke_link_later, args=(target_group_id, invite_link)).start()
        else:
            bot.send_message(
                user_id,
                "âœ… *You're Already a Member*\n\n"
                "You're already in our community group, so no need to join again.",
                parse_mode="Markdown"
            )
    except Exception as e:
        logging.error(f"Error generating group invite: {e}")
        bot.send_message(
            user_id,
            "âš ï¸ *Error Generating Invite*\n\n"
            "There was a problem generating your group invite link. Please contact an admin for assistance.",
            parse_mode="Markdown"
        )
    
    # Show main menu (which will now have the "Finish Forms" button)
    show_main_menu(chat_id, user_id)

@bot.callback_query_handler(func=lambda call: call.data.startswith("approve_xm_trial_"))
def callback_approve_xm_trial(call):
    """Handle admin approval of XM trial access"""
    user_id = int(call.data.split("_")[3])
    admin_id = call.from_user.id
    
    # Verify admin status
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to use this action.")
        return
    
    try:
        # Get user info for notifications
        try:
            user_info = bot.get_chat(user_id)
            username = user_info.username or "No Username"
            if username != "No Username":
                username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)
        except Exception:
            username = f"User {user_id}"
            
        # Get trial end date from PENDING_USERS
        trial_end_date_str = PENDING_USERS[user_id].get('trial_end_date')
        if trial_end_date_str:
            trial_end_date = datetime.strptime(trial_end_date_str, '%Y-%m-%d %H:%M:%S')
        else:
            # Default to 1 week if not found
            trial_end_date = datetime.now() + timedelta(days=7)
            
        # Calculate reminder date (2 days before expiry)
        reminder_date = trial_end_date - timedelta(days=2)
        
        # Update payment data for trial access
        if str(user_id) in PAYMENT_DATA:
            # Update existing data
            PAYMENT_DATA[str(user_id)].update({
                "username": username,
                "payment_plan": "Trial",
                "payment_mode": "XM Partnership",
                "mentorship_type": "Regular",
                "due_date": trial_end_date.strftime('%Y-%m-%d %H:%M:%S'),
                "haspayed": True,
                "xm_verified": True,
                "xm_verified_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "xm_verified_by": admin_id,
                "xm_trial": True,
                "forms_needed": True,
                "reminder_date": reminder_date.strftime('%Y-%m-%d %H:%M:%S')
            })
        else:
            # Create new entry
            PAYMENT_DATA[str(user_id)] = {
                "username": username,
                "payment_plan": "Trial",
                "payment_mode": "XM Partnership",
                "mentorship_type": "Regular",
                "due_date": trial_end_date.strftime('%Y-%m-%d %H:%M:%S'),
                "haspayed": True,
                "xm_verified": True, 
                "xm_verified_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "xm_verified_by": admin_id,
                "terms_accepted": True,
                "privacy_accepted": True,
                "xm_trial": True,
                "forms_needed": True,
                "reminder_date": reminder_date.strftime('%Y-%m-%d %H:%M:%S')
            }
        
        save_payment_data()
        
        # Update PENDING_USERS status
        PENDING_USERS[user_id]['status'] = 'xm_trial_approved'
        save_pending_users()
        
        # Notify admins
        admin_username = call.from_user.username or f"Admin {admin_id}"
        for admin in ADMIN_IDS:
            if admin != admin_id:  # Don't notify the admin who took action
                bot.send_message(
                    admin,
                    f"ðŸ“ *Activity Log*\n\n@{admin_username} has approved trial access for @{username}.",
                    parse_mode="Markdown"
                )
        
        # Edit admin message
        bot.edit_message_text(
            f"âœ… *XM Trial Access APPROVED*\n\n"
            f"You have granted @{username} trial access until {trial_end_date.strftime('%Y-%m-%d')}.\n\n"
            f"They will be reminded 2 days before expiration to complete their registration forms.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        bot.answer_callback_query(call.id, "âœ… Trial access approved")
        
        # Send invite link to the user
        # Check if user is already in the group
        target_group_id = PAID_GROUP_ID  # Trial is for regular membership only
        already_in_group = False
        
        try:
            chat_member = bot.get_chat_member(target_group_id, user_id)
            already_in_group = True
            logging.info(f"User {user_id} is already in group {target_group_id}, will skip invite")
        except ApiException:
            already_in_group = False
            logging.info(f"User {user_id} is not in group {target_group_id}, will generate invite")
        
        if not already_in_group:
            try:
                # Generate a new invite link
                new_invite = bot.create_chat_invite_link(
                    target_group_id,
                    name=f"User {user_id} trial access",
                    creates_join_request=False,
                    member_limit=1,
                    expire_date=int((datetime.now() + timedelta(minutes=15)).timestamp())
                )
                invite_link = new_invite.invite_link
                
                # Send approval and invite to user
                bot.send_message(
                    user_id,
                    f"âœ… *Trial Access Approved!*\n\n"
                    f"You have been granted trial access to Prodigy Trading Academy until {trial_end_date.strftime('%Y-%m-%d')}.\n\n"
                    f"Please join our community here: {invite_link}\n\n"
                    f"âš ï¸ Important: You will need to complete the registration forms before your trial expires.",
                    parse_mode="Markdown"
                )
                
                # Revoke link after 15 minutes
                def revoke_link_later(chat_id, invite_link):
                    time.sleep(900)  # 15 minutes
                    try:
                        bot.revoke_chat_invite_link(chat_id, invite_link)
                    except Exception as e:
                        logging.error(f"Error revoking invite link: {e}")
                        
                threading.Thread(target=revoke_link_later, args=(target_group_id, invite_link)).start()
                
            except Exception as e:
                logging.error(f"Error generating invite link: {e}")
                bot.send_message(
                    user_id,
                    f"âœ… *Trial Access Approved!*\n\n"
                    f"You have been granted trial access to Prodigy Trading Academy until {trial_end_date.strftime('%Y-%m-%d')}.\n\n"
                    f"Please contact an admin to get access to the group.\n\n"
                    f"âš ï¸ Important: You will need to complete the registration forms before your trial expires.",
                    parse_mode="Markdown"
                )
        else:
            # User is already in the group
            bot.send_message(
                user_id,
                f"âœ… *Trial Access Approved!*\n\n"
                f"You have been granted trial access to Prodigy Trading Academy until {trial_end_date.strftime('%Y-%m-%d')}.\n\n"
                f"You're already a member of our group, so no need to join again.\n\n"
                f"âš ï¸ Important: You will need to complete the registration forms before your trial expires.",
                parse_mode="Markdown"
            )
        
    except Exception as e:
        logging.error(f"Error in XM trial approval: {e}")
        bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}")
        bot.send_message(call.message.chat.id, f"Error processing XM trial approval: {e}")

def check_trial_reminders():
    """Check for users needing form completion reminders"""
    while True:
        try:
            now = datetime.now()
            
            for user_id_str, data in PAYMENT_DATA.items():
                # Check if this is a trial account needing forms
                if data.get('xm_trial') and data.get('forms_needed'):
                    # Check if it's time for the reminder
                    reminder_date_str = data.get('reminder_date')
                    
                    if reminder_date_str:
                        reminder_date = datetime.strptime(reminder_date_str, '%Y-%m-%d %H:%M:%S')
                        
                        # If it's time to send the reminder (current time >= reminder time)
                        if now >= reminder_date:
                            try:
                                user_id = int(user_id_str)
                                
                                # Get due date for message
                                due_date = datetime.strptime(data.get('due_date', '2099-01-01 00:00:00'), '%Y-%m-%d %H:%M:%S')
                                days_remaining = (due_date - now).days
                                
                                # Create registration buttons
                                markup = InlineKeyboardMarkup(row_width=1)
                                markup.add(InlineKeyboardButton("ðŸ“ Complete Registration Forms", callback_data="start_xm_forms"))
                                
                                # Send reminder
                                bot.send_message(
                                    user_id,
                                    f"âš ï¸ *Trial Expiring Soon!*\n\n"
                                    f"Your trial access expires in {days_remaining} days.\n\n"
                                    f"To maintain your access to Prodigy Trading Academy, you need to complete the registration forms.",
                                    parse_mode="Markdown",
                                    reply_markup=markup
                                )
                                
                                # Update to prevent sending reminder again
                                PAYMENT_DATA[user_id_str]['reminder_sent'] = True
                                save_payment_data()
                                
                            except Exception as e:
                                logging.error(f"Error sending trial reminder to user {user_id_str}: {e}")
            
            # Check every hour
            time.sleep(3600)
            
        except Exception as e:
            logging.error(f"Error in trial reminder checker: {e}")
            time.sleep(3600)  # Sleep and try again in an hour

@bot.callback_query_handler(func=lambda call: call.data == "start_xm_forms")
def start_xm_forms(call):
    """Start registration forms from trial reminder"""
    user_id = call.from_user.id
    
    # Answer callback
    bot.answer_callback_query(call.id, "Starting registration forms...")
    
    # Send registration form graphic
    try:
        with open('graphics/registration_form.jpeg', 'rb') as form_img:
            bot.send_photo(
                user_id,
                form_img,
                caption="Please complete the registration form to continue"
            )
    except FileNotFoundError:
        logging.error("Registration form image not found at graphics/registration_form.jpeg")
    except Exception as e:
        logging.error(f"Error sending registration form image: {e}")
    
    # Add a small delay for better UX
    time.sleep(1.5)
    
    # Start onboarding forms
    send_onboarding_form(user_id)

@bot.callback_query_handler(func=lambda call: call.data == "xm_comply_yes")
def handle_xm_comply_yes(call):
    """Handle user agreeing to add partner code"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    
    # Answer callback
    bot.answer_callback_query(call.id, "Processing...")
    
    # Update user status to await screenshot
    PENDING_USERS[user_id]['status'] = 'xm_awaiting_screenshot'
    save_pending_users()
    
    # Create back button
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(InlineKeyboardButton("â¬…ï¸ Back", callback_data="xm_existing_yes"))
    
    # Edit message to request screenshot
    bot.edit_message_text(
        "ðŸ“¸ *XM Partnership Verification*\n\n"
        "Please send a screenshot showing that you have applied the partner code *PTAPARTNER* in your XM account.\n\n"
        "The screenshot should clearly show:\n"
        "â€¢ Your XM account ID\n"
        "â€¢ The partner code field with our code\n\n"
        "Please upload your screenshot now.",
        chat_id,
        message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.callback_query_handler(func=lambda call: call.data == "xm_comply_no")
def handle_xm_comply_no(call):
    """Handle user not agreeing to add partner code"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    
    # Answer callback
    bot.answer_callback_query(call.id, "Processing...")
    
    # Create back to main menu button
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
    
    # Edit message to show inability to proceed
    bot.edit_message_text(
        "âŒ *Process Terminated*\n\n"
        "To qualify for free mentorship, you must add our partner code to your XM account.\n\n"
        "If you change your mind, you can restart the process from the main menu.",
        chat_id,
        message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.callback_query_handler(func=lambda call: call.data.startswith("approve_xm_"))
def callback_approve_xm(call):
    """Handle admin approval of XM verification"""
    user_id = int(call.data.split("_")[2])
    admin_id = call.from_user.id
    
    # Verify admin status
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to use this action.")
        return
    
    try:
        # Get user info for notifications
        try:
            user_info = bot.get_chat(user_id)
            username = user_info.username or "No Username"
            if username != "No Username":
                username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)
        except Exception:
            username = f"User {user_id}"
            
        # Calculate membership due date (1 month for XM partnership)
        due_date = datetime.now() + timedelta(days=30)
        
        # Update payment data - note it's free but still tracked
        if str(user_id) in PAYMENT_DATA:
            # Update existing record
            PAYMENT_DATA[str(user_id)].update({
                "username": username,
                "payment_plan": "Monthly",
                "payment_mode": "XM Partnership",
                "mentorship_type": "Regular",  # XM partnership is only for Regular membership
                "due_date": due_date.strftime('%Y-%m-%d %H:%M:%S'),
                "haspayed": True,
                "xm_verified": True,  # Add flag for XM verification
                "xm_verified_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "xm_verified_by": admin_id
            })
        else:
            # Create new entry
            PAYMENT_DATA[str(user_id)] = {
                "username": username,
                "payment_plan": "Monthly",
                "payment_mode": "XM Partnership",
                "mentorship_type": "Regular",  # XM partnership is only for Regular membership
                "due_date": due_date.strftime('%Y-%m-%d %H:%M:%S'),
                "haspayed": True,
                "xm_verified": True,  # Add flag for XM verification
                "xm_verified_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "xm_verified_by": admin_id,
                "terms_accepted": True,
                "privacy_accepted": True
            }
        
        logging.info(f"XM verification approved for user {user_id}")
        save_payment_data()
        
        # Remove user from pending status
        PENDING_USERS.pop(user_id, None)
        delete_pending_user(user_id)
        
        # Notify admins of the approval
        admin_username = call.from_user.username or f"Admin {admin_id}"
        for admin in ADMIN_IDS:
            if admin != admin_id:  # Don't notify the admin who took action
                bot.send_message(
                    admin,
                    f"ðŸ“ *Activity Log*\n\n@{admin_username} has approved XM verification for @{username}.",
                    parse_mode="Markdown"
                )
        
        # Update message to confirming approval
        bot.edit_message_text(
            f"âœ… *XM Verification APPROVED*\n\n"
            f"You have approved @{username}'s XM partnership verification.\n\n"
            f"They will now receive free access to Regular mentorship.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        bot.answer_callback_query(call.id, "âœ… XM verification approved")
        
        # Now notify the user and start onboarding process
        # First send success graphic if available
        try:
            with open('graphics/xm_verified.jpeg', 'rb') as approval_img:
                bot.send_photo(
                    user_id,
                    approval_img,
                    caption="Your XM Partnership verification has been approved!"
                )
        except FileNotFoundError:
            logging.error("XM approval image not found at graphics/xm_verified.jpeg")
            bot.send_message(
                user_id,
                "âœ… *XM Verification Successful!*\n\n"
                "Your XM account verification has been approved. Welcome to the Prodigy Trading Academy!",
                parse_mode="Markdown"
            )
        except Exception as e:
            logging.error(f"Error sending XM approval image: {e}")
            bot.send_message(
                user_id,
                "âœ… *XM Verification Successful!*\n\n"
                "Your XM account verification has been approved. Welcome to the Prodigy Trading Academy!",
                parse_mode="Markdown"
            )
        
        # Add a small delay for better UX
        time.sleep(1.5)
        
        # Send membership details message
        bot.send_message(
            user_id, 
            f"ðŸ“… *Your free mentorship is active!*\n\n"
            f"â€¢ Membership Type: Regular Mentorship\n"
            f"â€¢ Duration: 1 Month (Free through XM Partnership)\n"
            f"â€¢ Expiration: {due_date.strftime('%Y-%m-%d')}\n\n"
            f"To maintain your free membership, you must keep your XM account active with our partner code.\n"
            f"Monthly verification may be required.",
            parse_mode="Markdown"
        )
        
        # Send registration form graphic first
        try:
            with open('graphics/registration_form.jpeg', 'rb') as form_img:
                bot.send_photo(
                    user_id,
                    form_img,
                    caption="Please complete the registration form to continue"
                )
        except FileNotFoundError:
            logging.error("Registration form image not found at graphics/registration_form.jpeg")
        
        # Add a small delay for better UX
        time.sleep(1.5)
            
        # Start onboarding process
        target_group_id = PAID_GROUP_ID  # XM partnership is for regular mentorship only
        
        # Check if user is already in the group
        already_in_group = False
        try:
            chat_member = bot.get_chat_member(target_group_id, user_id)
            already_in_group = True
            logging.info(f"User {user_id} is already in group {target_group_id}, will skip invite during onboarding")
        except ApiException:
            already_in_group = False
            logging.info(f"User {user_id} is not in group {target_group_id}, will generate invite during onboarding")
            
        # Setup for onboarding form
        PENDING_USERS[user_id] = {
            'form_answers': {},  # Initialize empty dict to store responses
            'invite_link': None,  # Will store the invite link to use after form completion
            'target_group_id': target_group_id,
            'already_in_group': already_in_group,
            'membership_type': "Regular"  # XM partnership is for regular mentorship only
        }
        save_pending_users()
        
        # Start onboarding form process
        send_onboarding_form(user_id)
        
    except Exception as e:
        logging.error(f"Error in XM verification approval: {e}")
        bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}")
        bot.send_message(call.message.chat.id, f"Error processing XM verification: {e}")

@bot.callback_query_handler(func=lambda call: call.data.startswith("reject_xm_"))
def callback_reject_xm(call):
    """Handle admin rejection of XM verification"""
    user_id = int(call.data.split("_")[2])
    admin_id = call.from_user.id
    
    # Verify admin status
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to use this action.")
        return
    
    try:
        # Get user info for notifications
        try:
            user_info = bot.get_chat(user_id)
            username = user_info.username or "No Username"
            if username != "No Username":
                username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)
        except Exception:
            username = f"User {user_id}"
        
        # Remove user from pending status to allow them to try again
        PENDING_USERS.pop(user_id, None)
        delete_pending_user(user_id)
        
        # Notify admins of the rejection
        admin_username = call.from_user.username or f"Admin {admin_id}"
        for admin in ADMIN_IDS:
            if admin != admin_id:  # Don't notify the admin who took action
                bot.send_message(
                    admin,
                    f"ðŸ“ *Activity Log*\n\n@{admin_username} has rejected XM verification for @{username}.",
                    parse_mode="Markdown"
                )
        
        # Update message confirming rejection
        bot.edit_message_text(
            f"âŒ *XM Verification REJECTED*\n\n"
            f"You have rejected @{username}'s XM partnership verification.\n\n"
            f"They have been notified and can try again if needed.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        bot.answer_callback_query(call.id, "âŒ XM verification rejected")
        
        # Notify user of rejection with instructions to try again
        bot.send_message(
            user_id,
            "âŒ *XM Verification Declined*\n\n"
            "We couldn't verify your XM partnership code from the screenshot provided.\n\n"
            "Possible reasons:\n"
            "â€¢ The partner code wasn't visible or didn't match PTAPARTNER\n"
            "â€¢ Account details weren't clearly shown\n"
            "â€¢ Image quality was too low\n\n"
            "You can try again by returning to the main menu (/start) and selecting 'Free Mentorship by XM' again.",
            parse_mode="Markdown"
        )
        
    except Exception as e:
        logging.error(f"Error in XM verification rejection: {e}")
        bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}")
        bot.send_message(call.message.chat.id, f"Error processing XM rejection: {e}")

@bot.callback_query_handler(func=lambda call: call.data == "menu_aichat")
def handle_menu_aichat(call):
    """Start AI chat from the main menu"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    
    bot.answer_callback_query(call.id, "Starting AI chat...")
    
    # Edit the message to show AI chat intro
    bot.edit_message_text(
        "ðŸ¤– *AI Assistant Mode Activated*\n\n"
        "âœ¨ I'm here to help with your trading questions! Just type your question and I'll respond.\n\n"
        "ðŸ’¬ You can ask about:\n"
        "â€¢ ðŸ“Š Trading strategies and analysis\n"
        "â€¢ ðŸ§  Trading psychology\n"
        "â€¢ ðŸ› ï¸ How to use this bot\n"
        "â€¢ ðŸ“š PTA resources and membership\n\n"
        "When you're done, simply type `/exit` to end our conversation.",
        chat_id,
        message_id,
        parse_mode="Markdown",
        reply_markup=None  # Remove buttons
    )
    
    # Set user state to conversation_active (directly going to conversation mode)
    PENDING_USERS[call.from_user.id] = {'status': 'conversation_active'}
    save_pending_users()

def remove_inline_keyboard(chat_id, message_id):
    """Remove the inline keyboard from a message to prevent further clicks"""
    try:
        bot.edit_message_reply_markup(chat_id, message_id, reply_markup=None)
        return True
    except Exception as e:
        logging.error(f"Error removing inline keyboard: {e}")
        return False

@bot.callback_query_handler(func=lambda call: call.data.startswith("menu_"))
def handle_main_menu_selection(call):
    """Handle main menu inline button selections with editing instead of new messages"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    option = call.data.split("_")[1]
    
    # Set a default pending status
    PENDING_USERS[user_id] = {'status': 'choosing_option'}
    save_pending_users()
    
    # Handle different menu options
    if option == "buy":
        # Answer callback for feedback to user
        bot.answer_callback_query(call.id, "Starting membership purchase...")
        
        # Check enrollment status for new purchases
        regular_enrollment_open = BOT_SETTINGS.get('regular_enrollment_open', True)
        supreme_enrollment_open = BOT_SETTINGS.get('supreme_enrollment_open', True)
        
        # If both enrollment types are closed, edit message to show notification
        if not regular_enrollment_open and not supreme_enrollment_open:
            markup = InlineKeyboardMarkup(row_width=2)
            markup.add(
                InlineKeyboardButton("ðŸ”” Get Notified", callback_data="update_yes"),
                InlineKeyboardButton("â“ FAQ", callback_data="faq_back")
            )
            markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
            
            bot.edit_message_text(
                "âš ï¸ *Enrollment is currently closed*\n\n"
                "New memberships are not available at this time. Please wait for the next "
                "announcement from the admins about when enrollment will open again.\n\n"
                "â€¢ Click *Get Notified* to receive updates when enrollment opens\n"
                "â€¢ Check our *FAQ* section for more information\n\n"
                "Thank you for your interest in Prodigy Trading Academy!",
                chat_id,
                message_id,
                parse_mode="Markdown",
                reply_markup=markup
            )
            return
        
        # Update status and edit message to show mentorship options
        PENDING_USERS[chat_id]['status'] = 'choosing_mentorship_type'
        PENDING_USERS[chat_id]['is_renewal'] = False
        save_pending_users()
        
        # Show mentorship type options with inline keyboard
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineKeyboardButton("Regular Mentorship (FREE)", callback_data="menu_xm_free"),
            InlineKeyboardButton("Supreme Mentorship", callback_data="mentorship_supreme")
        )
        markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
        
        # Edit the current message instead of sending a new one
        bot.edit_message_text(
            "Please select your preferred mentorship level:",
            chat_id,
            message_id,
            reply_markup=markup
        )
        
    elif option == "renew":
        bot.answer_callback_query(call.id, "Starting membership renewal...")
        
        # Check if they can renew
        can_renew, message_text = can_renew_membership(user_id)
        if not can_renew:
            # Create back button
            markup = InlineKeyboardMarkup(row_width=1)
            markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
            
            # Edit message to show error with back button
            bot.edit_message_text(
                message_text,
                chat_id,
                message_id,
                reply_markup=markup
            )
            return
        
        # Update status and edit message to show mentorship options
        PENDING_USERS[chat_id]['status'] = 'choosing_mentorship_type'
        PENDING_USERS[chat_id]['is_renewal'] = True
        save_pending_users()
        
        # Show mentorship type options with inline keyboard
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineKeyboardButton("Regular Mentorship (FREE)", callback_data="menu_xm_free"),
            InlineKeyboardButton("Supreme Mentorship", callback_data="mentorship_supreme")
        )
        markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
        
        # Edit the message
        bot.edit_message_text(
            "Please select your preferred mentorship level to renew:",
            chat_id,
            message_id,
            reply_markup=markup
        )
        
    elif option == "cancel":
        bot.answer_callback_query(call.id, "Starting membership cancellation...")
        
        # Check if user has an active membership
        if str(user_id) not in PAYMENT_DATA:
            markup = InlineKeyboardMarkup(row_width=1)
            markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
            
            bot.edit_message_text(
                "âŒ You don't have an active membership to cancel.",
                chat_id,
                message_id,
                reply_markup=markup
            )
            return
            
        # Check if already cancelled
        if str(user_id) in PAYMENT_DATA and PAYMENT_DATA[str(user_id)].get('cancelled', False):
            due_date = PAYMENT_DATA[str(user_id)].get('due_date', 'Unknown')
            markup = InlineKeyboardMarkup(row_width=1)
            markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
            
            try:
                due_date_obj = datetime.strptime(due_date, '%Y-%m-%d %H:%M:%S')
                days_remaining = (due_date_obj - datetime.now()).days
                
                bot.edit_message_text(
                    f"â„¹ï¸ Your membership is already cancelled.\n\n"
                    f"You will still have access until {due_date_obj.strftime('%Y-%m-%d')} "
                    f"({days_remaining} days remaining).\n\n"
                    f"If you wish to reactivate your membership, please use /start and select 'Renew Membership'.",
                    chat_id,
                    message_id,
                    reply_markup=markup
                )
            except Exception as e:
                # Fallback if date parsing fails
                bot.edit_message_text(
                    "â„¹ï¸ Your membership is already cancelled. You will still have access until the next payment cycle.",
                    chat_id,
                    message_id,
                    reply_markup=markup
                )
            return

        # Update user status and show confirmation buttons
        PENDING_USERS[chat_id]['status'] = 'cancel_membership'
        save_pending_users()
        
        # Use InlineKeyboardMarkup for Yes/No confirmation
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineKeyboardButton("Yes", callback_data="cancel_confirm_yes"),
            InlineKeyboardButton("No", callback_data="cancel_confirm_no")
        )
        markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
        
        bot.edit_message_text(
            "Are you sure you want to cancel your membership? You will still have access until next month/year, but you will not be charged. Please confirm.", 
            chat_id,
            message_id,
            reply_markup=markup
        )
        
    elif option == "dashboard":
        # Show user dashboard
        bot.answer_callback_query(call.id, "Loading your dashboard...")
            
        # Create a modified version of the message with the correct user info
        modified_message = call.message
        # Important: Set the correct user info from the callback
        modified_message.from_user = call.from_user
            
        # Now call the dashboard function with the corrected message object
        show_user_dashboard(modified_message)
        
    elif option == "faq":
        # Show FAQ options
        bot.answer_callback_query(call.id, "Loading FAQ...")
        
        # Show the FAQ categories with inline buttons
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(
            InlineKeyboardButton("ðŸŽ“ Join Academy", callback_data="faq_join_academy"),
            InlineKeyboardButton("ðŸ“… Opening of Admissions", callback_data="faq_admissions"),
            InlineKeyboardButton("ðŸ’² Mentorship Plans & Pricings", callback_data="faq_plans_pricing"),
            InlineKeyboardButton("ðŸ“¦ Products & Services", callback_data="faq_products_services"),
            InlineKeyboardButton("ðŸŒŸ Benefits in Enrollment", callback_data="faq_enrollment_benefits"),
            InlineKeyboardButton("ðŸ“ Terms & Conditions", callback_data="faq_terms"),
            InlineKeyboardButton("ðŸ”’ Privacy Policy", callback_data="faq_privacy")
        )
        markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
        
        bot.edit_message_text(
            "ðŸ” *Frequently Asked Questions*\n\n"
            "Select a category to view related questions:",
            chat_id,
            message_id,
            reply_markup=markup,
            parse_mode="Markdown"
        )

    elif option == "redeem":
        bot.answer_callback_query(call.id, "Starting serial redemption...")
        
        # Update user status
        PENDING_USERS[user_id] = {'status': 'awaiting_serial'}
        save_pending_users()
        
        # Create back button
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
        
        # Edit message to prompt for serial number
        bot.edit_message_text(
            "ðŸ”‘ *Serial Redemption*\n\n"
            "Please enter the serial number you received for your membership.\n\n"
            "Note: Serial numbers are case-sensitive.",
            chat_id,
            message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
        
    elif option == "howto":
        # Show "How to Use" information
        bot.answer_callback_query(call.id, "Loading usage guide...")
        
        # Create buttons with AI help option
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(
            InlineKeyboardButton("ðŸ¤– Ask AI Assistant for Help", callback_data="menu_aichat"),
            InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu")
        )
        
        bot.edit_message_text(
            "ðŸ“– *How to Use Prodigy Trading Academy Bot*\n\n"
            "Welcome to our academy bot! Here's a quick guide to get you started:\n\n"
            "1ï¸âƒ£ *Buy Membership*: Follow these steps to join our community:\n"
            "   â€¢ Select 'Buy Membership' from main menu\n"
            "   â€¢ Choose Regular or Supreme membership\n"
            "   â€¢ Select your preferred plan\n"
            "   â€¢ Complete payment using your preferred method\n"
            "   â€¢ Submit verification proof\n\n"
            "2ï¸âƒ£ *Redeem Serial*: Have a serial number?\n"
            "   â€¢ Select 'Redeem Serial' from main menu\n"
            "   â€¢ Enter your serial number when prompted\n"
            "   â€¢ Get instant access to your membership\n\n"
            "3ï¸âƒ£ *Using AI Assistant*: Need help or have questions?\n"
            "   â€¢ Press the 'Ask AI Assistant' button below\n"
            "   â€¢ Ask any question about trading or our academy\n"
            "   â€¢ Type /exit when done to return to main menu\n\n"
            "4ï¸âƒ£ *Common Commands*:\n"
            "   â€¢ /start - Return to the main menu\n"
            "   â€¢ /dashboard - View your membership status\n"
            "   â€¢ /chat - Start a conversation with AI\n\n"
            "Still need help? Click 'ðŸ¤– Ask AI Assistant' below or contact our Tech Support at @FujiPTA",
            chat_id,
            message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )

@bot.callback_query_handler(func=lambda call: call.data == "back_to_main_menu")
def back_to_main_menu(call):
    """Return to the main menu from anywhere"""
    bot.answer_callback_query(call.id, "Returning to main menu...")
    
    # Show main menu by editing the current message
    show_main_menu(call.message.chat.id, call.from_user.id, message_id=call.message.message_id)

@bot.callback_query_handler(func=lambda call: call.data.startswith("cancel_confirm_"))
def handle_cancel_confirmation_callback(call):
    """Handle confirmation for membership cancellation"""
    chat_id = call.message.chat.id
    user_id = call.from_user.id
    response = call.data.split("_")[2]  # Either "yes" or "no"
    
    # Check if user has an active membership
    if str(user_id) not in PAYMENT_DATA or not PAYMENT_DATA[str(user_id)].get('haspayed', False):
        bot.answer_callback_query(call.id, "âŒ No active membership found")
        bot.send_message(chat_id, "âŒ You don't have an active membership to cancel.")
        PENDING_USERS.pop(user_id, None)
        delete_pending_user(user_id)
        return

    # Get membership details for better context
    plan = PAYMENT_DATA[str(user_id)].get('payment_plan', 'Unknown')
    due_date = PAYMENT_DATA[str(user_id)].get('due_date', 'Unknown')
    
    if response == "yes":
        # User confirmed cancellation
        bot.answer_callback_query(call.id, "Processing cancellation...")
        PENDING_USERS[user_id]['status'] = 'membership_cancelled'
        save_pending_users()

        # Set cancellation flags in payment data
        PAYMENT_DATA[str(user_id)]['cancelled'] = True
        PAYMENT_DATA[str(user_id)]['reminder_sent'] = True  # Prevent future reminders
        PAYMENT_DATA[str(user_id)]['cancellation_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        save_payment_data()

        # Get user's information
        try:
            username = call.from_user.username or f"User {user_id}"
            username = safe_markdown_escape(username)
        except Exception:
            username = f"User {user_id}"  # Fallback if can't get username
            
        # Send cancellation graphic
        try:
            with open('graphics/confirm_cancel.jpeg', 'rb') as cancellation_img:
                bot.send_photo(
                    chat_id,
                    cancellation_img,
                )
        except FileNotFoundError:
            logging.error("Cancellation image not found at graphics/confirm_cancel.jpeg")
        except Exception as e:
            logging.error(f"Error sending cancellation image: {e}")
            
        time.sleep(1)  # Small delay for better UX

        # Forward the cancellation request to admins with additional context
        for admin in ADMIN_IDS:
            bot.send_message(
                admin, 
                f"ðŸš« *MEMBERSHIP CANCELLATION*\n\n"
                f"ðŸ‘¤ Username: @{username}\n"
                f"ðŸ†” User ID: `{user_id}`\n"
                f"ðŸ“… Plan: {plan}\n"
                f"â° Due date: {due_date}\n"
                f"ðŸ“ Cancelled on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                parse_mode="Markdown"
            )

        # Provide better information to the user
        try:
            due_date_obj = datetime.strptime(due_date, '%Y-%m-%d %H:%M:%S')
            days_remaining = (due_date_obj - datetime.now()).days
            
            bot.send_message(
                chat_id, 
                f"âœ… Your membership is cancelled. You will still have access until {due_date_obj.strftime('%Y-%m-%d')} "
                f"({days_remaining} days remaining), but will not be renewed.\n\n"
                f"Thank you for being with us! If you change your mind before expiration, use /start to reactivate.",
                parse_mode="Markdown"
            )
        except Exception as e:
            # Fallback if date parsing fails
            bot.send_message(chat_id, "âœ… Your membership is cancelled. You will still have access until the next payment cycle, but will not be charged anymore. Thank you for being with us!")
            logging.error(f"Error parsing due date during cancellation: {e}")
    
    elif response == "no":
        # User did not confirm cancellation
        bot.answer_callback_query(call.id, "Cancellation aborted")
        bot.send_message(chat_id, "âŒ No changes have been made to your membership. You will continue with the current payment plan.")
    
    # Clean up pending user data
    PENDING_USERS.pop(user_id, None)  # Remove from dictionary
    delete_pending_user(user_id)  # Remove from MongoDB

@bot.callback_query_handler(func=lambda call: call.data.startswith("mentorship_"))
def handle_mentorship_selection(call):
    """Handle mentorship type selection from inline buttons"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    mentorship_type = call.data.split("_")[1]
    
    # Get renewal status from pending users
    is_renewal = PENDING_USERS[chat_id].get('is_renewal', False)
    
    if mentorship_type == "regular":
        # Process Regular Mentorship selection
        bot.answer_callback_query(call.id, "Selected Regular Mentorship")
        
        # Check enrollment status for Regular membership if this is a new purchase
        if not is_renewal and not BOT_SETTINGS.get('regular_enrollment_open', True):
            # Create inline keyboard with Get Notified and FAQ buttons
            markup = InlineKeyboardMarkup(row_width=2)
            markup.add(
                InlineKeyboardButton("ðŸ”” Get Notified", callback_data="update_yes"),
                InlineKeyboardButton("â“ FAQ", callback_data="faq_back")
            )
            markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
            
            bot.edit_message_text(
                "âš ï¸ *Regular Membership enrollment is currently CLOSED*\n\n"
                "New Regular Membership purchases are temporarily unavailable.\n"
                "Please wait for the next announcement about when enrollment will open again.\n\n"
                "â€¢ Click *Get Notified* to receive updates when enrollment opens\n"
                "â€¢ Check our *FAQ* section for more information\n\n"
                "Existing members can still renew their memberships.",
                chat_id,
                message_id,
                parse_mode="Markdown",
                reply_markup=markup
            )
            return
        
        # Get regular discount specifically
        applicable_discount = DISCOUNTS.get('regular')
        discount_active = applicable_discount and applicable_discount.get('active', False)
        
        # Check if discount applies to this transaction type (new or renewal)
        if discount_active:
            transaction_type = "renewal" if is_renewal else "new"
            discount_transaction_type = applicable_discount.get('transaction_type', 'both')
            
            # Determine if discount applies to this transaction type
            applies_to_transaction = (
                discount_transaction_type == 'both' or 
                discount_transaction_type == transaction_type
            )
            
            # Only consider discount active if it applies to this transaction type
            if not applies_to_transaction:
                discount_active = False
                
        discount_percentage = applicable_discount.get('percentage', 0) if discount_active else 0
        discount_name = applicable_discount.get('name', '') if discount_active else ''
        
        PENDING_USERS[chat_id]['mentorship_type'] = 'regular'
        PENDING_USERS[chat_id]['status'] = 'choosing_mentorship_plan'
        save_pending_users()
        
        # Original prices
        trial_price = 7.99
        momentum_price = 20.99
        legacy_price = 89.99
        
        # Create markup for plan selection
        markup = InlineKeyboardMarkup(row_width=1)
        
        if discount_active:
            # Calculate discounted prices
            trial_discounted = round(trial_price * (1 - discount_percentage / 100), 2)
            momentum_discounted = round(momentum_price * (1 - discount_percentage / 100), 2)
            legacy_discounted = round(legacy_price * (1 - discount_percentage / 100), 2)
            
            # Add buttons with discounted prices
            markup.add(
                InlineKeyboardButton(f"Trial (${trial_discounted:.2f}) / Monthly", callback_data="plan_trial"),
                InlineKeyboardButton(f"Momentum (${momentum_discounted:.2f}) / 3 Months", callback_data="plan_momentum"),
                InlineKeyboardButton(f"Legacy (${legacy_discounted:.2f}) / Yearly", callback_data="plan_regular_legacy")
            )
            
            # ADD BACK BUTTON
            markup.add(InlineKeyboardButton("â¬…ï¸ Go Back", callback_data="back_to_mentorship_type"))
            
            # Customize message based on whether this is a renewal
            intro_text = "Renewal options" if is_renewal else "Please select your Regular Mentorship plan"
            
            # Edit the current message with plans
            bot.edit_message_text( 
                f"{intro_text}:\n\n"
                f"ðŸŽ‰ <b>{discount_name}: {discount_percentage}% OFF!</b>\n\n"
                f"ðŸ’° <b>Trial</b> - <s>${trial_price:.2f}</s> ${trial_discounted:.2f} / Monthly\n"
                f"ðŸ’° <b>Momentum</b> - <s>${momentum_price:.2f}</s> ${momentum_discounted:.2f} / 3 Months\n"
                f"ðŸ’° <b>Legacy</b> - <s>${legacy_price:.2f}</s> ${legacy_discounted:.2f} / Yearly", 
                chat_id,
                message_id,
                reply_markup=markup, 
                parse_mode="HTML"
            )
        else:
            # No discount - show regular prices without strikethrough
            markup.add(
                InlineKeyboardButton(f"Trial (${trial_price:.2f}) / Monthly", callback_data="plan_trial"),
                InlineKeyboardButton(f"Momentum (${momentum_price:.2f}) / 3 Months", callback_data="plan_momentum"),
                InlineKeyboardButton(f"Legacy (${legacy_price:.2f}) / Yearly", callback_data="plan_regular_legacy")
            )
            
            # ADD BACK BUTTON
            markup.add(InlineKeyboardButton("â¬…ï¸ Go Back", callback_data="back_to_mentorship_type"))
            
            # Customize message based on whether this is a renewal
            intro_text = "Renewal options" if is_renewal else "Please select your Regular Mentorship plan"
            
            # Edit the current message
            bot.edit_message_text(
                f"{intro_text}:\n\n"
                f"ðŸ’° <b>Trial</b> - ${trial_price:.2f} / Monthly\n"
                f"ðŸ’° <b>Momentum</b> - ${momentum_price:.2f} / 3 Months\n"
                f"ðŸ’° <b>Legacy</b> - ${legacy_price:.2f} / Yearly", 
                chat_id,
                message_id,
                reply_markup=markup, 
                parse_mode="HTML"
            )
        
    elif mentorship_type == "supreme":
        # Process Supreme Mentorship selection
        bot.answer_callback_query(call.id, "Selected Supreme Mentorship")
        
        # Check enrollment status for Supreme membership if this is a new purchase
        if not is_renewal and not BOT_SETTINGS.get('supreme_enrollment_open', True):
            # Create inline keyboard with Get Notified and FAQ buttons
            markup = InlineKeyboardMarkup(row_width=2)
            markup.add(
                InlineKeyboardButton("ðŸ”” Get Notified", callback_data="update_yes"),
                InlineKeyboardButton("â“ FAQ", callback_data="faq_back")
            )
            markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
            
            # FIXED: Edit the current message instead of sending new one
            bot.edit_message_text(
                "âš ï¸ *Supreme Membership enrollment is currently CLOSED*\n\n"
                "New Supreme Membership purchases are temporarily unavailable.\n"
                "Please wait for the next announcement about when enrollment will open again.\n\n"
                "â€¢ Click *Get Notified* to receive updates when enrollment opens\n"
                "â€¢ Check our *FAQ* section for more information\n\n"
                "Existing members can still renew their memberships.",
                chat_id,
                message_id,
                parse_mode="Markdown",
                reply_markup=markup
            )
            return
        
        # Get supreme discount specifically
        applicable_discount = DISCOUNTS.get('supreme')
        discount_active = applicable_discount and applicable_discount.get('active', False)
        
        # Check if discount applies to this transaction type (new or renewal)
        if discount_active:
            transaction_type = "renewal" if is_renewal else "new"
            discount_transaction_type = applicable_discount.get('transaction_type', 'both')
            
            # Determine if discount applies to this transaction type
            applies_to_transaction = (
                discount_transaction_type == 'both' or 
                discount_transaction_type == transaction_type
            )
            
            # Only consider discount active if it applies to this transaction type
            if not applies_to_transaction:
                discount_active = False
                
        discount_percentage = applicable_discount.get('percentage', 0) if discount_active else 0
        discount_name = applicable_discount.get('name', '') if discount_active else ''
        
        PENDING_USERS[chat_id]['mentorship_type'] = 'supreme'
        PENDING_USERS[chat_id]['status'] = 'choosing_mentorship_plan'
        save_pending_users()
        
        # Original prices
        apprentice_price = 309.99
        disciple_price = 524.99
        legacy_price = 899.99
        
        # Create markup for plan selection
        markup = InlineKeyboardMarkup(row_width=1)
        
        if discount_active:
            # Calculate discounted prices
            apprentice_discounted = round(apprentice_price * (1 - discount_percentage / 100), 2)
            disciple_discounted = round(disciple_price * (1 - discount_percentage / 100), 2)
            legacy_discounted = round(legacy_price * (1 - discount_percentage / 100), 2)
            
            # Add buttons with discounted prices
            markup.add(
                InlineKeyboardButton(f"Apprentice (${apprentice_discounted:.2f}) / 3 Months", callback_data="plan_apprentice"),
                InlineKeyboardButton(f"Disciple (${disciple_discounted:.2f}) / 6 Months", callback_data="plan_disciple"),
                InlineKeyboardButton(f"Legacy (${legacy_discounted:.2f}) / Lifetime", callback_data="plan_supreme_legacy")
            )
            
            # ADD BACK BUTTON BEFORE SENDING THE MESSAGE
            markup.add(InlineKeyboardButton("â¬…ï¸ Go Back", callback_data="back_to_mentorship_type"))
            
            # Customize message based on whether this is a renewal
            intro_text = "Renewal options" if is_renewal else "Please select your Supreme Mentorship plan"
            
            # FIXED: Use edit_message_text instead of send_message
            bot.edit_message_text(
                f"{intro_text}:\n\n"
                f"ðŸŽ‰ <b>{discount_name}: {discount_percentage}% OFF!</b>\n\n"
                f"ðŸ’° <b>Apprentice</b> - <s>${apprentice_price:.2f}</s> ${apprentice_discounted:.2f} / 3 Months\n"
                f"ðŸ’° <b>Disciple</b> - <s>${disciple_price:.2f}</s> ${disciple_discounted:.2f} / 6 Months\n"
                f"ðŸ’° <b>Legacy</b> - <s>${legacy_price:.2f}</s> ${legacy_discounted:.2f} / Lifetime", 
                chat_id,
                message_id,
                reply_markup=markup, 
                parse_mode="HTML"
            )
        else:
            # No discount - show regular prices without strikethrough
            markup.add(
                InlineKeyboardButton(f"Apprentice (${apprentice_price:.2f}) / 3 Months", callback_data="plan_apprentice"),
                InlineKeyboardButton(f"Disciple (${disciple_price:.2f}) / 6 Months", callback_data="plan_disciple"),
                InlineKeyboardButton(f"Legacy (${legacy_price:.2f}) / Lifetime", callback_data="plan_supreme_legacy")
            )
            
            # ADD BACK BUTTON
            markup.add(InlineKeyboardButton("â¬…ï¸ Go Back", callback_data="back_to_mentorship_type"))
            
            # Customize message based on whether this is a renewal
            intro_text = "Renewal options" if is_renewal else "Please select your Supreme Mentorship plan"
            
            # FIXED: Use edit_message_text instead of send_message
            bot.edit_message_text(
                f"{intro_text}:\n\n"
                f"ðŸ’° <b>Apprentice</b> - ${apprentice_price:.2f} / 3 Months\n"
                f"ðŸ’° <b>Disciple</b> - ${disciple_price:.2f} / 6 Months\n"
                f"ðŸ’° <b>Legacy</b> - ${legacy_price:.2f} / Lifetime", 
                chat_id,
                message_id,
                reply_markup=markup, 
                parse_mode="HTML"
            )

@bot.callback_query_handler(func=lambda call: call.data.startswith("plan_"))
def handle_plan_selection(call):
    """Handle plan selection from inline buttons"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    selected_plan = call.data.split("_", 1)[1]  # Split only on first underscore
    
    # Extract plan data based on selection
    if selected_plan == "trial":
        plan = "Trial"
        price_usd = 7.99  # Store numeric USD price for currency conversion
        duration = "Monthly"
        mentorship_type = "regular"
    elif selected_plan == "momentum":
        plan = "Momentum"
        price_usd = 20.99
        duration = "3 Months"
        mentorship_type = "regular"
    elif selected_plan == "apprentice":
        plan = "Apprentice"
        price_usd = 309.99
        duration = "3 Months"
        mentorship_type = "supreme"
    elif selected_plan == "disciple":
        plan = "Disciple"
        price_usd = 524.99
        duration = "6 Months" 
        mentorship_type = "supreme"
    elif selected_plan == "regular_legacy":
        plan = "Legacy"
        price_usd = 89.99
        duration = "Yearly"
        mentorship_type = "regular"
    elif selected_plan == "supreme_legacy":
        plan = "Legacy"
        price_usd = 899.99
        duration = "Lifetime"
        mentorship_type = "supreme"
    else:
        # If we get an unknown plan, respond with an error and return
        bot.answer_callback_query(call.id, "âŒ Invalid plan selection.")
        return

    # Acknowledge the selection
    bot.answer_callback_query(call.id, f"Selected {plan} plan")

    # Store original price before any discounts
    original_price_usd = price_usd
    original_price = f"${original_price_usd:.2f}"
    
    # Apply appropriate discount based on mentorship type
    discount_applied = False
    applicable_discount = DISCOUNTS.get(mentorship_type)
    
    if applicable_discount and applicable_discount.get('active', False):
        # Get transaction type - renewal or new purchase
        is_renewal = PENDING_USERS[chat_id].get('is_renewal', False)
        transaction_type = "renewal" if is_renewal else "new"
        
        # Debug log to identify the issue
        logging.info(f"Discount check for {mentorship_type}: is_renewal={is_renewal}, user_transaction={transaction_type}, discount_type={applicable_discount.get('transaction_type', 'both')}")
        
        # Check if discount applies to this transaction type
        applies_to_transaction = (
            applicable_discount.get('transaction_type', 'both') == 'both' or 
            applicable_discount.get('transaction_type', 'both') == transaction_type
        )
        
        # Only apply if the discount is valid for this transaction type
        if applies_to_transaction:
            # Get discount percentage and user limit
            discount_percentage = applicable_discount.get('percentage', 0)
            user_limit = applicable_discount.get('user_limit')
            users_used = applicable_discount.get('users_used', [])
            discount_name = applicable_discount.get('name', f'Special {mentorship_type.capitalize()} Discount')
            
            # Check if user limit is reached and if this user already used the discount
            if (user_limit is None or len(users_used) < user_limit) and str(user_id) not in users_used:
                # Apply the discount
                price_usd = original_price_usd * (1 - discount_percentage / 100)
                
                # Store discount info for later use
                PENDING_USERS[chat_id]['discount_percentage'] = discount_percentage
                PENDING_USERS[chat_id]['discount_name'] = discount_name
                PENDING_USERS[chat_id]['original_price_usd'] = original_price_usd
                
                # Mark discount as applied
                discount_applied = True
                
                # Add user to the list of users who used this discount
                DISCOUNTS[mentorship_type]['users_used'].append(str(user_id))
                save_discount(DISCOUNTS[mentorship_type], mentorship_type)
                
                # Log the discount application
                logging.info(f"Applied {discount_percentage}% {mentorship_type} discount to user {user_id}, price reduced from ${original_price_usd} to ${price_usd}")
            elif user_limit is not None and len(users_used) >= user_limit and str(user_id) not in users_used:
                # User limit reached
                bot.send_message(chat_id, f"âŒ The {discount_name} discount has reached its user limit. Your purchase will continue at the regular price.")
            elif str(user_id) in users_used:
                # User already used the discount
                bot.send_message(chat_id, f"â„¹ï¸ You've already used the {discount_name} discount. This purchase will be at the regular price.")
        else:
            # Not applicable - inform user if this is a specific discount type
            if applicable_discount.get('transaction_type') == "renewal" and not is_renewal:
                bot.send_message(chat_id, f"â„¹ï¸ The current {mentorship_type.capitalize()} discount is only available for renewals, not new purchases.")
            elif applicable_discount.get('transaction_type') == "new" and is_renewal:
                bot.send_message(chat_id, f"â„¹ï¸ The current {mentorship_type.capitalize()} discount is only available for new purchases, not renewals.")
    
    # Format the price for display (always .99)
    price = f"${price_usd:.2f}"

    # Store the plan details including the numeric USD price
    PENDING_USERS[chat_id]['plan'] = plan
    PENDING_USERS[chat_id]['price'] = price
    PENDING_USERS[chat_id]['price_usd'] = price_usd  # Store numeric price for conversion
    PENDING_USERS[chat_id]['mentorship_type'] = mentorship_type  # Store for later use
    PENDING_USERS[chat_id]['duration'] = duration
    PENDING_USERS[chat_id]['status'] = 'choosing_payment_method'
    save_pending_users()

    # Customize message based on discount
    if discount_applied:
        plan_message = (
            f"ðŸ™Œ You've selected the {mentorship_type.capitalize()} {plan} plan:\n\n"
            f"â€¢ Original price: {original_price}\n"
            f"ðŸ”– Discount: {PENDING_USERS[chat_id]['discount_percentage']}% OFF ({PENDING_USERS[chat_id]['discount_name']})\n\n"
            f"â€¢ You pay: {price}\n"
            f"â€¢ Duration: {duration}\n\n"
            f"Please select your payment method:"
        )
    else:
        plan_message = (
            f"ðŸ™Œ You've selected the {mentorship_type.capitalize()} {plan} plan:\n\n"
            f"â€¢ Price: {price}\n"
            f"â€¢ Duration: {duration}\n\n"
            f"Please select your payment method:"
        )
    
    # Create payment method markup using inline buttons
    markup = InlineKeyboardMarkup(row_width=2)
    markup.add(
        InlineKeyboardButton("ðŸ’³ Paypal", callback_data="payment_paypal"),
        InlineKeyboardButton("ðŸ“± GCash", callback_data="payment_gcash")
    )
    markup.add(
        InlineKeyboardButton("ðŸ’¸ Exness Direct", callback_data="payment_exness"),
        InlineKeyboardButton("ðŸ¦ Bank Transfer", callback_data="payment_bank")
    )
    markup.add(InlineKeyboardButton("â¬…ï¸ Go Back", callback_data="back_to_plan_selection"))
    
    # Edit current message with payment options
    bot.edit_message_text(plan_message, chat_id, message_id, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("payment_"))
def handle_payment_method(call):
    """Handle payment method selection from inline buttons"""
    chat_id = call.message.chat.id
    user_id = call.from_user.id
    method_code = call.data.split("_")[1]
    
    # First, remove the keyboard from the existing message to prevent multiple clicks
    remove_inline_keyboard(chat_id, call.message.message_id)
    
    # Map method codes to full method names with emoji
    method_map = {
        "paypal": "ðŸ’³ Paypal",
        "gcash": "ðŸ“± GCash",
        "exness": "ðŸ’¸ Exness Direct",
        "bank": "ðŸ¦ Bank Transfer"
    }
    
    method = method_map.get(method_code, "Unknown")
    
    # Acknowledge selection
    bot.answer_callback_query(call.id, f"Selected {method}")

    # Store the user's payment method choice
    PENDING_USERS[chat_id]['method'] = method
    PENDING_USERS[chat_id]['status'] = 'awaiting_payment'
    save_pending_users()
    
    # Determine which payment graphic to show based on the selected method
    payment_image = None
    if "Paypal" in method:
        payment_image = "graphics/paypal.jpeg"
    elif "Bank Transfer" in method:
        payment_image = "graphics/bank.jpeg"
    elif "GCash" in method:
        payment_image = "graphics/gcash.jpeg"
    elif "Exness Direct" in method:
        payment_image = "graphics/exness.jpeg"
    
    # Send the payment method-specific graphic if available
    if payment_image:
        try:
            with open(payment_image, 'rb') as img:
                bot.send_photo(
                    chat_id,
                    img,
                    caption=f"You've selected {method.strip('ðŸ’³ðŸ“±ðŸ¦ðŸŒðŸ’¸ ')} as your payment method"
                )
        except FileNotFoundError:
            logging.error(f"Payment method image not found at {payment_image}")
        except Exception as e:
            logging.error(f"Error sending payment method image: {e}")
        
        # Add a small delay for better UX
        time.sleep(1.5)
    
    # Get plan details for the message
    mentorship_type = PENDING_USERS[chat_id].get('mentorship_type', '')
    plan = PENDING_USERS[chat_id].get('plan', '')
    price = PENDING_USERS[chat_id].get('price', '')
    price_usd = PENDING_USERS[chat_id].get('price_usd', 0)  # Get numeric discounted price
    original_price_usd = PENDING_USERS[chat_id].get('original_price_usd')  # Original price before discount
    duration = PENDING_USERS[chat_id].get('duration', '')
    
    # Calculate payment method fee based on the ALREADY DISCOUNTED price
    fee_percentage = PAYMENT_FEES.get(method, 0.0)
    if fee_percentage > 0:
        fee_adjusted_price = price_usd * (1 + fee_percentage/100)
        fee_adjusted_price_str = f"${fee_adjusted_price:.2f}"
    else:
        fee_adjusted_price = price_usd  # No fee adjustment
        fee_adjusted_price_str = price
    
    # Fetch real-time exchange rates - use fee_adjusted_price for conversions
    exchange_rates = get_exchange_rates()
    
    # Prepare currency conversion information based on the FINAL price
    currency_info = ""
    if exchange_rates:
        currency_info = "\n\nðŸ’± *Equivalent Prices*:\n"
        
        # Define currency symbols for better display
        currency_symbols = {
            'USD': '$', 'GBP': 'Â£', 'EUR': 'â‚¬', 'IDR': 'Rp', 'PHP': 'â‚±'
        }
        
        for currency, rate in exchange_rates.items():
            if currency == 'USD':
                continue  # Skip USD as it's already shown
            
            # Use the final fee-adjusted price for all currency conversions
            equivalent_price = fee_adjusted_price * rate
            
            symbol = currency_symbols.get(currency, '')
            
            # Format based on currency
            if currency in ['IDR', 'PHP']:
                formatted_price = f"{symbol}{equivalent_price:,.0f}"
            else:
                whole_price = int(equivalent_price)
                formatted_price = f"{symbol}{whole_price:,}.99"
            
            currency_info += f"â€¢ *{currency}*: {formatted_price}\n"
    
    # Format plan details based on discount and payment fee combinations
    discount_percentage = PENDING_USERS[chat_id].get('discount_percentage')
    discount_name = PENDING_USERS[chat_id].get('discount_name')

    # Create a comprehensive message showing all calculations clearly
    if original_price_usd and discount_percentage:
        # Case: Has discount
        if fee_percentage > 0:
            # Has both discount AND payment fee - show complete calculation
            plan_details = (
                f"*Plan Details*:\n"
                f"- Type: {mentorship_type.capitalize()} Mentorship\n"
                f"- Plan: {plan}\n"
                f"- Original Price: ${original_price_usd:.2f} USD\n"
                f"- *{discount_name}: {discount_percentage}% OFF* (-${(original_price_usd - price_usd):.2f})\n"
                f"- Discounted Price: ${price_usd:.2f} USD\n"
                f"- *{method.strip('ðŸ’³ðŸ“±ðŸ¦ðŸŒðŸ’¸ ')} Fee: {fee_percentage}%* (+${(fee_adjusted_price - price_usd):.2f})\n"
                f"- **Final Price: {fee_adjusted_price_str} USD**\n"
                f"- Duration: {duration}{currency_info}\n\n"
                f"*Note: A {fee_percentage}% fee applies to cover {method.strip('ðŸ’³ðŸ“±ðŸ¦ðŸŒðŸ’¸ ')} transaction costs.*\n\n"
            )
        else:
            # Has discount but NO payment fee
            plan_details = (
                f"*Plan Details*:\n"
                f"- Type: {mentorship_type.capitalize()} Mentorship\n"
                f"- Plan: {plan}\n"
                f"- Original Price: ${original_price_usd:.2f} USD\n"
                f"- *{discount_name}: {discount_percentage}% OFF* (-${(original_price_usd - price_usd):.2f})\n"
                f"- **Final Price: ${price_usd:.2f} USD**\n"
                f"- Duration: {duration}{currency_info}\n\n"
            )
    else:
        # Case: No discount applied
        if fee_percentage > 0:
            # No discount, but has payment fee
            plan_details = (
                f"*Plan Details*:\n"
                f"- Type: {mentorship_type.capitalize()} Mentorship\n"
                f"- Plan: {plan}\n"
                f"- Base Price: ${price_usd:.2f} USD\n"
                f"- *{method.strip('ðŸ’³ðŸ“±ðŸ¦ðŸŒðŸ’¸ ')} Fee: {fee_percentage}%* (+${(fee_adjusted_price - price_usd):.2f})\n"
                f"- **Final Price: {fee_adjusted_price_str} USD**\n"
                f"- Duration: {duration}{currency_info}\n\n"
                f"*Note: A {fee_percentage}% fee applies to cover {method.strip('ðŸ’³ðŸ“±ðŸ¦ðŸŒðŸ’¸ ')} transaction costs.*\n\n"
            )
        else:
            # No discount, no payment fee - simplest case
            plan_details = (
                f"*Plan Details*:\n"
                f"- Type: {mentorship_type.capitalize()} Mentorship\n"
                f"- Plan: {plan}\n"
                f"- **Price: ${price_usd:.2f} USD**\n"
                f"- Duration: {duration}{currency_info}\n\n"
            )

    # Also update the payment information in PENDING_USERS to track the final price
    PENDING_USERS[chat_id]['final_price_usd'] = fee_adjusted_price
    save_pending_users()

    # Send payment credentials based on the selected method
    payment_details = {
        "ðŸ’³ Paypal": "PayPal:\nOption 1: https://paypal.me/daintyrich\n\nOption 2: \nName: R Mina\nEmail: romeomina061109@gmail.com",
        "ðŸ¦ Bank Transfer": "ðŸ¦ **Bank Details:**\nBank: BDO\nDebit Number: 5210 6912 0103 9329\nAccount Name: Romeo B. Mina III",
        "ðŸ“± GCash": "ðŸ“± **GCash Number:** 09274478330 (R. U.)",
        "ðŸ’¸ Exness Direct": {
            "account_id_1": "108377569",
            "email_1": "diamondchay626@gmail.com",
            "account_id_2": "217136604",
            "email_2": "romeomina061109@gmail.com"
        }
    }

    # Format the message properly with plan details
    if method == "ðŸ’¸ Exness Direct":
        message = (
            "ðŸ’° **Payment Instructions:**\n\n"
            f"{plan_details}"
            "For Exness Direct:\n"
            f"ðŸ“Œ **Exness Account ID 1:** {payment_details['ðŸ’¸ Exness Direct']['account_id_1']}\n"
            f"ðŸ“§ **Email 1:** {payment_details['ðŸ’¸ Exness Direct']['email_1']}\n\n"
            f"ðŸ“Œ **Exness Account ID 2:** {payment_details['ðŸ’¸ Exness Direct']['account_id_2']}\n"
            f"ðŸ“§ **Email 2:** {payment_details['ðŸ’¸ Exness Direct']['email_2']}\n\n"
            "After completing your transaction, please use /verify to confirm your payment.\n\n"
            "Note: Your Telegram ID and name will be securely recorded."
        )
    else:
        message = (
            "ðŸ’° **Payment Instructions:**\n\n"
            f"{plan_details}"
            f"{payment_details[method]}\n\n"
            "After completing your transaction, please use /verify to confirm your payment.\n\n"
            "Note: Your Telegram ID and name will be securely recorded."
        )

    # Send the message
    bot.send_message(chat_id, message, parse_mode="Markdown")

@bot.callback_query_handler(func=lambda call: call.data == "back_to_plan_selection")
def back_to_plan_selection(call):
    """Handle going back to plan selection"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    
    bot.answer_callback_query(call.id, "Returning to plan selection...")
    
    # Get mentorship type from the stored data
    mentorship_type = PENDING_USERS[chat_id].get('mentorship_type', 'regular')
    is_renewal = PENDING_USERS[chat_id].get('is_renewal', False)
    
    # Update the status
    PENDING_USERS[chat_id]['status'] = 'choosing_mentorship_plan'
    save_pending_users()
    
    # Create new markup based on the mentorship type
    markup = InlineKeyboardMarkup(row_width=1)
    
    # Use stored mentorship type to determine which plans to show
    if mentorship_type == 'regular':
        # Get regular discount if applicable
        applicable_discount = DISCOUNTS.get('regular')
        discount_active = applicable_discount and applicable_discount.get('active', False)
        
        # Check if discount applies to this transaction type
        if discount_active:
            transaction_type = "renewal" if is_renewal else "new"
            discount_transaction_type = applicable_discount.get('transaction_type', 'both')
            
            # Determine if discount applies to this transaction type
            applies_to_transaction = (
                discount_transaction_type == 'both' or 
                discount_transaction_type == transaction_type
            )
            
            # Only consider discount active if it applies to this transaction type
            if not applies_to_transaction:
                discount_active = False
                
        discount_percentage = applicable_discount.get('percentage', 0) if discount_active else 0
        discount_name = applicable_discount.get('name', '') if discount_active else ''
        
        # Original prices
        trial_price = 7.99
        momentum_price = 20.99
        legacy_price = 89.99
        
        if discount_active:
            # Calculate discounted prices
            trial_discounted = round(trial_price * (1 - discount_percentage / 100), 2)
            momentum_discounted = round(momentum_price * (1 - discount_percentage / 100), 2)
            legacy_discounted = round(legacy_price * (1 - discount_percentage / 100), 2)
            
            # Add buttons with discounted prices
            markup.add(
                InlineKeyboardButton(f"Trial (${trial_discounted:.2f}) / Monthly", callback_data="plan_trial"),
                InlineKeyboardButton(f"Momentum (${momentum_discounted:.2f}) / 3 Months", callback_data="plan_momentum"),
                InlineKeyboardButton(f"Legacy (${legacy_discounted:.2f}) / Yearly", callback_data="plan_regular_legacy")
            )
            
            # Add back button
            markup.add(InlineKeyboardButton("â¬…ï¸ Go Back", callback_data="back_to_mentorship_type"))
            
            # Customize message based on whether this is a renewal
            intro_text = "Renewal options" if is_renewal else "Please select your Regular Mentorship plan"
            
            # FIXED: Use edit_message_text instead of send_message
            bot.edit_message_text(
                f"{intro_text}:\n\n"
                f"ðŸŽ‰ <b>{discount_name}: {discount_percentage}% OFF!</b>\n\n"
                f"ðŸ’° <b>Trial</b> - <s>${trial_price:.2f}</s> ${trial_discounted:.2f} / Monthly\n"
                f"ðŸ’° <b>Momentum</b> - <s>${momentum_price:.2f}</s> ${momentum_discounted:.2f} / 3 Months\n"
                f"ðŸ’° <b>Legacy</b> - <s>${legacy_price:.2f}</s> ${legacy_discounted:.2f} / Yearly", 
                chat_id,
                message_id,
                reply_markup=markup, 
                parse_mode="HTML"
            )
        else:
            # No discount - show regular prices without strikethrough
            markup.add(
                InlineKeyboardButton(f"Trial (${trial_price:.2f}) / Monthly", callback_data="plan_trial"),
                InlineKeyboardButton(f"Momentum (${momentum_price:.2f}) / 3 Months", callback_data="plan_momentum"),
                InlineKeyboardButton(f"Legacy (${legacy_price:.2f}) / Yearly", callback_data="plan_regular_legacy")
            )
            
            # Add back button
            markup.add(InlineKeyboardButton("â¬…ï¸ Go Back", callback_data="back_to_mentorship_type"))
            
            # Customize message based on whether this is a renewal
            intro_text = "Renewal options" if is_renewal else "Please select your Regular Mentorship plan"
            
            # FIXED: Use edit_message_text instead of send_message
            bot.edit_message_text(
                f"{intro_text}:\n\n"
                f"ðŸ’° <b>Trial</b> - ${trial_price:.2f} / Monthly\n"
                f"ðŸ’° <b>Momentum</b> - ${momentum_price:.2f} / 3 Months\n"
                f"ðŸ’° <b>Legacy</b> - ${legacy_price:.2f} / Yearly", 
                chat_id,
                message_id,
                reply_markup=markup, 
                parse_mode="HTML"
            )
    else:
        # Supreme membership
        # Get supreme discount if applicable
        applicable_discount = DISCOUNTS.get('supreme')
        discount_active = applicable_discount and applicable_discount.get('active', False)
        
        # Check if discount applies to this transaction type
        if discount_active:
            transaction_type = "renewal" if is_renewal else "new"
            discount_transaction_type = applicable_discount.get('transaction_type', 'both')
            
            # Determine if discount applies to this transaction type
            applies_to_transaction = (
                discount_transaction_type == 'both' or 
                discount_transaction_type == transaction_type
            )
            
            # Only consider discount active if it applies to this transaction type
            if not applies_to_transaction:
                discount_active = False
                
        discount_percentage = applicable_discount.get('percentage', 0) if discount_active else 0
        discount_name = applicable_discount.get('name', '') if discount_active else ''
        
        # Original prices
        apprentice_price = 309.99
        disciple_price = 524.99
        legacy_price = 899.99
        
        if discount_active:
            # Calculate discounted prices
            apprentice_discounted = round(apprentice_price * (1 - discount_percentage / 100), 2)
            disciple_discounted = round(disciple_price * (1 - discount_percentage / 100), 2)
            legacy_discounted = round(legacy_price * (1 - discount_percentage / 100), 2)
            
            # Add buttons with discounted prices
            markup.add(
                InlineKeyboardButton(f"Apprentice (${apprentice_discounted:.2f}) / 3 Months", callback_data="plan_apprentice"),
                InlineKeyboardButton(f"Disciple (${disciple_discounted:.2f}) / 6 Months", callback_data="plan_disciple"),
                InlineKeyboardButton(f"Legacy (${legacy_discounted:.2f}) / Lifetime", callback_data="plan_supreme_legacy")
            )
            
            # Add back button
            markup.add(InlineKeyboardButton("â¬…ï¸ Go Back", callback_data="back_to_mentorship_type"))
            
            # Customize message based on whether this is a renewal
            intro_text = "Renewal options" if is_renewal else "Please select your Supreme Mentorship plan"
            
            # FIXED: Use edit_message_text instead of send_message
            bot.edit_message_text(
                f"{intro_text}:\n\n"
                f"ðŸŽ‰ <b>{discount_name}: {discount_percentage}% OFF!</b>\n\n"
                f"ðŸ’° <b>Apprentice</b> - <s>${apprentice_price:.2f}</s> ${apprentice_discounted:.2f} / 3 Months\n"
                f"ðŸ’° <b>Disciple</b> - <s>${disciple_price:.2f}</s> ${disciple_discounted:.2f} / 6 Months\n"
                f"ðŸ’° <b>Legacy</b> - <s>${legacy_price:.2f}</s> ${legacy_discounted:.2f} / Lifetime", 
                chat_id,
                message_id,
                reply_markup=markup, 
                parse_mode="HTML"
            )
        else:
            # No discount - show regular prices without strikethrough
            markup.add(
                InlineKeyboardButton(f"Apprentice (${apprentice_price:.2f}) / 3 Months", callback_data="plan_apprentice"),
                InlineKeyboardButton(f"Disciple (${disciple_price:.2f}) / 6 Months", callback_data="plan_disciple"),
                InlineKeyboardButton(f"Legacy (${legacy_price:.2f}) / Lifetime", callback_data="plan_supreme_legacy")
            )
            
            # Add back button
            markup.add(InlineKeyboardButton("â¬…ï¸ Go Back", callback_data="back_to_mentorship_type"))
            
            # Customize message based on whether this is a renewal
            intro_text = "Renewal options" if is_renewal else "Please select your Supreme Mentorship plan"
            
            # FIXED: Use edit_message_text instead of send_message
            bot.edit_message_text(
                f"{intro_text}:\n\n"
                f"ðŸ’° <b>Apprentice</b> - ${apprentice_price:.2f} / 3 Months\n"
                f"ðŸ’° <b>Disciple</b> - ${disciple_price:.2f} / 6 Months\n"
                f"ðŸ’° <b>Legacy</b> - ${legacy_price:.2f} / Lifetime", 
                chat_id,
                message_id,
                reply_markup=markup, 
                parse_mode="HTML"
            )

# Add handler for back button if needed
@bot.callback_query_handler(func=lambda call: call.data == "back_to_mentorship_type")
def back_to_mentorship_type(call):
    """Handle back button to return to mentorship type selection"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    
    # Update status to go back to mentorship type selection
    PENDING_USERS[chat_id]['status'] = 'choosing_mentorship_type'
    save_pending_users()
    
    # Show mentorship type options with inline keyboard
    markup = InlineKeyboardMarkup(row_width=2)
    markup.add(
        InlineKeyboardButton("Regular Mentorship", callback_data="mentorship_regular"),
        InlineKeyboardButton("Supreme Mentorship", callback_data="mentorship_supreme")
    )
    markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
    
    bot.edit_message_text(
        "Please select your preferred mentorship level:",
        chat_id,
        message_id,
        reply_markup=markup
    )
    
    bot.answer_callback_query(call.id, "Returning to mentorship selection")

def has_user_paid(user_id):
    return str(user_id) in PAYMENT_DATA and PAYMENT_DATA[str(user_id)]['haspayed']

def can_renew_membership(user_id):
    """Check if user can renew their membership based on expiration date"""
    user_id_str = str(user_id)
    
    # First check if user exists in PAYMENT_DATA
    if user_id_str not in PAYMENT_DATA:
        return False, "âŒ You don't have an existing membership to renew. Please use 'Purchase Membership' option instead."
    
    data = PAYMENT_DATA[user_id_str]
    
    # Check for any evidence of a previous valid membership
    has_ever_paid = data.get('haspayed') or data.get('ever_paid', False)
    has_due_date = 'due_date' in data
    
    # If no evidence of ever having had a real membership, reject renewal
    if not has_ever_paid and not has_due_date:
        return False, "âŒ You don't have a previous membership to renew. Please use 'Purchase Membership' option instead."
    
    # If not currently paid, they can renew (this is a genuine expired membership case)
    if not data.get('haspayed', False):
        return True, None
    
    # If they're currently paid, check expiration date
    try:
        due_date = datetime.strptime(data['due_date'], '%Y-%m-%d %H:%M:%S')
        current_date = datetime.now()
        days_remaining = (due_date - current_date).days
        
        # Allow renewal if within 3 days of expiration
        if days_remaining <= 3:
            return True, None
        else:
            return False, f"âš ï¸ You still have {days_remaining} days remaining on your current membership. Early renewal is only available within 3 days of expiration."
    except Exception as e:
        logging.error(f"Error checking renewal eligibility: {e}")
        # If there's an error, default to no
        return False, "âŒ Error checking renewal eligibility. Please try again later or contact an admin."

@bot.callback_query_handler(func=lambda call: call.data == "_redeem")
def handle_redeem_serial_callback(call):
    """Handle redeem serial button click from main menu"""
    chat_id = call.message.chat.id
    user_id = call.from_user.id
    
    # First, remove the keyboard from the existing message
    remove_inline_keyboard(chat_id, call.message.message_id)
    
    # Update user status - CHANGED to match the process_serial_number handler
    PENDING_USERS[user_id] = {'status': 'awaiting_serial'}
    save_pending_users()
    
    # Create back button
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(InlineKeyboardButton("â¬…ï¸ Back to Main Menu", callback_data="back_to_main_menu"))
    
    # Send message prompting for serial number
    bot.send_message(
        chat_id,
        "ðŸ”‘ *Serial Redemption*\n\n"
        "Please enter the serial number you received for your membership.\n\n"
        "Note: Serial numbers are case-sensitive.",
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    # Answer the callback query
    bot.answer_callback_query(call.id, "Please enter your serial number")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'awaiting_serial')
def process_serial_number(message):
    """Process the serial number entered by the user"""
    chat_id = message.chat.id
    user_id = message.from_user.id
    serial = message.text.strip()
    
    # Check if the serial number exists in our database
    if serial not in SERIAL_NUMBERS:
        bot.send_message(
            chat_id,
            "âŒ Invalid serial number. Please check and try again, or contact support if you believe this is an error.",
            parse_mode="Markdown"
        )
        return
    
    # Check if the serial has already been used
    if SERIAL_NUMBERS[serial].get('used', False):
        bot.send_message(
            chat_id,
            "âŒ This serial number has already been used. Each serial can only be redeemed once.",
            parse_mode="Markdown"
        )
        return
    
    # Valid serial number - get the associated plan details
    mentorship_type = SERIAL_NUMBERS[serial].get('mentorship_type', 'regular')
    plan = SERIAL_NUMBERS[serial].get('plan', 'Trial')
    
    # Send confirmation message
    bot.send_message(
        chat_id,
        f"âœ… *Valid Serial Number!*\n\n"
        f"You are redeeming a serial for:\n"
        f"â€¢ Mentorship Type: *{mentorship_type.capitalize()}*\n"
        f"â€¢ Plan: *{plan}*\n\n"
        f"Processing your redemption...",
        parse_mode="Markdown"
    )
    
    # Calculate due date based on the plan
    if plan == "Trial":
        due_date = datetime.now() + timedelta(days=30)  # 1 month
    elif plan == "Momentum" or plan == "Apprentice":
        due_date = datetime.now() + timedelta(days=90)  # 3 months
    elif plan == "Disciple":
        due_date = datetime.now() + timedelta(days=180)  # 6 months
    elif plan == "Regular Legacy":
        due_date = datetime.now() + timedelta(days=365)  # 1 year
    elif plan == "Supreme Legacy":
        due_date = datetime.now() + timedelta(days=3650)  # ~10 years (lifetime)
    else:
        due_date = datetime.now() + timedelta(days=30)  # Default to 1 month
    
    # Mark the serial as used
    SERIAL_NUMBERS[serial]['used'] = True
    SERIAL_NUMBERS[serial]['used_by'] = user_id
    SERIAL_NUMBERS[serial]['used_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    save_serial_number(serial, SERIAL_NUMBERS[serial])
    
    # Save user membership data
    username = message.from_user.username or "No Username"
    
    if str(user_id) in PAYMENT_DATA:
        # Update existing data
        PAYMENT_DATA[str(user_id)].update({
            "username": username,
            "payment_plan": plan,
            "payment_mode": "Serial Redemption",
            "mentorship_type": mentorship_type,
            "due_date": due_date.strftime('%Y-%m-%d %H:%M:%S'),
            "haspayed": True,
            "serial_number": serial
        })
    else:
        # Create new entry
        PAYMENT_DATA[str(user_id)] = {
            "username": username,
            "payment_plan": plan,
            "payment_mode": "Serial Redemption",
            "mentorship_type": mentorship_type,
            "due_date": due_date.strftime('%Y-%m-%d %H:%M:%S'),
            "haspayed": True,
            "terms_accepted": True,
            "privacy_accepted": True,
            "serial_number": serial
        }
    
    save_payment_data()
    
    # Clear pending status
    PENDING_USERS[user_id]['status'] = 'serial_redeemed'
    save_pending_users()
    
    # Send success message with celebration GIF or image
    try:
        with open('graphics/serial_success.jpeg', 'rb') as success_img:
            bot.send_photo(
                chat_id,
                success_img,
                caption=f"ðŸŽ‰ *SERIAL REDEMPTION SUCCESSFUL!*\n\n"
                        f"Your {mentorship_type.capitalize()} Membership ({plan}) has been activated!\n\n"
                        f"Membership expires: {due_date.strftime('%Y-%m-%d')}\n\n"
                        f"We're preparing your onboarding process...",
                parse_mode="Markdown"
            )
    except FileNotFoundError:
        # Fallback to text message if image not found
        bot.send_message(
            chat_id,
            f"ðŸŽ‰ *SERIAL REDEMPTION SUCCESSFUL!*\n\n"
            f"Your {mentorship_type.capitalize()} Membership ({plan}) has been activated!\n\n"
            f"Membership expires: {due_date.strftime('%Y-%m-%d')}\n\n"
            f"We're preparing your onboarding process...",
            parse_mode="Markdown"
        )
    
    # Add a small delay for better UX
    time.sleep(1.5)
    
    # Start the onboarding process
    send_onboarding_form(user_id)
    
    # Notify admins about the serial redemption
    for admin_id in ADMIN_IDS:
        bot.send_message(
            admin_id,
            f"ðŸ”‘ *SERIAL REDEMPTION*\n\n"
            f"â€¢ User: @{username} (ID: `{user_id}`)\n"
            f"â€¢ Serial: `{serial}`\n"
            f"â€¢ Mentorship: {mentorship_type.capitalize()}\n"
            f"â€¢ Plan: {plan}\n"
            f"â€¢ Expires: {due_date.strftime('%Y-%m-%d')}\n\n"
            f"Generated by: {SERIAL_NUMBERS[serial].get('generated_by', 'Unknown')}",
            parse_mode="Markdown"
        )

@bot.callback_query_handler(func=lambda call: call.data.startswith("faq_"))
def handle_faq_category(call):
    """Handle FAQ category selection"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    category = call.data.split("_", 1)[1]  # Split at first underscore to get category
    
    if category == "join_academy":
        # Send specific welcome message for Join Academy
        welcome_message = (
            "ðŸŽ“ *Welcome to Prodigy Trading Academy!*\n\n"
            "You're one step away from accessing our full mentorship and community resources on Telegram. "
            "Here's how to join:\n\n"
            "1ï¸âƒ£ Purchase your chosen mentorship plan by accessing our official Telegram bot down below.\n\n"
            "2ï¸âƒ£ Follow the step-by-step instructions provided after payment.\n\n"
            "3ï¸âƒ£ You'll be added to the main channel and receive access to all relevant group chats.\n\n"
            "If you need help choosing a plan or navigating the bot for joining, we're here to assist!"
        )
        
        # Add back buttons with improved layout
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(InlineKeyboardButton("Â« Back to Categories", callback_data="faq_back"))
        markup.add(InlineKeyboardButton("ðŸ  Back to Main Menu", callback_data="faq_main_menu"))
        
        bot.edit_message_text(
            welcome_message,
            chat_id,
            message_id,
            reply_markup=markup,
            parse_mode="Markdown"
        )
    
    elif category == "admissions":
        # Send information about admissions schedule
        admissions_message = (
            "ðŸ“… *Opening of Admissions*\n\n"
            "Admissions are open every two weeks (bi-weekly), and each opening period lasts 1 full week only for enrollment.\n\n"
            "Make sure to check our post updates regarding enrollment! If you see no posts related, we are likely to be closed.\n\n"
            "Once admissions close, we do not accept late entries until the next opening â€” this is to keep things focused "
            "and structured for our current students.\n\n"
            "Stay tuned to our page for announcements on the next enrollment period, and don't miss your chance to join!"
        )
        
        # Add back buttons with improved layout
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(InlineKeyboardButton("Â« Back to Categories", callback_data="faq_back"))
        markup.add(InlineKeyboardButton("ðŸ  Back to Main Menu", callback_data="faq_main_menu"))
        
        bot.edit_message_text(
            admissions_message,
            chat_id,
            message_id,
            reply_markup=markup,
            parse_mode="Markdown"
        )
    
    elif category == "plans_pricing":
        # Send mentorship plans and pricing information
        pricing_message = (
            "ðŸ’² *Mentorship Plans & Pricing*\n\n"
            "âœ¨ *Supreme Mentorship:*\n"
            "â€¢ Apprentice: $309.99 / 3 months\n"
            "â€¢ Disciple: $524.99 / 6 months\n"
            "â€¢ Master: $899.99 / lifetime\n\n"
            "ðŸ”„ *Regular Mentorship:*\n"
            "â€¢ Trial: $7.99 / month\n"
            "â€¢ Momentum: $20.99 / 3 months - *save 11%*\n"
            "â€¢ Legacy: $89.99 / year - *save 7%*\n\n"
            "ðŸ’³ *Payment Options*\n"
            "GCash â€¢ PayPal â€¢ Bank Transfer\n"
            "_(Stripe & Crypto coming soon)_\n\n"
            "â†’ Click \"Join Academy\" in the menu for enrollment details!\n\n"
            "Let us know if you need any help! We will be able to respond as soon as possible."
        )
        
        # Add back buttons with improved layout
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(InlineKeyboardButton("Â« Back to Categories", callback_data="faq_back"))
        markup.add(InlineKeyboardButton("ðŸ  Back to Main Menu", callback_data="faq_main_menu"))
        
        bot.edit_message_text(
            pricing_message,
            chat_id,
            message_id,
            reply_markup=markup,
            parse_mode="Markdown"
        )
        
    elif category == "products_services":
        # Send products and services information with fixed formatting
        products_message = (
            "ðŸ“š *Products & Services*\n\n"
            "ðŸ“• *Learning Materials:*\n"
            "â€¢ Prodigy Path E-books - Exclusive to mentorship students\n"
            "â€¢ Bundles 1-3 - Public access (Coming soon!)\n\n"
            "ðŸ› ï¸ *Services:*\n"
            "â€¢ Fund Management - We trade for you with our excellent strategies!\n"
            "  Get to know about our expertise in fund management through\n"
            "  Telegram and we will send you proofs of backtesting and journaling.\n\n"
            "ðŸ’¬ *Contact Our Founders:*\n"
            "â€¢ @rom\\_pta\n"
            "â€¢ @fiftysevenrupees\n\n"
            "ðŸ”® *More tools and offers launching soon. Stay tuned!*"
        )
        
        # Add back buttons with improved layout
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(InlineKeyboardButton("Â« Back to Categories", callback_data="faq_back"))
        markup.add(InlineKeyboardButton("ðŸ  Back to Main Menu", callback_data="faq_main_menu"))
        
        bot.edit_message_text(
            products_message,
            chat_id,
            message_id,
            reply_markup=markup,
            parse_mode="Markdown"
        )
        
    elif category == "enrollment_benefits":
        # Send benefits information with improved formatting
        benefits_message = (
            "ðŸŒŸ *Mentorship Benefits*\n\n"
            "âœ¨ *SUPREME MENTORSHIP*\n"
            "â€¢ 1-on-1 private coaching sessions\n"
            "â€¢ Personalized teaching approach\n"
            "â€¢ Deep learning sessions + priority support\n"
            "â€¢ Includes everything in Regular membership\n\n"
            "ðŸ”¹ *REGULAR MENTORSHIP*\n"
            "â€¢ Full access to core trading lessons\n"
            "â€¢ Livestreams & educational discussions\n"
            "â€¢ Access to community group chat\n\n"
            "Need help choosing the best plan for your needs? Send us a message and we'll guide you through the options!"
        )
        
        # Add back buttons with improved layout
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(InlineKeyboardButton("Â« Back to Categories", callback_data="faq_back"))
        markup.add(InlineKeyboardButton("ðŸ  Back to Main Menu", callback_data="faq_main_menu"))
        
        bot.edit_message_text(
            benefits_message,
            chat_id,
            message_id,
            reply_markup=markup,
            parse_mode="Markdown"
        )
        
    elif category == "terms":
        # First try sending the PDF if available
        pdf_sent = False
        try:
            with open('pdf/terms_and_conditions.pdf', 'rb') as doc:
                # Send as separate message and store the message ID
                pdf_message = bot.send_document(
                    call.message.chat.id, 
                    doc, 
                    caption="ðŸ“ Terms and Conditions"
                )
                # Store the PDF message ID for later deletion
                user_id = call.from_user.id
                PDF_MESSAGE_IDS[user_id] = {
                    'message_id': pdf_message.message_id,
                    'chat_id': call.message.chat.id
                }
                pdf_sent = True
        except FileNotFoundError:
            pass  # PDF not found, continue with text-only version
        except Exception as e:
            logging.error(f"Error sending T&C PDF in FAQ: {e}")
            
        # Now edit the original message with the text content
        terms_text = (
            "ðŸ“ *TERMS & CONDITIONS*\n\n"
            "By using the Prodigy Trading Academy Bot and services, you agree to the following:\n\n"
            "1. Membership fees are non-refundable once access is granted.\n"
            "2. Trading involves risk; we do not guarantee financial returns.\n"
            "3. Your membership is for personal use only and cannot be shared.\n"
            "4. We may revoke access for policy violations without refund.\n"
            "5. All educational content provided is property of Prodigy Trading Academy.\n"
            "6. You must be at least 18 years old to use our services.\n"
            "7. We reserve the right to modify these terms at any time."
        )
        
        # Add back buttons with improved layout
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(InlineKeyboardButton("Â« Back to Categories", callback_data="faq_back"))
        markup.add(InlineKeyboardButton("ðŸ  Back to Main Menu", callback_data="faq_main_menu"))
        
        # Edit the original message instead of sending a new one
        bot.edit_message_text(
            terms_text,
            chat_id,
            message_id,
            reply_markup=markup,
            parse_mode="Markdown"
        )
        
        # Answer the callback to remove loading indicator
        bot.answer_callback_query(call.id, "Viewing Terms & Conditions" + (" (PDF sent)" if pdf_sent else ""))
    
    elif category == "privacy":
        # First try sending the PDF if available
        pdf_sent = False
        try:
            with open('pdf/privacy_policy.pdf', 'rb') as doc:
                # Send as separate message and store the message ID
                pdf_message = bot.send_document(
                    call.message.chat.id, 
                    doc, 
                    caption="ðŸ”’ Privacy Policy"
                )
                # Store the PDF message ID for later deletion
                user_id = call.from_user.id
                PDF_MESSAGE_IDS[user_id] = {
                    'message_id': pdf_message.message_id,
                    'chat_id': call.message.chat.id
                }
                pdf_sent = True
        except FileNotFoundError:
            pass  # PDF not found, continue with text-only version
        except Exception as e:
            logging.error(f"Error sending Privacy Policy PDF in FAQ: {e}")
            
        # Edit the original message with the text content
        privacy_text = (
            "ðŸ”’ *PRIVACY POLICY*\n\n"
            "This policy explains how we handle your personal data:\n\n"
            "1. We collect your Telegram ID, username, and payment information.\n"
            "2. Your data is used to manage your membership and provide services.\n"
            "3. We may send payment reminders and service updates.\n"
            "4. Your information is not sold to third parties.\n"
            "5. Your data is stored securely for the duration of your membership.\n"
            "6. You may request access to or deletion of your data at any time.\n"
            "7. We use encryption to protect your payment information."
        )
        
        # Add back buttons with improved layout
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(InlineKeyboardButton("Â« Back to Categories", callback_data="faq_back"))
        markup.add(InlineKeyboardButton("ðŸ  Back to Main Menu", callback_data="faq_main_menu"))
        
        # Edit the original message instead of sending a new one
        bot.edit_message_text(
            privacy_text,
            chat_id,
            message_id,
            reply_markup=markup,
            parse_mode="Markdown"
        )
        
        # Answer the callback to remove loading indicator
        bot.answer_callback_query(call.id, "Viewing Privacy Policy" + (" (PDF sent)" if pdf_sent else ""))
    
    elif category == "back":
        # Go back to main FAQ categories
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(
            InlineKeyboardButton("ðŸŽ“ Join Academy", callback_data="faq_join_academy"),
            InlineKeyboardButton("ðŸ“… Opening of Admissions", callback_data="faq_admissions"),
            InlineKeyboardButton("ðŸ’² Mentorship Plans & Pricings", callback_data="faq_plans_pricing"),
            InlineKeyboardButton("ðŸ“¦ Products & Services", callback_data="faq_products_services"),
            InlineKeyboardButton("ðŸŒŸ Benefits in Enrollment", callback_data="faq_enrollment_benefits"),
            InlineKeyboardButton("ðŸ“ Terms & Conditions", callback_data="faq_terms"),
            InlineKeyboardButton("ðŸ”’ Privacy Policy", callback_data="faq_privacy")
        )
        markup.add(InlineKeyboardButton("ðŸ  Back to Main Menu", callback_data="faq_main_menu"))
        
        bot.edit_message_text(
            "ðŸ” *Frequently Asked Questions*\n\n"
            "Select a category to view related questions:",
            chat_id,
            message_id,
            reply_markup=markup,
            parse_mode="Markdown"
        )
        
        bot.answer_callback_query(call.id, "Back to FAQ categories")
    
    elif category == "main_menu":
        # Return to the main menu
        show_main_menu(chat_id, user_id=call.from_user.id, message_id=message_id)
        bot.answer_callback_query(call.id, "Returning to main menu")

# Verify Command - Asks for proof of payment
@bot.message_handler(commands=['verify'])
def request_payment_proof(message):
    if message.chat.type != 'private':
        return  # Ignore if not in private chat
    chat_id = message.chat.id

    if chat_id not in PENDING_USERS or PENDING_USERS[chat_id]['status'] != 'awaiting_payment':
        bot.send_message(chat_id, "âŒ You haven't selected a payment plan and method. Please start with /start.")
        return

    PENDING_USERS[chat_id]['status'] = 'awaiting_proof'
    save_pending_users()
    bot.send_message(chat_id, "ðŸ“¸ Please upload a screenshot of your payment proof.")

# Handle Screenshot Upload
@bot.message_handler(content_types=['photo'])
def handle_payment_screenshot(message):
    if message.chat.type != 'private':
        return  # Ignore if not in private chat
    
    user_id = message.from_user.id
    chat_id = message.chat.id
    
    # First check if user is in confession mode - if so, skip this handler
    if user_id in USERS_CONFESSING and USERS_CONFESSING[user_id]['status'] == 'awaiting_confession':
        # Process as a confession photo - call the handler directly
        handle_photo_confession(message)
        return
        
    # Check if user is in announcement mode - if so, skip this handler
    if user_id in ADMIN_ANNOUNCING and ADMIN_ANNOUNCING[user_id]['status'] == 'waiting_for_announcement':
        # Process as an announcement photo
        handle_announcement_message(message)
        return

    # Add new condition for awaiting XM account screenshot
    if user_id in PENDING_USERS and PENDING_USERS[user_id].get('status') == 'xm_awaiting_account_screenshot':
        # Store screenshot details
        PENDING_USERS[user_id]['screenshot_message_id'] = message.message_id
        PENDING_USERS[user_id]['status'] = 'xm_awaiting_account_id'
        save_pending_users()
        
        # Ask for account ID
        bot.send_message(
            chat_id,
            "ðŸ“ *XM Account ID*\n\n"
            "Thank you for the screenshot. Now, please provide your XM Trading Account ID:",
            parse_mode="Markdown"
        )
        return
    
    # Check if waiting for XM screenshot
    if user_id in PENDING_USERS and PENDING_USERS[user_id].get('status') == 'xm_awaiting_screenshot':
        # Handle XM verification screenshot
        username = message.from_user.username or "No Username"
        if username != "No Username":
            username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)
        
        # Forward screenshot to admins WITH enhanced context and inline buttons
        for admin in ADMIN_IDS:
            # Forward the actual screenshot
            bot.forward_message(admin, chat_id, message.message_id)
            
            # Send explanatory message with verification buttons
            markup = InlineKeyboardMarkup()
            markup.add(
                InlineKeyboardButton("âœ… Approve XM", callback_data=f"approve_xm_{user_id}"),
                InlineKeyboardButton("âŒ Reject XM", callback_data=f"reject_xm_{user_id}")
            )

            bot.send_message(
                admin,
                f"ðŸ”” *XM Partnership Verification Request:*\n\n"
                f"User @{username} (ID: `{user_id}`) has submitted a screenshot showing their XM account with partner code.\n\n"
                f"Please verify that:\n"
                f"â€¢ The partner code PTAPARTNER is correctly applied\n"
                f"â€¢ The screenshot is from a valid XM account\n\n"
                f"Then approve or reject this verification request.",
                reply_markup=markup,
                parse_mode="Markdown"
            )
        
        # Update user status
        PENDING_USERS[user_id]['status'] = 'xm_waiting_approval'
        save_pending_users()
        
        # Send confirmation message to user
        bot.send_message(
            chat_id,
            "âœ… *XM Verification Submitted*\n\n"
            "Your XM account verification screenshot has been sent to our admins for review.\n\n"
            "You will be notified once your verification is approved or rejected.",
            parse_mode="Markdown"
        )
        return
    
    # Otherwise, continue with original payment verification logic
    if chat_id not in PENDING_USERS or PENDING_USERS[chat_id]['status'] != 'awaiting_proof':
        bot.send_message(chat_id, "âŒ Please start verification with `/verify`.")
        return

    user_id = message.from_user.id
    username = message.from_user.username or "No Username"
    
    # Escape Markdown characters in username
    if username != "No Username":
        username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)

    # Get mentorship type and plan details from PENDING_USERS
    if user_id in PENDING_USERS:
        plan = PENDING_USERS[chat_id]['plan']
        mentorship_type = PENDING_USERS[chat_id].get('mentorship_type', 'Regular')
        price_usd = PENDING_USERS[chat_id].get('price_usd', 0)  # Base discounted price
        original_price_usd = PENDING_USERS[chat_id].get('original_price_usd', price_usd)  # Original price
        price = PENDING_USERS[chat_id].get('price', '')
        method = PENDING_USERS[chat_id]['method']
        
        # Get discount information for context
        discount_percentage = PENDING_USERS[chat_id].get('discount_percentage', 0)
        discount_applied = discount_percentage > 0
        
        # Calculate adjusted price if payment method has a fee
        fee_percentage = PAYMENT_FEES.get(method, 0.0)
        
        if fee_percentage > 0:
            # Apply fee to the already-discounted price
            adjusted_price_usd = price_usd * (1 + fee_percentage/100)
            adjusted_price = f"${adjusted_price_usd:.2f}"
            
            # Create complete price display showing both discount and fee if applicable
            if discount_applied:
                price_display = f"${original_price_usd:.2f} â†’ ${price_usd:.2f} ({discount_percentage}% off) â†’ {adjusted_price} (+{fee_percentage}% fee)"
            else:
                price_display = f"{price} â†’ {adjusted_price} (+{fee_percentage}% fee)"
        else:
            # No fee, but might have discount
            if discount_applied:
                price_display = f"${original_price_usd:.2f} â†’ {price} ({discount_percentage}% off)"
            else:
                price_display = price
        
        # Calculate due date based on the plan duration
        if "Yearly" in plan or plan == "Regular Legacy":
            due_date = datetime.now() + timedelta(days=365)
        elif "3 Months" in plan or plan == "Momentum" or plan == "Apprentice":
            due_date = datetime.now() + timedelta(days=90)  # 3 months
        elif "6 Months" in plan or plan == "Disciple":
            due_date = datetime.now() + timedelta(days=180)  # 6 months
        elif "Lifetime" in plan or plan == "Supreme Legacy":
            # For lifetime, set a very long expiration (10 years)
            due_date = datetime.now() + timedelta(days=3650)  # ~10 years
        else:
            due_date = datetime.now() + timedelta(days=30)  # Default to monthly
    else:
        plan = "Monthly"
        mentorship_type = "Regular"
        price_display = ""
        method = "Unknown"
        due_date = datetime.now() + timedelta(days=30)  # Default to monthly if no plan is found

    USER_PAYMENT_DUE[user_id] = due_date

    # Forward the screenshot to Admins WITH enhanced payment details and inline buttons
    for admin in ADMIN_IDS:
        markup = InlineKeyboardMarkup()
        markup.add(InlineKeyboardButton("Approve", callback_data=f"approve_payment_{user_id}"))
        markup.add(InlineKeyboardButton("Reject", callback_data=f"reject_payment_{user_id}"))

        bot.forward_message(admin, chat_id, message.message_id)

        price_usd = PENDING_USERS[chat_id].get('price_usd', 0)  # Base price
        fee_adjusted_price = PENDING_USERS[chat_id].get('final_price_usd', price_usd)  # With payment fee if applicable

        exchange_rates = get_exchange_rates()

        currency_info = ""
        if exchange_rates:
            currency_info = "\n\nðŸ’± *Currency Equivalents*:\n"
            
            # Define currency symbols for better display
            currency_symbols = {
                'USD': '$', 'GBP': 'Â£', 'EUR': 'â‚¬', 'IDR': 'Rp', 'PHP': 'â‚±'
            }
            
            for currency, rate in exchange_rates.items():
                if currency == 'USD':
                    continue  # Skip USD as it's already shown
                
                # Calculate equivalent price in this currency
                equivalent_price = fee_adjusted_price * rate
                
                symbol = currency_symbols.get(currency, '')
                
                # Format based on currency
                if currency in ['IDR', 'PHP']:
                    formatted_price = f"{symbol}{equivalent_price:,.0f}"
                else:
                    # For currencies like GBP and EUR, always use .99 ending for consistent pricing
                    whole_price = int(equivalent_price)
                    formatted_price = f"{symbol}{whole_price}.99"
                
                currency_info += f"â€¢ *{currency}*: {formatted_price}\n"

        bot.send_message(admin,
            f"ðŸ”” *Payment Request:*\n"
            f"Someone is waiting for a payment approval. Here are the details:\n\n"
            f"ðŸ†” @{username} (ID: `{user_id}`)\n"
            f"ðŸ’³ Mode of Payment: {method}\n"
            f"ðŸ“… Mentorship Type: {mentorship_type.capitalize()}\n"
            f"ðŸ“… Plan: {plan} ({price_display})\n"  # Now shows adjusted price if needed
            f"ðŸ“… Due Date: {USER_PAYMENT_DUE[user_id].strftime('%Y-%m-%d %H:%M:%S')}"
            f"{currency_info}",
            reply_markup=markup,
            parse_mode="Markdown"
        )

    PENDING_USERS[chat_id]['status'] = 'waiting_approval'
    PENDING_USERS[chat_id]['request_time'] = datetime.now()  # Add timestamp
    delete_pending_user(user_id)
    save_pending_users()
    bot.send_message(chat_id, random.choice(payment_review_messages), parse_mode="Markdown")

# Admin Approves Payment
@bot.callback_query_handler(func=lambda call: call.data.startswith("approve_payment_"))
def callback_approve_payment(call):
    user_id = int(call.data.split("_")[2])
    if call.message.chat.id not in ADMIN_IDS:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to use this action.")
        return

    try:
        # Check if this is a renewal (special case)
        is_renewal = False
        if user_id in PENDING_USERS and PENDING_USERS[user_id].get('status') in ['waiting_approval'] and str(user_id) in PAYMENT_DATA:
            # This is likely a renewal if they went through the workflow and already have payment data
            is_renewal = True

        # Only check if already paid when it's NOT a renewal
        if not is_renewal and str(user_id) in PAYMENT_DATA and PAYMENT_DATA[str(user_id)]['haspayed']:
            bot.answer_callback_query(call.id, "âš ï¸ This user has already been approved.")
            return

        # Determine the plan and payment mode
        if user_id in PENDING_USERS:
            plan = PENDING_USERS[user_id].get('plan', 'Monthly')  # Default to 'Monthly' if not found
            payment_mode = PENDING_USERS[user_id].get('method', 'Unknown')  # Default to 'Unknown' if not found
            mentorship_type = PENDING_USERS[user_id].get('mentorship_type', 'Regular')  # Get mentorship type with fallback
            
            # Calculate due date based on the plan duration
            if "Yearly" in plan or plan == "Regular Legacy":
                due_date = datetime.now() + timedelta(days=365)
            elif "3 Months" in plan or plan == "Momentum" or plan == "Apprentice":
                due_date = datetime.now() + timedelta(days=90)  # 3 months
            elif "6 Months" in plan or plan == "Disciple":
                due_date = datetime.now() + timedelta(days=180)  # 6 months
            elif "Lifetime" in plan or plan == "Supreme Legacy":
                # For lifetime, set a very long expiration (10 years)
                due_date = datetime.now() + timedelta(days=3650)  # ~10 years
            else:
                due_date = datetime.now() + timedelta(days=30)  # Default to monthly

            PENDING_USERS.pop(user_id, None)  # Remove from pending list
            delete_pending_user(user_id)
        else:
            plan = 'Monthly'
            payment_mode = 'Unknown'
            mentorship_type = 'Regular'  # Default to Regular if not in PENDING_USERS
            due_date = datetime.now() + timedelta(days=30)  # Default to monthly if no plan is found

        save_pending_users()

        # Get user info directly from Telegram to ensure correct username
        try:
            user_info = bot.get_chat(user_id)
            username = user_info.username or "No Username"
        except Exception:
            username = "No Username"  # Fallback if can't get username

        # FIX: Preserve existing data when saving payment approval
        if str(user_id) in PAYMENT_DATA:
            # Get existing data first to preserve terms/privacy acceptance
            existing_data = PAYMENT_DATA[str(user_id)]
            
            # Update only the payment-related fields
            existing_data.update({
                "username": username,  # Use the user's username instead of admin's
                "payment_plan": plan,
                "payment_mode": payment_mode,
                "mentorship_type": mentorship_type,
                "due_date": due_date.strftime('%Y-%m-%d %H:%M:%S'),
                "haspayed": True
            })
            
            # The existing terms_accepted and privacy_accepted will be preserved
        else:
            # No existing data - create new entry
            PAYMENT_DATA[str(user_id)] = {
                "username": username,
                "payment_plan": plan,
                "payment_mode": payment_mode,
                "mentorship_type": mentorship_type,
                "due_date": due_date.strftime('%Y-%m-%d %H:%M:%S'),
                "haspayed": True,
                "terms_accepted": False,  # Default values
                "privacy_accepted": False  # Default values
            }
        
        logging.info(f"Saving payment data for user {user_id}: {PAYMENT_DATA[str(user_id)]}")
        save_payment_data()  # Ensure this function is called to save the data

        # âœ… Check if the bot can message the user
        try:
            bot.send_chat_action(user_id, 'typing')  # Check if the user exists
        except ApiException:
            bot.answer_callback_query(call.id, "âŒ Error: I can't message this user. They need to start the bot first.")
            return

        # Log admin activity and notify all admins
        admin_username = call.from_user.username or f"Admin ({call.message.chat.id})"
        user_info = bot.get_chat(user_id)
        username = user_info.username or f"ID: {user_id}"

        # Escape Markdown characters in the usernames
        username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)
        admin_username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', admin_username)

        for admin_id in ADMIN_IDS:
            bot.send_message(admin_id, f"ðŸ“ *Activity Log*\n\n{admin_username} has approved payment from PTA member @{username}.", parse_mode="Markdown")

        # After successful payment approval:
        try:
            # First send the payment approved graphic
            try:
                with open('graphics/payment_approved.jpeg', 'rb') as approval_img:
                    bot.send_photo(
                        user_id,
                        approval_img,
                        caption="Your payment has been approved!"
                    )
            except FileNotFoundError:
                logging.error("Payment approval image not found at graphics/payment_approved.jpeg")
            except Exception as e:
                logging.error(f"Error sending payment approval image: {e}")
                
            # Add a small delay for better UX
            time.sleep(1.5)
            
            # âœ… Step 1: Verification successful (keep the text confirmation too)
            bot.send_message(user_id, random.choice(payment_approval_messages), parse_mode="Markdown")
            bot.answer_callback_query(call.id, "âœ… Payment approved successfully.")

            # ðŸ“… Step 2: Determine and send due date
            USER_PAYMENT_DUE[user_id] = due_date
            bot.send_message(user_id, f"ðŸ“… *Your next payment is due on:* {due_date.strftime('%Y/%m/%d %I:%M:%S %p')}.")

            # Send registration form graphic first
            try:
                with open('graphics/registration_form.jpeg', 'rb') as form_img:
                    bot.send_photo(
                        user_id,
                        form_img,
                        caption="Please complete the registration form to continue"
                    )
            except FileNotFoundError:
                logging.error("Registration form image not found at graphics/registration_form.jpeg")
            except Exception as e:
                logging.error(f"Error sending registration form image: {e}")
                
            # Add a small delay for better UX
            time.sleep(1.5)

            # Add transition message before moving to onboarding
            bot.send_chat_action(user_id, 'typing')
            transition_msg = bot.send_message(user_id, "â³ Loading Registration Forms...")
            
            # Add a slight delay for better UX
            time.sleep(1)
            
            try:
                # Delete the transition message after delay
                bot.delete_message(user_id, transition_msg.message_id)
            except Exception as e:
                logging.error(f"Error deleting transition message: {e}")
            
                # ðŸ“ Step 3: Send onboarding form
                target_group_id = SUPREME_GROUP_ID if mentorship_type.lower() == 'supreme' else PAID_GROUP_ID

                # --- Add check if user is already in the target group ---
                already_in_group = False
                try:
                    chat_member = bot.get_chat_member(target_group_id, user_id)
                    # If we get here without exception, user is already in the group
                    already_in_group = True
                    logging.info(f"User {user_id} is already in group {target_group_id}, will skip invite during onboarding")
                except ApiException:
                    # User is not in the chat yet
                    already_in_group = False
                    logging.info(f"User {user_id} is not in group {target_group_id}, will generate invite during onboarding")

                PENDING_USERS[user_id] = {
                    'form_answers': {},  # Initialize empty dict to store responses
                    'invite_link': None,  # Will store the invite link to use after form completion
                    'target_group_id': target_group_id,  # Store which group to invite to
                    'already_in_group': already_in_group,  # Add this flag
                    'membership_type': mentorship_type  # Make sure membership type is carried through
                }
                save_pending_users()
            
        except Exception as e:
            bot.answer_callback_query(call.id, f"âŒ Error sending onboarding form: {e}")
            logging.error(f"Error in onboarding form: {e}")

        # ðŸ”’ Step 6: Ensure bot is an admin before adding restrictions
        try:
            # Check if user is a chat member first
            chat_member = None
            try:
                chat_member = bot.get_chat_member(PAID_GROUP_ID, user_id)
            except ApiException:
                # User is not in the chat yet, which is fine
                pass
            
            # Only try to restrict if user is not the chat owner
            if not chat_member or chat_member.status != "creator":
                bot.restrict_chat_member(PAID_GROUP_ID, user_id, can_send_messages=True)
        except ApiException as e:
            bot.send_message(call.message.chat.id, f"âš ï¸ Note: User permissions were not modified. Error: {e}")

        # Then call send_onboarding_form which will set the proper status
        send_onboarding_form(user_id)

    except Exception as e:
        bot.answer_callback_query(call.id, f"âŒ Unexpected error approving payment: {e}")

def send_onboarding_form(user_id):
    """Send the first question of the onboarding form to the user"""
    # Check if the user has already completed the form previously
    user_id_str = str(user_id)
    
    # Look for existing form_answers in PAYMENT_DATA
    if user_id_str in PAYMENT_DATA and PAYMENT_DATA[user_id_str].get('form_answers'):
        logging.info(f"User {user_id} has already completed forms - skipping to final steps")
        
        # Get existing data for passing to complete_onboarding
        membership_type = PAYMENT_DATA[user_id_str].get('mentorship_type', 'regular').lower()
        
        # Get is_renewal flag from existing PENDING_USERS data or default to True
        # This ensures admin notifications are skipped for users with existing forms
        is_renewal = PENDING_USERS.get(user_id, {}).get('is_renewal', True)
        
        # Prepare minimal PENDING_USERS data - just enough for the complete_onboarding function
        PENDING_USERS[user_id] = {
            'status': 'completed_onboarding',
            'form_answers': PAYMENT_DATA[user_id_str]['form_answers'],
            'membership_type': membership_type,
            'invite_link': PENDING_USERS.get(user_id, {}).get('invite_link') if user_id in PENDING_USERS else None,
            'target_group_id': SUPREME_GROUP_ID if membership_type.lower() == 'supreme' else PAID_GROUP_ID,
            'is_renewal': is_renewal  # Always set is_renewal when reusing forms
        }
        save_pending_users()
        
        # Send a message that we're using existing form data
        bot.send_message(
            user_id,
            "âœ… *Welcome back!*\n\nWe noticed you've already filled out our onboarding form. We'll use your previous responses.",
            parse_mode="Markdown"
        )
        
        # Skip directly to the final step of onboarding
        complete_onboarding(user_id)
        return
    
    # Original function continues here for new users
    # Show typing indicator for better UX
    bot.send_chat_action(user_id, 'typing')
    time.sleep(1)
    
    # Determine which form to send based on membership type
    membership_type = "regular"  # Default to regular
    if user_id_str in PAYMENT_DATA:
        membership_type = PAYMENT_DATA[user_id_str].get('mentorship_type', 'regular').lower()
    
    # Initialize form data structure - Store with BOTH user_id (int) and the string version
    # This ensures we have consistent keys no matter how they're accessed
    PENDING_USERS[user_id] = {
        'status': 'onboarding_form_regular_step1' if membership_type == 'regular' else 'onboarding_form_supreme_step1',
        'form_answers': {},
        'membership_type': membership_type,
        'invite_link': PENDING_USERS.get(user_id, {}).get('invite_link') if user_id in PENDING_USERS else None,
        'target_group_id': PENDING_USERS.get(user_id, {}).get('target_group_id', PAID_GROUP_ID) if user_id in PENDING_USERS else PAID_GROUP_ID,
        'user_id_str': user_id_str  # Add this for clarity
    }
    save_pending_users()
    
    # Log user state to help debug
    logging.info(f"Initialized onboarding form for user {user_id} with status: {PENDING_USERS[user_id]['status']}")
    
    # Send welcome message with form explanation
    welcome_msg = "ðŸŽ‰ *Welcome to Prodigy Trading Academy!*\n\n"
    
    if membership_type == 'supreme':
        welcome_msg += "As a *Supreme Mentorship* member, we'd like to learn more about you to personalize your experience. Please take a few moments to complete this 10-question onboarding survey:"
        
        # Send the first question for supreme membership with example
        bot.send_message(
            user_id,
            welcome_msg + "\n\nðŸ“ *Question 1:* Full Name (ex. Juan Dela Cruz)",
            parse_mode="Markdown"
        )
    else:
        welcome_msg += "Before we add you to our community group, we'd like to learn a bit about you. Please take a moment to answer these 5 quick questions:"
        
        # Send the first question for regular membership (Full Name) with example
        bot.send_message(
            user_id,
            welcome_msg + "\n\nðŸ“ *Question 1:* Full Name (ex. Juan Dela Cruz)",
            parse_mode="Markdown"
        )


# REGULAR MEMBERSHIP FORM HANDLERS

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_regular_step1')
def handle_regular_form_step1(message):
    user_id = message.from_user.id
    
    # Store the answer - this is the full name question
    PENDING_USERS[user_id]['form_answers']['full_name'] = message.text
    PENDING_USERS[user_id]['status'] = 'onboarding_form_regular_step2'
    save_pending_users()
    
    # Show typing indicator
    bot.send_chat_action(user_id, 'typing')
    time.sleep(1.5)
    
    # Send second question for Birthday (changed format to MM/DD/YYYY)
    bot.send_message(
        user_id,
        "ðŸ“ *Question 2:* Birthday (MM/DD/YYYY format, e.g., 06/15/1990)",
        parse_mode="Markdown"
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_regular_step2')
def handle_regular_form_step2(message):
    user_id = message.from_user.id
    birthday_input = message.text
    
    # Validate birthday format (MM/DD/YYYY)
    try:
        # Check if input matches MM/DD/YYYY pattern
        if not re.match(r"^\d{2}/\d{2}/\d{4}$", birthday_input):
            bot.send_message(
                user_id,
                "âŒ Invalid birthday format. Please use MM/DD/YYYY format (e.g., 06/15/1990)."
            )
            return
            
        # Try to parse the date to ensure it's valid
        month, day, year = map(int, birthday_input.split('/'))
        birthday_date = datetime(year, month, day)
        
        # Check if birthday is in the past
        if birthday_date > datetime.now():
            bot.send_message(
                user_id,
                "âŒ Birthday cannot be in the future. Please enter a valid date."
            )
            return
        
        # Store the birthday temporarily and move to confirmation step
        PENDING_USERS[user_id]['temp_birthday'] = birthday_input
        PENDING_USERS[user_id]['status'] = 'onboarding_form_regular_step2_confirm'
        save_pending_users()
        
        # Format the date in a human-readable format for confirmation
        formatted_date = birthday_date.strftime("%B %d, %Y")  # Format as "June 15, 1990"
        
        # Create keyboard for confirmation
        markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        markup.row("âœ… Yes, that's correct", "âŒ No, I need to re-enter")
        
        # Send confirmation message
        bot.send_message(
            user_id,
            f"ðŸ“… You entered: *{formatted_date}*\n\nIs this correct?",
            parse_mode="Markdown",
            reply_markup=markup
        )
        
    except ValueError:
        bot.send_message(
            user_id,
            "âŒ Invalid date. Please enter a valid birthday in MM/DD/YYYY format (e.g., 06/15/1990)."
        )

# Add a new handler for the birthday confirmation step
@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_regular_step2_confirm')
def handle_regular_form_step2_confirm(message):
    user_id = message.from_user.id
    response = message.text.lower()
    
    # Check the user's response
    if "yes" in response or "correct" in response:
        # User confirmed the birthday is correct
        
        # Get the temporary birthday we stored
        birthday = PENDING_USERS[user_id].get('temp_birthday')
        
        # Store the answer in form_answers
        PENDING_USERS[user_id]['form_answers']['birthday'] = birthday
        
        # Remove temporary birthday storage
        if 'temp_birthday' in PENDING_USERS[user_id]:
            del PENDING_USERS[user_id]['temp_birthday']
        
        # Move to step 3
        PENDING_USERS[user_id]['status'] = 'onboarding_form_regular_step3'
        save_pending_users()
        
        # Show typing indicator
        bot.send_chat_action(user_id, 'typing')
        time.sleep(1.5)
        
        # Remove keyboard
        markup = ReplyKeyboardRemove()
        
        # Send third question with multiple choice options for experience level
        experience_markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        experience_markup.row("a. Completely new")
        experience_markup.row("b. Beginner")
        experience_markup.row("c. Intermediate")
        experience_markup.row("d. Expert")
        experience_markup.row("e. Master")
        
        bot.send_message(
            user_id,
            "ðŸ“ *Question 3:* Are you completely new to trading or do you have some experience?",
            parse_mode="Markdown",
            reply_markup=experience_markup
        )
    else:
        # User wants to re-enter the birthday
        PENDING_USERS[user_id]['status'] = 'onboarding_form_regular_step2'
        save_pending_users()
        
        # Remove keyboard
        markup = ReplyKeyboardRemove()
        
        # Send message to retry
        bot.send_message(
            user_id,
            "ðŸ“… Please re-enter your birthday in MM/DD/YYYY format (e.g., 06/15/1990):",
            reply_markup=markup
        )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_regular_step3')
def handle_regular_form_step3(message):
    user_id = message.from_user.id
    
    # Store the answer - this is the experience level
    PENDING_USERS[user_id]['form_answers']['experience_level'] = message.text
    PENDING_USERS[user_id]['status'] = 'onboarding_form_regular_step4'
    save_pending_users()
    
    # Show typing indicator
    bot.send_chat_action(user_id, 'typing')
    time.sleep(1.5)
    
    # Send fourth question about learning goals
    bot.send_message(
        user_id,
        "ðŸ“ *Question 4:* What do you hope to learn inside the academy?",
        parse_mode="Markdown"
    )
    
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row("a. Technical and Fundamental Analysis")
    markup.row("b. Trading Psychology")
    markup.row("c. Risk Management")
    markup.row("d. Full Path to Financial Freedom")
    markup.row("e. All of the above")
    
    bot.send_message(
        user_id,
        "Please select what you hope to learn:",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_regular_step4')
def handle_regular_form_step4(message):
    user_id = message.from_user.id
    
    # Store the answer - this is what they hope to learn
    PENDING_USERS[user_id]['form_answers']['learning_goals'] = message.text
    PENDING_USERS[user_id]['status'] = 'onboarding_form_regular_step5'
    save_pending_users()
    
    # Show typing indicator
    bot.send_chat_action(user_id, 'typing')
    time.sleep(1.5)
    
    # Send fifth question about how they found PTA
    bot.send_message(
        user_id,
        "ðŸ“ *Question 5:* How did you find out about Prodigy Trading Academy?",
        parse_mode="Markdown"
    )
    
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row("a. Socials (Telegram, Facebook, Instagram)")
    markup.row("b. Referred to by a friend, family or expert")
    markup.row("c. Invited by executives or founders themselves")
    markup.row("d. Other")
    
    bot.send_message(
        user_id,
        "Please select how you found us:",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_regular_step5')
def handle_regular_form_step5(message):
    user_id = message.from_user.id
    
    # Store the answer - this is how they found PTA
    PENDING_USERS[user_id]['form_answers']['source'] = message.text
    
    # Remove custom keyboard
    markup = ReplyKeyboardRemove()
    
    # Show typing indicator
    bot.send_chat_action(user_id, 'typing')
    time.sleep(1.5)
    
    # Send registration complete graphic with the completion message as caption
    try:
        with open('graphics/registration_done.jpeg', 'rb') as complete_img:
            bot.send_photo(
                user_id,
                complete_img,
                caption="âœ… *Form Complete!*\n\n"
                "Thank you for sharing this information with us. This helps us "
                "better understand your needs and tailor our community support.\n\n"
                "Now, let's get you connected with the Prodigy Trading Academy community!",
                parse_mode="Markdown",
                reply_markup=markup
            )
    except FileNotFoundError:
        logging.error("Registration complete image not found at graphics/registration_done.jpeg")
        # Fallback to text-only message
        bot.send_message(
            user_id,
            "âœ… *Form Complete!*\n\n"
            "Thank you for sharing this information with us. This helps us "
            "better understand your needs and tailor our community support.\n\n"
            "Now, let's get you connected with the Prodigy Trading Academy community!",
            parse_mode="Markdown",
            reply_markup=markup
        )
    except Exception as e:
        logging.error(f"Error sending registration complete image: {e}")
        # Fallback to text-only message
        bot.send_message(
            user_id,
            "âœ… *Form Complete!*\n\n"
            "Thank you for sharing this information with us. This helps us "
            "better understand your needs and tailor our community support.\n\n"
            "Now, let's get you connected with the Prodigy Trading Academy community!",
            parse_mode="Markdown",
            reply_markup=markup
        )
    
    # Now proceed with the welcome package and group invite
    complete_onboarding(user_id)

# SUPREME MEMBERSHIP FORM HANDLERS

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_supreme_step1')
def handle_supreme_form_step1(message):
    user_id = message.from_user.id
    
    # Store the answer - this is the full name question
    PENDING_USERS[user_id]['form_answers']['full_name'] = message.text
    PENDING_USERS[user_id]['status'] = 'onboarding_form_supreme_step2'
    save_pending_users()
    
    # Show typing indicator
    bot.send_chat_action(user_id, 'typing')
    time.sleep(1.5)
    
    # Send second question for Birthday (changed format to MM/DD/YYYY)
    bot.send_message(
        user_id,
        "ðŸ“ *Question 2:* Birthday (MM/DD/YYYY format, e.g., 06/15/1990)",
        parse_mode="Markdown"
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_supreme_step2')
def handle_supreme_form_step2(message):
    user_id = message.from_user.id
    birthday_input = message.text
    
    # Validate birthday format (MM/DD/YYYY)
    try:
        # Check if input matches MM/DD/YYYY pattern
        if not re.match(r"^\d{2}/\d{2}/\d{4}$", birthday_input):
            bot.send_message(
                user_id,
                "âŒ Invalid birthday format. Please use MM/DD/YYYY format (e.g., 06/15/1990)."
            )
            return
            
        # Try to parse the date to ensure it's valid
        month, day, year = map(int, birthday_input.split('/'))
        birthday_date = datetime(year, month, day)
        
        # Check if birthday is in the past
        if birthday_date > datetime.now():
            bot.send_message(
                user_id,
                "âŒ Birthday cannot be in the future. Please enter a valid date."
            )
            return
        
        # Store the birthday temporarily and move to confirmation step
        PENDING_USERS[user_id]['temp_birthday'] = birthday_input
        PENDING_USERS[user_id]['status'] = 'onboarding_form_supreme_step2_confirm'
        save_pending_users()
        
        # Format the date in a human-readable format for confirmation
        formatted_date = birthday_date.strftime("%B %d, %Y")  # Format as "June 15, 1990"
        
        # Create keyboard for confirmation
        markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        markup.row("âœ… Yes, that's correct", "âŒ No, I need to re-enter")
        
        # Send confirmation message
        bot.send_message(
            user_id,
            f"ðŸ“… You entered: *{formatted_date}*\n\nIs this correct?",
            parse_mode="Markdown",
            reply_markup=markup
        )
        
    except ValueError:
        bot.send_message(
            user_id,
            "âŒ Invalid date. Please enter a valid birthday in MM/DD/YYYY format (e.g., 06/15/1990)."
        )

# Add a new handler for the birthday confirmation step
@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_supreme_step2_confirm')
def handle_supreme_form_step2_confirm(message):
    user_id = message.from_user.id
    response = message.text.lower()
    
    # Check the user's response
    if "yes" in response or "correct" in response:
        # User confirmed the birthday is correct
        
        # Get the temporary birthday we stored
        birthday = PENDING_USERS[user_id].get('temp_birthday')
        
        # Store the answer in form_answers
        PENDING_USERS[user_id]['form_answers']['birthday'] = birthday
        
        # Remove temporary birthday storage
        if 'temp_birthday' in PENDING_USERS[user_id]:
            del PENDING_USERS[user_id]['temp_birthday']
        
        # Move to step 3
        PENDING_USERS[user_id]['status'] = 'onboarding_form_supreme_step3'
        save_pending_users()
        
        # Show typing indicator
        bot.send_chat_action(user_id, 'typing')
        time.sleep(1.5)
        
        # Remove keyboard
        markup = ReplyKeyboardRemove()
        
        # Send third question for Phone Number
        bot.send_message(
            user_id,
            "ðŸ“ *Question 3:* Phone Number (include country code, ex. +63 917 123 4567 for Philippines or +1 555 123 4567 for US)",
            parse_mode="Markdown",
            reply_markup=markup
        )
    else:
        # User wants to re-enter the birthday
        PENDING_USERS[user_id]['status'] = 'onboarding_form_supreme_step2'
        save_pending_users()
        
        # Remove keyboard
        markup = ReplyKeyboardRemove()
        
        # Send message to retry
        bot.send_message(
            user_id,
            "ðŸ“… Please re-enter your birthday in MM/DD/YYYY format (e.g., 06/15/1990):",
            reply_markup=markup
        )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_supreme_step3')
def handle_supreme_form_step3(message):
    user_id = message.from_user.id
    phone = message.text
    
    # Store the answer - this is the phone number
    PENDING_USERS[user_id]['form_answers']['phone_number'] = phone
    PENDING_USERS[user_id]['status'] = 'onboarding_form_supreme_step4'
    save_pending_users()
    
    # Show typing indicator
    bot.send_chat_action(user_id, 'typing')
    time.sleep(1.5)
    
    # Send fourth question for Time Zone with preset options
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row("PHT (Philippines)", "SGT (Singapore)")
    markup.row("PST (Pacific)", "EST (Eastern)")
    markup.row("Other (please specify)")
    
    bot.send_message(
        user_id,
        "ðŸ“ *Question 4:* Current Time Zone",
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_supreme_step4')
def handle_supreme_form_step4(message):
    user_id = message.from_user.id
    timezone = message.text
    
    # Check if the user selected "Other (please specify)"
    if timezone == "Other (please specify)":
        # Set a different status for custom timezone entry
        PENDING_USERS[user_id]['status'] = 'onboarding_form_supreme_step4_custom'
        save_pending_users()
        
        # Show typing indicator
        bot.send_chat_action(user_id, 'typing')
        time.sleep(1.5)
        
        # Ask for the specific time zone
        bot.send_message(
            user_id,
            "ðŸ“ Please enter your specific time zone:",
            parse_mode="Markdown"
        )
        return
    
    # Regular flow for predefined options
    # Store the answer - this is the time zone
    PENDING_USERS[user_id]['form_answers']['time_zone'] = timezone
    PENDING_USERS[user_id]['status'] = 'onboarding_form_supreme_step5'
    save_pending_users()
    
    # Show typing indicator
    bot.send_chat_action(user_id, 'typing')
    time.sleep(1.5)
    
    # Send fifth question for Trading Expertise Level
    bot.send_message(
        user_id,
        "ðŸ“ *Question 5:* Trading Expertise Level",
        parse_mode="Markdown"
    )
    
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row("a. Completely New")
    markup.row("b. Beginner")
    markup.row("c. Intermediate")
    markup.row("d. Expert")
    markup.row("e. Master")
    
    bot.send_message(
        user_id,
        "Please select your expertise level:",
        reply_markup=markup
    )

# Add a new handler for custom timezone entry
@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_supreme_step4_custom')
def handle_supreme_form_step4_custom(message):
    user_id = message.from_user.id
    custom_timezone = message.text
    
    # Store the answer - this is the custom time zone
    PENDING_USERS[user_id]['form_answers']['time_zone'] = f"Custom: {custom_timezone}"
    PENDING_USERS[user_id]['status'] = 'onboarding_form_supreme_step5'
    save_pending_users()
    
    # Show typing indicator
    bot.send_chat_action(user_id, 'typing')
    time.sleep(1.5)
    
    # Send fifth question for Trading Expertise Level
    bot.send_message(
        user_id,
        "ðŸ“ *Question 5:* Trading Expertise Level",
        parse_mode="Markdown"
    )
    
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row("a. Completely New")
    markup.row("b. Beginner")
    markup.row("c. Intermediate")
    markup.row("d. Expert")
    markup.row("e. Master")
    
    bot.send_message(
        user_id,
        "Please select your expertise level:",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_supreme_step5')
def handle_supreme_form_step5(message):
    user_id = message.from_user.id
    expertise = message.text
    
    # Store the answer - this is the trading expertise level
    PENDING_USERS[user_id]['form_answers']['expertise_level'] = expertise
    PENDING_USERS[user_id]['status'] = 'onboarding_form_supreme_step6'
    save_pending_users()
    
    # Show typing indicator
    bot.send_chat_action(user_id, 'typing')
    time.sleep(1.5)
    
    # Send sixth question for Part-time vs. Full-time
    bot.send_message(
        user_id,
        "ðŸ“ *Question 6:* Will you be a part-time or full-time trader? Why?",
        parse_mode="Markdown"
    )
    
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row("a. Part-time, I'm currently employed and trading on the side.")
    markup.row("b. Part-time, I have other responsibilities but want to gradually build my trading career.")
    markup.row("c. Full-time, I'm fully committed and have the time and resources to focus on trading.")
    markup.row("d. Full-time, I'm a student and want to dedicate my free time to becoming a funded trader.")
    
    bot.send_message(
        user_id,
        "Please select the option that best describes you:",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_supreme_step6')
def handle_supreme_form_step6(message):
    user_id = message.from_user.id
    trading_time = message.text
    
    # Store the answer - this is the part-time vs. full-time answer
    PENDING_USERS[user_id]['form_answers']['trading_time_commitment'] = trading_time
    PENDING_USERS[user_id]['status'] = 'onboarding_form_supreme_step7'
    save_pending_users()
    
    # Show typing indicator
    bot.send_chat_action(user_id, 'typing')
    time.sleep(1.5)
    
    # Send seventh question - Interest in Supreme mentorship
    bot.send_message(
        user_id,
        "ðŸ“ *Question 7:* What drew your interest to join the Supreme mentorship specifically?",
        parse_mode="Markdown"
    )
    
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row("a. The premium content and structured education.")
    markup.row("b. The fun and engaging way you teach and discuss.")
    markup.row("c. It's affordable and still high quality.")
    markup.row("d. The uniqueness and overall vibe of the group.")
    markup.row("e. All of the above.")
    
    bot.send_message(
        user_id,
        "Please select your reason:",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_supreme_step7')
def handle_supreme_form_step7(message):
    user_id = message.from_user.id
    interest_reason = message.text
    
    # Store the answer - interest in supreme mentorship
    PENDING_USERS[user_id]['form_answers']['interest_reason'] = interest_reason
    PENDING_USERS[user_id]['status'] = 'onboarding_form_supreme_step8'
    save_pending_users()
    
    # Show typing indicator
    bot.send_chat_action(user_id, 'typing')
    time.sleep(1.5)
    
    # Send eighth question - Personal expectations/goals
    bot.send_message(
        user_id,
        "ðŸ“ *Question 8:* What are your personal expectations or goals inside the Supreme group?",
        parse_mode="Markdown"
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_supreme_step8')
def handle_supreme_form_step8(message):
    user_id = message.from_user.id
    goals = message.text
    
    # Store the answer - personal goals
    PENDING_USERS[user_id]['form_answers']['personal_goals'] = goals
    PENDING_USERS[user_id]['status'] = 'onboarding_form_supreme_step9'
    save_pending_users()
    
    # Show typing indicator
    bot.send_chat_action(user_id, 'typing')
    time.sleep(1.5)
    
    # Send ninth question - Comfort with 1:1 calls
    bot.send_message(
        user_id,
        "ðŸ“ *Question 9:* Are you comfortable with regular 1:1 private calls or mentorship check-ins?",
        parse_mode="Markdown"
    )
    
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row("a. No, please do not schedule 1:1 mentoring sessions for me.")
    markup.row("b. Yes, but please only a few times or scheduled times in a week as I have a hectic schedule.")
    markup.row("c. Yes, I am willing and comfortable to do so anytime.")
    
    bot.send_message(
        user_id,
        "Please select your preference:",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_supreme_step9')
def handle_supreme_form_step9(message):
    user_id = message.from_user.id
    call_preference = message.text
    
    # Store the answer - 1:1 call preference
    PENDING_USERS[user_id]['form_answers']['call_preference'] = call_preference
    PENDING_USERS[user_id]['status'] = 'onboarding_form_supreme_step10'
    save_pending_users()
    
    # Show typing indicator
    bot.send_chat_action(user_id, 'typing')
    time.sleep(1.5)
    
    # Send tenth question - Trading challenges
    bot.send_message(
        user_id,
        "ðŸ“ *Question 10:* Are there any trading topics or challenges you've struggled with that you'd like us to focus on during mentorship?",
        parse_mode="Markdown"
    )
    
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row("a. Trading Psychology")
    markup.row("b. Risk Management")
    markup.row("c. Strategy and Technical Analysis")
    markup.row("d. Fundamental Analysis")
    markup.row("e. I am at ease with all of the choices above, please think of what is best to focus for me.")
    markup.row("f. I lack ability in all of the choices above as I am a beginner.")
    
    bot.send_message(
        user_id,
        "Please select your answer:",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'onboarding_form_supreme_step10')
def handle_supreme_form_step10(message):
    user_id = message.from_user.id
    challenges = message.text
    
    # Store the answer
    PENDING_USERS[user_id]['form_answers']['challenges'] = challenges
    
    # Remove custom keyboard
    markup = ReplyKeyboardRemove()
    
    # Show typing indicator
    bot.send_chat_action(user_id, 'typing')
    time.sleep(1.5)
    
    # Send registration complete graphic with completion message as caption
    try:
        with open('graphics/registration_done.jpeg', 'rb') as complete_img:
            bot.send_photo(
                user_id,
                complete_img,
                caption="âœ… *Supreme Mentorship Form Complete!*\n\n"
                "Thank you for taking the time to share this detailed information with us. "
                "As a Supreme member, this will help us tailor your 1-on-1 coaching sessions "
                "and personalized support to your specific trading goals and needs.\n\n"
                "Now, let's get you connected with the Prodigy Trading Academy Supreme community!",
                parse_mode="Markdown",
                reply_markup=markup
            )
    except FileNotFoundError:
        logging.error("Registration complete image not found at graphics/registration_done.jpeg")
        # Fallback to text-only message
        bot.send_message(
            user_id,
            "âœ… *Supreme Mentorship Form Complete!*\n\n"
            "Thank you for taking the time to share this detailed information with us. "
            "As a Supreme member, this will help us tailor your 1-on-1 coaching sessions "
            "and personalized support to your specific trading goals and needs.\n\n"
            "Now, let's get you connected with the Prodigy Trading Academy Supreme community!",
            parse_mode="Markdown",
            reply_markup=markup
        )
    except Exception as e:
        logging.error(f"Error sending registration complete image: {e}")
        # Fallback to text-only message
        bot.send_message(
            user_id,
            "âœ… *Supreme Mentorship Form Complete!*\n\n"
            "Thank you for taking the time to share this detailed information with us. "
            "As a Supreme member, this will help us tailor your 1-on-1 coaching sessions "
            "and personalized support to your specific trading goals and needs.\n\n"
            "Now, let's get you connected with the Prodigy Trading Academy Supreme community!",
            parse_mode="Markdown",
            reply_markup=markup
        )
    
    # Now proceed with the welcome package and group invite
    complete_onboarding(user_id)

def complete_onboarding(user_id):
    """Complete the onboarding process by sending welcome package and group invite"""
    try:
        # Add a transition message before sending welcome package
        bot.send_chat_action(user_id, 'typing')
        transition_msg = bot.send_message(user_id, "âœ¨ Finalizing your enrollment...")
        
        # Add a short delay for better UX
        time.sleep(1)
        
        try:
            # Delete the transition message
            bot.delete_message(user_id, transition_msg.message_id)
        except Exception as e:
            logging.error(f"Error deleting transition message: {e}")

        # IMPORTANT: Generate and send certificate BEFORE the welcome package
        send_completion_certificate(user_id)

        time.sleep(1)
            
        # Send the Welcome to Academy graphic
        try:
            with open('graphics/welcomePTA.jpeg', 'rb') as welcome_img:
                bot.send_photo(
                    user_id,
                    welcome_img,
                    caption="ðŸŽ‰ Welcome to Prodigy Trading Academy!"
                )
        except FileNotFoundError:
            logging.error("Welcome academy image not found at graphics/welcome_academy.jpeg")
        except Exception as e:
            logging.error(f"Error sending welcome academy image: {e}")

        # Add another small delay for better UX
        time.sleep(1.5)
        
        # 1. Send social media connections
        send_welcome_package(user_id)
        
        # 2. Send group invite - IMPROVED VERSION WITH ALREADY-IN-GROUP CHECK
        invite_link = PENDING_USERS[user_id].get('invite_link')
        already_in_group = PENDING_USERS[user_id].get('already_in_group', False)
        membership_type = PENDING_USERS[user_id].get('membership_type', 'regular')
        
        # Determine target group based on membership type
        target_group_id = SUPREME_GROUP_ID if membership_type == 'supreme' else PAID_GROUP_ID
        group_name = "Supreme Mentorship" if target_group_id == SUPREME_GROUP_ID else "Prodigy Trading Academy"
        
        # Check if user is already in the group - ADDED CHECK
        if already_in_group:
            # Notify user they're already in the group
            bot.send_message(
                user_id,
                f"âœ… *You're already a member!*\n\n"
                f"You're already in our {group_name} group. No need to join again.",
                parse_mode="Markdown"
            )
            logging.info(f"User {user_id} is already in group {target_group_id}, skipping invite")
        else:
            # If we don't have an invite link yet, generate one
            if not invite_link:
                try:
                    # Generate a new invite link - explicitly note it's a ChatInviteLink object
                    new_invite: ChatInviteLink = bot.create_chat_invite_link(
                        target_group_id,
                        name=f"User {user_id} onboarding",
                        creates_join_request=False,
                        member_limit=1,
                        expire_date=int((datetime.now() + timedelta(minutes=15)).timestamp())
                    )
                    invite_link = new_invite.invite_link
                    
                    # Save the new invite link
                    PENDING_USERS[user_id]['invite_link'] = invite_link
                    PENDING_USERS[user_id]['target_group_id'] = target_group_id
                    save_pending_users()
                    
                    logging.info(f"Generated new invite link for user {user_id} joining group {target_group_id}")
                except Exception as e:
                    logging.error(f"Failed to create invite link for user {user_id}: {e}")
                    
                    # Send error notification to admins
                    for admin_id in ADMIN_IDS:
                        bot.send_message(
                            admin_id,
                            f"âš ï¸ *Error generating invite link*\n\n"
                            f"Could not create invite link for user {user_id} during serial redemption flow.\n"
                            f"Error: {e}\n\n"
                            f"Please manually send an invite to this user.",
                            parse_mode="Markdown"
                        )
                    
                    # Notify user about the issue
                    bot.send_message(
                        user_id,
                        "âš ï¸ *Group Invite Pending*\n\n"
                        "We're currently having trouble generating your group invite link. "
                        "An admin has been notified and will send you an invite shortly.\n\n"
                        "Thank you for your patience!",
                        parse_mode="Markdown"
                    )
                    
            # If we have a valid invite link, send it along with disclaimer
            if invite_link:
                # ADDED: Send an important disclaimer about trading before group invite
                disclaimer_message = (
                    "âš ï¸ *IMPORTANT TRADING DISCLAIMER*\n\n"
                    "Before you join our trading community, please understand:\n\n"
                    "â€¢ Trading is NOT a 'get rich quick' scheme\n"
                    "â€¢ Trading does NOT guarantee easy money\n"
                    "â€¢ Our content is purely EDUCATIONAL in nature\n"
                    "â€¢ We are NOT financial advisors and do not provide personalized financial advice\n"
                    "â€¢ Any decisions you make are your own responsibility\n\n"
                    "Success in trading requires dedication, discipline, and continuous learning. "
                    "Our academy provides education and community support, but trading always involves risk."
                )
                
                # Send disclaimer first
                bot.send_message(
                    user_id,
                    disclaimer_message,
                    parse_mode="Markdown"
                )
                
                # Wait 1 second for user to read disclaimer before sending invite
                time.sleep(1)
                
                # Add another transition message before sending group invite
                group_transition_msg = bot.send_message(user_id, "ðŸ”— Generating your exclusive group invite link...")
                time.sleep(1.5)
                
                try:
                    bot.delete_message(user_id, group_transition_msg.message_id)
                except Exception as e:
                    logging.error(f"Error deleting group transition message: {e}")
                
                # Then send group invite
                expiration_time = datetime.now() + timedelta(seconds=60)  # Link expires in 60 seconds
                formatted_expiration = expiration_time.strftime("%I:%M:%S %p")  # Format as hour:minute:second AM/PM

                # Escape special characters in the invite link
                safe_invite_link = invite_link.replace("_", "\\_").replace("*", "\\*").replace("`", "\\`").replace("[", "\\[").replace("]", "\\]")
                
                bot.send_message(
                    user_id,
                    f"ðŸŽ‰ *Welcome to {group_name}!*\n\n"
                    f"Please join our community here: {safe_invite_link}\n\n"
                    f"â° *Link expires at:* {formatted_expiration} (60 seconds from now)",
                    parse_mode="Markdown"
                )
                
                # Set up delayed link revocation with longer timeout
                def revoke_link_later(chat_id, invite_link, admin_ids):
                    time.sleep(60)  # Wait 60 seconds before revoking
                    try:
                        bot.revoke_chat_invite_link(chat_id, invite_link)
                        for admin_id in admin_ids:
                            pass
                    except Exception as e:
                        logging.error(f"âš ï¸ Failed to revoke invite link: {e}")

                threading.Thread(target=revoke_link_later, args=(target_group_id, invite_link, ADMIN_IDS)).start()
        
        # 3. Record user form responses for admin reference
        try:
            form_answers = PENDING_USERS[user_id]['form_answers']
            # Update membership_type if it wasn't already set
            membership_type = PENDING_USERS[user_id].get('membership_type', 'regular')
            
            # Check if this is a renewal - if so, skip admin notifications
            is_renewal = PENDING_USERS[user_id].get('is_renewal', False)
            
            # NEW: Save form responses to PAYMENT_DATA for persistence
            user_id_str = str(user_id)
            if user_id_str in PAYMENT_DATA:
                # Add form answers to payment data
                PAYMENT_DATA[user_id_str]['form_answers'] = form_answers
                PAYMENT_DATA[user_id_str]['form_completion_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                # IMPORTANT FIX: Remove the forms_needed flag after completion
                PAYMENT_DATA[user_id_str]['forms_needed'] = False
                save_payment_data()  # Save to database
                logging.info(f"Form responses saved to PAYMENT_DATA for user {user_id}")
            
            # Only send admin notifications if this is NOT a renewal
            if not is_renewal:
                # Get username with proper escaping
                username = safe_markdown_escape(PAYMENT_DATA[str(user_id)].get('username', f'ID: {user_id}'))
                
                # Create different notification formats based on membership type
                if membership_type == 'supreme':
                    # Notify admins about the Supreme form responses (more detailed)
                    for admin_id in ADMIN_IDS:
                        bot.send_message(
                            admin_id,
                            f"ðŸ“‹ *New SUPREME Member Form Responses*\n\n"
                            f"User: {username}\n"
                            f"Full Name: {safe_markdown_escape(form_answers.get('full_name', 'Not provided'))}\n"
                            f"Birthday: {safe_markdown_escape(form_answers.get('birthday', 'Not provided'))}\n"
                            f"Phone Number: {(form_answers.get('phone_number', 'Not provided'))}\n"
                            f"Time Zone: {(form_answers.get('time_zone', 'Not provided'))}\n"
                            f"Expertise Level: {(form_answers.get('expertise_level', 'Not provided'))}\n"
                            f"Trading Time: {(form_answers.get('trading_time_commitment', 'Not provided'))}\n"
                            f"Interest Reason: {(form_answers.get('interest_reason', 'Not provided'))}\n"
                            f"Personal Goals: {(form_answers.get('personal_goals', 'Not provided'))}\n"
                            f"Call Preference: {(form_answers.get('call_preference', 'Not provided'))}\n"
                            f"Trading Challenges: {(form_answers.get('challenges', 'Not provided'))}",
                            parse_mode="Markdown"
                        )
                else:
                    # Notify admins about the Regular form responses with new fields
                    for admin_id in ADMIN_IDS:
                        bot.send_message(
                            admin_id,
                            f"ðŸ“‹ *New Regular Member Form Responses*\n\n"
                            f"User: {username}\n"
                            f"Full Name: {safe_markdown_escape(form_answers.get('full_name', 'Not provided'))}\n"
                            f"Age/Birth Year: {safe_markdown_escape(form_answers.get('birthday', 'Not provided'))}\n"
                            f"Experience Level: {(form_answers.get('experience_level', 'Not provided'))}\n"
                            f"Learning Goals: {(form_answers.get('learning_goals', 'Not provided'))}\n"
                            f"Found us via: {(form_answers.get('source', 'Not provided'))}",
                            parse_mode="Markdown"
                        )
            else:
                # Log that we're skipping admin notification for renewal
                logging.info(f"Skipping form notification for renewal user {user_id}")
        except Exception as e:
            logging.error(f"Error sending form responses to admins: {e}")
        
        # NEW: Notify user of successful registration completion
        try:
            bot.send_message(
                user_id,
                "âœ… *Registration Complete!*\n\n"
                "Thank you for completing your registration forms. Your membership is now fully activated.\n\n"
                "Use the /start command anytime to access your membership options and features.",
                parse_mode="Markdown"
            )
        except Exception as e:
            logging.error(f"Error sending completion notification to user: {e}")
        
        # NEW: Prompt user to subscribe to updates
        time.sleep(1)  # Add a small delay for better UX
        
        # Create inline keyboard with Yes/No options
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineKeyboardButton("âœ… Yes, Keep Me Updated", callback_data="update_yes"),
            InlineKeyboardButton("âŒ No Thanks", callback_data="update_no")
        )
        
        bot.send_message(
            user_id,
            "ðŸ”” *Stay Updated!*\n\n"
            "Would you like to receive notifications about:\n\n"
            "â€¢ ðŸ”“ When enrollment opens or closes\n"
            "â€¢ ðŸ’° Special discount promotions & offers\n"
            "â€¢ ðŸŽ Other academy opportunities\n\n"
            "You can change this setting anytime with /update command.",
            parse_mode="Markdown",
            reply_markup=markup
        )
        
        # 4. Clean up pending user data - THIS IS THE IMPORTANT CHANGE
        # Instead of deleting the user, set a completed status
        PENDING_USERS[user_id]['status'] = 'completed_onboarding'
        save_pending_users()
        
    except Exception as e:
        logging.error(f"Error in complete_onboarding for user {user_id}: {e}")

def send_completion_certificate(user_id):
    """Generate and send personalized certificates to the user"""
    try:
        # Get user's form answers and membership type
        form_answers = PENDING_USERS[user_id].get('form_answers', {})
        membership_type = PENDING_USERS[user_id].get('membership_type', 'regular').lower()
        
        # Get user's full name from form answers
        full_name = form_answers.get('full_name', 'Member')

        # Format current date
        current_date = datetime.now().strftime('%B %d, %Y')
        
        # Send Certificate of Completion first (no customization needed)
        cert1_path = 'cert/cert1.jpeg'
        if os.path.exists(cert1_path):
            with open(cert1_path, 'rb') as cert:
                bot.send_photo(
                    user_id,
                    cert,
                    caption=f"ðŸŽ“ *CERTIFICATE OF COMPLETION*\n\n"
                           f"Congratulations on completing your onboarding for the "
                           f"{'Supreme' if membership_type == 'supreme' else 'Regular'} "
                           f"Membership of Prodigy Trading Academy!",
                    parse_mode="Markdown"
                )
            
            # Add a small delay between certificates for better UX
            time.sleep(1.5)
        
        # Now send Certificate of Enrollment (needs customization)
        cert2_path = 'cert/cert2.jpeg'
        if os.path.exists(cert2_path):
            try:
                from PIL import Image, ImageDraw, ImageFont
                
                # Open the certificate template
                img = Image.open(cert2_path)
                draw = ImageDraw.Draw(img)

                # Load fonts (make sure these font files exist)
                try:
                    # Start with preferred font size
                    name_font_size = 40
                    name_font = ImageFont.truetype('fonts/Garet-Book.ttf', name_font_size)
                    date_font = ImageFont.truetype('fonts/Garet-Book.ttf', 18)
                    
                    # Get certificate width
                    img_width = img.width
                    
                    # Calculate name width with initial font size
                    name_width = draw.textlength(full_name, font=name_font)
                    
                    # Auto-adjust font size if name is too long (reduce until it fits)
                    while name_width > (img_width * 0.8) and name_font_size > 20:  # Allow name to use up to 80% of width
                        name_font_size -= 2
                        name_font = ImageFont.truetype('fonts/Garet-Book.ttf', name_font_size)
                        name_width = draw.textlength(full_name, font=name_font)
                    
                    # Calculate center position for the name (this will auto-center regardless of name length)
                    name_x = (img_width - name_width) / 2
                    name_position = (name_x, 320)  # Keep Y position at 320
                    
                except OSError:
                    # Fallback to default font
                    name_font = ImageFont.load_default()
                    date_font = ImageFont.load_default()
                    
                    # Even with default font, still center the name
                    name_width = draw.textlength(full_name, font=name_font)
                    name_x = (img_width - name_width) / 2
                    name_position = (name_x, 320)

                # Set exact position for date (unchanged)
                date_position = (650, 490)  # Exact X and Y coordinates for date

                # Draw text on image
                draw.text(name_position, full_name, fill=(0, 0, 0), font=name_font)  # Black text
                draw.text(date_position, current_date, fill=(0, 0, 0), font=date_font)
                
                # Save the customized certificate
                temp_certificate = f'cert/temp_{user_id}.jpg'
                img.save(temp_certificate)
                
                # Send the certificate with a nice message
                with open(temp_certificate, 'rb') as cert:
                    bot.send_photo(
                        user_id,
                        cert,
                        caption=f"ðŸ“œ *CERTIFICATE OF ENROLLMENT*\n\n"
                               f"This certifies that {full_name} is officially enrolled in the "
                               f"{'Supreme' if membership_type == 'supreme' else 'Regular'} "
                               f"Membership of Prodigy Trading Academy as of {current_date}.",
                        parse_mode="Markdown"
                    )
                
                # Clean up temporary file
                try:
                    os.remove(temp_certificate)
                except:
                    pass
                    
            except ImportError:
                # If PIL is not available, send the static certificate with explanation message
                with open(cert2_path, 'rb') as cert:
                    bot.send_photo(
                        user_id,
                        cert,
                        caption=f"ðŸ“œ *CERTIFICATE OF ENROLLMENT*\n\n"
                               f"This certifies that {full_name} is officially enrolled in the "
                               f"{'Supreme' if membership_type == 'supreme' else 'Regular'} "
                               f"Membership of Prodigy Trading Academy as of {current_date}.\n\n"
                               f"(Note: This is a generic certificate - your personalized one will be available soon)",
                        parse_mode="Markdown"
                    )
                    
        logging.info(f"Sent certificates to user {user_id}")
        
    except Exception as e:
        logging.error(f"Error sending certificates: {e}")
        # Fallback message if certificate generation fails
        bot.send_message(
            user_id,
            f"ðŸŽ“ *MEMBERSHIP CONFIRMED*\n\n"
            f"Congratulations on completing your enrollment in Prodigy Trading Academy!",
            parse_mode="Markdown"
        )

def send_welcome_package(user_id):
    """Send social media links and welcome info after form completion"""
    try:
        # Show typing indicator for better UX
        bot.send_chat_action(user_id, 'typing')
        time.sleep(1)
        
        # Send social media connections message with clickable buttons
        social_text = (
            "ðŸŒ *CONNECT WITH OUR COMMUNITY*\n\n"
            "To get the most from your membership, follow us on our social platforms:\n\n"
            "â€¢ *Instagram:* Daily tips and market insights\n"
            "â€¢ *Facebook Page:* Connect with fellow traders\n"
            "We share exclusive content and community highlights you won't want to miss!"
        )
        
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineKeyboardButton("ðŸ“± Instagram", url="https://instagram.com/prodigytradingacademy"),
            InlineKeyboardButton("ðŸ‘¥ Facebook", url="https://www.facebook.com/prodigytradingacademy"),
        )
        
        bot.send_message(
            user_id,
            social_text,
            parse_mode="Markdown",
            reply_markup=markup,
            disable_web_page_preview=True
        )
        
        logging.info(f"Sent welcome package with social links to user {user_id}")
        
    except Exception as e:
        logging.error(f"Failed to send welcome package to user {user_id}: {e}")

# Admin Rejects Payment
@bot.callback_query_handler(func=lambda call: call.data.startswith("reject_payment_"))
def callback_reject_payment(call):
    user_id = int(call.data.split("_")[2])
    if call.message.chat.id not in ADMIN_IDS:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to use this action.")
        return

    try:
        # Check if the user already has an approved payment
        if str(user_id) in PAYMENT_DATA and PAYMENT_DATA[str(user_id)]['haspayed']:
            bot.answer_callback_query(call.id, "âš ï¸ This user has already been approved. Cannot reject.")
            return

        # Check if user is actually waiting for payment verification
        if user_id not in PENDING_USERS or PENDING_USERS[user_id].get('status') != 'waiting_approval':
            bot.answer_callback_query(call.id, "âŒ This user is not waiting for payment verification.")
            return

        bot.send_message(user_id, random.choice(payment_rejection_messages), parse_mode="Markdown")
        PENDING_USERS.pop(user_id, None)
        delete_pending_user(user_id)  # Remove from MongoDB
        save_pending_users()
        bot.answer_callback_query(call.id, "âŒ Payment rejected successfully.")

        # Log admin activity and notify all admins
        admin_username = call.from_user.username or f"Admin ({call.message.chat.id})"
        user_info = bot.get_chat(user_id)
        username = user_info.username or f"ID: {user_id}"

        # Escape Markdown characters in the usernames
        admin_username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', admin_username)
        username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)

        for admin_id in ADMIN_IDS:
            bot.send_message(admin_id, f"ðŸ“ *Activity Log*\n\n{admin_username} has rejected payment from PTA member @{username}.", parse_mode="Markdown")

    except Exception as e:
        bot.answer_callback_query(call.id, f"âŒ Unexpected error rejecting payment: {e}")

@bot.message_handler(content_types=['new_chat_members'])
def welcome_new_members(message):
    """Welcome new members when they join the group"""
    try:
        # Check which group the message is coming from
        if message.chat.id == PAID_GROUP_ID:
            group_type = "regular"
        elif message.chat.id == SUPREME_GROUP_ID:
            group_type = "supreme"
        else:
            # Not a monitored group
            return
            
        # Process all new members (in case multiple users join at once)
        for new_member in message.new_chat_members:
            # Skip the bot itself
            if new_member.id == bot.get_me().id:
                continue
                
            # Get user's name - use first name, or username as fallback
            user_name = new_member.first_name or new_member.username or "New member"
            # Escape any special Markdown characters
            user_name = safe_markdown_escape(user_name)
            
            # Different welcome messages based on group type
            if group_type == "supreme":
                welcome_message = (
                    f"ðŸŒŸ *Welcome to the SUPREME Mentorship Group, {user_name}!* ðŸŒŸ\n\n"
                    f"You've joined our exclusive premium community of dedicated traders. "
                    f"With your Supreme membership, you'll have access to:\n\n"
                    f"ðŸ”¹ *Advanced trading strategies* and premium analysis\n"
                    f"ðŸ”¹ *Priority support* from our expert mentors\n"
                    f"ðŸ”¹ *Exclusive learning resources* not available elsewhere\n" 
                    f"ðŸ”¹ *Daily challenges* and premium accountability features\n\n"
                    f"Your investment in yourself shows your commitment to excellence. "
                    f"We're here to help you maximize your trading potential!\n\n"
                    f"If you have any questions, our mentors are ready to assist you.\n"
                    f"Happy Trading! ðŸ“ˆ"
                )
            else:  # Regular group
                welcome_message = (
                    f"ðŸŽ‰ *Welcome to Prodigy Trading Academy, {user_name}!* ðŸŽ‰\n\n"
                    f"We're excited to have you join our trading community. "
                    f"Here you'll find valuable insights, daily challenges, and a supportive network of fellow traders.\n\n"
                    f"ðŸ“Š *Daily challenges* are posted at 8:00 AM PH time\n"
                    f"ðŸ’¡ *Expert guidance* from our community\n"
                    f"ðŸ“š *Learning resources* to improve your skills\n\n"
                    f"If you have any questions, our mentors are here to help!\n"
                    f"Happy Trading! ðŸ“ˆ"
                )
            
            # Send the welcome message directly to the group
            try:
                bot.send_message(
                    message.chat.id, 
                    welcome_message,
                    parse_mode="Markdown"
                )
                    
                logging.info(f"Sent {group_type} welcome message for new member {new_member.id} ({user_name})")
            except ApiException as e:
                # If Markdown fails, try without formatting
                logging.error(f"Failed to send welcome with Markdown: {e}")
                bot.send_message(
                    message.chat.id, 
                    welcome_message.replace('*', '')
                )
    except Exception as e:
        logging.error(f"Error in welcome_new_members: {e}")

def safe_markdown_escape(text):
    """
    Comprehensive function to safely escape ANY text for Telegram Markdown
    Returns the escaped text or plain text if the input contains problematic characters
    """
    if text is None:
        return "None"
        
    try:
        # First try with standard escaping pattern
        special_chars = r'_*[]()~`>#+-=|{}.!'
        escaped_text = text
        for char in special_chars:
            escaped_text = escaped_text.replace(char, f"\\{char}")
        return escaped_text
    except Exception:
        # If anything fails, sanitize by removing problematic characters
        return ''.join(c for c in text if c.isalnum() or c.isspace() or c in '.-_')    

# Function to remind users 3 days before payment deadline
def escape_markdown(text):
    return re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', text)

def send_payment_reminder():
    """Payment reminder function using scheduled times for consistency."""
    logging.info("Payment reminder thread started")
    
    # Define specific times of day to send reminders (24-hour format in Philippines timezone)
    REMINDER_TIMES = ["09:00"]  # 9:00 AM
    
    # Track the last day we sent reminders to avoid duplicate sends
    last_reminder_dates = {time: None for time in REMINDER_TIMES}
    
    # Load reminder messages from MongoDB for persistence
    global reminder_messages
    reminder_messages = load_reminder_messages()
    
    while True:
        try:
            # Get current time in Philippines timezone
            now = datetime.now(pytz.timezone('Asia/Manila'))
            current_time = now.strftime('%H:%M')
            current_date = now.strftime('%Y-%m-%d')
            
            # Check if it's time to send reminders and we haven't sent them today at this time
            if current_time in REMINDER_TIMES and last_reminder_dates[current_time] != current_date:
                logging.info(f"Scheduled time {current_time} reached - sending payment reminders...")
                
                # Process all users for payment reminders
                for user_id_str, data in PAYMENT_DATA.items():
                    try:
                        user_id = int(user_id_str)

                        # Check if due_date exists before trying to access it
                        if 'due_date' not in data:
                            logging.error(f"Error processing payment reminder for {user_id_str}: 'due_date' field missing")
                            continue

                        # Get the naive datetime first
                        try:
                            naive_due_date = datetime.strptime(data['due_date'], '%Y-%m-%d %H:%M:%S')
                        except ValueError as e:
                            logging.error(f"Error processing payment reminder for {user_id_str}: invalid due_date format - {e}")
                            continue
                        
                        # Make it timezone-aware by adding Manila timezone
                        manila_tz = pytz.timezone('Asia/Manila')
                        due_date = manila_tz.localize(naive_due_date)
                        
                        username = data.get('username', None)
                        
                        if username:
                            username = escape_markdown(username)
                            user_display = f"@{username}"
                        else:
                            user_display = f"User {user_id}"

                        # Now both dates are timezone-aware, subtraction will work
                        days_until_due = (due_date - now).days
                        
                        # Check for users in grace period
                        if data.get('grace_period', False):
                            grace_end_date = datetime.strptime(data.get('grace_end_date'), '%Y-%m-%d %H:%M:%S')
                            grace_end_date = manila_tz.localize(grace_end_date)
                            
                            # If grace period has expired
                            if now >= grace_end_date:
                                # Delete previous reminders for this user
                                if user_id in reminder_messages:
                                    try:
                                        # Delete previous user reminder
                                        if 'user_msg_id' in reminder_messages[user_id]:
                                            bot.delete_message(user_id, reminder_messages[user_id]['user_msg_id'])
                                    except Exception as e:
                                        logging.error(f"Failed to delete previous user reminder: {e}")
                                    
                                    # Delete previous admin reminders
                                    for admin_id, msg_id in reminder_messages[user_id].get('admin_msg_ids', {}).items():
                                        try:
                                            bot.delete_message(admin_id, msg_id)
                                        except Exception as e:
                                            logging.error(f"Failed to delete previous admin reminder: {e}")
                                
                                # Notify admins about expired grace period
                                admin_messages = {}
                                for admin_id in ADMIN_IDS:
                                    markup = InlineKeyboardMarkup()
                                    markup.add(
                                        InlineKeyboardButton("âœ“ Kick Member", callback_data=f"kick_{user_id}"),
                                        InlineKeyboardButton("âœ— Keep Member", callback_data=f"keep_{user_id}")
                                    )
                                    
                                    sent_msg = bot.send_message(
                                        admin_id,
                                        f"âš ï¸ *GRACE PERIOD EXPIRED*\n\n"
                                        f"{user_display}'s grace period has now expired. "
                                        f"Their membership expired on {due_date.strftime('%Y/%m/%d')}.\n\n"
                                        f"What would you like to do with this member?",
                                        parse_mode="Markdown",
                                        reply_markup=markup
                                    )
                                    admin_messages[admin_id] = sent_msg.message_id
                                
                                # Store new message IDs
                                reminder_messages[user_id] = {
                                    'admin_msg_ids': admin_messages
                                }
                                # Save to MongoDB
                                save_reminder_message(user_id, reminder_messages[user_id])
                                
                                # Remove grace period flag after notifying
                                PAYMENT_DATA[user_id_str]['grace_period'] = False
                                PAYMENT_DATA[user_id_str]['grace_end_date'] = None
                                save_payment_data()
                                
                                # Skip to next user since we've handled this case
                                continue
                        
                        # Send reminders for all users within 3 days of expiry
                        if 0 <= days_until_due <= 3 and data['haspayed'] and not data.get('cancelled', False):
                            # Debug information about the existing messages for this user
                            logging.info(f"Processing payment reminder for user {user_id} with {days_until_due} days until due")
                            if user_id in reminder_messages:
                                logging.info(f"Found existing reminder messages for user {user_id}: {reminder_messages[user_id]}")
                            else:
                                logging.info(f"No existing reminder messages for user {user_id}")

                            # Delete previous reminders for this user
                            if user_id in reminder_messages:
                                try:
                                    # Delete previous user reminder
                                    if 'user_msg_id' in reminder_messages[user_id]:
                                        msg_id = reminder_messages[user_id]['user_msg_id']
                                        logging.info(f"Attempting to delete user message {msg_id} for user {user_id}")
                                        try:
                                            bot.delete_message(user_id, msg_id)
                                            logging.info(f"Successfully deleted message {msg_id} for user {user_id}")
                                        except ApiException as e:
                                            error_msg = str(e)
                                            if "message to delete not found" in error_msg:
                                                logging.warning(f"Message {msg_id} for user {user_id} already deleted")
                                            elif "bot was blocked by the user" in error_msg:
                                                logging.warning(f"Cannot delete message for user {user_id} - user blocked the bot")
                                            else:
                                                logging.error(f"Failed to delete previous user reminder: {e}")
                                except Exception as e:
                                    logging.error(f"General error in user message deletion for {user_id}: {e}")
                                
                                # Delete previous admin reminders
                                for admin_id, msg_id in reminder_messages[user_id].get('admin_msg_ids', {}).items():
                                    try:
                                        logging.info(f"Attempting to delete admin message {msg_id} for admin {admin_id}")
                                        bot.delete_message(admin_id, msg_id)
                                        logging.info(f"Successfully deleted admin message {msg_id} for admin {admin_id}")
                                    except ApiException as e:
                                        error_msg = str(e)
                                        if "message to delete not found" in error_msg:
                                            logging.warning(f"Admin message {msg_id} for admin {admin_id} already deleted")
                                        else:
                                            logging.error(f"Failed to delete admin reminder for admin {admin_id}: {e}")
                                    except Exception as e:
                                        logging.error(f"General error in admin message deletion for admin {admin_id}: {e}")
                            
                            try:
                                # Send reminder to user
                                bot.send_chat_action(user_id, 'typing')
                                user_msg = bot.send_message(
                                    user_id, 
                                    f"â³ Reminder: Your next payment is due in {days_until_due} days: {due_date.strftime('%Y/%m/%d %I:%M:%S %p')}."
                                )
                                logging.info(f"Sent payment reminder to user {user_id}, message ID: {user_msg.message_id}")
                                
                                # Send notification to admins
                                admin_messages = {}
                                for admin_id in ADMIN_IDS:
                                    admin_msg = bot.send_message(
                                        admin_id, 
                                        f"Admin Notice: {user_display} has an upcoming payment due in {days_until_due} days."
                                    )
                                    admin_messages[admin_id] = admin_msg.message_id
                                    logging.info(f"Sent admin notification to {admin_id}, message ID: {admin_msg.message_id}")
                                
                                # Store new message IDs with explicit logging
                                reminder_messages[user_id] = {
                                    'user_msg_id': user_msg.message_id,
                                    'admin_msg_ids': admin_messages
                                }
                                # Save to MongoDB
                                save_reminder_message(user_id, reminder_messages[user_id])
                                logging.info(f"Updated reminder_messages for user {user_id}: {reminder_messages[user_id]}")
                            
                            except ApiException as e:
                                logging.error(f"Failed to send payment reminder to user {user_id}: {e}")
                                
                                # For failed user notifications, still notify admins
                                admin_messages = {}
                                for admin_id in ADMIN_IDS:
                                    admin_msg = bot.send_message(
                                        admin_id, 
                                        f"âš ï¸ *Failed to send payment reminder*\n\n"
                                        f"Could not send payment reminder to {user_display}.\n"
                                        f"The user hasn't started a conversation with the bot or has blocked it.\n\n"
                                        f"Their payment is due in {days_until_due} days: {due_date.strftime('%Y/%m/%d')}\n\n"
                                        f"Please contact them manually.",
                                        parse_mode="Markdown"
                                    )
                                    admin_messages[admin_id] = admin_msg.message_id
                                
                                # Store only admin message IDs
                                reminder_messages[user_id] = {
                                    'admin_msg_ids': admin_messages
                                }
                                # Save to MongoDB
                                save_reminder_message(user_id, reminder_messages[user_id])
                        
                        # Check if membership has expired
                        elif due_date < now and (data['haspayed'] or data.get('admin_action_pending', False)) and not data.get('grace_period', False):
                            # Delete previous reminders for this user
                            if user_id in reminder_messages:
                                try:
                                    # Delete previous user reminder
                                    if 'user_msg_id' in reminder_messages[user_id]:
                                        bot.delete_message(user_id, reminder_messages[user_id]['user_msg_id'])
                                except Exception as e:
                                    logging.error(f"Failed to delete previous user reminder: {e}")
                                
                                # Delete previous admin reminders
                                for admin_id, msg_id in reminder_messages[user_id].get('admin_msg_ids', {}).items():
                                    try:
                                        bot.delete_message(admin_id, msg_id)
                                    except Exception as e:
                                        logging.error(f"Failed to delete previous admin reminder: {e}")

                                # Update payment data
                                PAYMENT_DATA[user_id_str]['haspayed'] = False
                                PAYMENT_DATA[user_id_str]['admin_action_pending'] = True
                                PAYMENT_DATA[user_id_str]['reminder_sent'] = False
                                save_payment_data()

                                # Calculate days since expiration
                                days_expired = (now - due_date).days
                                
                                # Send notification to admins with action buttons based on expiration duration
                                admin_messages = {}
                                for admin_id in ADMIN_IDS:
                                    markup = InlineKeyboardMarkup()
                                    
                                    # If expired more than 3 days, only offer kick or keep (no grace period)
                                    if days_expired > 3:
                                        markup.add(
                                            InlineKeyboardButton("âŒ Kick Member", callback_data=f"kick_{user_id}"),
                                            InlineKeyboardButton("âœ“ Keep Member", callback_data=f"keep_{user_id}")
                                        )
                                        
                                        admin_msg = bot.send_message(
                                            admin_id, 
                                            f"âš ï¸ *LONG-EXPIRED MEMBERSHIP*\n\n"
                                            f"{user_display}'s membership has been expired for {days_expired} days.\n\n"
                                            f"What would you like to do with this member?",
                                            parse_mode="Markdown",
                                            reply_markup=markup
                                        )
                                    else:
                                        # For recently expired members, offer grace period
                                        markup.add(
                                            InlineKeyboardButton("â³ Give 2 Days Grace", callback_data=f"grace_{user_id}"),
                                            InlineKeyboardButton("âŒ Kick Member", callback_data=f"kick_{user_id}")
                                        )
                                        
                                        admin_msg = bot.send_message(
                                            admin_id, 
                                            f"âš ï¸ *MEMBERSHIP EXPIRED*\n\n"
                                            f"{user_display}'s membership has expired and has been marked as unpaid in the system.\n\n"
                                            f"What would you like to do with this member?",
                                            parse_mode="Markdown",
                                            reply_markup=markup
                                        )
                                        
                                    admin_messages[admin_id] = admin_msg.message_id
                            try:
                                # Send expiry notice to user
                                bot.send_chat_action(user_id, 'typing')
                                user_msg = bot.send_message(
                                    user_id, 
                                    "âŒ Your membership has expired. Please renew your membership to continue accessing our services."
                                )
                                logging.info(f"Sent expiry notice to user {user_id}")
                                
                                # Update payment data - mark as pending admin action instead of just setting haspayed=False
                                PAYMENT_DATA[user_id_str]['haspayed'] = False
                                PAYMENT_DATA[user_id_str]['admin_action_pending'] = True
                                PAYMENT_DATA[user_id_str]['reminder_sent'] = False
                                save_payment_data()
                                
                                # Send notification to admins with action buttons
                                admin_messages = {}
                                for admin_id in ADMIN_IDS:
                                    markup = InlineKeyboardMarkup()
                                    markup.add(
                                        InlineKeyboardButton("â³ Give 2 Days Grace", callback_data=f"grace_{user_id}"),
                                        InlineKeyboardButton("âŒ Kick Member", callback_data=f"kick_{user_id}")
                                    )
                                    
                                    admin_msg = bot.send_message(
                                        admin_id, 
                                        f"âš ï¸ *MEMBERSHIP EXPIRED*\n\n"
                                        f"{user_display}'s membership has expired and has been marked as unpaid in the system.\n\n"
                                        f"What would you like to do with this member?",
                                        parse_mode="Markdown",
                                        reply_markup=markup
                                    )
                                    admin_messages[admin_id] = admin_msg.message_id
                                
                                # Store new message IDs
                                reminder_messages[user_id] = {
                                    'user_msg_id': user_msg.message_id,
                                    'admin_msg_ids': admin_messages
                                }
                                # Save to MongoDB
                                save_reminder_message(user_id, reminder_messages[user_id])
                            
                            except ApiException as e:
                                logging.error(f"Failed to send expiry notice to user {user_id}: {e}")
                                PAYMENT_DATA[user_id_str]['haspayed'] = False
                                PAYMENT_DATA[user_id_str]['reminder_sent'] = False
                                save_payment_data()
                                
                                # Still notify admins with action buttons
                                admin_messages = {}
                                for admin_id in ADMIN_IDS:
                                    markup = InlineKeyboardMarkup()
                                    markup.add(
                                        InlineKeyboardButton("â³ Give 2 Days Grace", callback_data=f"grace_{user_id}"),
                                        InlineKeyboardButton("âŒ Kick Member", callback_data=f"kick_{user_id}")
                                    )
                                    
                                    admin_msg = bot.send_message(
                                        admin_id, 
                                        f"âš ï¸ *FAILED TO NOTIFY USER & MEMBERSHIP EXPIRED*\n\n"
                                        f"Could not notify {user_display} about their expired membership.\n"
                                        f"The user hasn't started a conversation with the bot or has blocked it.\n\n"
                                        f"Their membership has been marked as expired in the system.\n\n"
                                        f"What would you like to do with this member?",
                                        parse_mode="Markdown",
                                        reply_markup=markup
                                    )
                                    admin_messages[admin_id] = admin_msg.message_id
                                
                                # Store only admin message IDs
                                reminder_messages[user_id] = {
                                    'admin_msg_ids': admin_messages
                                }
                                # Save to MongoDB
                                save_reminder_message(user_id, reminder_messages[user_id])
                                    
                    except Exception as e:
                        logging.error(f"Error processing payment reminder for user {user_id_str}: {e}")
                        for admin_id in ADMIN_IDS:
                            bot.send_message(admin_id, f"âš ï¸ Error processing payment reminder for {user_id_str}: {str(e)}")
                
                # Record that we've sent reminders for this scheduled time today
                last_reminder_dates[current_time] = current_date
                logging.info(f"Completed sending payment reminders at scheduled time {current_time}")
            
            # Calculate the time to sleep until the start of the next minute
            now = datetime.now(pytz.timezone('Asia/Manila'))
            sleep_time = 60 - now.second - now.microsecond / 1_000_000
            time.sleep(sleep_time)
            
        except Exception as e:
            logging.error(f"Error in payment reminder main loop: {e}")
            time.sleep(60)  # Wait a minute on error before trying again

def safe_markdown_escape_v2(text):
    """Comprehensive and reliable function to escape text for Telegram Markdown V2"""
    if text is None:
        return "None"
        
    # First replace literal backslashes
    text = text.replace('\\', '\\\\')
    
    # Then escape all special characters
    special_chars = '_*[]()~`>#+-=|{}.!'
    for char in special_chars:
        text = text.replace(char, f"\\{char}")
    
    return text


# Handle admin clicking "Give Grace Period" button
@bot.callback_query_handler(func=lambda call: call.data.startswith("grace_"))
def handle_grace_period(call):
    """Handle admin clicking the 'Give 2 Days Grace' button"""
    admin_id = call.from_user.id
    
    # Verify the user is an admin
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    # Extract user ID
    user_id = int(call.data.split("_")[1])
    user_id_str = str(user_id)
    
    try:
        # Get username for display
        if user_id_str in PAYMENT_DATA:
            username = PAYMENT_DATA[user_id_str].get('username', None)
        else:
            try:
                user_info = bot.get_chat(user_id)
                username = user_info.username
            except:
                username = None
                
        # Safely escape the username for Markdown
        if username:
            user_display = f"@{safe_markdown_escape_v2(username)}"
        else:
            user_display = f"User {user_id}"
        
        # Create confirmation markup
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineKeyboardButton("âœ… Yes, give grace", callback_data=f"confirm_grace_{user_id}"),
            InlineKeyboardButton("âŒ No, cancel", callback_data=f"cancel_action_{call.message.message_id}")
        )
        
        # Send confirmation message
        bot.edit_message_text(
            f"âš ï¸ *CONFIRMATION REQUIRED*\n\n"
            f"Are you sure you want to give {user_display} a 2-day grace period?\n\n"
            f"This will temporarily mark them as paid and prevent them from being kicked.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
        
        bot.answer_callback_query(call.id)
        
    except Exception as e:
        logging.error(f"Error in handle_grace_period confirmation: {e}")
        bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}", show_alert=True)

# Handle admin clicking "Kick Member" button
@bot.callback_query_handler(func=lambda call: call.data.startswith("kick_"))
def handle_kick_member(call):
    """Handle admin clicking the 'Kick Member' button"""
    admin_id = call.from_user.id
    
    # Verify the user is an admin
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    # Extract user ID to kick
    user_id = int(call.data.split("_")[1])
    user_id_str = str(user_id)
    
    try:
        # Get username for display
        if user_id_str in PAYMENT_DATA:
            username = PAYMENT_DATA[user_id_str].get('username', None)
        else:
            try:
                user_info = bot.get_chat(user_id)
                username = user_info.username
            except:
                username = None
                
        # Safely escape the username for Markdown
        if username:
            user_display = f"@{safe_markdown_escape_v2(username)}"
        else:
            user_display = f"User {user_id}"
        
        # Create confirmation markup
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineKeyboardButton("âœ… Yes, kick member", callback_data=f"confirm_kick_{user_id}"),
            InlineKeyboardButton("âŒ No, cancel", callback_data=f"cancel_action_{call.message.message_id}")
        )
        
        # Send confirmation message
        bot.edit_message_text(
            f"âš ï¸ *CONFIRMATION REQUIRED*\n\n"
            f"Are you sure you want to kick {user_display} from the group?\n\n"
            f"They will be notified and removed from the group immediately.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
        
        bot.answer_callback_query(call.id)
        
    except Exception as e:
        logging.error(f"Error in handle_kick_member confirmation: {e}")
        bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}", show_alert=True)

# Handle admin decision to keep member after grace period expires
@bot.callback_query_handler(func=lambda call: call.data.startswith("keep_"))
def handle_keep_member(call):
    """Handle admin clicking 'Keep Member' after grace period expiry"""
    admin_id = call.from_user.id
    
    # Verify the user is an admin
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    # Extract user ID
    user_id = int(call.data.split("_")[1])
    user_id_str = str(user_id)
    
    try:
        # Get username for display
        if user_id_str in PAYMENT_DATA:
            username = PAYMENT_DATA[user_id_str].get('username', None)
        else:
            try:
                user_info = bot.get_chat(user_id)
                username = user_info.username
            except:
                username = None
                
        # Safely escape the username for Markdown
        if username:
            user_display = f"@{safe_markdown_escape_v2(username)}"
        else:
            user_display = f"User {user_id}"
        
        # Create confirmation markup
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineKeyboardButton("âœ… Yes, keep member", callback_data=f"confirm_keep_{user_id}"),
            InlineKeyboardButton("âŒ No, cancel", callback_data=f"cancel_action_{call.message.message_id}")
        )
        
        # Send confirmation message
        bot.edit_message_text(
            f"âš ï¸ *CONFIRMATION REQUIRED*\n\n"
            f"Are you sure you want to keep {user_display} in the group despite expired membership?\n\n"
            f"Their account will remain marked as unpaid in the system.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
        
        bot.answer_callback_query(call.id)
        
    except Exception as e:
        logging.error(f"Error in handle_keep_member confirmation: {e}")
        bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}", show_alert=True)

# Now add handlers for the confirmation actions
@bot.callback_query_handler(func=lambda call: call.data.startswith("confirm_grace_"))
def confirm_grace_period(call):
    """Handle admin confirming to give grace period"""
    # Copy the existing grace period handler code here, but with user_id from call.data.split("_")[2]
    admin_id = call.from_user.id
    
    # Verify the user is an admin
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    # Extract user ID
    user_id = int(call.data.split("_")[2])
    user_id_str = str(user_id)
    
    try:
        # Rest of the original grace period handler code
        # Get username for display - WITH PROPER ESCAPING
        if user_id_str in PAYMENT_DATA:
            username = PAYMENT_DATA[user_id_str].get('username', None)
        else:
            try:
                user_info = bot.get_chat(user_id)
                username = user_info.username
            except:
                username = None
                
        # Safely escape the username for Markdown
        if username:
            user_display = f"@{safe_markdown_escape_v2(username)}"
        else:
            user_display = f"User {user_id}"
        
        # Calculate grace period end date (2 days from now)
        now = datetime.now(pytz.timezone('Asia/Manila'))
        grace_end_date = now + timedelta(days=2)
        
        # Update payment data to add grace period
        if user_id_str in PAYMENT_DATA:
            PAYMENT_DATA[user_id_str]['grace_period'] = True
            PAYMENT_DATA[user_id_str]['grace_end_date'] = grace_end_date.strftime('%Y-%m-%d %H:%M:%S')
            PAYMENT_DATA[user_id_str]['haspayed'] = True  # Temporarily mark as paid during grace period
            PAYMENT_DATA[user_id_str]['admin_action_pending'] = False  # Clear the pending flag
            save_payment_data()
        
        # Notify the user about the grace period
        try:
            bot.send_message(
                user_id,
                "â³ *Grace Period Granted*\n\n"
                "You have been given a 2-day grace period to renew your membership. "
                f"Please renew before {grace_end_date.strftime('%Y-%m-%d %I:%M:%S %p')} to avoid being removed from the group.",
                parse_mode="Markdown"
            )
            user_notified = True
        except:
            user_notified = False
            
        # Get safely escaped admin username
        admin_username = call.from_user.username or f"Admin {admin_id}"
        admin_username = safe_markdown_escape_v2(admin_username)
        
        # Update the button for the admin who took action (USING SAFE ESCAPING)
        bot.edit_message_text(
            f"âœ… *ACTION TAKEN: GRACE PERIOD*\n\n"
            f"{user_display} has been given a 2-day grace period until {grace_end_date.strftime('%Y-%m-%d')}.\n"
            f"User notification {'sent' if user_notified else 'FAILED'}.\n"
            f"Action taken by @{admin_username}",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        
        # Update messages for ALL other admins too (USING SAFE ESCAPING)
        if user_id in reminder_messages and 'admin_msg_ids' in reminder_messages[user_id]:
            other_admin_messages = reminder_messages[user_id]['admin_msg_ids']
            
            for other_admin_id, msg_id in other_admin_messages.items():
                if int(other_admin_id) == admin_id:
                    continue
                    
                try:
                    bot.edit_message_text(
                        f"âœ… *ACTION ALREADY TAKEN: GRACE PERIOD*\n\n"
                        f"{user_display} has been given a 2-day grace period until {grace_end_date.strftime('%Y-%m-%d')}.\n"
                        f"Action was taken by @{admin_username}.",
                        other_admin_id,
                        msg_id,
                        parse_mode="Markdown"
                    )
                except Exception as e:
                    logging.error(f"Failed to update message for admin {other_admin_id}: {e}")
        
        # Notify ALL admins (WITH SAFE ESCAPING)
        for admin_id in ADMIN_IDS:
            bot.send_message(admin_id, 
                f"ðŸ“ *Activity Log*\n\n@{admin_username} gave {user_display} a 2-day grace period until {grace_end_date.strftime('%Y-%m-%d')}.", 
                parse_mode="Markdown"
            )
        
        bot.answer_callback_query(call.id, f"Grace period granted to {username or user_id} until {grace_end_date.strftime('%Y-%m-%d')}")
        
    except Exception as e:
        logging.error(f"Error in confirm_grace_period: {e}")
        bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}", show_alert=True)

@bot.callback_query_handler(func=lambda call: call.data.startswith("confirm_kick_"))
def confirm_kick_member(call):
    """Handle admin confirming to kick member"""
    admin_id = call.from_user.id
    
    # Verify the user is an admin
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    # Extract user ID to kick
    user_id = int(call.data.split("_")[2])
    user_id_str = str(user_id)
    
    try:
        # Get username for display - WITH PROPER ESCAPING
        if user_id_str in PAYMENT_DATA:
            username = PAYMENT_DATA[user_id_str].get('username', None)
        else:
            try:
                user_info = bot.get_chat(user_id)
                username = user_info.username
            except:
                username = None
                
        # Safely escape the username for Markdown
        if username:
            user_display = f"@{safe_markdown_escape_v2(username)}"
        else:
            user_display = f"User {user_id}"
        
        # First notify the user that they're being removed
        try:
            bot.send_message(
                user_id,
                "âŒ *Your membership has expired*\n\n"
                "You are being removed from the group because your membership has expired. "
                "To rejoin, please renew your membership using the /start command.",
                parse_mode="Markdown"
            )
            user_notified = True
        except:
            user_notified = False
        
        # Try to kick the user from the group
        try:
            bot.ban_chat_member(PAID_GROUP_ID, user_id)
            bot.unban_chat_member(PAID_GROUP_ID, user_id)  # Immediately unban so they can rejoin later
            kick_successful = True
        except Exception as e:
            logging.error(f"Failed to kick user {user_id}: {e}")
            kick_successful = False
        
        # Update the button to show action was taken
        bot.edit_message_text(
            f"{'âœ…' if kick_successful else 'âŒ'} *ACTION TAKEN: KICK MEMBER*\n\n"
            f"{user_display} has {'been removed from' if kick_successful else 'FAILED to be removed from'} the group.\n"
            f"User notification {'sent' if user_notified else 'FAILED'}.\n"
            f"Action taken by admin @{safe_markdown_escape_v2(call.from_user.username or f'ID:{admin_id}')}",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )

        # Update messages for ALL other admins too
        admin_username = call.from_user.username or f"Admin {admin_id}"
        admin_username = safe_markdown_escape_v2(admin_username)
        
        # Find all admin messages for this user from reminder_messages
        if user_id in reminder_messages and 'admin_msg_ids' in reminder_messages[user_id]:
            other_admin_messages = reminder_messages[user_id]['admin_msg_ids']
            
            # Update each admin's message except the one who took action
            for other_admin_id, msg_id in other_admin_messages.items():
                # Skip the admin who took the action (their message was already updated)
                if int(other_admin_id) == admin_id:
                    continue
                    
                try:
                    bot.edit_message_text(
                        f"{'âœ…' if kick_successful else 'âŒ'} *ACTION ALREADY TAKEN: KICK MEMBER*\n\n"
                        f"{user_display} has {'been removed from' if kick_successful else 'FAILED to be removed from'} the group.\n"
                        f"Action was taken by @{admin_username}.",
                        other_admin_id,
                        msg_id,
                        parse_mode="Markdown"
                    )
                    logging.info(f"Updated message for admin {other_admin_id} about kick action taken by {admin_id}")
                except Exception as e:
                    logging.error(f"Failed to update message for admin {other_admin_id}: {e}")

        # Mark user as no longer pending admin action if kick was successful
        if kick_successful:
            PAYMENT_DATA[user_id_str]['admin_action_pending'] = False
            save_payment_data()
        
        # Notify ALL admins about this action
        for admin_id in ADMIN_IDS:
            bot.send_message(admin_id, 
                f"ðŸ“ *Activity Log*\n\n@{admin_username} kicked {user_display} from the group.", 
                parse_mode="Markdown"
            )
        
        bot.answer_callback_query(
            call.id, 
            f"User {username or user_id} has {'been kicked' if kick_successful else 'FAILED to be kicked'} from the group."
        )
        
    except Exception as e:
        logging.error(f"Error in confirm_kick_member: {e}")
        bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}", show_alert=True)

@bot.callback_query_handler(func=lambda call: call.data.startswith("confirm_keep_"))
def confirm_keep_member(call):
    """Handle admin confirming to keep member"""
    admin_id = call.from_user.id
    
    # Verify the user is an admin
    if admin_id not in ADMIN_IDS and admin_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    # Extract user ID
    user_id = int(call.data.split("_")[2])
    user_id_str = str(user_id)
    
    try:
        # Get username for display - WITH PROPER ESCAPING
        if user_id_str in PAYMENT_DATA:
            username = PAYMENT_DATA[user_id_str].get('username', None)
        else:
            try:
                user_info = bot.get_chat(user_id)
                username = user_info.username
            except:
                username = None
                
        # Safely escape the username for Markdown
        if username:
            user_display = f"@{safe_markdown_escape_v2(username)}"
        else:
            user_display = f"User {user_id}"
        
        # Mark user as no longer pending admin action, but still unpaid
        PAYMENT_DATA[user_id_str]['admin_action_pending'] = False
        save_payment_data()
        
        # Get safely escaped admin username
        admin_username = call.from_user.username or f"Admin {admin_id}"
        admin_username = safe_markdown_escape_v2(admin_username)
        
        # Update the button to show action was taken
        bot.edit_message_text(
            f"âœ… *ACTION TAKEN: KEPT MEMBER*\n\n"
            f"{user_display} has been allowed to remain in the group despite expired membership.\n"
            f"Their account is still marked as unpaid in the system.\n"
            f"Action taken by @{admin_username}",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        
        # Update messages for ALL other admins too
        if user_id in reminder_messages and 'admin_msg_ids' in reminder_messages[user_id]:
            other_admin_messages = reminder_messages[user_id]['admin_msg_ids']
            
            # Update each admin's message except the one who took action
            for other_admin_id, msg_id in other_admin_messages.items():
                # Skip the admin who took the action (their message was already updated)
                if int(other_admin_id) == admin_id:
                    continue
                    
                try:
                    bot.edit_message_text(
                        f"âœ… *ACTION ALREADY TAKEN: KEPT MEMBER*\n\n"
                        f"{user_display} has been allowed to remain in the group despite expired membership.\n"
                        f"Action was taken by @{admin_username}.",
                        other_admin_id,
                        msg_id,
                        parse_mode="Markdown"
                    )
                    logging.info(f"Updated message for admin {other_admin_id} about keep action taken by {admin_id}")
                except Exception as e:
                    logging.error(f"Failed to update message for admin {other_admin_id}: {e}")
        
        # Notify ALL admins about this action
        for admin_id in ADMIN_IDS:
            bot.send_message(admin_id, 
                f"ðŸ“ *Activity Log*\n\n@{admin_username} allowed {user_display} to remain in the group despite expired membership.", 
                parse_mode="Markdown"
            )
        
        bot.answer_callback_query(call.id, f"Decision recorded: {username or user_id} will remain in the group")
        
    except Exception as e:
        logging.error(f"Error in confirm_keep_member: {e}")
        bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}", show_alert=True)

@bot.callback_query_handler(func=lambda call: call.data.startswith("cancel_action_"))
def handle_action_cancellation(call):
    """Handle admin cancelling a confirmation"""
    try:
        # Parse the data to get message ID and user ID
        parts = call.data.split("_")
        message_id = parts[2]
        
        # Get the original message content to determine which action was being taken
        message_text = call.message.text.lower()
        
        # Determine which action was being cancelled and extract user ID from the message text
        # The user ID is mentioned in the text like "...give {user_display} a 2-day grace period..."
        # or "...kick {user_display} from the group..."
        
        # Find user ID in the message text - it's between 'user' and next space or line break
        user_id_match = re.search(r'user\s+(\d+)', message_text)
        if user_id_match:
            user_id = int(user_id_match.group(1))
        else:
            # Get username from the text if possible and try to find them
            username_match = re.search(r'@([a-zA-Z0-9_]+)', message_text)
            if username_match:
                username = username_match.group(1)
                try:
                    # Try to get user from username (won't work for display names)
                    chat = bot.get_chat(f"@{username}")
                    user_id = chat.id
                except:
                    # If we can't get the user, we'll show a generic message
                    user_id = None
            else:
                user_id = None
        
        # If we have the user ID, restore the original action buttons
        if user_id:
            # Check user membership status to determine appropriate actions
            is_expired = False
            days_expired = 0
            
            if str(user_id) in PAYMENT_DATA:
                due_date = datetime.strptime(PAYMENT_DATA[str(user_id)]['due_date'], '%Y-%m-%d %H:%M:%S')
                now = datetime.now()
                if due_date < now:
                    is_expired = True
                    days_expired = (now - due_date).days
            
            # Create appropriate markup based on expiration status
            markup = InlineKeyboardMarkup()
            
            if is_expired and days_expired > 3:
                # Long expired - show kick or keep only
                markup.add(
                    InlineKeyboardButton("âŒ Kick Member", callback_data=f"kick_{user_id}"),
                    InlineKeyboardButton("âœ“ Keep Member", callback_data=f"keep_{user_id}")
                )
                bot.edit_message_text(
                    f"âš ï¸ *MEMBERSHIP EXPIRED*\n\nThis member's membership has expired for {days_expired} days.\n\nWhat would you like to do?",
                    call.message.chat.id,
                    call.message.message_id,
                    parse_mode="Markdown",
                    reply_markup=markup
                )
            else:
                # Recently expired or upcoming expiry - show grace period option
                markup.add(
                    InlineKeyboardButton("â³ Give 2 Days Grace", callback_data=f"grace_{user_id}"),
                    InlineKeyboardButton("âŒ Kick Member", callback_data=f"kick_{user_id}")
                )
                bot.edit_message_text(
                    f"âš ï¸ *MEMBERSHIP EXPIRED*\n\nThis member's membership has expired or will expire soon.\n\nWhat would you like to do?",
                    call.message.chat.id,
                    call.message.message_id,
                    parse_mode="Markdown",
                    reply_markup=markup
                )
        else:
            # If we couldn't determine the user ID, show generic cancellation message
            bot.edit_message_text(
                "ðŸš« *Action cancelled*\n\n"
                "The admin action was cancelled. Please use the original notification to try again.",
                call.message.chat.id,
                call.message.message_id,
                parse_mode="Markdown"
            )
        
        bot.answer_callback_query(call.id, "âœ… Action cancelled")
    except Exception as e:
        logging.error(f"Error handling action cancellation: {e}")
        bot.answer_callback_query(call.id, "âŒ Error cancelling action", show_alert=True)


def delete_all_reminders():
    """Function to delete all payment reminder messages at midnight."""
    logging.info("Midnight cleanup: Deleting all payment reminder messages")
    
    global reminder_messages
    
    # Make a copy of the keys to avoid modifying dictionary during iteration
    user_ids = list(reminder_messages.keys())
    
    deleted_count = 0
    failed_count = 0
    
    for user_id in user_ids:
        try:
            # Delete user message if it exists
            if 'user_msg_id' in reminder_messages[user_id]:
                try:
                    bot.delete_message(user_id, reminder_messages[user_id]['user_msg_id'])
                    logging.info(f"Midnight cleanup: Deleted user message for {user_id}")
                    deleted_count += 1
                except Exception as e:
                    logging.error(f"Midnight cleanup: Failed to delete user message for {user_id}: {e}")
                    failed_count += 1
            
            # Delete admin messages if they exist
            for admin_id, msg_id in reminder_messages[user_id].get('admin_msg_ids', {}).items():
                try:
                    bot.delete_message(admin_id, msg_id)
                    logging.info(f"Midnight cleanup: Deleted admin message for admin {admin_id}")
                    deleted_count += 1
                except Exception as e:
                    logging.error(f"Midnight cleanup: Failed to delete admin message for admin {admin_id}: {e}")
                    failed_count += 1
                    
            # Delete from MongoDB
            delete_reminder_message(user_id)
            
            # Check if this user's membership has expired
            user_id_str = str(user_id)
            if user_id_str in PAYMENT_DATA:
                due_date = datetime.strptime(PAYMENT_DATA[user_id_str]['due_date'], '%Y-%m-%d %H:%M:%S')
                now = datetime.now()
                if due_date < now and PAYMENT_DATA[user_id_str].get('haspayed', False):
                    # Reset admin_action_pending flag to ensure fresh admin notifications will be sent
                    # at the next scheduled reminder time
                    PAYMENT_DATA[user_id_str]['admin_action_pending'] = True
                    logging.info(f"Midnight cleanup: Marked user {user_id} for admin notification")
            
        except Exception as e:
            logging.error(f"Midnight cleanup: Error processing user {user_id}: {e}")
            failed_count += 1
    
    # Save updates to PAYMENT_DATA
    save_payment_data()
    
    # Clear the reminder_messages dictionary
    reminder_messages.clear()
    
    # Also clear the entire MongoDB collection for a fresh start
    try:
        reminder_messages_collection.delete_many({})
        logging.info("Cleared all reminder messages from MongoDB")
    except Exception as e:
        logging.error(f"Error clearing reminder messages from MongoDB: {e}")
        
    logging.info(f"Midnight cleanup complete: {deleted_count} messages deleted, {failed_count} failures")

def midnight_cleanup_thread():
    """Thread to run at midnight and delete all reminder messages."""
    logging.info("Midnight cleanup thread started")
    
    # Track the last day we performed cleanup
    last_cleanup_date = None
    
    while True:
        try:
            # Get current time in Philippines timezone
            now = datetime.now(pytz.timezone('Asia/Manila'))
            current_time = now.strftime('%H:%M')
            current_date = now.strftime('%Y-%m-%d')
            
            # Check if it's midnight and we haven't cleaned up today
            if current_time == '00:00' and last_cleanup_date != current_date:
                logging.info("Midnight reached - cleaning up all reminder messages")
                delete_all_reminders()
                last_cleanup_date = current_date
            
            # Calculate the time to sleep until the start of the next minute
            sleep_time = 60 - now.second - now.microsecond / 1_000_000
            time.sleep(sleep_time)
            
        except Exception as e:
            logging.error(f"Error in midnight cleanup thread: {e}")
            time.sleep(60)  # Wait a minute on error before trying again

@bot.message_handler(commands=['admin_dashboard'])
def admin_dashboard(message):
    """Send link to the web-based admin dashboard"""
    # Check if user is authorized (admin or creator)
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.send_message(message.chat.id, "âŒ You are not authorized to use this command.")
        return

    # Get username for logging purposes
    username = message.from_user.username or f"User {message.from_user.id}"
    
    # Send the dashboard link with a simple button
    markup = InlineKeyboardMarkup()
    markup.add(InlineKeyboardButton("ðŸ” Open Admin Dashboard", url="https://ptabot.up.railway.app/"))
    
    bot.send_message(
        message.chat.id,
        "Click the button below to access the admin dashboard:",
        reply_markup=markup
    )
    
    # Log the access
    logging.info(f"Admin dashboard accessed by {username} ({message.from_user.id})")

@bot.message_handler(commands=['ping'])
def handle_ping_command(message):
    if message.chat.type in ['group', 'supergroup']:
        bot.send_message(message.chat.id, "ðŸ“ Pong!")
    else:
        bot.send_message(message.chat.id, "âŒ This command can only be used in group chats.")

# Define the scheduled times and corresponding GIF URLs
SCHEDULED_TIMES = {
    # Asia Open and Close
    "07:30": "gifs/AsiaOpen.mp4",
    "15:59": "gifs/AsiaClose.mp4",
    # London Open and Close
    "16:00": "gifs/LondonOpen.mp4",
    "00:30": "gifs/LondonClose.mp4",
    # New York Open and Close
    "21:30": "gifs/nyamopen.mp4",
    "23:00": "gifs/nyamclose.mp4",
    # New York Afternoon Open and Close
    "01:30": "gifs/nypmopen.mp4",
    "04:00": "gifs/nypmclose.mp4"
}

def send_scheduled_gifs():
    """Send scheduled GIFs to the group at specific times, deleting previous GIFs before sending new ones"""
    last_message_id = get_last_gif_message()
    logging.info(f"Starting GIF scheduler, last message ID: {last_message_id}")
    
    while True:
        now = datetime.now(pytz.timezone('Asia/Manila'))
        current_time = now.strftime('%H:%M')
        
        # Only proceed if it's a weekday (Monday=0, Sunday=6)
        is_weekday = now.weekday() < 5  # 0-4 are Monday to Friday
        
        if current_time in SCHEDULED_TIMES and is_weekday:
            # First, delete the previous GIF if available
            if last_message_id:
                try:
                    bot.delete_message(PAID_GROUP_ID, last_message_id)
                    logging.info(f"Deleted previous GIF message ID: {last_message_id}")
                except ApiException as e:
                    if "message to delete not found" in str(e):
                        logging.warning(f"Previous GIF message {last_message_id} already deleted")
                    elif "bot was blocked by the user" in str(e):
                        logging.warning("Cannot delete previous GIF - bot was blocked")
                    else:
                        logging.error(f"Failed to delete previous GIF: {e}")
                except Exception as e:
                    logging.error(f"General error deleting previous GIF: {e}")
            
            # Now send the new GIF
            file_path_or_url = SCHEDULED_TIMES[current_time]
            try:
                message = None
                if file_path_or_url.startswith('https'):
                    message = bot.send_animation(PAID_GROUP_ID, file_path_or_url)
                else:
                    with open(file_path_or_url, 'rb') as file:
                        if file_path_or_url.endswith('.gif'):
                            message = bot.send_animation(PAID_GROUP_ID, file)
                        elif file_path_or_url.endswith('.mp4'):
                            message = bot.send_video(PAID_GROUP_ID, file, supports_streaming=True)
                
                if message:
                    # Store the new message ID for future deletion
                    last_message_id = message.message_id
                    save_last_gif_message(last_message_id)
                    
                logging.info(f"Sent scheduled file at {current_time} Philippine time. New message ID: {last_message_id}")
            except Exception as e:
                logging.error(f"Failed to send scheduled file at {current_time}: {e}")
        elif current_time in SCHEDULED_TIMES and not is_weekday:
            logging.info(f"Skipped scheduled file at {current_time}: Weekend.")
        
        # Calculate the time to sleep until the start of the next minute
        now = datetime.now(pytz.timezone('Asia/Manila'))
        sleep_time = 60 - now.second - now.microsecond / 1_000_000
        time.sleep(sleep_time)

CREATOR_USERNAME = "FujiPTA" 

@bot.message_handler(commands=['tip'])
def handle_tip_command(message):
    if message.chat.type in ['group', 'supergroup']:
        tip_message = (
            f"â¤ï¸ Love the bot? Give a tip to the creator! @{CREATOR_USERNAME}!\n\n"
            "ðŸ’¸ *Crypto Payments*:\n\n"
            "*USDT (TRC20)*: `TV9K3DwWLufYU5yeyXZYCtB3QNX1s983wD`\n\n"
            "*Bitcoin*: `3H7uF4H29cqDiUGNd7M9tpWashEfN8a3wP`\n\n"
            "ðŸ“± *GCash*:\n"
            "GCash Number: `09763624531` (J. M.)"
        )
        bot.send_message(message.chat.id, tip_message, parse_mode='Markdown')
    else:
        bot.send_message(message.chat.id, "âŒ This command can only be used in group chats.")

@bot.message_handler(commands=['dashboard', 'status'])
def show_user_dashboard(message):
    """Display the user's membership dashboard with status and details"""
    chat_id = message.chat.id
    user_id = str(message.from_user.id)
    
    # Check if this is a private chat
    if message.chat.type != 'private':
        bot.send_message(chat_id, "ðŸ”’ Please use this command in a private message with the bot.")
        return
        
    # Check if the user has membership data
    if user_id in PAYMENT_DATA:
        data = PAYMENT_DATA[user_id]
        username = message.from_user.username or "No Username"
        
        # Escape special markdown characters in username to prevent parsing errors
        username = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', username)
        
        # Calculate days remaining until expiration
        try:
            due_date = datetime.strptime(data['due_date'], '%Y-%m-%d %H:%M:%S')
            current_date = datetime.now()
            days_remaining = (due_date - current_date).days
            hours_remaining = int((due_date - current_date).seconds / 3600)
            
            # Check if membership is cancelled first
            if data.get('cancelled', False):
                status_icon = "ðŸš«"
                status_text = "Cancelled"
            # If not cancelled, format status based on days remaining
            elif days_remaining > 7:
                status_icon = "âœ…"
                status_text = "Active"
            elif days_remaining > 0:
                status_icon = "âš ï¸"
                status_text = "Expiring Soon"
            else:
                status_icon = "âŒ"
                status_text = "Expired"
                
            # Get mentorship type - added in this update
            mentorship_type = data.get('mentorship_type', 'Regular')
            
            # Get challenge points if available
            user_points = 0
            user_rank = "N/A"
            
            # Check for points in the current month
            current_month = datetime.now().strftime('%Y-%m')
            try:
                monthly_scores = get_monthly_leaderboard(current_month)
                for i, entry in enumerate(monthly_scores):
                    if entry.get('user_id') == int(user_id):
                        user_points = entry.get('total_points', 0)
                        user_rank = i + 1
                        break
            except Exception as e:
                logging.error(f"Error getting user points for dashboard: {e}")
            
            # Create progress bar for time remaining
            if days_remaining > 0:
                # Create progress bar based on plan duration
                plan = data.get('payment_plan', '')
                
                if "Yearly" in plan or "Legacy" in plan:
                    total_days = 365
                elif "3 Months" in plan or "Momentum" in plan or "Apprentice" in plan:
                    total_days = 90
                elif "6 Months" in plan or "Disciple" in plan:
                    total_days = 180
                elif "Lifetime" in plan:
                    total_days = 3650  # ~10 years
                else:
                    total_days = 30  # Default to monthly
                    
                # Calculate days elapsed 
                elapsed = total_days - days_remaining
                if elapsed < 0:  # Handle edge cases
                    elapsed = 0
                    
                # Cap elapsed at total days
                if elapsed > total_days:
                    elapsed = total_days
                    
                # Create visual progress bar (15 segments) 
                filled = int((elapsed / total_days) * 15)
                progress_bar = "â–ˆ" * filled + "â–’" * (15 - filled)
                percentage = round((elapsed / total_days) * 100)
                
                time_progress = f"{progress_bar} {percentage}% used"
            elif data.get('cancelled', False):
                time_progress = "ðŸš« Membership cancelled"
            else:
                time_progress = "âŒ Membership expired"
            
            # Escape any special characters in payment_plan and payment_mode
            payment_plan = data.get('payment_plan', 'Unknown')
            payment_mode = data.get('payment_mode', 'Unknown')
            
            # Create and send the dashboard message with improved formatting
            dashboard_message = (
                f"ðŸ“Š *MEMBERSHIP DASHBOARD*\n\n"
                f"â”Œâ”€â”€â”€â”€â”€â”€â”€ ðŸ‘¤ *MEMBER INFO* â”€â”€â”€â”€â”€â”€â”€â”€â”\n"
                f"â”‚ *Username:* @{username}\n"
                f"â”‚ *Member ID:* `{user_id}`\n"
                f"â””â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”˜\n\n"
                
                f"â”Œâ”€â”€â”€â”€â”€â”€â”€ ðŸ’³ *SUBSCRIPTION* â”€â”€â”€â”€â”€â”€â”€â”€â”\n"
                f"â”‚ *Plan:* {payment_plan}\n"
                f"â”‚ *Type:* {mentorship_type} Mentorship\n"
                f"â”‚ *Payment Method:* {payment_mode}\n"
                f"â””â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”˜\n\n"
                
                f"â”Œâ”€â”€â”€â”€â”€â”€â”€ â±ï¸ *STATUS* â”€â”€â”€â”€â”€â”€â”€â”€â”\n"
                f"â”‚ *Status:* {status_icon} {status_text}\n"
                f"â”‚ *Expiration:* {due_date.strftime('%Y-%m-%d')}\n"
                f"â”‚ *Time Left:* {days_remaining} days, {hours_remaining} hours\n"
                f"â”‚ {time_progress}\n"
                f"â””â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”˜\n"
            )
            
            # Add challenge stats if user has participated
            if user_points > 0:
                dashboard_message += (
                    f"\nâ”Œâ”€â”€â”€â”€â”€â”€â”€ ðŸ† *CHALLENGES* â”€â”€â”€â”€â”€â”€â”€â”€â”\n"
                    f"â”‚ *Monthly Points:* {user_points} points\n"
                    f"â”‚ *Current Rank:* #{user_rank}\n"
                    f"â””â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”˜\n"
                )
            
            # Add renewal instructions if expiring soon (and not cancelled)
            if days_remaining < 7 and days_remaining >= 0 and not data.get('cancelled', False):
                dashboard_message += (
                    f"\nâš ï¸ *Your membership expires soon!*\n"
                    f"Use /start and select 'Renew Membership' to continue access.\n"
                )
            # Add special message for cancelled memberships
            elif data.get('cancelled', False):
                dashboard_message += (
                    f"\nðŸš« *Your membership has been cancelled*\n"
                    f"You will still have access until {due_date.strftime('%Y-%m-%d')}.\n"
                    f"To reactivate before expiration, use /start and select 'Renew Membership'.\n"
                )
                
            # Add help information with some quick commands
            dashboard_message += (
                f"\nðŸ“‹ *QUICK COMMANDS*\n"
                f"â€¢ /start - Main menu\n"
                f"â€¢ /confess - Submit anonymous confession\n"
                f"â€¢ /commands - View all available commands"
            )
            
            # Create helpful inline buttons
            markup = InlineKeyboardMarkup(row_width=2)
            
            if days_remaining < 7 and days_remaining >= 0 and not data.get('cancelled', False):
                # Show renewal button for expiring memberships
                markup.add(InlineKeyboardButton("ðŸ”„ Renew Membership", callback_data="start_renewal"))
            
            # Add general buttons
            markup.add(InlineKeyboardButton("â“ FAQ", callback_data="faq_back"))
            markup.add(InlineKeyboardButton("ðŸ“Š Leaderboard", callback_data="view_leaderboard"))
            
            bot.send_message(chat_id, dashboard_message, parse_mode="Markdown", reply_markup=markup)
            
        except Exception as e:
            bot.send_message(chat_id, f"âŒ Error retrieving dashboard: {str(e)}")
            logging.error(f"Dashboard error for user {user_id}: {str(e)}")
    else:
        # User doesn't have membership data
        bot.send_message(
            chat_id, 
            "âŒ *No active membership found*\n\n"
            "You don't appear to have an active membership. Use /start to enroll in Prodigy Trading Academy.",
            parse_mode="Markdown"
        )

@bot.message_handler(commands=['supreme_dashboard'])
def show_supreme_dashboard(message, edit_message=False, call_obj=None):
    """Display the Supreme Membership dashboard with detailed student information and task overview"""
    if edit_message and call_obj:
        # We're editing an existing message (returning from a sub-menu)
        chat_id = call_obj.message.chat.id
        message_id = call_obj.message.message_id
        user_id = call_obj.from_user.id
    else:
        # Normal command invocation
        chat_id = message.chat.id
        user_id = message.from_user.id
    
    user_id_str = str(user_id)
    
    # Only allow in private chats for privacy (only check for normal command invocation)
    if not edit_message and message.chat.type != 'private':
        bot.send_message(chat_id, "ðŸ”’ Please use this command in a private message with the bot.")
        return
    
    # Check if the user has Supreme membership
    if user_id_str in PAYMENT_DATA and PAYMENT_DATA[user_id_str].get('mentorship_type', '').lower() == 'supreme':
        try:
            data = PAYMENT_DATA[user_id_str]
            
            if edit_message:
                username = call_obj.from_user.username or "No Username"
            else:
                username = message.from_user.username or "No Username"
                
            full_name = data.get('form_answers', {}).get('full_name', username)
            
            # Get membership details
            plan = data.get('payment_plan', 'Unknown')
            
            # Determine the mentor (you can customize this logic)
            mentors = {'Apprentice': 'Rom', 'Disciple': 'Konfu', 'Legacy': 'Nath'}
            mentor = mentors.get(plan, '???')
            
            # Calculate enrollment date and expiry date
            enrollment_date = datetime.strptime(data.get('enrollment_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')), '%Y-%m-%d %H:%M:%S')
            due_date = datetime.strptime(data.get('due_date', '2099-12-31 23:59:59'), '%Y-%m-%d %H:%M:%S')
            
            # Calculate days remaining until expiration
            current_date = datetime.now()
            days_remaining = (due_date - current_date).days
            hours_remaining = int((due_date - current_date).seconds / 3600)
            
            # Get mentorship type
            mentorship_type = data.get('mentorship_type', 'Supreme')
            
            # Check if membership is cancelled
            if data.get('cancelled', False):
                status_icon = "ðŸš«"
                status_text = "Cancelled"
            # If not cancelled, format status based on days remaining
            elif days_remaining > 7:
                status_icon = "âœ…"
                status_text = "Active"
            elif days_remaining > 0:
                status_icon = "âš ï¸"
                status_text = "Expiring Soon"
            else:
                status_icon = "âŒ"
                status_text = "Expired"
            
            # Determine user ranking based on custom logic (placeholder)
            rank_points = data.get('rank_points', 0)
            
            ranking_tiers = {
                (0, 100): "Rookie I",
                (101, 200): "Rookie II",
                (201, 300): "Rookie III",
                (301, 500): "Elite I",
                (501, 700): "Elite II",
                (701, 900): "Elite III",
                (901, 1200): "Master I",
                (1201, 1500): "Master II",
                (1501, 1800): "Master III",
                (1801, 2200): "Legend I",
                (2201, 2600): "Legend II",
                (2601, 3000): "Legend III",
                (3001, float('inf')): "Sovereign"
            }
            
            ranking = "Rookie I"  # Default
            for (min_points, max_points), rank in ranking_tiers.items():
                if min_points <= rank_points <= max_points:
                    ranking = rank
                    break
            
            # Get user expertise level from form answers
            expertise_level = "Beginner"
            
            if 'form_answers' in data and 'expertise_level' in data['form_answers']:
                expertise = data['form_answers']['expertise_level'].lower()
                
                if 'beginner' in expertise or 'a.' in expertise:
                    expertise_level = "Beginner"
                elif 'intermediate' in expertise or 'c.' in expertise:
                    expertise_level = "Intermediate"  
                elif 'expert' in expertise or 'd.' in expertise:
                    expertise_level = "Expert"
                elif 'master' in expertise or 'e.' in expertise:
                    expertise_level = "Master"
            
            # Create nice formatting for the dashboard with box-style formatting
            dashboard_message = (
                f"âœ¨ *SUPREME DASHBOARD*\n\n"
                
                f"â”Œâ”€â”€â”€â”€â”€â”€â”€ ðŸ‘¤ *STUDENT INFO* â”€â”€â”€â”€â”€â”€â”€â”€â”\n"
                f" *Name:* {full_name}\n"
                f" *Member ID:* `{user_id_str}`\n"
                f" *Ranking:* {ranking}\n"
                f" *Expertise Level:* {expertise_level}\n"
                f"â””â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”˜\n\n"
                
                f"â”Œâ”€â”€â”€â”€â”€â”€â”€ ðŸ’³ *SUBSCRIPTION* â”€â”€â”€â”€â”€â”€â”€â”€â”\n"
                f" *Plan:* {plan}\n"
                f" *Type:* {mentorship_type} Mentorship\n"
                f" *Mentor:* {mentor}\n"
                f" *Enrollment:* {enrollment_date.strftime('%B %d, %Y')}\n"
                f"â””â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”˜\n\n"
                
                f"â”Œâ”€â”€â”€â”€â”€â”€â”€ â±ï¸ *STATUS* â”€â”€â”€â”€â”€â”€â”€â”€â”\n"
                f" *Status:* {status_icon} {status_text}\n"
                f" *Expiration:* {due_date.strftime('%Y-%m-%d')}\n"
                f" *Time Left:* {days_remaining} days, {hours_remaining} hours\n"
                f"â””â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”˜\n\n"
                
                f"â”Œâ”€â”€â”€â”€â”€â”€â”€ ðŸ“ *LEARNING TASKS* â”€â”€â”€â”€â”€â”€â”€â”€â”\n"
                f" â€¢ Complete Market Structure Lesson\n"
                f"   *Due: Friday*\n"
                f" â€¢ Submit Weekly Chart Analysis\n"
                f"   *Due: Sunday*\n"
                f" â€¢ Schedule Next 1:1 Session\n"
                f"   *Due: Tomorrow*\n"
                f"â””â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”˜\n"
            )
            
            # Add renewal instructions if expiring soon (and not cancelled)
            if days_remaining < 7 and days_remaining >= 0 and not data.get('cancelled', False):
                dashboard_message += (
                    f"\nâš ï¸ *Your membership expires soon!*\n"
                    f"Use /start and select 'Renew Membership' to continue access.\n"
                )
            # Add special message for cancelled memberships
            elif data.get('cancelled', False):
                dashboard_message += (
                    f"\nðŸš« *Your membership has been cancelled*\n"
                    f"You will still have access until {due_date.strftime('%Y-%m-%d')}.\n"
                    f"To reactivate before expiration, use /start and select 'Renew Membership'.\n"
                )
            
            # Create navigation buttons for the Supreme dashboard - UPDATED PER REQUIREMENTS
            markup = InlineKeyboardMarkup(row_width=2)
            
            # First row
            markup.add(
                InlineKeyboardButton("ðŸ’Ž Supreme Vault", callback_data="supreme_vault"),
                InlineKeyboardButton("ðŸ‘¨â€ðŸ« Mentor Line", callback_data="mentor_line")
            )
            
            # Second row
            markup.add(
                InlineKeyboardButton("ðŸ›£ï¸ Trading Journey", callback_data="trading_journey"),
                InlineKeyboardButton("ðŸ“” Trading Journal", callback_data="trading_journal")
            )
            
            # Third row
            markup.add(
                InlineKeyboardButton("ðŸ† Leaderboard", callback_data="supreme_leaderboard"),
                InlineKeyboardButton("ðŸ‘¥ Supreme Network", callback_data="supreme_network")
            )
            
            # Add renewal button if needed
            if days_remaining < 7 and days_remaining >= 0 and not data.get('cancelled', False):
                markup.add(InlineKeyboardButton("ðŸ”„ Renew Membership", callback_data="start_renewal"))
            
            # Send or edit the dashboard message with buttons
            if edit_message:
                bot.edit_message_text(
                    dashboard_message, 
                    chat_id,
                    message_id,
                    parse_mode="Markdown", 
                    reply_markup=markup
                )
            else:
                bot.send_message(chat_id, dashboard_message, parse_mode="Markdown", reply_markup=markup)
            
        except Exception as e:
            error_msg = f"âŒ Error retrieving Supreme dashboard: {str(e)}"
            if edit_message and call_obj:
                bot.edit_message_text(error_msg, chat_id, message_id)
            else:
                bot.send_message(chat_id, error_msg)
            logging.error(f"Supreme dashboard error for user {user_id}: {str(e)}")
    else:
        # User is not a Supreme member
        error_msg = (
            "â­ *Supreme Dashboard Access Restricted*\n\n"
            "This feature is exclusively for Supreme membership holders. "
            "To upgrade your membership, use /start and select 'Purchase Membership'."
        )
        
        if edit_message and call_obj:
            bot.edit_message_text(error_msg, chat_id, message_id, parse_mode="Markdown")
        else:
            bot.send_message(chat_id, error_msg, parse_mode="Markdown")

# Add a callback handler for Supreme dashboard buttons
@bot.callback_query_handler(func=lambda call: call.data.startswith("supreme_") or call.data in ["mentor_line", "trading_journey", "trading_journal", "supreme_vault", "supreme_leaderboard"])
def handle_supreme_dashboard_buttons(call):
    """Handle navigation buttons on the Supreme dashboard"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    
    # Handle different button actions based on callback data
    if call.data == "supreme_vault":
        # Show Supreme Vault options
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineKeyboardButton("ðŸ“š E-books", callback_data="vault_ebooks"),
            InlineKeyboardButton("ðŸ“” Trading Journal", callback_data="vault_journal")
        )
        markup.add(
            InlineKeyboardButton("ðŸŽ¬ Video Lessons", callback_data="vault_videos"),
            InlineKeyboardButton("ðŸ“– Prodigy Codex", callback_data="vault_codex")
        )
        markup.add(
            InlineKeyboardButton("Â« Back to Dashboard", callback_data="back_to_supreme_dashboard")
        )
        
        bot.edit_message_text(
            "ðŸ’Ž *SUPREME VAULT*\n\n"
            "Access exclusive premium resources curated for Supreme members. "
            "Select a category to explore:",
            chat_id,
            message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
        
    elif call.data == "mentor_line":
        # Check if user has a mentor
        user_id = call.from_user.id
        user_id_str = str(user_id)
        has_mentor = False
        
        try:
            if user_id_str in PAYMENT_DATA:
                mentor = PAYMENT_DATA[user_id_str].get('assigned_mentor')
                if mentor:
                    has_mentor = True
                    
                    # Show mentor details - ENHANCED WITH AVAILABILITY AND STATUS
                    markup = InlineKeyboardMarkup(row_width=2)
                    
                    # Main mentor interaction buttons - now in 2x2 grid
                    markup.add(
                        InlineKeyboardButton("ðŸ’¬ Message", callback_data=f"message_mentor_{mentor}"),
                        InlineKeyboardButton("ðŸ“… Schedule", callback_data=f"schedule_mentor_{mentor}")
                    )
                    markup.add(
                        InlineKeyboardButton("ðŸ“š Resources", callback_data=f"mentor_resources_{mentor}"),
                        InlineKeyboardButton("â“ Support", callback_data=f"mentor_support_{mentor}")
                    )
                    
                    # History and back buttons
                    markup.add(
                        InlineKeyboardButton("ðŸ“‹ Session History", callback_data=f"mentor_history_{mentor}")
                    )
                    markup.add(
                        InlineKeyboardButton("Â« Back to Dashboard", callback_data="back_to_supreme_dashboard")
                    )
                    
                    # Check if MENTORS is initialized and contains this mentor
                    if 'MENTORS' in globals() and MENTORS and mentor in MENTORS:
                        mentor_info = MENTORS.get(mentor, {})
                    else:
                        # Fallback mentor details if database isn't loaded
                        mentor_info = {
                            "name": mentor,
                            "expertise": "Trading",
                            "description": "Your dedicated Supreme mentor.",
                            "availability": "Contact for availability",
                            "current_status": "Available",
                            "success_rate": "90%"
                        }
                        logging.warning(f"Using fallback mentor info for {mentor} - MENTORS may not be properly initialized")
                    
                    # Create a status emoji based on current status
                    status_emoji = "ðŸŸ¢" if mentor_info.get('current_status') == "Available" else "ðŸ”´" if mentor_info.get('current_status') == "In Session" else "ðŸŸ¡"
                    
                    # Next scheduled session (placeholder - would be from actual data)
                    next_session = "Tomorrow at 3:00 PM PHT" 
                    
                    bot.edit_message_text(
                        f"ðŸ‘¨â€ðŸ« *YOUR MENTOR*\n\n"
                        f"*Name:* {mentor_info.get('name', mentor)}\n"
                        f"*Status:* {status_emoji} {mentor_info.get('current_status', 'Unknown')}\n"
                        f"*Expertise:* {mentor_info.get('expertise', 'Trading')}\n"
                        f"*Success Rate:* {mentor_info.get('success_rate', 'N/A')}\n\n"
                        f"*About:* {mentor_info.get('description', 'Your dedicated Supreme mentor.')}\n\n"
                        f"*Availability:* {mentor_info.get('availability', 'Contact for availability')}\n"
                        f"*Next Session:* {next_session}\n\n"
                        f"What would you like to do?",
                        chat_id,
                        message_id,
                        parse_mode="Markdown",
                        reply_markup=markup
                    )
                else:
                    # No mentor assigned yet, show enhanced find mentor options
                    markup = InlineKeyboardMarkup(row_width=1)
                    markup.add(
                        InlineKeyboardButton("ðŸ” Browse Available Mentors", callback_data="view_mentors"),
                        InlineKeyboardButton("â“ How Mentorship Works", callback_data="mentor_info"),
                        InlineKeyboardButton("Â« Back to Dashboard", callback_data="back_to_supreme_dashboard")
                    )
                    
                    bot.edit_message_text(
                        "ðŸ‘¨â€ðŸ« *SUPREME MENTOR LINE*\n\n"
                        "Finding the right mentor can transform your trading journey. Our expert mentors are "
                        "here to provide personalized guidance tailored to your style and goals.\n\n"
                        "âœ… *Benefits of having a mentor:*\n"
                        "â€¢ Personalized trading strategies\n"
                        "â€¢ Faster improvement curve\n"
                        "â€¢ Accountability partner for consistent growth\n"
                        "â€¢ Direct feedback on your trades\n\n"
                        "What would you like to do?",
                        chat_id,
                        message_id,
                        parse_mode="Markdown",
                        reply_markup=markup
                    )
            else:
                # User doesn't have payment data, show "no mentor" options
                markup = InlineKeyboardMarkup(row_width=1)
                markup.add(
                    InlineKeyboardButton("ðŸ” Browse Available Mentors", callback_data="view_mentors"),
                    InlineKeyboardButton("ðŸ§ª 1-Week Mentor Trial", callback_data="mentor_trial"),
                    InlineKeyboardButton("â“ How Mentorship Works", callback_data="mentor_info"),
                    InlineKeyboardButton("Â« Back to Dashboard", callback_data="back_to_supreme_dashboard")
                )
                
                bot.edit_message_text(
                    "ðŸ‘¨â€ðŸ« *SUPREME MENTOR LINE*\n\n"
                    "Finding the right mentor can transform your trading journey. Our expert mentors are "
                    "here to provide personalized guidance tailored to your style and goals.\n\n"
                    "âœ… *Benefits of having a mentor:*\n"
                    "â€¢ Personalized trading strategies\n"
                    "â€¢ Faster improvement curve\n"
                    "â€¢ Accountability partner for consistent growth\n"
                    "â€¢ Direct feedback on your trades\n\n"
                    "What would you like to do?",
                    chat_id,
                    message_id,
                    parse_mode="Markdown",
                    reply_markup=markup
                )
        except Exception as e:
            logging.error(f"Error in mentor_line handler: {e}")
            # Provide a fallback UI in case of errors
            markup = InlineKeyboardMarkup(row_width=1)
            markup.add(InlineKeyboardButton("Â« Back to Dashboard", callback_data="back_to_supreme_dashboard"))
            
            bot.edit_message_text(
                "âš ï¸ *Something went wrong*\n\n"
                "We encountered a problem displaying your mentor information. Please try again later or contact support.",
                chat_id,
                message_id,
                parse_mode="Markdown",
                reply_markup=markup
            )
                
    elif call.data == "trading_journey":
        # Show Trading Journey options
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(
            InlineKeyboardButton("ðŸ“Š Market Structure Path", callback_data="journey_market_structure"),
            InlineKeyboardButton("ðŸ’° Smart Money Path", callback_data="journey_smart_money"),
            InlineKeyboardButton("âš™ï¸ Automated Trading Path", callback_data="journey_automated"),
            InlineKeyboardButton("ðŸ§  Trading Psychology Path", callback_data="journey_psychology"),
            InlineKeyboardButton("Â« Back to Dashboard", callback_data="back_to_supreme_dashboard")
        )
        
        bot.edit_message_text(
            "ðŸ›£ï¸ *TRADING JOURNEY*\n\n"
            "Track your progress and access personalized learning paths "
            "based on your trading interests and mentor's expertise.\n\n"
            "Select a journey to explore:",
            chat_id,
            message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
        
    elif call.data == "trading_journal":
        # Show Trading Journal features
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineKeyboardButton("ðŸ“ New Entry", callback_data="journal_new_entry"),
            InlineKeyboardButton("ðŸ“ˆ Performance", callback_data="journal_performance")
        )
        markup.add(
            InlineKeyboardButton("ðŸ“Š Statistics", callback_data="journal_stats"),
            InlineKeyboardButton("ðŸ” Analysis", callback_data="journal_analysis")
        )
        markup.add(
            InlineKeyboardButton("Â« Back to Dashboard", callback_data="back_to_supreme_dashboard")
        )
        
        bot.edit_message_text(
            "ðŸ“” *TRADING JOURNAL*\n\n"
            "Track your trades, analyze patterns, and improve your performance "
            "with our comprehensive trading journal.\n\n"
            "What would you like to do?",
            chat_id,
            message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
        
    elif call.data == "supreme_leaderboard":
        # Show Supreme Leaderboard
        # Get current month leaderboard data (placeholder logic)
        now = datetime.now(pytz.timezone('Asia/Manila'))
        month_year = now.strftime('%Y-%m')
        leaderboard_data = get_monthly_leaderboard(month_year)
        
        # Format leaderboard text
        leaderboard_text = f"ðŸ† *SUPREME LEADERBOARD: {now.strftime('%B %Y')}*\n\n"
        
        if not leaderboard_data:
            leaderboard_text += "No entries yet for this month!"
        else:
            # Keep track of the current rank and last score for tie detection
            current_rank = 1
            last_score = None
            
            for i, entry in enumerate(leaderboard_data[:10]):  # Show top 10
                total_points = entry.get('total_points', 0)
                username = entry.get('username', 'No_Username')
                
                # If this score is different from the previous one, update the rank
                if last_score is not None and total_points != last_score:
                    current_rank = i + 1
                    
                last_score = total_points
                
                # Create emoji for ranks
                if current_rank == 1:
                    rank_emoji = "ðŸ¥‡"
                elif current_rank == 2:
                    rank_emoji = "ðŸ¥ˆ"
                elif current_rank == 3:
                    rank_emoji = "ðŸ¥‰"
                else:
                    rank_emoji = f"{current_rank}."
                
                leaderboard_text += f"{rank_emoji} @{username}: *{total_points} points*\n"
        
        # Create back button
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(InlineKeyboardButton("Â« Back to Dashboard", callback_data="back_to_supreme_dashboard"))
        
        bot.edit_message_text(
            leaderboard_text,
            chat_id,
            message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
        
    elif call.data == "supreme_network":
        # Show Supreme Network options
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(
            InlineKeyboardButton("ðŸ’¬ Supreme Group Chat", url="https://t.me/+abc123"),  # Replace with actual group link
            InlineKeyboardButton("ðŸ‘¥ Members Directory", callback_data="network_directory"),
            InlineKeyboardButton("ðŸ“„ Group Guidelines", callback_data="network_guidelines"),
            InlineKeyboardButton("Â« Back to Dashboard", callback_data="back_to_supreme_dashboard")
        )
        
        bot.edit_message_text(
            "ðŸ‘¥ *SUPREME NETWORK*\n\n"
            "Connect with fellow Supreme members for collaboration, growth, and "
            "exclusive networking opportunities.\n\n"
            "Select an option to begin:",
            chat_id,
            message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
    
    elif call.data == "back_to_supreme_dashboard":
        # Return to main supreme dashboard (re-display it)
        show_supreme_dashboard(call.message, edit_message=True, call_obj=call)
    
    # Answer the callback to remove the loading indicator
    bot.answer_callback_query(call.id)

# Add this after your existing callback handlers
@bot.callback_query_handler(func=lambda call: call.data == "view_mentors")
def view_available_mentors(call):
    """Show the list of available mentors to choose from"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    
    # Create a formatted list of mentors from database
    mentors_message = "ðŸ” *AVAILABLE MENTORS*\n\n"
    
    # Convert database dictionary to list for easier iteration
    mentor_list = []
    for mentor_id, mentor_data in MENTORS.items():
        mentor_data['id'] = mentor_id  # Add ID to the data
        mentor_list.append(mentor_data)
    
    # If no mentors found in database, show empty message
    if not mentor_list:
        mentors_message += "No mentors currently available. Please try again later."
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(InlineKeyboardButton("Â« Back to Mentor Line", callback_data="mentor_line"))
        
        bot.edit_message_text(
            mentors_message,
            chat_id,
            message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
        bot.answer_callback_query(call.id)
        return
    
    # Format each mentor entry
    for i, mentor in enumerate(mentor_list, 1):
        # Create availability indicators based on level in database
        avail_level = mentor.get("availability_level", "Unknown")
        if avail_level == "High":
            avail_indicator = "ðŸŸ¢ High"
        elif avail_level == "Medium":
            avail_indicator = "ðŸŸ¡ Medium"
        else:
            avail_indicator = "ðŸŸ  Limited"
            
        mentors_message += (
            f"*{i}. {mentor['name']}*\n"
            f"â€¢ Expertise: {mentor.get('expertise', 'Trading')}\n"
            f"â€¢ Experience: {mentor.get('experience', 'N/A')}\n"
            f"â€¢ Style: {mentor.get('style', 'N/A')}\n"
            f"â€¢ Availability: {avail_indicator}\n\n"
        )
    
    mentors_message += "Select a mentor to view their profile and request mentorship:"
    
    # Create buttons for each mentor
    markup = InlineKeyboardMarkup(row_width=2)
    
    # Add buttons in pairs for better layout
    buttons = []
    for mentor in mentor_list:
        buttons.append(InlineKeyboardButton(
            mentor["name"], 
            callback_data=f"mentor_profile_{mentor['id']}"
        ))
    
    # Add buttons in pairs
    for i in range(0, len(buttons), 2):
        if i + 1 < len(buttons):
            markup.add(buttons[i], buttons[i+1])
        else:
            markup.add(buttons[i])
    
    # Add back button
    markup.add(InlineKeyboardButton("Â« Back to Mentor Line", callback_data="mentor_line"))
    
    bot.edit_message_text(
        mentors_message,
        chat_id,
        message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    # Answer the callback to remove loading state
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda call: call.data.startswith("mentor_profile_"))
def view_mentor_profile(call):
    """Show detailed profile for a specific mentor"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    
    # Extract mentor ID from callback data
    mentor_id = call.data.split("_")[2]
    
    # Get mentor data from the database
    mentor = MENTORS.get(mentor_id)
    
    # Check if mentor exists in our database
    if not mentor:
        bot.answer_callback_query(call.id, "Mentor profile not found!")
        return
    
    # Create strengths bullet points
    strengths_text = ""
    for strength in mentor.get("strengths", ["No strengths listed"]):
        strengths_text += f"â€¢ {strength}\n"
    
    # Create the profile message
    profile_message = (
        f"ðŸ§  *{mentor['name'].upper()} - MENTOR PROFILE*\n\n"
        f"*{mentor.get('full_name', mentor['name'])}*\n"
        f"_{mentor.get('title', 'Trading Mentor')}_\n\n"
        f"*Expertise:* {mentor.get('expertise', 'Trading')}\n"
        f"*Experience:* {mentor.get('experience', 'N/A')}\n"
        f"*Trading Style:* {mentor.get('style', 'N/A')}\n"
        f"*Success Rate:* {mentor.get('success_rate', 'N/A')}\n\n"
        f"*About:*\n{mentor.get('description', 'No description available.')}\n\n"
        f"*Key Strengths:*\n{strengths_text}\n"
        f"*Teaching Style:*\n{mentor.get('teaching_style', 'No teaching style specified.')}\n\n"
        f"*Availability:*\n{mentor.get('availability', 'Contact for availability.')}\n\n"
        f"*Student Testimonial:*\n_{mentor.get('student_testimonial', 'No testimonials yet.')}_\n"
    )
    
    # Create buttons for profile actions
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(
        InlineKeyboardButton("ðŸ“† Request Mentorship", callback_data=f"request_mentor_{mentor_id}"),
        InlineKeyboardButton("Â« Back to Mentor List", callback_data="view_mentors"),
        InlineKeyboardButton("Â« Back to Mentor Line", callback_data="mentor_line")
    )
    
    bot.edit_message_text(
        profile_message,
        chat_id,
        message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    # Answer the callback to remove loading state
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda call: call.data.startswith("request_mentor_"))
def handle_mentor_request(call):
    """Handle a user requesting mentorship with a specific mentor"""
    chat_id = call.message.chat.id
    message_id = call.message.message_id
    user_id = call.from_user.id
    
    # Extract mentor ID from callback data
    mentor_id = call.data.split("_")[2]
    
    # Get user information
    username = call.from_user.username or f"User {user_id}"
    first_name = call.from_user.first_name or username
    
    # Get mentor data
    mentor = MENTORS.get(mentor_id)
    if not mentor:
        bot.answer_callback_query(call.id, "âŒ Mentor not found in database", show_alert=True)
        return
    
    mentor_name = mentor.get('name', mentor_id)
    
    try:
        # Find the mentor's user ID from PAYMENT_DATA
        mentor_user_id = None
        for user_id_str, data in PAYMENT_DATA.items():
            if data.get('is_mentor') and data.get('mentor_id') == mentor_id:
                mentor_user_id = int(user_id_str)
                break
        
        if not mentor_user_id:
            # Fallback for demo/development - use creator ID
            mentor_user_id = CREATOR_ID
            logging.warning(f"No mentor user ID found for {mentor_id}, using creator ID")
        
        # Send confirmation to user
        bot.edit_message_text(
            f"âœ… *Mentorship Request Sent*\n\n"
            f"Your request to have {mentor_name} as your mentor has been sent!\n\n"
            f"You will be notified once the mentor responds to your request.",
            chat_id,
            message_id,
            parse_mode="Markdown"
        )
        
        # Create acceptance/rejection buttons for mentor
        markup = InlineKeyboardMarkup(row_width=2)
        markup.add(
            InlineKeyboardButton("âœ… Accept", callback_data=f"accept_mentee_{user_id}_{mentor_id}"),
            InlineKeyboardButton("âŒ Decline", callback_data=f"decline_mentee_{user_id}_{mentor_id}")
        )
        
        # Send request notification to mentor
        bot.send_message(
            mentor_user_id,
            f"ðŸ”” *New Mentorship Request*\n\n"
            f"@{username} ({first_name}) would like to have you as their mentor.\n\n"
            f"Would you like to accept this mentorship request?",
            parse_mode="Markdown",
            reply_markup=markup
        )
        
        bot.answer_callback_query(call.id, "Request sent successfully!")
        
    except Exception as e:
        logging.error(f"Error handling mentor request: {e}")
        bot.answer_callback_query(call.id, "âŒ Error processing request", show_alert=True)

@bot.callback_query_handler(func=lambda call: call.data.startswith("accept_mentee_"))
def handle_accept_mentee(call):
    """Handle mentor accepting a mentee"""
    parts = call.data.split("_")
    mentee_id = int(parts[2])
    mentor_id = parts[3]
    
    # Get mentor data for display
    mentor = MENTORS.get(mentor_id)
    mentor_name = mentor.get('name', mentor_id)
    
    # Generate random session date (between 3-7 days from now)
    days_ahead = random.randint(3, 7)
    session_date = datetime.now() + timedelta(days=days_ahead)
    
    # Random time between 9 AM and 5 PM
    hour = random.randint(9, 17)
    minute = random.choice([0, 15, 30, 45])
    
    session_time = session_date.replace(hour=hour, minute=minute)
    formatted_date = session_time.strftime("%A, %B %d, %Y")
    formatted_time = session_time.strftime("%I:%M %p")
    
    # Update mentee's data to assign mentor
    user_id_str = str(mentee_id)
    if user_id_str in PAYMENT_DATA:
        PAYMENT_DATA[user_id_str]['assigned_mentor'] = mentor_id
        PAYMENT_DATA[user_id_str]['mentor_session_date'] = session_time.strftime('%Y-%m-%d %H:%M:%S')
        save_payment_data()
    
    try:
        # Notify mentee of acceptance
        bot.send_message(
            mentee_id,
            f"ðŸŽ‰ *Mentorship Request Accepted!*\n\n"
            f"Great news! {mentor_name} has accepted your mentorship request.\n\n"
            f"Your first 1:1 session is scheduled for:\n"
            f"ðŸ“† *Date:* {formatted_date}\n"
            f"â° *Time:* {formatted_time} PHT\n\n"
            f"Please be prepared for your session. Your mentor will contact you directly.",
            parse_mode="Markdown"
        )
        
        # Update the mentor's message
        bot.edit_message_text(
            f"âœ… *Mentorship Request Accepted*\n\n"
            f"You have accepted @{call.message.text.split('@')[1].split(' ')[0]} as your mentee.\n\n"
            f"ðŸ“† First session scheduled for:\n"
            f"*{formatted_date} at {formatted_time} PHT*\n\n"
            f"Please contact your mentee directly before the session.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        
        bot.answer_callback_query(call.id, "âœ… Mentee accepted successfully!")
        
    except Exception as e:
        logging.error(f"Error handling mentee acceptance: {e}")
        bot.answer_callback_query(call.id, "âŒ Error processing acceptance", show_alert=True)

@bot.callback_query_handler(func=lambda call: call.data.startswith("decline_mentee_"))
def handle_decline_mentee(call):
    """Handle mentor declining a mentee"""
    parts = call.data.split("_")
    mentee_id = int(parts[2])
    mentor_id = parts[3]
    
    # Get mentor data for display
    mentor = MENTORS.get(mentor_id)
    mentor_name = mentor.get('name', mentor_id)
    
    try:
        # Notify mentee of declination
        bot.send_message(
            mentee_id,
            f"â„¹ï¸ *Mentorship Request Update*\n\n"
            f"Unfortunately, {mentor_name} is unable to accept your mentorship request at this time.\n\n"
            f"This could be due to their current workload or availability constraints. "
            f"Please try requesting another mentor or try again at a later time.",
            parse_mode="Markdown"
        )
        
        # Update the mentor's message
        bot.edit_message_text(
            f"âŒ *Mentorship Request Declined*\n\n"
            f"You have declined the mentorship request from @{call.message.text.split('@')[1].split(' ')[0]}.\n\n"
            f"The user has been notified of your decision.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        
        bot.answer_callback_query(call.id, "Mentorship request declined")
        
    except Exception as e:
        logging.error(f"Error handling mentee declination: {e}")
        bot.answer_callback_query(call.id, "âŒ Error processing declination", show_alert=True)

@bot.message_handler(commands=['mentorstatus'])
def update_mentor_status(message):
    """Admin command to update a mentor's status"""
    user_id = message.from_user.id
    
    # Check if user is admin
    if user_id not in ADMIN_IDS and user_id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    args = message.text.split()
    if len(args) < 3:
        bot.reply_to(message, "âŒ Usage: `/mentorstatus [mentor_id] [status]`\n\nStatus can be: available, in_session, away", parse_mode="Markdown")
        return
    
    mentor_id = args[1]
    status = args[2].lower()
    
    # Check if mentor exists
    if mentor_id not in MENTORS:
        bot.reply_to(message, f"âŒ Mentor '{mentor_id}' not found in database.")
        return
    
    # Map status code to readable status
    status_map = {
        "available": "Available",
        "in_session": "In Session", 
        "away": "Away",
        "a": "Available",
        "i": "In Session",
        "s": "In Session"
    }
    
    if status not in status_map:
        bot.reply_to(message, "âŒ Invalid status. Use: available, in_session, or away")
        return
    
    # Update mentor status
    MENTORS[mentor_id]['current_status'] = status_map[status]
    save_mentor(mentor_id, MENTORS[mentor_id])
    
    bot.reply_to(message, f"âœ… Updated {mentor_id}'s status to: {status_map[status]}")

@bot.message_handler(commands=['mentors'])
def list_all_mentors(message):
    """List all mentors and their current status"""
    user_id = message.from_user.id
    
    # Check if user is admin or mentor
    is_admin = user_id in ADMIN_IDS or user_id == CREATOR_ID
    
    # Only admins can see full details
    if not is_admin:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    # Generate mentor list
    mentor_text = "ðŸ‘¨â€ðŸ« *MENTOR STATUS LIST*\n\n"
    
    if not MENTORS:
        mentor_text += "No mentors found in database."
    else:
        for mentor_id, mentor in MENTORS.items():
            # Get status emoji
            status = mentor.get('current_status', 'Unknown')
            if status == "Available":
                status_emoji = "ðŸŸ¢"
            elif status == "In Session":
                status_emoji = "ðŸ”´"
            else:
                status_emoji = "ðŸŸ¡"
                
            mentor_text += f"{status_emoji} *{mentor['name']}*: {status}\n"
    
    bot.reply_to(message, mentor_text, parse_mode="Markdown")

# Update the show_supreme_dashboard function to support editing messages

@bot.message_handler(commands=['addmentor'])
def start_add_mentor(message):
    """Start the process of adding a new mentor to the database"""
    user_id = message.from_user.id
    
    # Check if user is admin or creator
    if user_id not in ADMIN_IDS and user_id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    # Start the mentor creation process
    PENDING_USERS[user_id] = {
        'status': 'adding_mentor_id',
        'mentor_data': {}
    }
    save_pending_users()
    
    bot.send_message(user_id, 
        "ðŸ§  *Add New Mentor*\n\n"
        "Let's add a new mentor to the system. Please provide the following information:\n\n"
        "First, enter the unique ID for this mentor (e.g., Rom, Konfu). "
        "This will be used as the internal identifier:",
        parse_mode="Markdown"
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'adding_mentor_id')
def process_mentor_id(message):
    user_id = message.from_user.id
    mentor_id = message.text.strip()
    
    # Basic validation
    if not mentor_id or len(mentor_id) < 2:
        bot.send_message(user_id, "âŒ Mentor ID must be at least 2 characters. Please try again.")
        return
        
    # Check if mentor ID already exists
    if mentor_id in MENTORS:
        bot.send_message(user_id, 
            f"âš ï¸ A mentor with ID '{mentor_id}' already exists.\n"
            f"Please use a different ID or use the `/updatementor {mentor_id}` command to update the existing mentor."
        )
        return
    
    # Store the mentor ID and move to next step
    PENDING_USERS[user_id]['mentor_data']['id'] = mentor_id
    PENDING_USERS[user_id]['status'] = 'adding_mentor_name'
    save_pending_users()
    
    bot.send_message(user_id, "Enter the mentor's display name:")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'adding_mentor_name')
def process_mentor_name(message):
    user_id = message.from_user.id
    name = message.text.strip()
    
    if not name:
        bot.send_message(user_id, "âŒ Name cannot be empty. Please try again.")
        return
    
    # Store the name and move to next step
    PENDING_USERS[user_id]['mentor_data']['name'] = name
    PENDING_USERS[user_id]['status'] = 'adding_mentor_full_name'
    save_pending_users()
    
    bot.send_message(user_id, "Enter the mentor's full name:")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'adding_mentor_full_name')
def process_mentor_full_name(message):
    user_id = message.from_user.id
    full_name = message.text.strip()
    
    if not full_name:
        bot.send_message(user_id, "âŒ Full name cannot be empty. Please try again.")
        return
    
    # Store the full name and move to next step
    PENDING_USERS[user_id]['mentor_data']['full_name'] = full_name
    PENDING_USERS[user_id]['status'] = 'adding_mentor_title'
    save_pending_users()
    
    bot.send_message(user_id, "Enter the mentor's title (e.g., 'Senior Trading Strategist'):")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'adding_mentor_title')
def process_mentor_title(message):
    user_id = message.from_user.id
    title = message.text.strip()
    
    if not title:
        bot.send_message(user_id, "âŒ Title cannot be empty. Please try again.")
        return
    
    # Store the title and move to next step
    PENDING_USERS[user_id]['mentor_data']['title'] = title
    PENDING_USERS[user_id]['status'] = 'adding_mentor_photo_url'
    save_pending_users()
    
    bot.send_message(user_id, "Enter the mentor's photo URL (or type 'skip' to leave blank):")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'adding_mentor_photo_url')
def process_mentor_photo_url(message):
    user_id = message.from_user.id
    photo_url = message.text.strip()
    
    # Allow skipping this field
    if photo_url.lower() == 'skip':
        photo_url = ""
    
    # Store the photo URL and move to next step
    PENDING_USERS[user_id]['mentor_data']['photo_url'] = photo_url
    PENDING_USERS[user_id]['status'] = 'adding_mentor_expertise'
    save_pending_users()
    
    bot.send_message(user_id, "Enter the mentor's expertise (e.g., 'Price Action & Market Structure'):")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'adding_mentor_expertise')
def process_mentor_expertise(message):
    user_id = message.from_user.id
    expertise = message.text.strip()
    
    if not expertise:
        bot.send_message(user_id, "âŒ Expertise cannot be empty. Please try again.")
        return
    
    # Store the expertise and move to next step
    PENDING_USERS[user_id]['mentor_data']['expertise'] = expertise
    PENDING_USERS[user_id]['status'] = 'adding_mentor_experience'
    save_pending_users()
    
    bot.send_message(user_id, "Enter the mentor's experience (e.g., '7+ years'):")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'adding_mentor_experience')
def process_mentor_experience(message):
    user_id = message.from_user.id
    experience = message.text.strip()
    
    if not experience:
        bot.send_message(user_id, "âŒ Experience cannot be empty. Please try again.")
        return
    
    # Store the experience and move to next step
    PENDING_USERS[user_id]['mentor_data']['experience'] = experience
    PENDING_USERS[user_id]['status'] = 'adding_mentor_style'
    save_pending_users()
    
    bot.send_message(user_id, "Enter the mentor's trading style (e.g., 'Technical Analysis'):")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'adding_mentor_style')
def process_mentor_style(message):
    user_id = message.from_user.id
    style = message.text.strip()
    
    if not style:
        bot.send_message(user_id, "âŒ Trading style cannot be empty. Please try again.")
        return
    
    # Store the trading style and move to next step
    PENDING_USERS[user_id]['mentor_data']['style'] = style
    PENDING_USERS[user_id]['status'] = 'adding_mentor_description'
    save_pending_users()
    
    bot.send_message(user_id, "Enter a detailed description of the mentor:")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'adding_mentor_description')
def process_mentor_description(message):
    user_id = message.from_user.id
    description = message.text.strip()
    
    if not description:
        bot.send_message(user_id, "âŒ Description cannot be empty. Please try again.")
        return
    
    # Store the description and move to next step
    PENDING_USERS[user_id]['mentor_data']['description'] = description
    PENDING_USERS[user_id]['status'] = 'adding_mentor_strengths'
    save_pending_users()
    
    bot.send_message(user_id, "Enter the mentor's strengths, separated by commas (e.g., 'Clean chart analysis, Entry/exit precision, Risk management'):")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'adding_mentor_strengths')
def process_mentor_strengths(message):
    user_id = message.from_user.id
    strengths_text = message.text.strip()
    
    if not strengths_text:
        bot.send_message(user_id, "âŒ Strengths cannot be empty. Please try again.")
        return
    
    # Split by commas and strip whitespace
    strengths = [s.strip() for s in strengths_text.split(',')]
    
    # Store the strengths and move to next step
    PENDING_USERS[user_id]['mentor_data']['strengths'] = strengths
    PENDING_USERS[user_id]['status'] = 'adding_mentor_teaching_style'
    save_pending_users()
    
    bot.send_message(user_id, "Enter the mentor's teaching style:")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'adding_mentor_teaching_style')
def process_mentor_teaching_style(message):
    user_id = message.from_user.id
    teaching_style = message.text.strip()
    
    if not teaching_style:
        bot.send_message(user_id, "âŒ Teaching style cannot be empty. Please try again.")
        return
    
    # Store the teaching style and move to next step
    PENDING_USERS[user_id]['mentor_data']['teaching_style'] = teaching_style
    PENDING_USERS[user_id]['status'] = 'adding_mentor_availability'
    save_pending_users()
    
    bot.send_message(user_id, "Enter the mentor's availability schedule (e.g., 'Weekdays 9AM-5PM PHT'):")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'adding_mentor_availability')
def process_mentor_availability(message):
    user_id = message.from_user.id
    availability = message.text.strip()
    
    if not availability:
        bot.send_message(user_id, "âŒ Availability cannot be empty. Please try again.")
        return
    
    # Store the availability and move to next step
    PENDING_USERS[user_id]['mentor_data']['availability'] = availability
    PENDING_USERS[user_id]['status'] = 'adding_mentor_availability_level'
    save_pending_users()
    
    # Create a keyboard for availability level
    markup = ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
    markup.add(KeyboardButton("High"), KeyboardButton("Medium"), KeyboardButton("Limited"))
    
    bot.send_message(user_id, "Select the mentor's availability level:", reply_markup=markup)

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'adding_mentor_availability_level')
def process_mentor_availability_level(message):
    user_id = message.from_user.id
    availability_level = message.text.strip()
    
    # Validate availability level
    valid_levels = ["High", "Medium", "Limited"]
    if availability_level not in valid_levels:
        bot.send_message(user_id, "âŒ Please select a valid availability level: High, Medium, or Limited.")
        return
    
    # Store the availability level and move to next step
    PENDING_USERS[user_id]['mentor_data']['availability_level'] = availability_level
    PENDING_USERS[user_id]['status'] = 'adding_mentor_current_status'
    save_pending_users()
    
    # Create a keyboard for current status
    markup = ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
    markup.add(KeyboardButton("Available"), KeyboardButton("In Session"), KeyboardButton("Away"))
    
    bot.send_message(user_id, "Select the mentor's current status:", reply_markup=markup)

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'adding_mentor_current_status')
def process_mentor_current_status(message):
    user_id = message.from_user.id
    current_status = message.text.strip()
    
    # Validate current status
    valid_statuses = ["Available", "In Session", "Away"]
    if current_status not in valid_statuses:
        bot.send_message(user_id, "âŒ Please select a valid status: Available, In Session, or Away.")
        return
    
    # Store the current status and move to next step
    PENDING_USERS[user_id]['mentor_data']['current_status'] = current_status
    PENDING_USERS[user_id]['status'] = 'adding_mentor_testimonial'
    save_pending_users()
    
    # Remove the keyboard
    markup = ReplyKeyboardRemove()
    bot.send_message(user_id, "Enter a student testimonial (or type 'skip' to leave blank):", reply_markup=markup)

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'adding_mentor_testimonial')
def process_mentor_testimonial(message):
    user_id = message.from_user.id
    testimonial = message.text.strip()
    
    # Allow skipping this field
    if testimonial.lower() == 'skip':
        testimonial = "No testimonials yet."
    
    # Store the testimonial and move to next step
    PENDING_USERS[user_id]['mentor_data']['student_testimonial'] = testimonial
    PENDING_USERS[user_id]['status'] = 'adding_mentor_success_rate'
    save_pending_users()
    
    bot.send_message(user_id, "Enter the mentor's success rate (e.g., '92%'):")

# Final handler to save the mentor to database
@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'adding_mentor_success_rate')
def process_mentor_success_rate(message):
    user_id = message.from_user.id
    success_rate = message.text.strip()
    
    # Store the success rate
    PENDING_USERS[user_id]['mentor_data']['success_rate'] = success_rate
    
    # Get all the collected mentor data
    mentor_data = PENDING_USERS[user_id]['mentor_data']
    mentor_id = mentor_data.pop('id')  # Remove ID from the data dict
    
    # Save the mentor to database
    save_mentor(mentor_id, mentor_data)
    
    # Update the MENTORS global variable
    MENTORS[mentor_id] = mentor_data
    
    # Clear the pending state
    PENDING_USERS.pop(user_id, None)
    delete_pending_user(user_id)
    
    # Send confirmation
    bot.send_message(
        user_id, 
        f"âœ… Mentor '{mentor_id}' has been successfully added to the database!",
        parse_mode="Markdown"
    )

@bot.message_handler(commands=['deletementor'])
def delete_mentor(message):
    """Delete a mentor from the database"""
    user_id = message.from_user.id
    
    # Check if user is admin or creator
    if user_id not in ADMIN_IDS and user_id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    # Get mentor ID from command
    args = message.text.split()
    if len(args) < 2:
        bot.reply_to(message, "âŒ Usage: `/deletementor [mentor_id]`", parse_mode="Markdown")
        return
        
    mentor_id = args[1]
    
    # Check if mentor exists
    if mentor_id not in MENTORS:
        bot.reply_to(message, f"âŒ Mentor '{mentor_id}' not found.")
        return
    
    # Confirm deletion
    markup = InlineKeyboardMarkup()
    markup.add(
        InlineKeyboardButton("âœ… Yes, delete", callback_data=f"confirm_delete_mentor_{mentor_id}"),
        InlineKeyboardButton("âŒ No, cancel", callback_data="cancel_delete_mentor")
    )
    
    bot.reply_to(
        message,
        f"âš ï¸ Are you sure you want to delete mentor '{mentor_id}'? This action cannot be undone.",
        reply_markup=markup
    )

def show_supreme_dashboard(message, edit_message=False, call_obj=None):
    """Display the Supreme Membership dashboard with detailed student information and task overview"""
    if edit_message and call_obj:
        # We're editing an existing message (returning from a sub-menu)
        chat_id = call_obj.message.chat.id
        message_id = call_obj.message.message_id
        user_id = call_obj.from_user.id
    else:
        # Normal command invocation
        chat_id = message.chat.id
        user_id = message.from_user.id
    
    user_id_str = str(user_id)
    
    # Only allow in private chats for privacy (only check for normal command invocation)
    if not edit_message and message.chat.type != 'private':
        bot.send_message(chat_id, "ðŸ”’ Please use this command in a private message with the bot.")
        return
    
    # Check if the user has Supreme membership
    if user_id_str in PAYMENT_DATA and PAYMENT_DATA[user_id_str].get('mentorship_type', '').lower() == 'supreme':
        try:
            data = PAYMENT_DATA[user_id_str]
            
            if edit_message:
                username = call_obj.from_user.username or "No Username"
            else:
                username = message.from_user.username or "No Username"
                
            full_name = data.get('form_answers', {}).get('full_name', username)
            
            # Get membership details
            plan = data.get('payment_plan', 'Unknown')
            
            # Determine the mentor (you can customize this logic)
            mentors = {'Apprentice': 'Rom', 'Disciple': 'Konfu', 'Legacy': 'Nath'}
            mentor = mentors.get(plan, '???')
            
            # Calculate enrollment date and expiry date
            enrollment_date = datetime.strptime(data.get('enrollment_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')), '%Y-%m-%d %H:%M:%S')
            due_date = datetime.strptime(data.get('due_date', '2099-12-31 23:59:59'), '%Y-%m-%d %H:%M:%S')
            
            # Calculate days remaining until expiration
            current_date = datetime.now()
            days_remaining = (due_date - current_date).days
            hours_remaining = int((due_date - current_date).seconds / 3600)
            
            # Get mentorship type
            mentorship_type = data.get('mentorship_type', 'Supreme')
            
            # Check if membership is cancelled
            if data.get('cancelled', False):
                status_icon = "ðŸš«"
                status_text = "Cancelled"
            # If not cancelled, format status based on days remaining
            elif days_remaining > 7:
                status_icon = "âœ…"
                status_text = "Active"
            elif days_remaining > 0:
                status_icon = "âš ï¸"
                status_text = "Expiring Soon"
            else:
                status_icon = "âŒ"
                status_text = "Expired"
            
            # Determine user ranking based on custom logic (placeholder)
            rank_points = data.get('rank_points', 0)
            
            ranking_tiers = {
                (0, 100): "Rookie I",
                (101, 200): "Rookie II",
                (201, 300): "Rookie III",
                (301, 500): "Elite I",
                (501, 700): "Elite II",
                (701, 900): "Elite III",
                (901, 1200): "Master I",
                (1201, 1500): "Master II",
                (1501, 1800): "Master III",
                (1801, 2200): "Legend I",
                (2201, 2600): "Legend II",
                (2601, 3000): "Legend III",
                (3001, float('inf')): "Sovereign"
            }
            
            ranking = "Rookie I"  # Default
            for (min_points, max_points), rank in ranking_tiers.items():
                if min_points <= rank_points <= max_points:
                    ranking = rank
                    break
            
            # Get user expertise level from form answers
            expertise_level = "Beginner"
            
            if 'form_answers' in data and 'expertise_level' in data['form_answers']:
                expertise = data['form_answers']['expertise_level'].lower()
                
                if 'beginner' in expertise or 'a.' in expertise:
                    expertise_level = "Beginner"
                elif 'intermediate' in expertise or 'c.' in expertise:
                    expertise_level = "Intermediate"  
                elif 'expert' in expertise or 'd.' in expertise:
                    expertise_level = "Expert"
                elif 'master' in expertise or 'e.' in expertise:
                    expertise_level = "Master"
            
            # Create nice formatting for the dashboard with box-style formatting
            dashboard_message = (
                f"âœ¨ *SUPREME DASHBOARD*\n\n"
                
                f"â”Œâ”€â”€â”€â”€â”€â”€â”€ ðŸ‘¤ *STUDENT INFO* â”€â”€â”€â”€â”€â”€â”€â”€â”\n"
                f"â”‚ *Name:* {full_name}\n"
                f"â”‚ *Member ID:* `{user_id_str}`\n"
                f"â”‚ *Ranking:* {ranking}\n"
                f"â”‚ *Expertise Level:* {expertise_level}\n"
                f"â””â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”˜\n\n"
                
                f"â”Œâ”€â”€â”€â”€â”€â”€â”€ ðŸ’³ *SUBSCRIPTION* â”€â”€â”€â”€â”€â”€â”€â”€â”\n"
                f"â”‚ *Plan:* {plan}\n"
                f"â”‚ *Type:* {mentorship_type} Mentorship\n"
                f"â”‚ *Mentor:* {mentor}\n"
                f"â”‚ *Enrollment:* {enrollment_date.strftime('%B %d, %Y')}\n"
                f"â””â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”˜\n\n"
                
                f"â”Œâ”€â”€â”€â”€â”€â”€â”€ â±ï¸ *STATUS* â”€â”€â”€â”€â”€â”€â”€â”€â”\n"
                f"â”‚ *Status:* {status_icon} {status_text}\n"
                f"â”‚ *Expiration:* {due_date.strftime('%Y-%m-%d')}\n"
                f"â”‚ *Time Left:* {days_remaining} days, {hours_remaining} hours\n"
                f"â””â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”˜\n\n"
                
                f"â”Œâ”€â”€â”€â”€â”€â”€â”€ ðŸ“ *LEARNING TASKS* â”€â”€â”€â”€â”€â”€â”€â”€â”\n"
                f"â”‚ â€¢ Complete Market Structure Lesson\n"
                f"â”‚   *Due: Friday*\n"
                f"â”‚ â€¢ Submit Weekly Chart Analysis\n"
                f"â”‚   *Due: Sunday*\n"
                f"â”‚ â€¢ Schedule Next 1:1 Session\n"
                f"â”‚   *Due: Tomorrow*\n"
                f"â””â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”˜\n"
            )
            
            # Add renewal instructions if expiring soon (and not cancelled)
            if days_remaining < 7 and days_remaining >= 0 and not data.get('cancelled', False):
                dashboard_message += (
                    f"\nâš ï¸ *Your membership expires soon!*\n"
                    f"Use /start and select 'Renew Membership' to continue access.\n"
                )
            # Add special message for cancelled memberships
            elif data.get('cancelled', False):
                dashboard_message += (
                    f"\nðŸš« *Your membership has been cancelled*\n"
                    f"You will still have access until {due_date.strftime('%Y-%m-%d')}.\n"
                    f"To reactivate before expiration, use /start and select 'Renew Membership'.\n"
                )
            
            # Create navigation buttons for the Supreme dashboard - UPDATED PER REQUIREMENTS
            markup = InlineKeyboardMarkup(row_width=2)
            
            # First row
            markup.add(
                InlineKeyboardButton("ðŸ’Ž Supreme Vault", callback_data="supreme_vault"),
                InlineKeyboardButton("ðŸ‘¨â€ðŸ« Mentor Line", callback_data="mentor_line")
            )
            
            # Second row
            markup.add(
                InlineKeyboardButton("ðŸ›£ï¸ Trading Journey", callback_data="trading_journey"),
                InlineKeyboardButton("ðŸ“” Trading Journal", callback_data="trading_journal")
            )
            
            # Third row
            markup.add(
                InlineKeyboardButton("ðŸ† Leaderboard", callback_data="supreme_leaderboard"),
                InlineKeyboardButton("ðŸ‘¥ Supreme Network", callback_data="supreme_network")
            )
            
            # Add renewal button if needed
            if days_remaining < 7 and days_remaining >= 0 and not data.get('cancelled', False):
                markup.add(InlineKeyboardButton("ðŸ”„ Renew Membership", callback_data="start_renewal"))
            
            # Send or edit the dashboard message with buttons
            if edit_message:
                bot.edit_message_text(
                    dashboard_message, 
                    chat_id,
                    message_id,
                    parse_mode="Markdown", 
                    reply_markup=markup
                )
            else:
                bot.send_message(chat_id, dashboard_message, parse_mode="Markdown", reply_markup=markup)
            
        except Exception as e:
            error_msg = f"âŒ Error retrieving Supreme dashboard: {str(e)}"
            if edit_message and call_obj:
                bot.edit_message_text(error_msg, chat_id, message_id)
            else:
                bot.send_message(chat_id, error_msg)
            logging.error(f"Supreme dashboard error for user {user_id}: {str(e)}")
    else:
        # User is not a Supreme member
        error_msg = (
            "â­ *Supreme Dashboard Access Restricted*\n\n"
            "This feature is exclusively for Supreme membership holders. "
            "To upgrade your membership, use /start and select 'Purchase Membership'."
        )
        
        if edit_message and call_obj:
            bot.edit_message_text(error_msg, chat_id, message_id, parse_mode="Markdown")
        else:
            bot.send_message(chat_id, error_msg, parse_mode="Markdown")

# Add a callback handler for dashboard buttons
@bot.callback_query_handler(func=lambda call: call.data == "start_renewal")
def handle_renewal_button(call):
    """Handle renewal button from dashboard"""
    try:
        # Simulate the /start command followed by renewal choice
        PENDING_USERS[call.from_user.id] = {
            'status': 'choosing_mentorship_type',
            'is_renewal': True
        }
        save_pending_users()
        
        markup = ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
        markup.add(KeyboardButton("Regular Mentorship"), KeyboardButton("Supreme Mentorship"))
        
        bot.send_message(
            call.from_user.id, 
            "Thank you for choosing to renew! Please select your preferred mentorship level:",
            reply_markup=markup
        )
        
        # Answer the callback to close the loading indicator
        bot.answer_callback_query(call.id, "Starting renewal process...")
        
        # Edit the message to remove buttons
        bot.edit_message_reply_markup(
            call.message.chat.id, 
            call.message.message_id,
            reply_markup=None
        )
    except Exception as e:
        bot.answer_callback_query(call.id, f"Error: {str(e)}")
        logging.error(f"Error in renewal button handler: {e}")

@bot.callback_query_handler(func=lambda call: call.data == "view_leaderboard")
def handle_leaderboard_button(call):
    """Show leaderboard from dashboard button"""
    try:
        # Generate daily leaderboard for current date
        now = datetime.now(pytz.timezone('Asia/Manila'))
        leaderboard_text = generate_daily_leaderboard_text(now)
        
        # Send leaderboard
        bot.send_message(
            call.from_user.id,
            leaderboard_text,
            parse_mode="Markdown"
        )
        
        bot.answer_callback_query(call.id, "Showing current leaderboard")
    except Exception as e:
        bot.answer_callback_query(call.id, f"Error: {str(e)}")
        logging.error(f"Error in leaderboard button handler: {e}")

# Command to post a new changelog entry (creator only)
@bot.message_handler(commands=['post_changelog'])
def post_changelog_command(message):
    if message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to the bot creator.")
        return
        
    # Ask for changelog type
    markup = ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
    markup.add(KeyboardButton("Admin Changelog"), KeyboardButton("User Changelog"))
    bot.send_message(message.chat.id, "Select the changelog type:", reply_markup=markup)
    PENDING_USERS[message.chat.id] = {'status': 'selecting_changelog_type'}
    save_pending_users()

# Handle changelog type selection
@bot.message_handler(func=lambda message: PENDING_USERS.get(message.chat.id, {}).get('status') == 'selecting_changelog_type')
def select_changelog_type(message):
    if message.from_user.id != CREATOR_ID:
        return
        
    changelog_type = message.text.lower()
    chat_id = message.chat.id
    
    if "admin" in changelog_type:
        PENDING_USERS[chat_id]['changelog_type'] = 'admin'
        save_pending_users()
        bot.send_message(chat_id, "ðŸ“ Please enter the admin changelog entry with the following format:\n\n*Version*\nChangelog details")
    elif "user" in changelog_type:
        PENDING_USERS[chat_id]['changelog_type'] = 'user'
        save_pending_users()
        bot.send_message(chat_id, "ðŸ“ Please enter the user changelog entry with the following format:\n\n*Version*\nChangelog details")
    else:
        bot.send_message(chat_id, "âŒ Invalid option. Please select either 'Admin Changelog' or 'User Changelog'.")
        return
    
    PENDING_USERS[chat_id]['status'] = 'entering_changelog'
    save_pending_users()

def escape_markdown_v2(text):
    """
    Escape special characters for Markdown V2 format in Telegram
    This handles all special characters that need escaping
    """
    special_chars = r'_*[]()~`>#+-=|{}.!'
    for char in special_chars:
        text = text.replace(char, f"\\{char}")
    return text

# Handle changelog entry
@bot.message_handler(func=lambda message: PENDING_USERS.get(message.chat.id, {}).get('status') == 'entering_changelog')
def enter_changelog(message):
    if message.from_user.id != CREATOR_ID:
        return
        
    chat_id = message.chat.id
    changelog_text = message.text
    changelog_type = PENDING_USERS[chat_id]['changelog_type']
    save_pending_users()
    
    # Add timestamp to the changelog
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Store original text in the changelog
    changelog_entry = {
        "timestamp": timestamp,
        "content": changelog_text
    }
    
    CHANGELOGS[changelog_type].append(changelog_entry)
    save_changelogs(CHANGELOGS)
    
    # Confirmation message
    bot.send_message(chat_id, f"âœ… {changelog_type.capitalize()} changelog added successfully!")
    
    # For user changelogs, broadcast to all registered users
    if changelog_type == 'admin':
        # Send to admins only - with properly escaped markdown
        for admin_id in ADMIN_IDS:
            try:
                bot.send_chat_action(admin_id, 'typing')  # Check if user can be messaged
                # Try with parse_mode=None first if markdown fails
                try:
                    bot.send_message(
                        admin_id,
                        f"ðŸ“¢ *NEW ADMIN CHANGELOG*\n\n{changelog_text}\n\nðŸ•’ Posted: {timestamp}",
                        parse_mode="Markdown"
                    )
                except Exception:
                    # If Markdown parsing fails, send without formatting
                    bot.send_message(
                        admin_id,
                        f"ðŸ“¢ NEW ADMIN CHANGELOG\n\n{changelog_text}\n\nðŸ•’ Posted: {timestamp}",
                        parse_mode=None
                    )
                    logging.info(f"Sent admin changelog to {admin_id} without markdown formatting")
            except Exception as e:
                logging.error(f"Failed to send admin changelog to {admin_id}: {e}")
                bot.send_message(chat_id, f"âš ï¸ Warning: Could not send changelog to admin {admin_id}")
    else:
        # For user changelogs - add to pending notifications and broadcast
        # Track successful and failed deliveries
        success_count = 0
        failed_count = 0
        
        # Add a "last_changelog" field to track users who haven't seen latest changelog
        changelog_entry["seen_by"] = []
        
        for user_id_str in PAYMENT_DATA:
            if not PAYMENT_DATA[user_id_str]['haspayed']:
                continue
                
            try:
                user_id = int(user_id_str)
                bot.send_chat_action(user_id, 'typing')  # Check if user can be messaged
                
                # Try with parse_mode=Markdown first, if it fails, try without formatting
                try:
                    bot.send_message(
                        user_id,
                        f"ðŸ“¢ *NEW UPDATE*\n\n{changelog_text}\n\nðŸ•’ Posted: {timestamp}",
                        parse_mode="Markdown"
                    )
                except Exception:
                    bot.send_message(
                        user_id,
                        f"ðŸ“¢ NEW UPDATE\n\n{changelog_text}\n\nðŸ•’ Posted: {timestamp}",
                        parse_mode=None
                    )
                    
                changelog_entry["seen_by"].append(user_id_str)
                success_count += 1
            except Exception as e:
                logging.error(f"Failed to send user changelog to {user_id_str}: {e}")
                failed_count += 1
        
        # Show delivery stats
        bot.send_message(
            chat_id, 
            f"ðŸ“Š Changelog Delivery Stats:\nâœ… Successfully sent: {success_count}\nâŒ Failed: {failed_count}"
        )
        
        # Option to also post in group chat for maximum visibility
        markup = InlineKeyboardMarkup()
        markup.add(InlineKeyboardButton("Yes", callback_data=f"post_group_changelog_{len(CHANGELOGS[changelog_type])-1}"))
        markup.add(InlineKeyboardButton("No", callback_data="cancel_group_post"))
        
        bot.send_message(
            chat_id,
            "Would you like to also post this changelog in the main group chat?",
            reply_markup=markup
        )
    # Remove user from pending users after successfully processing the changelog
    PENDING_USERS.pop(chat_id, None)  # Remove from dictionary
    delete_pending_user(chat_id)  # Remove from MongoDB

# View changelogs command
@bot.message_handler(commands=['changelogs'])
def view_changelogs(message):
    chat_id = message.chat.id
    user_id = message.from_user.id
    
    # Load the appropriate changelogs based on user role
    is_admin = user_id in ADMIN_IDS
    is_creator = user_id == CREATOR_ID
    
    if is_admin or is_creator:
        # Admins and creator can see both changelogs
        markup = InlineKeyboardMarkup()
        markup.add(InlineKeyboardButton("Admin Changelogs", callback_data='view_admin_changelogs'))
        markup.add(InlineKeyboardButton("User Changelogs", callback_data='view_user_changelogs'))
        bot.send_message(chat_id, "Select which changelogs to view:", reply_markup=markup)
    else:
        # Regular users can only see user changelogs
        send_user_changelogs(chat_id)

@bot.callback_query_handler(func=lambda call: call.data.startswith("post_group_changelog_"))
def post_changelog_to_group(call):
    if call.from_user.id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ Only the creator can post changelogs to the group.")
        return
        
    changelog_index = int(call.data.split("_")[3])
    changelog = CHANGELOGS["user"][changelog_index]
    
    try:
        # Send to the announcement topic if configured, otherwise to main group
        if ANNOUNCEMENT_TOPIC_ID:
            bot.send_message(
                PAID_GROUP_ID,
                f"ðŸ“¢ *IMPORTANT UPDATE*\n\n{changelog['content']}\n\nðŸ•’ Posted: {changelog['timestamp']}",
                parse_mode="Markdown",
                message_thread_id=ANNOUNCEMENT_TOPIC_ID
            )
            bot.answer_callback_query(call.id, f"âœ… Posted to announcements topic (ID: {ANNOUNCEMENT_TOPIC_ID}) successfully!")
            bot.edit_message_text(
                f"Changelog posted to announcements topic (ID: {ANNOUNCEMENT_TOPIC_ID}) successfully!",
                call.message.chat.id,
                call.message.message_id
            )
            logging.info(f"Posted changelog to announcement topic {ANNOUNCEMENT_TOPIC_ID}")
        else:
            # Original behavior - post to main group
            bot.send_message(
                PAID_GROUP_ID,
                f"ðŸ“¢ *IMPORTANT UPDATE*\n\n{changelog['content']}\n\nðŸ•’ Posted: {changelog['timestamp']}",
                parse_mode="Markdown"
            )
            bot.answer_callback_query(call.id, "âœ… Posted to group successfully!")
            bot.edit_message_text(
                "Changelog posted to main group successfully!",
                call.message.chat.id,
                call.message.message_id
            )
            logging.info("Posted changelog to main group (no topic ID set)")
    except Exception as e:
        bot.answer_callback_query(call.id, "âŒ Failed to post to group.")
        bot.send_message(call.message.chat.id, f"Error: {str(e)}")
        logging.error(f"Error posting changelog: {e}")

@bot.callback_query_handler(func=lambda call: call.data == "cancel_group_post")
def cancel_group_post(call):
    bot.answer_callback_query(call.id, "âŒ Cancelled posting to group.")
    bot.edit_message_text(
        "Group posting cancelled.",
        call.message.chat.id,
        call.message.message_id
    )

# Callback handler for changelog viewing
@bot.callback_query_handler(func=lambda call: call.data.startswith('view_'))
def handle_changelog_view(call):
    chat_id = call.message.chat.id
    
    if call.data == 'view_admin_changelogs':
        if call.from_user.id in ADMIN_IDS or call.from_user.id == CREATOR_ID:
            send_admin_changelogs(chat_id)
        else:
            bot.answer_callback_query(call.id, "âŒ You don't have permission to view admin changelogs.")
    elif call.data == 'view_user_changelogs':
        send_user_changelogs(chat_id)
    
    bot.answer_callback_query(call.id)

def send_admin_changelogs(chat_id):
    if not CHANGELOGS['admin']:
        bot.send_message(chat_id, "No admin changelogs available yet.")
        return
    
    # Get the latest 5 changelogs
    recent_logs = CHANGELOGS['admin'][-5:]
    
    # First try plain text (no markdown) to be safe
    plain_message = "ðŸ“‹ ADMIN CHANGELOGS\n\n"
    for log in recent_logs:
        plain_message += f"ðŸ•’ {log['timestamp']}\n{log['content']}\n\n"
    
    bot.send_message(chat_id, plain_message)

def send_user_changelogs(chat_id):
    if not CHANGELOGS['user']:
        bot.send_message(chat_id, "No changelogs available yet.")
        return
    
    # Get the latest 5 changelogs
    recent_logs = CHANGELOGS['user'][-5:]
    
    # Send as plain text to avoid formatting issues
    plain_message = "ðŸ“‹ RECENT UPDATES\n\n"
    for log in recent_logs:
        plain_message += f"ðŸ•’ {log['timestamp']}\n{log['content']}\n\n"
    
    bot.send_message(chat_id, plain_message)

@bot.message_handler(commands=['setannouncementtopic'])
def set_announcement_topic(message):
    """Set or change the topic ID for announcements"""
    global ANNOUNCEMENT_TOPIC_ID
    
    # Only allow the creator to use this command
    if message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to the bot creator.")
        return
    
    # Extract topic ID from command arguments
    args = message.text.split()
    
    # Show current setting if no arguments provided
    if len(args) == 1:
        current_topic = ANNOUNCEMENT_TOPIC_ID if ANNOUNCEMENT_TOPIC_ID else "Not set (using main group)"
        bot.reply_to(message, f"Current announcement topic ID: `{current_topic}`\n\nTo change, use: `/setannouncementtopic ID`", parse_mode="Markdown")
        return
        
    try:
        # Handle "clear" or "reset" to remove topic ID
        if args[1].lower() in ["clear", "reset", "none"]:
            ANNOUNCEMENT_TOPIC_ID = None
            # Save to database
            BOT_SETTINGS['announcement_topic_id'] = None
            save_settings(BOT_SETTINGS)
            bot.reply_to(message, "âœ… Announcements will now be sent to the main group chat.")
            return
            
        # Try to parse as integer
        new_topic_id = int(args[1])
        ANNOUNCEMENT_TOPIC_ID = new_topic_id
        
        # Save to database
        BOT_SETTINGS['announcement_topic_id'] = new_topic_id
        save_settings(BOT_SETTINGS)
        
        bot.reply_to(message, f"âœ… Announcements will now be sent to topic ID: `{new_topic_id}`\nThis setting has been saved to the database.", parse_mode="Markdown")
        
    except ValueError:
        bot.reply_to(message, "âŒ Invalid topic ID. Please provide a numeric ID or 'clear' to reset.")
    except Exception as e:
        bot.reply_to(message, f"âŒ Error setting topic ID: {str(e)}")

# Modify the post_changelog_to_group function to use the announcement topic ID
@bot.callback_query_handler(func=lambda call: call.data.startswith("post_group_changelog_"))
def post_changelog_to_group(call):
    if call.from_user.id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ Only the creator can post changelogs to the group.")
        return
        
    changelog_index = int(call.data.split("_")[3])
    changelog = CHANGELOGS["user"][changelog_index]
    
    try:
        # Send to the announcement topic if configured, otherwise to main group
        if ANNOUNCEMENT_TOPIC_ID:
            bot.send_message(
                PAID_GROUP_ID,
                f"ðŸ“¢ *IMPORTANT UPDATE*\n\n{changelog['content']}\n\nðŸ•’ Posted: {changelog['timestamp']}",
                parse_mode="Markdown",
                message_thread_id=ANNOUNCEMENT_TOPIC_ID
            )
            bot.answer_callback_query(call.id, "âœ… Posted to announcements topic successfully!")
            bot.edit_message_text(
                f"Changelog posted to announcements topic (ID: {ANNOUNCEMENT_TOPIC_ID}) successfully!",
                call.message.chat.id,
                call.message.message_id
            )
        else:
            # Original behavior - post to main group
            bot.send_message(
                PAID_GROUP_ID,
                f"ðŸ“¢ *IMPORTANT UPDATE*\n\n{changelog['content']}\n\nðŸ•’ Posted: {changelog['timestamp']}",
                parse_mode="Markdown"
            )
            bot.answer_callback_query(call.id, "âœ… Posted to group successfully!")
            bot.edit_message_text(
                "Changelog posted to group successfully!",
                call.message.chat.id,
                call.message.message_id
            )
    except Exception as e:
        bot.answer_callback_query(call.id, "âŒ Failed to post to group.")
        bot.send_message(call.message.chat.id, f"Error: {str(e)}")


@bot.message_handler(commands=['check'])
def check_mongodb_connection(message):
    # Restrict access to Creator only
    if message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to the bot creator.")
        return
    
    try:
        # Test connection with ping
        client.admin.command('ping')
        
        # Get collection stats
        payment_count = payment_collection.count_documents({})
        members_count = old_members_collection.count_documents({})
        pending_count = pending_collection.count_documents({})
        
        status_message = (
            f"âœ… **MongoDB Connection Status**\n\n"
            f"ðŸ”Œ Connected to: `{MONGO_URI}`\n"
            f"ðŸ“‚ Database: `{DB_NAME}`\n\n"
            f"ðŸ“Š **Collection Stats**\n"
            f"- Payments: {payment_count} records\n"
            f"- Old members: {members_count} records\n"
            f"- Pending users: {pending_count} records\n"
            f"- Changelogs: {len(CHANGELOGS.get('user', []))} user, {len(CHANGELOGS.get('admin', []))} admin\n\n"
            f"ðŸ•’ Checked at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
        
        bot.reply_to(message, status_message, parse_mode="Markdown")
        logging.info(f"MongoDB connection check: SUCCESS (requested by Creator ID: {message.from_user.id})")
        
    except Exception as e:
        error_message = f"âŒ **MongoDB Connection Error**\n\n{str(e)}"
        bot.reply_to(message, error_message, parse_mode="Markdown")
        logging.error(f"MongoDB connection check: FAILED - {e}")

@bot.message_handler(commands=['remove'])
def remove_self_from_pending(message):
    user_id = message.from_user.id
    
    # Check if admin or creator
    if user_id not in ADMIN_IDS and user_id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators and the creator.")
        return
    
    # Remove self from pending users
    if user_id in PENDING_USERS:
        status = PENDING_USERS[user_id].get('status', 'unknown')
        PENDING_USERS.pop(user_id, None)
        
        # Remove from MongoDB too
        delete_pending_user(user_id)
        
        bot.reply_to(message, f"âœ… You've been removed from pending users. Previous status: {status}")
        logging.info(f"Admin {user_id} removed self from pending users (status: {status})")
    else:
        bot.reply_to(message, "âœ… You're not in the pending users list.")

# Function to handle payment proof and old member verification requests
def send_pending_request_reminders():
    while True:
        try:
            current_time = datetime.now()
            
            for user_id, data in PENDING_USERS.items():
                # Check for payment verification requests
                if data.get('status') == 'waiting_approval':
                    # Check if submission timestamp exists
                    if 'request_time' not in data:
                        # Add timestamp now for existing requests
                        PENDING_USERS[user_id]['request_time'] = current_time
                        save_pending_users()
                        continue
                        
                    # Calculate time elapsed since request
                    request_time = data['request_time']
                    if isinstance(request_time, str):
                        request_time = datetime.strptime(request_time, '%Y-%m-%d %H:%M:%S')
                    
                    time_elapsed = (current_time - request_time).total_seconds() / 60  # in minutes
                    
                    # Check if it's been more than 10 minutes and reminder not sent yet
                    if time_elapsed > 15 and not data.get('reminder_sent', False):
                        # Send reminder to user
                        try:
                            bot.send_message(
                                user_id,
                                "â³ Your payment verification request is still pending. The admins might be busy at the moment. "
                                "Please be patient as they review your submission."
                            )
                            logging.info(f"Sent waiting reminder to user {user_id} for payment verification")
                        except Exception as e:
                            logging.error(f"Failed to send wait reminder to user {user_id}: {e}")
                        
                        # Send reminder to all admins
                        for admin_id in ADMIN_IDS:
                            try:
                                user_info = bot.get_chat(user_id)
                                username = user_info.username or f"User {user_id}"
                                escaped_username = safe_markdown_escape(username)  # Properly escape the username
                                bot.send_message(
                                    admin_id,
                                    f"âš ï¸ *Reminder:* @{escaped_username} has been waiting for payment verification for over 10 minutes.",
                                    parse_mode="Markdown"
                                )
                            except Exception as e:
                                logging.error(f"Failed to send admin reminder to {admin_id} about user {user_id}: {e}")
                        
                        # Mark reminder as sent
                        PENDING_USERS[user_id]['reminder_sent'] = True
                        save_pending_users()
                
                # Check for old member verification requests
                if data.get('status') == 'old_member_request':
                    # Check if submission timestamp exists
                    if 'request_time' not in data:
                        # Add timestamp now for existing requests
                        PENDING_USERS[user_id]['request_time'] = current_time
                        save_pending_users()
                        continue
                    
                    # Calculate time elapsed since request
                    request_time = data['request_time']
                    if isinstance(request_time, str):
                        request_time = datetime.strptime(request_time, '%Y-%m-%d %H:%M:%S')
                    
                    time_elapsed = (current_time - request_time).total_seconds() / 60  # in minutes
                    
                    # Check if it's been more than 10 minutes and reminder not sent yet
                    if time_elapsed > 10 and not data.get('reminder_sent', False):
                        # Send reminder to user
                        try:
                            bot.send_message(
                                user_id,
                                "â³ Your old member verification request is still pending. The admins might be busy at the moment. "
                                "Please be patient as they review your submission."
                            )
                            logging.info(f"Sent waiting reminder to user {user_id} for old member verification")
                        except Exception as e:
                            logging.error(f"Failed to send wait reminder to user {user_id}: {e}")
                        
                        # Send reminder to all admins
                        for admin_id in ADMIN_IDS:
                            try:
                                user_info = bot.get_chat(user_id)
                                username = user_info.username or f"User {user_id}"
                                escaped_username = safe_markdown_escape(username)  # Properly escape the username
                                bot.send_message(
                                    admin_id,
                                    f"âš ï¸ *Reminder:* @{escaped_username} has been waiting for old member verification for over 10 minutes.",
                                    parse_mode="Markdown"
                                )
                            except Exception as e:
                                logging.error(f"Failed to send admin reminder to {admin_id} about user {user_id}: {e}")
                        
                        # Mark reminder as sent
                        PENDING_USERS[user_id]['reminder_sent'] = True
                        save_pending_users()
            
            # Sleep for 1 minute before next check
            time.sleep(60)
            
        except Exception as e:
            logging.error(f"Error in pending request reminder thread: {e}")
            time.sleep(60)  # Wait a minute on error before trying again

def refresh_mongodb_data():
    """Refresh all data from MongoDB to ensure it's up to date."""
    global PAYMENT_DATA, CONFIRMED_OLD_MEMBERS, PENDING_USERS, CHANGELOGS
    
    try:
        PAYMENT_DATA = load_payment_data()
        
        CONFIRMED_OLD_MEMBERS = load_confirmed_old_members()
        
        PENDING_USERS = load_pending_users()
        
        CHANGELOGS = load_changelogs()
        logging.info("MongoDB data refresh completed successfully")
    except Exception as e:
        logging.error(f"Error refreshing MongoDB data: {e}")

def mongodb_refresh_thread():
    """Background thread to periodically refresh MongoDB data."""
    while True:
        try:
            # Sleep first to avoid immediate refresh after startup
            time.sleep(1800)  # 30 minutes = 1800 seconds
            
            # Refresh all data from MongoDB
            refresh_mongodb_data()
            
        except Exception as e:
            logging.error(f"Error in MongoDB refresh thread: {e}")
            time.sleep(300)  # Wait 5 minutes on error before trying again

# Define challenge content
SELF_IMPROVEMENT_CHALLENGES = [
    {"type": "ACTION", "content": "Meditate for 10 minutes today"},
    {"type": "ACTION", "content": "Practice mindful breathing (5 minute deep breathing)"},
    {"type": "QUESTION", "content": "What is one habit you want to build, and why?"},
    {"type": "QUESTION", "content": "Give 5 things you are grateful for today (no repeats)"},
    {"type": "QUESTION", "content": "What did you learn about yourself this week?"},
    {"type": "QUESTION", "content": "What is one small win you can celebrate today?"}
]

TRADING_CHALLENGES = [
    {"type": "ACTION", "content": "Watch the movements of a pair for at least 5 minutes"},
    {"type": "ACTION", "content": "Revisit your trading rules for 3 minutes"},
    {"type": "QUESTION", "content": "Review your last trade (profits, pips, tell us how it went)"},
    {"type": "QUESTION", "content": "Journal one key takeaway from today or yesterday's session"},
    {"type": "QUESTION", "content": "Write about how you felt after your latest win / loss"},
    {"type": "QUESTION", "content": "Write about how you felt during your latest trade"},
    {"type": "QUESTION", "content": "Review your latest loss and what you can learn"}
]


def generate_daily_challenge():
    """Generate a truly random daily challenge from the predefined lists."""
    
    # Get today's date with enhanced formatting including day of week
    now = datetime.now(pytz.timezone('Asia/Manila'))
    today = now.strftime("%A, %B %d, %Y").upper()  # Adds day of week (e.g., "MONDAY, MARCH 18, 2025")
    
    # For true randomization, reseed the random generator each time
    random.seed(secrets.randbits(128))
    
    # Select a random self-improvement challenge
    self_improvement = random.choice(SELF_IMPROVEMENT_CHALLENGES)
    
    # Select a random trading challenge
    trading = random.choice(TRADING_CHALLENGES)
    
    # Format the message with enhanced styling
    message = f"ðŸ“† {today}\n"
    message += f"ð——ð—”ð—œð—Ÿð—¬ ð—–ð—›ð—”ð—Ÿð—Ÿð—˜ð—¡ð—šð—˜ð—¦\n\n"
    
    message += f"ðŸ’ª *SELF-IMPROVEMENT* (10 POINTS)\n"
    message += f"â–«ï¸ *{self_improvement['type']}:* {self_improvement['content']}\n\n"
    
    message += f"ðŸ“ˆ *TRADING-RELATED* (10 POINTS)\n"
    message += f"â–«ï¸ *{trading['type']}:* {trading['content']}\n\n"
    
    message += f"ðŸ’» *DAILY CHARTING* (20 POINTS)\n"
    message += f"Share a trade you took or analyzed with the following details:\n"
    message += f"â€¢ Strategy used\n"
    message += f"â€¢ Confluences that supported your analysis\n"
    message += f"â€¢ Pair/instrument traded\n"
    message += f"â€¢ Win-rate for this strategy (estimated)\n"
    message += f"â€¢ Risk-to-reward ratio (R:R)\n\n"
    
    message += f"â­ Complete all challenges for 40 TOTAL POINTS!\n"
    message += f"ðŸ“¸ Share your progress in the chat to earn points!"
    
    return message

last_reminder_date = None

def send_daily_challenges():
    """Send daily challenges to the group at a specified time."""
    global last_reminder_date
    logging.info("Daily challenges thread started")
    
    # Track the last day we sent a challenge to avoid duplicates
    last_challenge_date = None
    
    while True:
        try:
            now = datetime.now(pytz.timezone('Asia/Manila'))
            current_time = now.strftime('%H:%M')
            current_date = now.strftime('%Y-%m-%d')
            
            # Send challenge at 8:00 AM Philippine time every day
            # Only if we haven't already sent one today
            if current_time == '08:00' and last_challenge_date != current_date:
                # Only proceed if it's a weekday (Monday=0, Sunday=6)
                is_weekday = now.weekday() < 5  # 0-4 are Monday to Friday
                
                if is_weekday:
                    challenge_message = generate_daily_challenge()
                    
                    # Send to specific topic if configured, otherwise to main group
                    if DAILY_CHALLENGE_TOPIC_ID:
                        bot.send_message(
                            PAID_GROUP_ID, 
                            challenge_message, 
                            message_thread_id=DAILY_CHALLENGE_TOPIC_ID,
                            parse_mode="Markdown"
                        )
                        logging.info(f"Sent daily challenge to topic {DAILY_CHALLENGE_TOPIC_ID} at {current_time} Philippine time.")
                    else:
                        bot.send_message(
                            PAID_GROUP_ID, 
                            challenge_message,
                            parse_mode="Markdown"
                        )
                        logging.info(f"Sent daily challenge to main group at {current_time} Philippine time.")
                    
                    # Update the last challenge date
                    last_challenge_date = current_date
                    # Reset reminder flag for the new day
                    last_reminder_date = None
                else:
                    logging.info(f"Skipped daily challenge: Weekend.")
                    
            # Send reminder at 8:30 AM if we sent a challenge today but haven't sent a reminder
            if current_time == '08:30' and last_challenge_date == current_date and last_reminder_date != current_date:
                is_weekday = now.weekday() < 5
                
                if is_weekday:
                    # Create a friendly, conversational reminder
                    reminder_messages = [
                        "Hey everyone! ðŸ‘‹ Just a friendly reminder to complete today's challenge. Share your work in the accountability roster to earn points for the leaderboard! ðŸ“Š",
                        
                        "Good morning traders! â˜• Don't forget to tackle today's challenge - it only takes a few minutes and helps build consistent trading habits. Post your response in the accountability roster!",
                        
                        "Rise and shine, traders! âœ¨ Have you done today's challenge yet? Remember to post in the accountability roster to get your points for the day!",
                        
                        "Time check! â° The daily challenge is waiting for your participation! Share your insights in the accountability roster and climb the leaderboard."
                    ]
                    
                    reminder = random.choice(reminder_messages)
                    
                    # MODIFIED: Always send reminder to main group chat
                    bot.send_message(
                        PAID_GROUP_ID, 
                        reminder
                    )
                    logging.info(f"Sent challenge reminder to main group at {current_time}.")
                        
                    # Update reminder date
                    last_reminder_date = current_date
            
            # Calculate the time to sleep until the start of the next minute
            sleep_time = 60 - now.second - now.microsecond / 1_000_000
            time.sleep(sleep_time)
            
        except Exception as e:
            logging.error(f"Failed to send daily challenge or reminder: {e}")
            time.sleep(60)  # Wait a minute on error before trying again

# Command to set the daily challenge topic ID
@bot.message_handler(commands=['setchallengetopic'])
def set_challenge_topic(message):
    """Set or change the topic ID for daily challenges"""
    global DAILY_CHALLENGE_TOPIC_ID
    
    # Only allow the creator to use this command
    if message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to the bot creator.")
        return
    
    # Extract topic ID from command arguments
    args = message.text.split()
    
    # Show current setting if no arguments provided
    if len(args) == 1:
        current_topic = DAILY_CHALLENGE_TOPIC_ID if DAILY_CHALLENGE_TOPIC_ID else "Not set (using main group)"
        bot.reply_to(message, f"Current daily challenge topic ID: `{current_topic}`\n\nTo change, use: `/setchallengetopic ID`", parse_mode="Markdown")
        return
        
    try:
        # Handle "clear" or "reset" to remove topic ID
        if args[1].lower() in ["clear", "reset", "none"]:
            DAILY_CHALLENGE_TOPIC_ID = None
            # Save to database
            BOT_SETTINGS['daily_challenge_topic_id'] = None
            save_settings(BOT_SETTINGS)
            bot.reply_to(message, "âœ… Daily challenges will now be sent to the main group chat.")
            return
            
        # Try to parse as integer
        new_topic_id = int(args[1])
        DAILY_CHALLENGE_TOPIC_ID = new_topic_id
        
        # Save to database
        BOT_SETTINGS['daily_challenge_topic_id'] = new_topic_id
        save_settings(BOT_SETTINGS)
        
        bot.reply_to(message, f"âœ… Daily challenges will now be sent to topic ID: `{new_topic_id}`\nThis setting has been saved to the database.", parse_mode="Markdown")
        
    except ValueError:
        bot.reply_to(message, "âŒ Invalid topic ID. Please provide a numeric ID or 'clear' to reset.")
    except Exception as e:
        bot.reply_to(message, f"âŒ Error setting topic ID: {str(e)}")

@bot.message_handler(commands=['gettopic'])
def get_topic_id(message):
    """Command to get the topic ID of the current chat or topic."""
    # Check if user is the creator (not available to admins)
    if message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to the bot creator.")
        return
    
    try:
        # Check if message was sent in a group
        if message.chat.type not in ['group', 'supergroup']:
            bot.reply_to(message, "âŒ This command only works in group chats with topics enabled.")
            return
            
        # Get the chat ID and topic ID
        chat_id = message.chat.id
        topic_id = message.message_thread_id if hasattr(message, 'message_thread_id') else None
        
        # Get chat title
        chat_title = message.chat.title
        
        # Send a brief acknowledgment in the group
        bot.reply_to(message, "ðŸ“ Topic information has been sent to your DM for privacy.")
        
        # Send the detailed information in a direct message
        if topic_id:
            bot.send_message(
                message.from_user.id, 
                f"ðŸ“Œ *Topic Information*\n\n"
                f"*Chat Title:* {chat_title}\n"
                f"*Chat ID:* `{chat_id}`\n\n"
                f"*Topic ID:* `{topic_id}`\n\n",
                parse_mode="Markdown"
            )
        else:
            bot.send_message(
                message.from_user.id,
                f"ðŸ“Œ *Main Group Information*\n\n"
                f"*Chat Title:* {chat_title}\n"
                f"*Chat ID:* `{chat_id}`\n\n"
                f"This message is in the main group chat (not in a topic).\n\n",
                parse_mode="Markdown"
            )
            
    except ApiException as e:
        logging.error(f"Error in get_topic_id: {e}")
        if "bot can't initiate conversation with a user" in str(e):
            bot.reply_to(message, "âŒ I couldn't send you a DM. Please start a conversation with me first by messaging me directly.")
        else:
            bot.reply_to(message, f"âŒ Error retrieving topic information: {str(e)}")
    except Exception as e:
        logging.error(f"Error in get_topic_id: {e}")
        bot.reply_to(message, f"âŒ Error retrieving topic information: {str(e)}")
        
@bot.message_handler(commands=['jarvis'])
def handle_jarvis_command(message):
    """Send a Jarvis image to the group chat with global usage limits"""
    if message.chat.type not in ['group', 'supergroup']:
        bot.reply_to(message, "âŒ This command can only be used in group chats.")
        return
        
    current_time = datetime.now()
    
    try:
        # Get global usage data
        global_data = jarvis_usage_collection.find_one({"_id": "global_counter"})
        
        if not global_data:
            # First time the command is being used after implementation
            global_data = {
                "_id": "global_counter",
                "count": 0,
                "last_reset": current_time.strftime('%Y-%m-%d %H:%M:%S')
            }
        
        # Check if 6 hours have passed since last reset
        last_reset = datetime.strptime(global_data["last_reset"], '%Y-%m-%d %H:%M:%S')
        hours_passed = (current_time - last_reset).total_seconds() / 3600
        
        # Reset counter if 6 hours have passed
        if hours_passed >= 6:
            global_data = {
                "_id": "global_counter",
                "count": 0,
                "last_reset": current_time.strftime('%Y-%m-%d %H:%M:%S')
            }
        
        # Check if global limit has been reached
        if global_data["count"] >= 4:
            next_reset_time = last_reset + timedelta(hours=6)
            time_until_reset = next_reset_time - current_time
            hours, remainder = divmod(time_until_reset.seconds, 3600)
            minutes, _ = divmod(remainder, 60)
            
            # Send message and self-destruct after 5 seconds
            limit_msg = bot.reply_to(
                message, 
                f"â³ The Jarvis command has reached its global limit (4 uses per 6 hours).\nTry again in {hours} hours and {minutes} minutes."
            )
            
            # Create a thread to delete BOTH messages after 5 seconds
            def delete_after_delay(chat_id, message_ids):
                time.sleep(3)
                for msg_id in message_ids:
                    try:
                        bot.delete_message(chat_id, msg_id)
                        logging.info(f"Auto-deleted message ID {msg_id} in chat {chat_id}")
                    except Exception as e:
                        logging.error(f"Failed to auto-delete message ID {msg_id}: {e}")
            
            # Delete both the original command message and the rate limit message
            threading.Thread(target=delete_after_delay, 
                         args=(message.chat.id, [message.message_id, limit_msg.message_id])).start()
            return
        
        # Increment the global usage counter
        global_data["count"] += 1
        jarvis_usage_collection.replace_one({"_id": "global_counter"}, global_data, upsert=True)
        
        remaining = 4 - global_data["count"]
        
        # Path to the Jarvis image
        jarvis_image = "gifs/jarvis.png"  # Using existing GIFs directory
        
        # Send the image with usage info
        with open(jarvis_image, 'rb') as photo:
            bot.send_photo(
                message.chat.id, 
                photo,
            )
            logging.info(f"Sent Jarvis image in chat {message.chat.id} (requested by {message.from_user.id}, {remaining} global uses remaining)")
            
            # Also delete the original command after sending the image for cleaner chat
            try:
                bot.delete_message(message.chat.id, message.message_id)
            except Exception as e:
                logging.error(f"Failed to delete original command message: {e}")
                
    except FileNotFoundError:
        bot.reply_to(message, "âŒ Image not found.")
        logging.warning(f"Jarvis image not found (requested by {message.from_user.id})")
    except Exception as e:
        bot.reply_to(message, "âŒ Error sending image.")
        logging.error(f"Error in Jarvis command: {e}")

@bot.message_handler(commands=['commands'])
def list_available_commands(message):
    """Send the user a list of available commands based on their permission level"""
    user_id = message.from_user.id
    chat_id = message.chat.id
    
    # Define commands by permission level
    user_commands = [
        ("/start", "Begin the bot interaction or return to main menu"),
        ("/verify", "Submit payment proof for verification"),
        ("/dashboard", "View your membership status and details"),
        ("/confess", "Submit an anonymous confession"),
        ("/cancel", "Cancel an in-progress action"),
        ("/commands", "Show this list of commands"),
        ("/ping", "Check if the bot is online"),
        ("/tip", "Support the bot creator"),
        ("/jarvis", "Display the Jarvis AI image"),
        ("/changelogs", "View recent updates to the bot/academy"),
        ("/dm", "Send a direct message to a user")
    ]
    
    admin_commands = [
        ("/notify", "Send payment reminders to users near expiration"),
        ("/challenge", "Manually trigger a daily challenge"),
        ("/admin_dashboard", "Access admin controls"),
        ("/leaderboard", "Manually trigger leaderboard update"),
        ("/resend", "Force cleanup and resend of payment reminders"),
        ("/dm", "Send a direct message to a user")
    ]
    
    creator_commands = [
        ("/post_changelog", "Create and post a new changelog"),
        ("/gettopic", "Get the topic ID of the current chat topic"),
        ("/setchallengetopic", "Set the topic ID for daily challenges"),
        ("/setannouncementtopic", "Set the topic ID for announcements"),
        ("/setaccountabilitytopic", "Set the topic ID for accountability roster"),
        ("/setleaderboardtopic", "Set the topic ID for leaderboards"),
        ("/setconfessiontopic", "Set the topic ID for confessions"),
        ("/refreshexpired", "Refresh expired members status flags"),
        ("/remove", "Remove yourself from pending users list"),
        ("/check", "Check MongoDB connection status")
    ]
    
    # Check if we're in a group chat - if so, tell the user we've sent a DM
    if message.chat.type in ['group', 'supergroup']:
        bot.reply_to(message, "ðŸ“ I've sent you a list of available commands in a private message.")
    
    try:
        # Determine which commands to show based on permission level
        is_admin = user_id in ADMIN_IDS
        is_creator = user_id == CREATOR_ID
        
        # Format the command list message
        message_text = "ðŸ¤– *AVAILABLE COMMANDS*\n\n"
        
        # Add user commands for everyone
        message_text += "*General User Commands:*\n"
        for cmd, desc in user_commands:
            message_text += f"`{cmd}` - {desc}\n"
        
        # For admins and creator, show ALL commands for transparency
        if is_admin or is_creator:
            message_text += "\n*Admin Commands:*\n"
            for cmd, desc in admin_commands:
                message_text += f"`{cmd}` - {desc}\n"
            
            message_text += "\n*Creator Commands:*\n"
            for cmd, desc in creator_commands:
                message_text += f"`{cmd}` - {desc}\n"
            
            # Additional note for admins
            if is_admin and not is_creator:
                message_text += "\n*Note:* Creator commands are shown for transparency but can only be used by the bot creator."
        
        # Send the message
        bot.send_message(
            user_id,  # Send as DM to the user
            message_text,
            parse_mode="Markdown"
        )
        
        logging.info(f"Sent command list to user {user_id}")
        
    except ApiException as e:
        # Handle case where bot can't DM the user
        if "bot can't initiate conversation with a user" in str(e):
            bot.reply_to(message, "âŒ I couldn't send you a DM. Please start a conversation with me first by sending /start in a private chat.")
        else:
            bot.reply_to(message, f"âŒ Error sending command list: {str(e)}")
        logging.error(f"Failed to send command list to user {user_id}: {e}")
    except Exception as e:
        bot.reply_to(message, "âŒ An error occurred while processing your request.")
        logging.error(f"Error in list_available_commands: {e}")

@bot.message_handler(commands=['setaccountabilitytopic'])
def set_accountability_topic(message):
    """Set or change the topic ID for accountability posts"""
    global ACCOUNTABILITY_TOPIC_ID
    
    # Only allow the creator to use this command
    if message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to the bot creator.")
        return
    
    args = message.text.split()
    
    # Show current setting if no arguments provided
    if len(args) == 1:
        current_topic = ACCOUNTABILITY_TOPIC_ID if ACCOUNTABILITY_TOPIC_ID else "Not set"
        bot.reply_to(message, f"Current accountability topic ID: `{current_topic}`\n\nTo change, use: `/setaccountabilitytopic ID`", parse_mode="Markdown")
        return
        
    try:
        # Handle "clear" or "reset" to remove topic ID
        if args[1].lower() in ["clear", "reset", "none"]:
            ACCOUNTABILITY_TOPIC_ID = None
            BOT_SETTINGS['accountability_topic_id'] = None
            save_settings(BOT_SETTINGS)
            bot.reply_to(message, "âœ… Accountability topic ID has been cleared.")
            return
            
        # Try to parse as integer
        new_topic_id = int(args[1])
        ACCOUNTABILITY_TOPIC_ID = new_topic_id
        
        # Save to database
        BOT_SETTINGS['accountability_topic_id'] = new_topic_id
        save_settings(BOT_SETTINGS)
        
        bot.reply_to(message, f"âœ… Accountability submissions will now be monitored in topic ID: `{new_topic_id}`", parse_mode="Markdown")
        
    except ValueError:
        bot.reply_to(message, "âŒ Invalid topic ID. Please provide a numeric ID or 'clear' to reset.")
    except Exception as e:
        bot.reply_to(message, f"âŒ Error setting topic ID: {str(e)}")

@bot.message_handler(commands=['setleaderboardtopic'])
def set_leaderboard_topic(message):
    """Set or change the topic ID for leaderboard posts"""
    global LEADERBOARD_TOPIC_ID
    
    # Only allow the creator to use this command
    if message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to the bot creator.")
        return
    
    args = message.text.split()
    
    # Show current setting if no arguments provided
    if len(args) == 1:
        current_topic = LEADERBOARD_TOPIC_ID if LEADERBOARD_TOPIC_ID else "Not set"
        bot.reply_to(message, f"Current leaderboard topic ID: `{current_topic}`\n\nTo change, use: `/setleaderboardtopic ID`", parse_mode="Markdown")
        return
        
    try:
        # Handle "clear" or "reset" to remove topic ID
        if args[1].lower() in ["clear", "reset", "none"]:
            LEADERBOARD_TOPIC_ID = None
            BOT_SETTINGS['leaderboard_topic_id'] = None
            save_settings(BOT_SETTINGS)
            bot.reply_to(message, "âœ… Leaderboard topic ID has been cleared.")
            return
            
        # Try to parse as integer
        new_topic_id = int(args[1])
        LEADERBOARD_TOPIC_ID = new_topic_id
        
        # Save to database
        BOT_SETTINGS['leaderboard_topic_id'] = new_topic_id
        save_settings(BOT_SETTINGS)
        
        bot.reply_to(message, f"âœ… Leaderboard will now be posted in topic ID: `{new_topic_id}`", parse_mode="Markdown")
        
    except ValueError:
        bot.reply_to(message, "âŒ Invalid topic ID. Please provide a numeric ID or 'clear' to reset.")
    except Exception as e:
        bot.reply_to(message, f"âŒ Error setting topic ID: {str(e)}")

def save_user_score(user_id, username, first_name, message_id, points, submission_date):
    """Save a user's score for a submission"""
    try:
        # Create a unique identifier for each submission
        submission_id = f"{user_id}_{submission_date.strftime('%Y-%m-%d')}"
        
        # Store the submission data
        submission_data = {
            "_id": submission_id,
            "user_id": user_id,
            "username": username,
            "first_name": first_name,
            "message_id": message_id,
            "points": points,
            "date": submission_date.strftime('%Y-%m-%d'),
            "timestamp": submission_date.strftime('%Y-%m-%d %H:%M:%S'),
            "month_year": submission_date.strftime('%Y-%m')
        }
        
        # Use upsert to update if exists or insert if new
        scores_collection.replace_one({"_id": submission_id}, submission_data, upsert=True)
        logging.info(f"Saved score for user {user_id}: {points} points")
        return True
    except Exception as e:
        logging.error(f"Error saving score for user {user_id}: {e}")
        return False

def get_daily_leaderboard(date):
    """Get the leaderboard for a specific day"""
    try:
        # Find all scores for the given date
        date_str = date.strftime('%Y-%m-%d')
        scores = list(scores_collection.find({"date": date_str}))
        
        # Sort by points (highest first)
        scores.sort(key=lambda x: x.get('points', 0), reverse=True)
        
        return scores
    except Exception as e:
        logging.error(f"Error getting daily leaderboard: {e}")
        return []

def get_monthly_leaderboard(year_month):
    """Get the leaderboard for a specific month"""
    try:
        # Find all submissions for the given month
        submissions = list(scores_collection.find({"month_year": year_month}))
        
        # Group by user_id and sum points
        user_scores = {}
        for submission in submissions:
            user_id = submission.get('user_id')
            points = submission.get('points', 0)
            
            if user_id not in user_scores:
                user_scores[user_id] = {
                    'user_id': user_id,
                    'username': submission.get('username'),
                    'first_name': submission.get('first_name'),
                    'total_points': 0,
                    'submissions': 0
                }
            
            user_scores[user_id]['total_points'] += points
            user_scores[user_id]['submissions'] += 1
        
        # Convert to list and sort by total points
        leaderboard = list(user_scores.values())
        leaderboard.sort(key=lambda x: x.get('total_points', 0), reverse=True)
        
        return leaderboard
    except Exception as e:
        logging.error(f"Error getting monthly leaderboard: {e}")
        return []

@bot.message_handler(func=lambda message: message.chat.id == PAID_GROUP_ID and 
                    hasattr(message, 'message_thread_id') and 
                    message.message_thread_id == ACCOUNTABILITY_TOPIC_ID)
def handle_accountability_submission(message):
    """Handle messages posted in the accountability roster topic"""
    try:
        # Only process if accountability topic is configured
        if not ACCOUNTABILITY_TOPIC_ID:
            return
            
        user_id = message.from_user.id
        username = message.from_user.username or "No_Username"
        first_name = message.from_user.first_name or username
        
        # Save information about this submission to track it
        submission_date = datetime.now(pytz.timezone('Asia/Manila'))
        submission_id = f"{user_id}_{submission_date.strftime('%Y-%m-%d')}"
        
        # Check if user already submitted today
        existing_submission = accountability_collection.find_one({"_id": submission_id})
        if existing_submission:
            # If they've already been graded, don't allow another submission
            if existing_submission.get("graded", False):
                try:
                    # Try to delete the duplicate message
                    bot.delete_message(message.chat.id, message.message_id)
                    
                    # Send a DM to inform the user
                    bot.send_message(
                        user_id,
                        "âš ï¸ *You've already submitted today's challenge*\n\n"
                        "I noticed you tried to submit another entry for today's challenge, but you're already "
                        "graded for today. You can only submit once per day.\n\n"
                        "Your earlier submission has been counted and will appear in today's leaderboard.",
                        parse_mode="Markdown"
                    )
                    logging.info(f"Removed duplicate submission from user {user_id}")
                except ApiException as e:
                    # If we can't delete (maybe bot isn't admin), just log it
                    logging.error(f"Failed to delete duplicate submission: {e}")
                
                return
                
            # User already submitted but not graded - update their submission
            accountability_collection.update_one(
                {"_id": submission_id},
                {"$set": {"message_id": message.message_id, "timestamp": submission_date.strftime('%Y-%m-%d %H:%M:%S')}}
            )
            logging.info(f"Updated submission for user {user_id}")
        else:
            # New submission for today
            submission_data = {
                "_id": submission_id,
                "user_id": user_id,
                "username": username,
                "first_name": first_name,
                "message_id": message.message_id,
                "graded": False,
                "points": 0,
                "date": submission_date.strftime('%Y-%m-%d'),
                "timestamp": submission_date.strftime('%Y-%m-%d %H:%M:%S')
            }
            accountability_collection.insert_one(submission_data)
            logging.info(f"New submission from user {user_id}")
        
        # Add more compact grading buttons
        markup = InlineKeyboardMarkup(row_width=5)  # Make buttons fit in one row if possible
        markup.add(
            InlineKeyboardButton("âŒ", callback_data=f"grade_{user_id}_0"),
            InlineKeyboardButton("10", callback_data=f"grade_{user_id}_10"),
            InlineKeyboardButton("20", callback_data=f"grade_{user_id}_20"),
            InlineKeyboardButton("30", callback_data=f"grade_{user_id}_30"),
            InlineKeyboardButton("40", callback_data=f"grade_{user_id}_40")
        )
        
        # Send a more minimal message with just the grade buttons
        bot.send_message(
            message.chat.id,
            f"Grade @{username}'s submission: â¤´ï¸",  # Arrow pointing up to indicate this is for the message above
            reply_to_message_id=message.message_id,
            reply_markup=markup,
            message_thread_id=ACCOUNTABILITY_TOPIC_ID
        )
        
    except Exception as e:
        logging.error(f"Error handling accountability submission: {e}")

@bot.callback_query_handler(func=lambda call: call.data.startswith("grade_"))
def handle_grading(call):
    """Handle grading button presses"""
    try:
        # Parse the callback data (format: grade_userId_points)
        parts = call.data.split("_")
        if len(parts) != 3:
            bot.answer_callback_query(call.id, "âŒ Invalid callback data")
            return
            
        submission_user_id = int(parts[1])
        points = int(parts[2])
        grader_id = call.from_user.id
        
        # Check if grader is admin or creator
        if grader_id not in ADMIN_IDS and grader_id != CREATOR_ID:
            # User is not authorized - send troll message
            bot.answer_callback_query(call.id, "Nice try! You can't grade yourself, bozo! ðŸ¤¡", show_alert=True)
            return
        
        # Get the submission from the original message
        original_msg = call.message.reply_to_message
        if not original_msg:
            bot.answer_callback_query(call.id, "âŒ Could not find the original submission", show_alert=True)
            return
            
        user_info = original_msg.from_user
        username = user_info.username or "No_Username"
        first_name = user_info.first_name or username
        
        # Get today's date in Manila timezone for the submission ID
        manila_tz = pytz.timezone('Asia/Manila')
        submission_date = datetime.now(manila_tz)
        submission_id = f"{submission_user_id}_{submission_date.strftime('%Y-%m-%d')}"
        
        # Add more debugging
        logging.info(f"Updating accountability for user {submission_user_id}, submission ID: {submission_id}")

        # First, verify if the document exists
        existing_doc = accountability_collection.find_one({"_id": submission_id})
        if not existing_doc:
            logging.warning(f"No document found for submission ID: {submission_id}")
            # Try using the original document submission date
            msg_date = datetime.fromtimestamp(original_msg.date).replace(tzinfo=pytz.UTC)
            alt_submission_date = msg_date.astimezone(manila_tz)
            alt_submission_id = f"{submission_user_id}_{alt_submission_date.strftime('%Y-%m-%d')}"
            logging.info(f"Trying alternative submission ID: {alt_submission_id}")
            submission_id = alt_submission_id
        
        # Update with explicit result verification
        update_result = accountability_collection.update_one(
            {"_id": submission_id},
            {"$set": {"graded": True, "points": points, "graded_by": grader_id}}
        )
        
        if update_result.matched_count == 0:
            logging.error(f"Failed to match document with ID: {submission_id}")
            # Try to find any document for this user today
            today_str = submission_date.strftime('%Y-%m-%d')
            docs = list(accountability_collection.find({"user_id": submission_user_id, "date": today_str}))
            if docs:
                doc_id = docs[0].get("_id")
                logging.info(f"Found alternative document with ID: {doc_id}")
                update_result = accountability_collection.update_one(
                    {"_id": doc_id},
                    {"$set": {"graded": True, "points": points, "graded_by": grader_id}}
                )
                logging.info(f"Update result with alternative ID: matched={update_result.matched_count}, modified={update_result.modified_count}")
        else:
            logging.info(f"Updated accountability document: matched={update_result.matched_count}, modified={update_result.modified_count}")
        
        # Save the score
        save_result = save_user_score(
            submission_user_id, 
            username, 
            first_name,
            original_msg.message_id, 
            points,
            submission_date
        )
        logging.info(f"Score save result: {save_result}")
        
        # Update the accountability collection to mark as graded
        accountability_collection.update_one(
            {"_id": submission_id},
            {"$set": {"graded": True, "points": points, "graded_by": grader_id}}
        )
        
        # After successfully saving the grade, delete the grading buttons message
        try:
            # Delete the original grading message to reduce clutter
            bot.delete_message(call.message.chat.id, call.message.message_id)
            
            # Send a small confirmation in the thread that will auto-delete
            if points > 0:
                confirm_msg = f"âœ… @{username}'s submission graded: {points} pts"
            else:
                confirm_msg = f"âŒ @{username}'s submission rejected"
                
            temp_msg = bot.send_message(
                call.message.chat.id,
                confirm_msg,
                reply_to_message_id=original_msg.message_id,
                message_thread_id=ACCOUNTABILITY_TOPIC_ID
            )
            
            # Schedule this confirmation to be deleted after 5 seconds
            def delete_later(chat_id, message_id):
                time.sleep(5)
                try:
                    bot.delete_message(chat_id, message_id)
                except:
                    pass  # Ignore errors if message can't be deleted
                
            threading.Thread(target=delete_later, 
                         args=(temp_msg.chat.id, temp_msg.message_id)).start()
            
            bot.answer_callback_query(call.id, f"Successfully graded with {points} points!")
            
        except ApiException as e:
            logging.error(f"Error cleaning up grading UI: {e}")
            # If we can't delete, fall back to just editing the message
            markup = InlineKeyboardMarkup()
            if points > 0:
                btn_text = f"Graded: {points} pts âœ…"
            else:
                btn_text = "Rejected âŒ"
                
            markup.add(InlineKeyboardButton(btn_text, callback_data="already_graded"))
            
            bot.edit_message_text(
                f"@{username}'s submission graded: {points} pts",
                call.message.chat.id,
                call.message.message_id,
                reply_markup=markup
            )
            
            bot.answer_callback_query(call.id, f"Graded with {points} points")
        
        # Notify the user via DM
        try:
            # Random delay to simulate human typing (1-3 seconds)
            time.sleep(1 + random.random() * 2)
            
            # Show typing indicator first
            bot.send_chat_action(submission_user_id, 'typing')
            time.sleep(1.5)  # Simulate thinking time
            
            if points > 0:
                # Add variety to positive notifications
                positive_messages = [
                    f"âœ… *Great job on your daily challenge submission!*\n\n"
                    f"Your hard work has been noticed by our admin team. Keep up the excellent work and check the leaderboard at midnight to see where you rank!",
                    
                    f"âœ… *Challenge submission graded!*\n\n"
                    f"Thanks for participating in today's challenge! Your submission has been reviewed and points have been awarded. The daily rankings will be posted at midnight - good luck!",
                    
                    f"âœ… *Daily challenge completed!*\n\n"
                    f"One of our admins has reviewed your work for today's challenge. Well done on your submission! Daily rankings are posted at midnight."
                ]
                notification_message = random.choice(positive_messages)
            else:
                # Add variety to rejection messages
                rejection_messages = [
                    f"âŒ *About your challenge submission...*\n\n"
                    f"Unfortunately, your submission didn't meet all the requirements for today's challenge. Take another look at the instructions and try again tomorrow!",
                    
                    f"âŒ *Challenge submission feedback*\n\n"
                    f"It looks like there was an issue with your submission for today's challenge. Review the daily challenge criteria and give it another shot tomorrow!",
                    
                    f"âŒ *Daily challenge feedback*\n\n"
                    f"Thanks for participating, but your submission wasn't quite what we were looking for today. Check the challenge requirements and try again tomorrow."
                ]
                notification_message = random.choice(rejection_messages)
                
            bot.send_message(
                submission_user_id,
                notification_message,
                parse_mode="Markdown"
            )
        except ApiException:
            logging.error(f"Could not send grade notification to user {submission_user_id}")
            
    except Exception as e:
        logging.error(f"Error handling grading callback: {e}")
        bot.answer_callback_query(call.id, "âŒ Error processing grade", show_alert=True)

# Handle "already graded" button to prevent further clicks
@bot.callback_query_handler(func=lambda call: call.data == "already_graded")
def handle_already_graded(call):
    bot.answer_callback_query(call.id, "This submission has already been graded!")


def generate_daily_leaderboard_text(date):
    """Generate formatted text for daily leaderboard with proper tie handling"""
    scores = get_daily_leaderboard(date)
    
    if not scores:
        return f"ðŸ“Š *DAILY LEADERBOARD: {date.strftime('%B %d, %Y')}*\n\nNo entries for today!"
    
    # Format the leaderboard message
    leaderboard_text = f"ðŸ“Š *DAILY LEADERBOARD: {date.strftime('%B %d, %Y')}*\n\n"
    
    # Keep track of the current rank and last score for tie detection
    current_rank = 1
    last_score = None
    
    for i, entry in enumerate(scores):
        points = entry.get('points', 0)
        username = safe_markdown_escape(entry.get('username', 'No_Username'))  # <-- Add safe_markdown_escape
        
        # If this score is different from the previous one, update the rank
        if last_score is not None and points != last_score:
            current_rank = i + 1
        
        last_score = points
        
        # Create emoji for ranks
        if current_rank == 1:
            rank_emoji = "ðŸ¥‡"
        elif current_rank == 2:
            rank_emoji = "ðŸ¥ˆ"
        elif current_rank == 3:
            rank_emoji = "ðŸ¥‰"
        else:
            rank_emoji = f"{current_rank}."
        
        leaderboard_text += f"{rank_emoji} @{username}: *{points} points*\n"
    
    return leaderboard_text

def generate_monthly_leaderboard_text(year_month_str):
    """Generate formatted text for monthly leaderboard with proper tie handling"""
    # Parse year-month string into a date object to get month name
    try:
        date = datetime.strptime(year_month_str, '%Y-%m')
        month_name = date.strftime('%B %Y')
    except:
        month_name = year_month_str
    
    scores = get_monthly_leaderboard(year_month_str)
    
    if not scores:
        return f"ðŸ“Š *MONTHLY LEADERBOARD: {month_name}*\n\nNo entries this month!"
    
    # Format the leaderboard message
    leaderboard_text = f"ðŸ“Š *MONTHLY LEADERBOARD: {month_name}*\n\n"
    
    # Keep track of the current rank and last score for tie detection
    current_rank = 1
    last_score = None
    
    for i, entry in enumerate(scores):
        total_points = entry.get('total_points', 0)
        username = safe_markdown_escape(entry.get('username', 'No_Username'))  # <-- Add safe_markdown_escape
        submissions = entry.get('submissions', 0)
        
        # If this score is different from the previous one, update the rank
        if last_score is not None and total_points != last_score:
            current_rank = i + 1
            
        last_score = total_points
        
        # Create emoji for ranks
        if current_rank == 1:
            rank_emoji = "ðŸ¥‡"
        elif current_rank == 2:
            rank_emoji = "ðŸ¥ˆ"
        elif current_rank == 3:
            rank_emoji = "ðŸ¥‰"
        else:
            rank_emoji = f"{current_rank}."
        
        leaderboard_text += f"{rank_emoji} @{username}: *{total_points} points* ({submissions} submissions)\n"
    
    leaderboard_text += f"\nðŸ† Congratulations to all participants! Keep up the great work!"
    
    return leaderboard_text

def send_daily_leaderboard():
    """Send the daily leaderboard at midnight"""
    logging.info("Daily leaderboard thread started")
    
    # Track last leaderboard sent date
    last_leaderboard_date = None
    
    while True:
        try:
            now = datetime.now(pytz.timezone('Asia/Manila'))
            
            # Check if it's midnight (00:00) and we haven't sent a leaderboard today
            if now.strftime('%H:%M') == '00:00' and now.strftime('%Y-%m-%d') != last_leaderboard_date:
                logging.info("It's midnight - generating daily leaderboard")
                
                # Get yesterday's date (since we're sending at midnight)
                yesterday = now - timedelta(days=1)
                
                # Generate leaderboard text
                leaderboard_text = generate_daily_leaderboard_text(yesterday)
                
                # Send the leaderboard to the designated topic
                if LEADERBOARD_TOPIC_ID:
                    bot.send_message(
                        PAID_GROUP_ID, 
                        leaderboard_text,
                        parse_mode="Markdown",
                        message_thread_id=LEADERBOARD_TOPIC_ID
                    )
                    logging.info(f"Sent daily leaderboard to topic {LEADERBOARD_TOPIC_ID}")
                else:
                    logging.warning("No leaderboard topic ID configured - skipping leaderboard")
                
                # Check if it's also month-end
                if yesterday.day == calendar.monthrange(yesterday.year, yesterday.month)[1]:
                    # It's the last day of the month - send monthly leaderboard too
                    month_year = yesterday.strftime('%Y-%m')
                    monthly_leaderboard = generate_monthly_leaderboard_text(month_year)
                    
                    # Send after a short delay
                    time.sleep(3)
                    
                    # MODIFIED: Send monthly leaderboard to ANNOUNCEMENT_TOPIC_ID if available
                    if ANNOUNCEMENT_TOPIC_ID:
                        # Send to announcements topic for better visibility
                        bot.send_message(
                            PAID_GROUP_ID, 
                            monthly_leaderboard,
                            parse_mode="Markdown",
                            message_thread_id=ANNOUNCEMENT_TOPIC_ID
                        )
                        logging.info(f"Sent monthly leaderboard to announcements topic {ANNOUNCEMENT_TOPIC_ID}")
                    elif LEADERBOARD_TOPIC_ID:
                        # Fall back to regular leaderboard topic if announcements not configured
                        bot.send_message(
                            PAID_GROUP_ID, 
                            monthly_leaderboard,
                            parse_mode="Markdown",
                            message_thread_id=LEADERBOARD_TOPIC_ID
                        )
                        logging.info(f"Sent monthly leaderboard to leaderboard topic {LEADERBOARD_TOPIC_ID}")
                    else:
                        # Last resort: send to main group if no topics configured
                        bot.send_message(
                            PAID_GROUP_ID, 
                            monthly_leaderboard,
                            parse_mode="Markdown"
                        )
                        logging.info("Sent monthly leaderboard to main group (no topic IDs configured)")
                
                # Update last leaderboard date
                last_leaderboard_date = now.strftime('%Y-%m-%d')
            
            # Sleep until the next minute
            sleep_time = 60 - now.second - now.microsecond / 1_000_000
            time.sleep(sleep_time)
            
        except Exception as e:
            logging.error(f"Error sending leaderboard: {e}")
            time.sleep(60)  # Wait for a minute before trying again

# Command handler for /setconfessiontopic
@bot.message_handler(commands=['setconfessiontopic'])
def set_confession_topic(message):
    """Set or change the topic ID for confessions"""
    global CONFESSION_TOPIC_ID
    
    # Only allow the creator to use this command
    if message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to the bot creator.")
        return
    
    args = message.text.split()
    
    # Show current setting if no arguments provided
    if len(args) == 1:
        current_topic = CONFESSION_TOPIC_ID if CONFESSION_TOPIC_ID else "Not set (using main group)"
        bot.reply_to(message, f"Current confession topic ID: `{current_topic}`\n\nTo change, use: `/setconfessiontopic ID`", parse_mode="Markdown")
        return
        
    try:
        # Handle "clear" or "reset" to remove topic ID
        if args[1].lower() in ["clear", "reset", "none"]:
            CONFESSION_TOPIC_ID = None
            BOT_SETTINGS['confession_topic_id'] = None
            save_settings(BOT_SETTINGS)
            bot.reply_to(message, "âœ… Confessions will now be sent to the main group chat.")
            return
            
        # Try to parse as integer
        new_topic_id = int(args[1])
        CONFESSION_TOPIC_ID = new_topic_id
        
        # Save to database
        BOT_SETTINGS['confession_topic_id'] = new_topic_id
        save_settings(BOT_SETTINGS)
        
        bot.reply_to(message, f"âœ… Confessions will now be sent to topic ID: `{new_topic_id}`\nThis setting has been saved to the database.", parse_mode="Markdown")
        
    except ValueError:
        bot.reply_to(message, "âŒ Invalid topic ID. Please provide a numeric ID or 'clear' to reset.")
    except Exception as e:
        bot.reply_to(message, f"âŒ Error setting topic ID: {str(e)}")

# Command to initiate a confession
@bot.message_handler(commands=['confess'])
def start_confession(message):
    """Start the confession process"""
    # Only allow in private chats
    if message.chat.type != 'private':
        bot.reply_to(message, "ðŸ¤« Please send me a direct message to start your confession.")
        return
    
    user_id = message.from_user.id
    
    # Check if user is already confessing
    if user_id in USERS_CONFESSING:
        bot.send_message(user_id, "â³ You already have a confession in progress. Please complete it or send /cancel to stop.")
        return
    
    USERS_CONFESSING[user_id] = {'status': 'awaiting_confession'}
    
    # Personalize the instruction for better engagement - now with media support instructions
    welcome_messages = [
        "ðŸ”’ *Anonymous Confession*\n\nShare your trading frustrations, wins, or anything on your mind. Your identity will remain hidden.\n\n"
        "ðŸ“ *Send text* for a regular confession\n"
        "ðŸ“¸ *Send a photo* to share an image anonymously\n"
        "ðŸŽžï¸ *Send a GIF* or video to share media anonymously\n\n"
        "Type your confession now or send /cancel to stop.",
        
        "ðŸ¤« *Secret Sharing*\n\nGot something to get off your chest about your trading journey? No one will know it's you.\n\n"
        "You can now include:\n"
        "â€¢ Text messages\n"
        "â€¢ Photos/Screenshots\n"
        "â€¢ GIFs or Videos\n\n"
        "Send your confession now or use /cancel to stop.",
        
        "ðŸŽ­ *Anonymous Message*\n\nShare your trading experiences, market observations, or personal thoughts anonymously with the community.\n\n"
        "âœ¨ *NEW:* You can now include images, GIFs, and videos in your confessions!\n\n"
        "Send your confession now or use /cancel to stop."
    ]
    
    bot.send_message(user_id, random.choice(welcome_messages), parse_mode="Markdown")

# Command to cancel an in-progress confession
@bot.message_handler(commands=['cancel'])
def cancel_confession(message):
    """Cancel an in-progress confession"""
    # Only process in private chats
    if message.chat.type != 'private':
        return
        
    user_id = message.from_user.id
    
    if user_id in USERS_CONFESSING:
        USERS_CONFESSING.pop(user_id)
        bot.send_message(user_id, "âœ… Confession cancelled. Your message wasn't sent.")
    else:
        bot.send_message(user_id, "â“ You don't have any active confession to cancel.")

# Add handler for text confessions (existing handler)
@bot.message_handler(func=lambda message: message.chat.type == 'private' and 
                    message.from_user.id in USERS_CONFESSING and 
                    USERS_CONFESSING[message.from_user.id]['status'] == 'awaiting_confession')
def handle_text_confession(message):
    """Process a user's text confession"""
    user_id = message.from_user.id
    confession_text = message.text
    
    # Some basic moderation/filtering
    if not confession_text or len(confession_text) < 3:
        bot.send_message(user_id, "âŒ Your confession is too short. Please write something meaningful or use /cancel to stop.")
        return
        
    if len(confession_text) > 2000:
        bot.send_message(user_id, "âŒ Your confession is too long (max 2000 characters). Please shorten it or use /cancel to stop.")
        return
    
    # Check for offensive content (this is a very basic implementation)
    offensive_words = ["slur1", "slur2", "badword"]  # Replace with actual moderation list
    if any(word in confession_text.lower() for word in offensive_words):
        bot.send_message(user_id, "âŒ Your confession contains content that violates our community guidelines. Please revise it or use /cancel to stop.")
        return
    
    # Increment the confession counter
    global CONFESSION_COUNTER
    CONFESSION_COUNTER += 1
    save_confession_counter(CONFESSION_COUNTER)
    
    # Format the confession message
    confession_message = f"ðŸ” *Confession #{CONFESSION_COUNTER}*\n\n{confession_text}"
    
    try:
        # Send to the group or topic
        if CONFESSION_TOPIC_ID:
            sent_message = bot.send_message(
                PAID_GROUP_ID,
                confession_message,
                parse_mode="Markdown",
                message_thread_id=CONFESSION_TOPIC_ID
            )
        else:
            sent_message = bot.send_message(
                PAID_GROUP_ID, 
                confession_message,
                parse_mode="Markdown"
            )
        
        # Log the confession (not linking to the user for privacy)
        logging.info(f"Text confession #{CONFESSION_COUNTER} sent to group")
        
        # Send confirmation to user
        confirmation_messages = [
            "âœ… *Confession sent!*\n\nYour message has been posted anonymously. Thank you for sharing.",
            "ðŸ¤« *Secret shared!*\n\nYour anonymous confession has been posted to the group.",
            "ðŸ“¨ *Message delivered!*\n\nYour thoughts have been shared anonymously with the community."
        ]
        
        bot.send_message(user_id, random.choice(confirmation_messages), parse_mode="Markdown")
        
        # Get user info for admin records
        try:
            user_info = bot.get_chat(user_id)
            username = user_info.username
            first_name = user_info.first_name or ""
            last_name = user_info.last_name or ""
            
            if username:
                user_display = f"@{username}"
            else:
                user_display = f"{first_name} {last_name}".strip() or f"User ID: {user_id}"
                
            # Escape any Markdown characters in user_display
            user_display = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', user_display)
        except Exception as e:
            logging.error(f"Error getting user info for confession: {e}")
            user_display = f"User ID: {user_id}"

        # Keep admin record of who sent what confession (for moderation purposes)
        # Only send to CREATOR, not all admins
        admin_record = f"ðŸ“ *Admin Log*\n\nConfession #{CONFESSION_COUNTER} was submitted by {user_display}"
        bot.send_message(CREATOR_ID, admin_record, parse_mode="Markdown")
        
    except Exception as e:
        logging.error(f"Error sending confession: {e}")
        bot.send_message(user_id, "âŒ There was an error sending your confession. Please try again later.")
    
    # Remove user from confessing dict
    USERS_CONFESSING.pop(user_id, None)

# Add new handler for photo confessions
@bot.message_handler(content_types=['photo'], func=lambda message: message.chat.type == 'private' and 
                    message.from_user.id in USERS_CONFESSING and 
                    USERS_CONFESSING[message.from_user.id]['status'] == 'awaiting_confession')
def handle_photo_confession(message):
    """Process a user's photo confession"""
    user_id = message.from_user.id
    photo_id = message.photo[-1].file_id  # Get the highest resolution photo
    caption = message.caption or ""
    
    # Increment the confession counter
    global CONFESSION_COUNTER
    CONFESSION_COUNTER += 1
    save_confession_counter(CONFESSION_COUNTER)
    
    try:
        # Send anonymously to the group
        if CONFESSION_TOPIC_ID:
            sent_message = bot.send_photo(
                PAID_GROUP_ID,
                photo_id,
                caption=f"ðŸ” *Confession #{CONFESSION_COUNTER}*\n\n{caption}" if caption else f"ðŸ” *Confession #{CONFESSION_COUNTER}*",
                parse_mode="Markdown",
                message_thread_id=CONFESSION_TOPIC_ID
            )
        else:
            sent_message = bot.send_photo(
                PAID_GROUP_ID,
                photo_id,
                caption=f"ðŸ” *Confession #{CONFESSION_COUNTER}*\n\n{caption}" if caption else f"ðŸ” *Confession #{CONFESSION_COUNTER}*",
                parse_mode="Markdown"
            )
            
        # Log the confession
        logging.info(f"Photo confession #{CONFESSION_COUNTER} sent to group")
        
        # Send confirmation to user
        bot.send_message(user_id, "âœ… Your photo confession has been sent anonymously.")
        
        # Admin record
        try:
            user_info = bot.get_chat(user_id)
            username = user_info.username
            user_display = f"@{username}" if username else f"User ID: {user_id}"
            user_display = re.sub(r'([_*[\]()~`>#\+\-=|{}.!])', r'\\\1', user_display)
            
            admin_record = f"ðŸ“ *Admin Log*\n\nPhoto confession #{CONFESSION_COUNTER} was submitted by {user_display}"
            bot.send_message(CREATOR_ID, admin_record, parse_mode="Markdown")
        except Exception as e:
            logging.error(f"Error sending admin log for photo confession: {e}")
    
    except Exception as e:
        logging.error(f"Error sending photo confession: {e}")
        bot.send_message(user_id, "âŒ There was an error sending your photo confession.")
    
    # Remove user from confessing dict
    USERS_CONFESSING.pop(user_id, None)

# Add handler for GIF/animation confessions
@bot.message_handler(content_types=['animation', 'document'], func=lambda message: message.chat.type == 'private' and 
                    message.from_user.id in USERS_CONFESSING and 
                    USERS_CONFESSING[message.from_user.id]['status'] == 'awaiting_confession' and
                    (hasattr(message, 'animation') or (hasattr(message, 'document') and message.document.mime_type == 'image/gif')))
def handle_gif_confession(message):
    """Process a user's GIF confession"""
    user_id = message.from_user.id
    gif_id = message.animation.file_id if hasattr(message, 'animation') else message.document.file_id
    caption = message.caption or ""
    
    # Increment the confession counter
    global CONFESSION_COUNTER
    CONFESSION_COUNTER += 1
    save_confession_counter(CONFESSION_COUNTER)
    
    try:
        # Send anonymously to the group
        if CONFESSION_TOPIC_ID:
            sent_message = bot.send_animation(
                PAID_GROUP_ID,
                gif_id,
                caption=f"ðŸ” *Confession #{CONFESSION_COUNTER}*\n\n{caption}" if caption else f"ðŸ” *Confession #{CONFESSION_COUNTER}*",
                parse_mode="Markdown",
                message_thread_id=CONFESSION_TOPIC_ID
            )
        else:
            sent_message = bot.send_animation(
                PAID_GROUP_ID,
                gif_id,
                caption=f"ðŸ” *Confession #{CONFESSION_COUNTER}*\n\n{caption}" if caption else f"ðŸ” *Confession #{CONFESSION_COUNTER}*",
                parse_mode="Markdown"
            )
            
        # Log the confession
        logging.info(f"GIF confession #{CONFESSION_COUNTER} sent to group")
        
        # Send confirmation to user
        bot.send_message(user_id, "âœ… Your GIF confession has been sent anonymously.")
        
        # Admin record
        admin_record = f"ðŸ“ *Admin Log*\n\nGIF confession #{CONFESSION_COUNTER} was submitted by User ID: {user_id}"
        bot.send_message(CREATOR_ID, admin_record, parse_mode="Markdown")
    
    except Exception as e:
        logging.error(f"Error sending GIF confession: {e}")
        bot.send_message(user_id, "âŒ There was an error sending your GIF confession.")
    
    # Remove user from confessing dict
    USERS_CONFESSING.pop(user_id, None)

# Add handler for video confessions
@bot.message_handler(content_types=['video'], func=lambda message: message.chat.type == 'private' and 
                    message.from_user.id in USERS_CONFESSING and 
                    USERS_CONFESSING[message.from_user.id]['status'] == 'awaiting_confession')
def handle_video_confession(message):
    """Process a user's video confession"""
    user_id = message.from_user.id
    video_id = message.video.file_id
    caption = message.caption or ""
    
    # Increment the confession counter
    global CONFESSION_COUNTER
    CONFESSION_COUNTER += 1
    save_confession_counter(CONFESSION_COUNTER)
    
    try:
        # Send anonymously to the group
        if CONFESSION_TOPIC_ID:
            sent_message = bot.send_video(
                PAID_GROUP_ID,
                video_id,
                caption=f"ðŸ” *Confession #{CONFESSION_COUNTER}*\n\n{caption}" if caption else f"ðŸ” *Confession #{CONFESSION_COUNTER}*",
                parse_mode="Markdown",
                message_thread_id=CONFESSION_TOPIC_ID
            )
        else:
            sent_message = bot.send_video(
                PAID_GROUP_ID,
                video_id,
                caption=f"ðŸ” *Confession #{CONFESSION_COUNTER}*\n\n{caption}" if caption else f"ðŸ” *Confession #{CONFESSION_COUNTER}*",
                parse_mode="Markdown"
            )
            
        # Log the confession
        logging.info(f"Video confession #{CONFESSION_COUNTER} sent to group")
        
        # Send confirmation to user
        bot.send_message(user_id, "âœ… Your video confession has been sent anonymously.")
        
        # Admin record
        admin_record = f"ðŸ“ *Admin Log*\n\nVideo confession #{CONFESSION_COUNTER} was submitted by User ID: {user_id}"
        bot.send_message(CREATOR_ID, admin_record, parse_mode="Markdown")
    
    except Exception as e:
        logging.error(f"Error sending video confession: {e}")
        bot.send_message(user_id, "âŒ There was an error sending your video confession.")
    
    # Remove user from confessing dict
    USERS_CONFESSING.pop(user_id, None)

@bot.message_handler(commands=['refreshexpired'])
def refresh_expired_members(message):
    # Check if user is admin
    if message.from_user.id not in CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to Creator only.")
        return
        
    count = 0
    for user_id_str, data in PAYMENT_DATA.items():
        try:
            due_date = datetime.strptime(data['due_date'], '%Y-%m-%d %H:%M:%S')
            now = datetime.now()
            if due_date < now and not data.get('admin_action_pending', False):
                PAYMENT_DATA[user_id_str]['admin_action_pending'] = True
                count += 1
        except Exception as e:
            logging.error(f"Error processing user {user_id_str}: {e}")
    
    save_payment_data()  # Save changes to database
    bot.reply_to(message, f"âœ… Added admin_action_pending flag to {count} expired members.")

@bot.message_handler(commands=['discount'])
def start_discount_setup(message):
    """Start discount setup process"""
    # Check if user is admin or creator
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    # Start the discount setup process - first ask for the event name
    bot.reply_to(message, "ðŸ·ï¸ *Discount Setup*\n\nPlease enter the name of the discount event:", parse_mode="Markdown")
    
    # Set user state
    PENDING_USERS[message.from_user.id] = {'status': 'discount_event_name'}
    save_pending_users()

@bot.message_handler(commands=['remove_discount'])
def remove_discount(message):
    """Remove active discount"""
    # Check if user is admin or creator
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    args = message.text.split()
    
    if len(args) < 2:
        # Show both discounts status and ask which one to remove
        reg_status = "ðŸŸ¢ Active" if DISCOUNTS.get('regular') else "ðŸ”´ Not active"
        sup_status = "ðŸŸ¢ Active" if DISCOUNTS.get('supreme') else "ðŸ”´ Not active"
        
        markup = InlineKeyboardMarkup(row_width=1)
        if DISCOUNTS.get('regular'):
            markup.add(InlineKeyboardButton("Remove Regular Discount", callback_data="remove_discount_regular"))
        if DISCOUNTS.get('supreme'):
            markup.add(InlineKeyboardButton("Remove Supreme Discount", callback_data="remove_discount_supreme"))
        if DISCOUNTS.get('regular') and DISCOUNTS.get('supreme'):
            markup.add(InlineKeyboardButton("Remove Both Discounts", callback_data="remove_discount_both"))
        
        if not markup.keyboard:
            bot.reply_to(message, "âŒ There are no active discounts to remove.")
            return
            
        bot.reply_to(
            message, 
            f"ðŸ“Š *Current Discount Status*\n\n"
            f"â€¢ Regular Discount: {reg_status}\n"
            f"â€¢ Supreme Discount: {sup_status}\n\n"
            f"Please select which discount to remove:",
            parse_mode="Markdown",
            reply_markup=markup
        )
        return
    
    # If a specific discount type was specified in command
    discount_type = args[1].lower()
    
    if discount_type not in ['regular', 'supreme', 'both']:
        bot.reply_to(message, "âŒ Invalid discount type. Use 'regular', 'supreme', or 'both'.")
        return
    
    # Handle the removal based on type
    if discount_type == 'both':
        removed = []
        if DISCOUNTS.get('regular'):
            old_discount = DISCOUNTS['regular'].copy()
            DISCOUNTS['regular'] = None
            save_discount(None, 'regular')
            removed.append(f"Regular: {old_discount.get('name')} ({old_discount.get('percentage')}%)")
            
        if DISCOUNTS.get('supreme'):
            old_discount = DISCOUNTS['supreme'].copy()
            DISCOUNTS['supreme'] = None
            save_discount(None, 'supreme')
            removed.append(f"Supreme: {old_discount.get('name')} ({old_discount.get('percentage')}%)")
            
        if removed:
            bot.reply_to(message, f"âœ… Discounts removed successfully!\n\n{', '.join(removed)}")
        else:
            bot.reply_to(message, "âŒ There were no active discounts to remove.")
            
    else:  # regular or supreme
        if not DISCOUNTS.get(discount_type):
            bot.reply_to(message, f"âŒ There is no active {discount_type} discount to remove.")
            return
            
        # Remove the specific discount
        old_discount = DISCOUNTS[discount_type].copy()
        DISCOUNTS[discount_type] = None
        save_discount(None, discount_type)
        
        bot.reply_to(
            message, 
            f"âœ… {discount_type.capitalize()} discount removed successfully!\n\n"
            f"Removed discount: {old_discount['name']} ({old_discount['percentage']}% off)"
        )

# Add callback handler for discount removal buttons
@bot.callback_query_handler(func=lambda call: call.data.startswith("remove_discount_"))
def handle_remove_discount_callback(call):
    user_id = call.from_user.id
    
    # Check if user is admin or creator
    if user_id not in ADMIN_IDS and user_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    discount_type = call.data.split("_")[2]  # regular, supreme, or both
    
    if discount_type == "both":
        removed = []
        if DISCOUNTS.get('regular'):
            old_discount = DISCOUNTS['regular'].copy()
            DISCOUNTS['regular'] = None
            save_discount(None, 'regular')
            removed.append(f"Regular: {old_discount.get('name')} ({old_discount.get('percentage')}%)")
            
        if DISCOUNTS.get('supreme'):
            old_discount = DISCOUNTS['supreme'].copy()
            DISCOUNTS['supreme'] = None
            save_discount(None, 'supreme')
            removed.append(f"Supreme: {old_discount.get('name')} ({old_discount.get('percentage')}%)")
        
        # Update the message
        if removed:
            bot.edit_message_text(
                f"âœ… *Discounts Removed Successfully*\n\n" + "\n".join(removed),
                call.message.chat.id,
                call.message.message_id,
                parse_mode="Markdown"
            )
        else:
            bot.edit_message_text(
                "âŒ There were no active discounts to remove.",
                call.message.chat.id,
                call.message.message_id
            )
        
    else:  # regular or supreme
        if DISCOUNTS.get(discount_type):
            old_discount = DISCOUNTS[discount_type].copy()
            DISCOUNTS[discount_type] = None
            save_discount(None, discount_type)
            
            bot.edit_message_text(
                f"âœ… *{discount_type.capitalize()} Discount Removed*\n\n"
                f"Name: {old_discount['name']}\n"
                f"Percentage: {old_discount['percentage']}% off",
                call.message.chat.id,
                call.message.message_id,
                parse_mode="Markdown"
            )
        else:
            bot.edit_message_text(
                f"âŒ There is no active {discount_type} discount to remove.",
                call.message.chat.id,
                call.message.message_id
            )
    
    bot.answer_callback_query(call.id, f"Discount removal processed")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'discount_event_name')
def process_discount_event_name(message):
    user_id = message.from_user.id
    discount_name = message.text.strip()
    
    if len(discount_name) < 3:
        bot.send_message(user_id, "âŒ Discount name must be at least 3 characters long. Please try again:")
        return
    
    # Store the discount name
    PENDING_USERS[user_id]['discount_name'] = discount_name
    PENDING_USERS[user_id]['status'] = 'discount_regular_percentage'
    save_pending_users()
    
    # Ask for regular mentorship percentage
    bot.send_message(user_id, "ðŸ”¢ What is the discount percentage for *Regular Mentorship*?\n\nPlease enter a number between 1 and 99:", parse_mode="Markdown")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'discount_regular_percentage')
def process_discount_regular_percentage(message):
    user_id = message.from_user.id
    
    try:
        # Try to parse the percentage
        percentage = int(message.text.strip())
        
        # Validate percentage range
        if percentage < 1 or percentage > 99:
            bot.send_message(user_id, "âŒ Percentage must be between 1 and 99. Please try again:")
            return
        
        # Store the regular percentage
        PENDING_USERS[user_id]['regular_percentage'] = percentage
        PENDING_USERS[user_id]['status'] = 'discount_supreme_percentage'
        save_pending_users()
        
        # Ask for supreme membership percentage
        bot.send_message(user_id, "ðŸ”¢ What is the discount percentage for *Supreme Mentorship*?\n\nPlease enter a number between 1 and 99:", parse_mode="Markdown")
        
    except ValueError:
        bot.send_message(user_id, "âŒ Invalid percentage. Please enter a number between 1 and 99:")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'discount_supreme_percentage')
def process_discount_supreme_percentage(message):
    user_id = message.from_user.id
    
    try:
        # Try to parse the percentage
        percentage = int(message.text.strip())
        
        # Validate percentage range
        if percentage < 1 or percentage > 99:
            bot.send_message(user_id, "âŒ Percentage must be between 1 and 99. Please try again:")
            return
        
        # Store the supreme percentage
        PENDING_USERS[user_id]['supreme_percentage'] = percentage
        PENDING_USERS[user_id]['status'] = 'discount_regular_expiry'
        save_pending_users()
        
        # Ask for regular mentorship expiry
        bot.send_message(
            user_id, 
            "ðŸ“… When will the *Regular Mentorship* discount end?\n\nPlease select the duration:",
            parse_mode="Markdown",
            reply_markup=create_duration_keyboard()
        )
        
    except ValueError:
        bot.send_message(user_id, "âŒ Invalid percentage. Please enter a number between 1 and 99:")

# Helper function to create duration keyboard
def create_duration_keyboard():
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row("1 day", "3 days", "1 week")
    markup.row("2 weeks", "1 month", "3 months")
    markup.row("Custom date")
    return markup

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'discount_regular_expiry')
def process_discount_regular_expiry(message):
    user_id = message.from_user.id
    expiry_text = message.text.strip()
    
    # Calculate end date based on selection
    end_date = calculate_end_date(expiry_text)
    
    if end_date:
        # Store the regular expiry date
        PENDING_USERS[user_id]['regular_end_date'] = end_date.strftime('%Y-%m-%d %H:%M:%S')
        PENDING_USERS[user_id]['status'] = 'discount_supreme_expiry'
        save_pending_users()
        
        # Ask for supreme mentorship expiry
        bot.send_message(
            user_id, 
            "ðŸ“… When will the *Supreme Mentorship* discount end?\n\nPlease select the duration:",
            parse_mode="Markdown",
            reply_markup=create_duration_keyboard()
        )
    elif expiry_text == "Custom date":
        PENDING_USERS[user_id]['status'] = 'discount_regular_custom_date'
        save_pending_users()
        bot.send_message(user_id, "ðŸ“… Please enter the end date for Regular Membership in format YYYY-MM-DD HH:MM:SS\n\nFor example: 2025-05-15 23:59:59")
    else:
        bot.send_message(user_id, "âŒ Invalid selection. Please choose from the keyboard or enter 'Custom date'.")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'discount_regular_custom_date')
def process_discount_regular_custom_date(message):
    user_id = message.from_user.id
    end_date_str = message.text.strip()
    
    try:
        # Try to parse the date
        naive_end_date = datetime.strptime(end_date_str, '%Y-%m-%d %H:%M:%S')
        
        # Make it timezone-aware in Manila timezone
        manila_tz = pytz.timezone('Asia/Manila')
        end_date = manila_tz.localize(naive_end_date)
        
        # Check if date is in the future
        now = datetime.now(manila_tz)
        if end_date <= now:
            bot.send_message(user_id, "âŒ End date must be in the future. Please try again:")
            return
        
        # Store the regular expiry date
        PENDING_USERS[user_id]['regular_end_date'] = end_date.strftime('%Y-%m-%d %H:%M:%S')
        PENDING_USERS[user_id]['status'] = 'discount_supreme_expiry'
        save_pending_users()
        
        # Ask for supreme mentorship expiry
        bot.send_message(
            user_id, 
            "ðŸ“… When will the *Supreme Mentorship* discount end?\n\nPlease select the duration:",
            parse_mode="Markdown",
            reply_markup=create_duration_keyboard()
        )
        
    except ValueError:
        bot.send_message(user_id, "âŒ Invalid date format. Please enter in format YYYY-MM-DD HH:MM:SS")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'discount_supreme_expiry')
def process_discount_supreme_expiry(message):
    user_id = message.from_user.id
    expiry_text = message.text.strip()
    
    # Calculate end date based on selection
    end_date = calculate_end_date(expiry_text)
    
    if end_date:
        # Store the supreme expiry date
        PENDING_USERS[user_id]['supreme_end_date'] = end_date.strftime('%Y-%m-%d %H:%M:%S')
        PENDING_USERS[user_id]['status'] = 'discount_regular_limit'
        save_pending_users()
        
        # Ask for regular mentorship user limit
        bot.send_message(user_id, "ðŸ‘¥ How many users can use the *Regular Mentorship* discount? Enter a number, or type 'unlimited' for no limit:", parse_mode="Markdown")
    elif expiry_text == "Custom date":
        PENDING_USERS[user_id]['status'] = 'discount_supreme_custom_date'
        save_pending_users()
        bot.send_message(user_id, "ðŸ“… Please enter the end date for Supreme Membership in format YYYY-MM-DD HH:MM:SS\n\nFor example: 2025-05-15 23:59:59")
    else:
        bot.send_message(user_id, "âŒ Invalid selection. Please choose from the keyboard or enter 'Custom date'.")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'discount_supreme_custom_date')
def process_discount_supreme_custom_date(message):
    user_id = message.from_user.id
    end_date_str = message.text.strip()
    
    try:
        # Try to parse the date
        naive_end_date = datetime.strptime(end_date_str, '%Y-%m-%d %H:%M:%S')
        
        # Make it timezone-aware in Manila timezone
        manila_tz = pytz.timezone('Asia/Manila')
        end_date = manila_tz.localize(naive_end_date)
        
        # Check if date is in the future
        now = datetime.now(manila_tz)
        if end_date <= now:
            bot.send_message(user_id, "âŒ End date must be in the future. Please try again:")
            return
        
        # Store the supreme expiry date
        PENDING_USERS[user_id]['supreme_end_date'] = end_date.strftime('%Y-%m-%d %H:%M:%S')
        PENDING_USERS[user_id]['status'] = 'discount_regular_limit'
        save_pending_users()
        
        # Ask for regular mentorship user limit
        bot.send_message(user_id, "ðŸ‘¥ How many users can use the *Regular Mentorship* discount? Enter a number, or type 'unlimited' for no limit:", parse_mode="Markdown")
        
    except ValueError:
        bot.send_message(user_id, "âŒ Invalid date format. Please enter in format YYYY-MM-DD HH:MM:SS")

# Helper function to calculate end date based on duration
def calculate_end_date(duration_text):
    now = datetime.now(pytz.timezone('Asia/Manila'))
    
    if duration_text == "1 day":
        return now + timedelta(days=1)
    elif duration_text == "3 days":
        return now + timedelta(days=3)
    elif duration_text == "1 week":
        return now + timedelta(days=7)
    elif duration_text == "2 weeks":
        return now + timedelta(days=14)
    elif duration_text == "1 month":
        return now + timedelta(days=30)
    elif duration_text == "3 months":
        return now + timedelta(days=90)
    elif duration_text == "Custom date":
        return None
    else:
        return None

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'discount_regular_limit')
def process_discount_regular_limit(message):
    user_id = message.from_user.id
    limit_text = message.text.strip().lower()
    
    # Process the limit input
    if limit_text == 'unlimited':
        user_limit = None  # No limit
    else:
        try:
            user_limit = int(limit_text)
            if user_limit < 1:
                bot.send_message(user_id, "âŒ User limit must be at least 1. Please try again, or type 'unlimited':")
                return
        except ValueError:
            bot.send_message(user_id, "âŒ Invalid input. Please enter a number or type 'unlimited':")
            return
    
    # Store the regular user limit
    PENDING_USERS[user_id]['regular_user_limit'] = user_limit
    PENDING_USERS[user_id]['status'] = 'discount_supreme_limit'
    save_pending_users()
    
    # Ask for supreme mentorship user limit
    bot.send_message(user_id, "ðŸ‘¥ How many users can use the *Supreme Mentorship* discount? Enter a number, or type 'unlimited' for no limit:", parse_mode="Markdown")

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'discount_supreme_limit')
def process_discount_supreme_limit(message):
    user_id = message.from_user.id
    limit_text = message.text.strip().lower()
    
    # Process the limit input
    if limit_text == 'unlimited':
        user_limit = None  # No limit
    else:
        try:
            user_limit = int(limit_text)
            if user_limit < 1:
                bot.send_message(user_id, "âŒ User limit must be at least 1. Please try again, or type 'unlimited':")
                return
        except ValueError:
            bot.send_message(user_id, "âŒ Invalid input. Please enter a number or type 'unlimited':")
            return
    
    # Store the supreme user limit
    PENDING_USERS[user_id]['supreme_user_limit'] = user_limit
    PENDING_USERS[user_id]['status'] = 'discount_regular_transaction_type'
    save_pending_users()
    
    # Ask for regular mentorship transaction type with a keyboard
    markup = ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
    markup.add(KeyboardButton("New Purchases Only"))
    markup.add(KeyboardButton("Renewals Only"))
    markup.add(KeyboardButton("Both New Purchases & Renewals"))
    
    bot.send_message(
        user_id, 
        "ðŸ›ï¸ Who can use the *Regular Mentorship* discount?\n\nPlease select which transaction types this discount applies to:",
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'discount_regular_transaction_type')
def process_discount_regular_transaction_type(message):
    user_id = message.from_user.id
    transaction_type = message.text.strip()
    
    # Validate the input
    valid_types = ["New Purchases Only", "Renewals Only", "Both New Purchases & Renewals"]
    if transaction_type not in valid_types:
        bot.send_message(user_id, "âŒ Invalid option. Please select from the keyboard.")
        return
    
    # Determine transaction type code
    if transaction_type == "New Purchases Only":
        transaction_code = "new"
    elif transaction_type == "Renewals Only":
        transaction_code = "renewal"
    else:  # Both
        transaction_code = "both"
    
    # Store the regular transaction type
    PENDING_USERS[user_id]['regular_transaction_type'] = transaction_code
    PENDING_USERS[user_id]['status'] = 'discount_supreme_transaction_type'
    save_pending_users()
    
    # Ask for supreme mentorship transaction type with a keyboard
    markup = ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
    markup.add(KeyboardButton("New Purchases Only"))
    markup.add(KeyboardButton("Renewals Only"))
    markup.add(KeyboardButton("Both New Purchases & Renewals"))
    
    bot.send_message(
        user_id, 
        "ðŸ›ï¸ Who can use the *Supreme Mentorship* discount?\n\nPlease select which transaction types this discount applies to:",
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'discount_supreme_transaction_type')
def process_discount_supreme_transaction_type(message):
    user_id = message.from_user.id
    transaction_type = message.text.strip()
    
    # Validate the input
    valid_types = ["New Purchases Only", "Renewals Only", "Both New Purchases & Renewals"]
    if transaction_type not in valid_types:
        bot.send_message(user_id, "âŒ Invalid option. Please select from the keyboard.")
        return
    
    # Determine transaction type code
    if transaction_type == "New Purchases Only":
        transaction_code = "new"
    elif transaction_type == "Renewals Only":
        transaction_code = "renewal"
    else:  # Both
        transaction_code = "both"
    
    # Store the supreme transaction type
    PENDING_USERS[user_id]['supreme_transaction_type'] = transaction_code
    PENDING_USERS[user_id]['status'] = 'discount_custom_message'
    save_pending_users()
    
    # Ask for custom announcement message
    bot.send_message(
        user_id, 
        "ðŸ“¢ *Add Custom Announcement Message* (Optional)\n\n"
        "Please enter a custom message to include in the discount announcement to all members.\n\n"
        "This could be additional details, special instructions, or promotional text.\n\n"
        "Type your message or send 'skip' to use the default announcement format.",
        parse_mode="Markdown"
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'discount_custom_message')
def process_discount_custom_message(message):
    user_id = message.from_user.id
    custom_message = message.text.strip()
    
    # Handle skip option
    if custom_message.lower() == 'skip':
        custom_message = ""
    
    # Store the custom message
    PENDING_USERS[user_id]['custom_message'] = custom_message
    PENDING_USERS[user_id]['status'] = 'discount_finalize'
    save_pending_users()
    
    # Create discount objects for both membership types
    discount_name = PENDING_USERS[user_id]['discount_name']
    custom_message = PENDING_USERS[user_id]['custom_message']
    
    # Regular Membership Discount
    regular_discount = {
        'name': discount_name,
        'end_date': PENDING_USERS[user_id]['regular_end_date'],
        'percentage': PENDING_USERS[user_id]['regular_percentage'],
        'user_limit': PENDING_USERS[user_id]['regular_user_limit'],
        'transaction_type': PENDING_USERS[user_id]['regular_transaction_type'],
        'users_used': [],
        'active': True,
        'custom_message': custom_message,
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'created_by': user_id
    }
    
    # Supreme Membership Discount
    supreme_discount = {
        'name': discount_name,
        'end_date': PENDING_USERS[user_id]['supreme_end_date'],
        'percentage': PENDING_USERS[user_id]['supreme_percentage'],
        'user_limit': PENDING_USERS[user_id]['supreme_user_limit'],
        'transaction_type': PENDING_USERS[user_id]['supreme_transaction_type'],
        'users_used': [],
        'active': True,
        'custom_message': custom_message,
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'created_by': user_id
    }
    
    # Update global discounts
    global DISCOUNTS
    DISCOUNTS['regular'] = regular_discount
    DISCOUNTS['supreme'] = supreme_discount
    
    # Save to database
    save_discount(regular_discount, 'regular')
    save_discount(supreme_discount, 'supreme')
    
    # Clear user from pending
    PENDING_USERS.pop(user_id, None)
    delete_pending_user(user_id)
    
    # Remove custom keyboard
    markup = ReplyKeyboardRemove()
    
    # Send confirmation message
    reg_end_date = datetime.strptime(regular_discount['end_date'], '%Y-%m-%d %H:%M:%S')
    sup_end_date = datetime.strptime(supreme_discount['end_date'], '%Y-%m-%d %H:%M:%S')
    
    reg_limit = "No limit" if regular_discount['user_limit'] is None else f"{regular_discount['user_limit']} users"
    sup_limit = "No limit" if supreme_discount['user_limit'] is None else f"{supreme_discount['user_limit']} users"
    
    # Get transaction type displays
    reg_type_display = get_transaction_type_display(regular_discount['transaction_type'])
    sup_type_display = get_transaction_type_display(supreme_discount['transaction_type'])
    
    confirmation_message = (
        f"ðŸŽ‰ *Discount Created Successfully!*\n\n"
        f"*Discount Name:* {discount_name}\n\n"
        f"*REGULAR MENTORSHIP*\n"
        f"â€¢ Discount: {regular_discount['percentage']}% off\n"
        f"â€¢ User Limit: {reg_limit}\n"
        f"â€¢ Applies to: {reg_type_display}\n"
        f"â€¢ Ends: {reg_end_date.strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        f"*SUPREME MENTORSHIP*\n"
        f"â€¢ Discount: {supreme_discount['percentage']}% off\n"
        f"â€¢ User Limit: {sup_limit}\n"
        f"â€¢ Applies to: {sup_type_display}\n"
        f"â€¢ Ends: {sup_end_date.strftime('%Y-%m-%d %H:%M:%S')}\n\n"
    )
    
    # Add custom message in the confirmation if provided
    if custom_message:
        confirmation_message += f"*Custom Message:*\n\"{custom_message}\"\n\n"
    
    confirmation_message += "Both discounts are now active!"
    
    bot.send_message(
        user_id, 
        confirmation_message,
        parse_mode="Markdown",
        reply_markup=markup
    )
    notify_discount_created(discount_name, regular_discount, supreme_discount)
    # Generate announcement message for preview
    announcement = create_discount_announcement(discount_name, regular_discount, supreme_discount)
    
    # Send announcement preview
    bot.send_message(
        user_id,
        f"*Preview of announcement:*\n\n{announcement}",
        parse_mode="Markdown"
    )
    
    # Automatically announce the discount to the announcement topic
    try:
        if ANNOUNCEMENT_TOPIC_ID:
            sent_msg = bot.send_message(
                PAID_GROUP_ID,
                announcement,
                parse_mode="Markdown",
                message_thread_id=ANNOUNCEMENT_TOPIC_ID
            )
            bot.send_message(
                user_id, 
                f"âœ… Discount has been automatically announced in the announcements topic!"
            )
            logging.info(f"Discount '{discount_name}' announced to announcement topic automatically")
        else:
            # If no announcement topic is set, send to main group
            sent_msg = bot.send_message(
                PAID_GROUP_ID,
                announcement,
                parse_mode="Markdown"
            )
            bot.send_message(
                user_id, 
                f"âœ… Discount has been announced in the main group! (No announcement topic configured)"
            )
            logging.info(f"Discount '{discount_name}' announced to main group (no topic configured)")
    except Exception as e:
        bot.send_message(
            user_id,
            f"âŒ Failed to automatically announce the discount: {str(e)}\n\nThe discount is still active, but you'll need to announce it manually."
        )
        logging.error(f"Failed to announce discount: {e}")
    
    # Notify all admins about the new discount
    admin_username = "Unknown"
    try:
        user_info = bot.get_chat(user_id)
        admin_username = user_info.username or f"Admin {user_id}"
    except:
        pass
        
    for admin_id in ADMIN_IDS:
        if admin_id != user_id:  # Don't notify the admin who created it
            try:
                bot.send_message(
                    admin_id,
                    f"ðŸ“ *New Discount Created*\n\n"
                    f"@{admin_username} has created a new discount:\n"
                    f"â€¢ Name: {discount_name}\n"
                    f"â€¢ Regular: {regular_discount['percentage']}% off (ends {reg_end_date.strftime('%Y-%m-%d')})\n"
                    f"â€¢ Supreme: {supreme_discount['percentage']}% off (ends {sup_end_date.strftime('%Y-%m-%d')})\n\n"
                    f"The discount is now active and has been automatically announced.",
                    parse_mode="Markdown"
                )
            except Exception as e:
                logging.error(f"Failed to notify admin {admin_id} about new discount: {e}")

# Helper function to create announcement message
def create_discount_announcement(discount_name, regular_discount, supreme_discount):
    reg_end_date = datetime.strptime(regular_discount['end_date'], '%Y-%m-%d %H:%M:%S')
    sup_end_date = datetime.strptime(supreme_discount['end_date'], '%Y-%m-%d %H:%M:%S')
    
    reg_limit = "Unlimited" if regular_discount['user_limit'] is None else f"Limited to {regular_discount['user_limit']} users"
    sup_limit = "Unlimited" if supreme_discount['user_limit'] is None else f"Limited to {supreme_discount['user_limit']} users"
    
    # Get transaction type info
    reg_transaction_info = get_transaction_note(regular_discount['transaction_type'])
    sup_transaction_info = get_transaction_note(supreme_discount['transaction_type'])
    
    # Get custom message (will be the same for both discounts)
    custom_message = regular_discount.get('custom_message', '')
    
    announcement = (
        f"ðŸ”¥ *SPECIAL DISCOUNT ALERT!* ðŸ”¥\n\n"
        f"ðŸ“¢ **{discount_name}**\n\n"
    )
    
    # Add custom message if provided
    if custom_message:
        announcement += f"{custom_message}\n\n"
    
    announcement += (
        f"*REGULAR MEMBERSHIP PLANS*\n"
        f"ðŸ’° **{regular_discount['percentage']}% OFF** on all regular membership plans!\n"
        f"â° Valid until: {reg_end_date.strftime('%B %d, %Y at %I:%M %p')}\n"
        f"ðŸ‘¥ {reg_limit}\n"
    )
    
    # Add regular transaction info if not "both"
    if reg_transaction_info:
        announcement += f"{reg_transaction_info}\n"
    
    announcement += (
        f"\n*SUPREME MEMBERSHIP PLANS*\n"
        f"ðŸ’° **{supreme_discount['percentage']}% OFF** on all supreme membership plans!\n"
        f"â° Valid until: {sup_end_date.strftime('%B %d, %Y at %I:%M %p')}\n"
        f"ðŸ‘¥ {sup_limit}\n"
    )
    
    # Add supreme transaction info if not "both"
    if sup_transaction_info:
        announcement += f"{sup_transaction_info}\n"
    
    announcement += (
        f"\nðŸ”¸ Regular plans include: Trial, Momentum & Legacy\n"
        f"ðŸ”¸ Supreme plans include: Apprentice, Disciple & Lifetime\n\n"
        f"Act fast! Use the /start command in a private message with the bot to take advantage of this limited-time offer!"
    )
    
    return announcement

# Helper functions for transaction type display
def get_transaction_type_display(transaction_type):
    if transaction_type == "new":
        return "New Purchases Only"
    elif transaction_type == "renewal":
        return "Renewals Only"
    else:
        return "Both New Purchases & Renewals"

def get_transaction_note(transaction_type):
    if transaction_type == "new":
        return "ðŸ†• *Available for new purchases only*"
    elif transaction_type == "renewal":
        return "ðŸ”„ *Available for membership renewals only*"
    return ""  # Empty for "both"

# Update the discount expiry check function to handle both discount types
def check_discount_expiry():
    """Check if the current discounts have expired and remove them if needed"""
    global DISCOUNTS
    
    for discount_type in ['regular', 'supreme']:
        discount = DISCOUNTS.get(discount_type)
        if discount and discount.get('active'):
            try:
                # Parse the stored end date (naive datetime)
                naive_end_date = datetime.strptime(discount.get('end_date'), '%Y-%m-%d %H:%M:%S')
                
                # Make it timezone-aware by adding Manila timezone
                manila_tz = pytz.timezone('Asia/Manila')
                end_date = manila_tz.localize(naive_end_date)
                
                # Compare with current time in Manila timezone
                now = datetime.now(manila_tz)
                
                if now >= end_date:
                    # Discount has expired
                    old_discount = discount.copy()
                    DISCOUNTS[discount_type] = None
                    save_discount(None, discount_type)
                    logging.info(f"{discount_type.capitalize()} discount '{old_discount.get('name')}' has expired and has been removed at {now.strftime('%Y-%m-%d %H:%M:%S %Z')}")
                    
                    # Notify admins about expired discount
                    for admin_id in ADMIN_IDS:
                        try:
                            bot.send_message(admin_id, 
                                f"ðŸ•’ *{discount_type.capitalize()} Discount Expired*\n\n"
                                f"The '{old_discount.get('name')}' discount ({old_discount.get('percentage')}% off) "
                                f"for {discount_type.capitalize()} memberships has ended and has been automatically removed.\n\n"
                                f"Expired at: {end_date.strftime('%Y-%m-%d %I:%M:%S %p')} Manila time",
                                parse_mode="Markdown"
                            )
                        except Exception as e:
                            logging.error(f"Failed to notify admin {admin_id} about expired discount: {e}")
            except Exception as e:
                logging.error(f"Error checking {discount_type} discount expiry: {e}")

def discount_expiry_thread():
    """Thread to check for expired discounts"""
    logging.info("Discount expiry check thread started")
    while True:
        try:
            check_discount_expiry()
            time.sleep(60)  # Check every minute
        except Exception as e:
            logging.error(f"Error in discount expiry thread: {e}")
            time.sleep(60)  # Wait a minute on error before trying again

@bot.message_handler(commands=['export_forms'])
def export_form_responses(message):
    """Export onboarding form responses to a professionally formatted Excel file"""
    # Check if user is admin or creator
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    try:
        # Send processing message with more detail
        processing_msg = bot.reply_to(message, "ðŸ“Š *Generating Form Responses Export*\n\nCollecting data and creating beautiful report...", parse_mode="Markdown")
        
        # Collect form data from PENDING_USERS and completed forms
        form_records = []
        
        # First check PENDING_USERS for users with form responses
        for user_id, data in PENDING_USERS.items():
            if 'form_answers' in data and data['form_answers']:
                # Get user info
                try:
                    user_info = bot.get_chat(user_id)
                    username = user_info.username or "No Username"
                    first_name = user_info.first_name or ""
                    last_name = user_info.last_name or ""
                except:
                    username = "Unknown"
                    first_name = ""
                    last_name = ""
                
                # Get membership type to determine which fields to extract
                membership_type = data.get('membership_type', 'regular').lower()
                form_answers = data['form_answers']
                
                # Create base record with user info
                record = {
                    'user_id': user_id,
                    'username': username,
                    'first_name': first_name,
                    'last_name': last_name,
                    'mentorship_type': membership_type,
                    'form_completion_date': datetime.now().strftime('%Y-%m-%d')
                }
                
                # Add plan info if available
                if str(user_id) in PAYMENT_DATA:
                    record['plan'] = PAYMENT_DATA[str(user_id)].get('payment_plan', 'Unknown')
                else:
                    record['plan'] = 'Unknown'
                
                # Extract form answers based on membership type
                if membership_type == 'supreme':
                    # Supreme form fields
                    record.update({
                        'full_name': form_answers.get('full_name', 'Not provided'),
                        'birthday': form_answers.get('birthday', 'Not provided'),
                        'phone_number': form_answers.get('phone_number', 'Not provided'),
                        'time_zone': form_answers.get('time_zone', 'Not provided'),
                        'expertise_level': form_answers.get('expertise_level', 'Not provided'),
                        'trading_time': form_answers.get('trading_time_commitment', 'Not provided'),
                        'interest_reason': form_answers.get('interest_reason', 'Not provided'),
                        'personal_goals': form_answers.get('personal_goals', 'Not provided'),
                        'call_preference': form_answers.get('call_preference', 'Not provided'),
                        'challenges': form_answers.get('challenges', 'Not provided')
                    })
                else:
                    # Regular form fields
                    record.update({
                        'full_name': form_answers.get('full_name', 'Not provided'),
                        'age_birth_year': form_answers.get('age_birth_year', 'Not provided'),
                        'experience_level': form_answers.get('experience_level', 'Not provided'),
                        'learning_goals': form_answers.get('learning_goals', 'Not provided'),
                        'source': form_answers.get('source', 'Not provided')
                    })
                
                form_records.append(record)
        
        # Now check for any users in PAYMENT_DATA that might have form responses stored
        for user_id_str, payment_data in PAYMENT_DATA.items():
            # Check if we already have this user from PENDING_USERS
            user_id = int(user_id_str)
            if user_id in [r['user_id'] for r in form_records]:
                continue
                
            # Check if this user has form data stored in their payment record
            if 'form_answers' in payment_data:
                # Get user info
                username = payment_data.get('username', 'No Username')
                
                # Get membership type
                membership_type = payment_data.get('mentorship_type', 'regular').lower()
                form_answers = payment_data['form_answers']
                
                # Create base record
                record = {
                    'user_id': user_id,
                    'username': username,
                    'first_name': '',  # Not available in payment data
                    'last_name': '',   # Not available in payment data
                    'mentorship_type': membership_type,
                    'plan': payment_data.get('payment_plan', 'Unknown'),
                    'form_completion_date': payment_data.get('form_completion_date', 'Unknown')
                }
                
                # Extract form answers based on membership type
                if membership_type == 'supreme':
                    # Supreme form fields
                    record.update({
                        'full_name': form_answers.get('full_name', 'Not provided'),
                        'birthday': form_answers.get('birthday', 'Not provided'),
                        'phone_number': form_answers.get('phone_number', 'Not provided'),
                        'time_zone': form_answers.get('time_zone', 'Not provided'),
                        'expertise_level': form_answers.get('expertise_level', 'Not provided'),
                        'trading_time': form_answers.get('trading_time_commitment', 'Not provided'),
                        'interest_reason': form_answers.get('interest_reason', 'Not provided'),
                        'personal_goals': form_answers.get('personal_goals', 'Not provided'),
                        'call_preference': form_answers.get('call_preference', 'Not provided'),
                        'challenges': form_answers.get('challenges', 'Not provided')
                    })
                else:
                    # Regular form fields
                    record.update({
                        'full_name': form_answers.get('full_name', 'Not provided'),
                        'age_birth_year': form_answers.get('age_birth_year', 'Not provided'),
                        'experience_level': form_answers.get('experience_level', 'Not provided'),
                        'learning_goals': form_answers.get('learning_goals', 'Not provided'),
                        'source': form_answers.get('source', 'Not provided')
                    })
                
                form_records.append(record)
        
        # Check if we have data to export
        if not form_records:
            bot.edit_message_text("âŒ No form responses found to export.", 
                                 chat_id=message.chat.id, 
                                 message_id=processing_msg.message_id)
            return
        
        # Create DataFrame for processing
        df = pd.DataFrame(form_records)
        
        # Sort by membership type and username
        df = df.sort_values(by=['mentorship_type', 'username'])
        
        # Update processing message
        bot.edit_message_text("ðŸ“Š Creating beautifully formatted Excel workbook...", 
                             chat_id=message.chat.id, 
                             message_id=processing_msg.message_id)
        
        # Create workbook
        wb = Workbook()
        
        # Create cover sheet (Dashboard)
        dashboard = wb.active
        dashboard.title = "Dashboard"
        
        # Define modern color scheme
        colors = {
            'primary': '4472C4',      # Modern blue
            'secondary': '5B9BD5',     # Lighter blue
            'accent1': '70AD47',      # Green
            'accent2': 'ED7D31',      # Orange
            'accent3': 'FFC000',      # Yellow
            'light_bg': 'F2F2F2',     # Light gray
            'dark_text': '262626',    # Dark gray
            'header_bg': '203864',    # Dark blue
            'alt_row': 'E6F0FF'       # Very light blue
        }
        
        # ---- DASHBOARD SHEET ----
        
        # Add logo placeholder and title
        dashboard.merge_cells('A1:H1')
        title_cell = dashboard['A1']
        title_cell.value = "PRODIGY TRADING ACADEMY"
        title_cell.font = Font(size=28, bold=True, color=colors['dark_text'], name='Calibri')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        dashboard.row_dimensions[1].height = 45
        
        # Add subtitle
        dashboard.merge_cells('A2:H2')
        subtitle_cell = dashboard['A2']
        subtitle_cell.value = "Member Onboarding Form Responses"
        subtitle_cell.font = Font(size=16, italic=True, color=colors['dark_text'], name='Calibri')
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add export date
        dashboard.merge_cells('A3:H3')
        date_cell = dashboard['A3']
        date_cell.value = f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        date_cell.font = Font(size=12, color=colors['dark_text'], name='Calibri')
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add some space
        dashboard.row_dimensions[4].height = 20
        
        # Calculate summary statistics
        regular_members = len([r for r in form_records if r.get('mentorship_type') == 'regular'])
        supreme_members = len([r for r in form_records if r.get('mentorship_type') == 'supreme'])
        total_members = len(form_records)
        
        experience_levels = Counter([r.get('expertise_level', '').split('.')[0].strip() if r.get('expertise_level') else '' 
                                   for r in form_records])
        
        # Add summary section header
        dashboard.merge_cells('B5:G5')
        summary_header = dashboard['B5']
        summary_header.value = "MEMBERSHIP SUMMARY"
        summary_header.font = Font(size=14, bold=True, color="FFFFFF", name='Calibri')
        summary_header.alignment = Alignment(horizontal='center', vertical='center')
        summary_header.fill = PatternFill(start_color=colors['header_bg'], end_color=colors['header_bg'], fill_type="solid")
        
        # Add key statistics in a modern card layout
        stats = [
            ["Total Members", total_members, colors['primary']],
            ["Regular Members", regular_members, colors['secondary']],
            ["Supreme Members", supreme_members, colors['accent1']]
        ]
        
        # Create stat cards
        for i, (label, value, color) in enumerate(stats):
            col = chr(ord('B') + i*2)  # B, D, F
            
            # Merge cells for each stat card
            dashboard.merge_cells(f'{col}6:{chr(ord(col)+1)}7')
            dashboard.merge_cells(f'{col}8:{chr(ord(col)+1)}9')
            
            # Add label
            label_cell = dashboard[f'{col}6']
            label_cell.value = label
            label_cell.font = Font(size=12, color="FFFFFF", name='Calibri')
            label_cell.alignment = Alignment(horizontal='center', vertical='center')
            label_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            # Add value
            value_cell = dashboard[f'{col}8']
            value_cell.value = value
            value_cell.font = Font(size=24, bold=True, color=colors['dark_text'], name='Calibri')
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add border to value cell
            thin_border = Border(
                left=Side(style='thin', color=color),
                right=Side(style='thin', color=color),
                top=Side(style='thin', color=color),
                bottom=Side(style='thin', color=color)
            )
            value_cell.border = thin_border
        
        # Add some space
        dashboard.row_dimensions[10].height = 30
        
        # Add navigation section
        dashboard.merge_cells('B11:G11')
        nav_header = dashboard['B11']
        nav_header.value = "NAVIGATION"
        nav_header.font = Font(size=14, bold=True, color="FFFFFF", name='Calibri')
        nav_header.alignment = Alignment(horizontal='center', vertical='center')
        nav_header.fill = PatternFill(start_color=colors['header_bg'], end_color=colors['header_bg'], fill_type="solid")
        
        # Add sheet links (note: these aren't actual hyperlinks but act as a visual guide)
        nav_items = [
            ["Regular Members", "View detailed information for all regular members"],
            ["Supreme Members", "View detailed information for all supreme members"]
        ]
        
        for i, (name, desc) in enumerate(nav_items):
            row = 12 + i
            
            # Sheet name
            cell_b = dashboard[f'B{row}']
            cell_b.value = name
            cell_b.font = Font(size=12, bold=True, color=colors['primary'], underline="single", name='Calibri')
            
            # Description
            dashboard.merge_cells(f'C{row}:G{row}')
            cell_c = dashboard[f'C{row}']
            cell_c.value = desc
            cell_c.font = Font(size=11, color=colors['dark_text'], name='Calibri')
        
        # Add some space and copyright/footer
        dashboard.row_dimensions[15].height = 30
        
        dashboard.merge_cells('B16:G16')
        footer = dashboard['B16']
        footer.value = "Confidential - For Administrative Use Only"
        footer.font = Font(size=10, italic=True, color=colors['dark_text'], name='Calibri')
        footer.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths for dashboard
        for col in ['A', 'H']:
            dashboard.column_dimensions[col].width = 2  # Margin columns
        
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            dashboard.column_dimensions[col].width = 15
        
        # ---- REGULAR MEMBERS SHEET ----
        ws_regular = wb.create_sheet(title="Regular Members")
        
        # Regular membership headers
        regular_headers = [
            'User ID', 'Username', 'Full Name', 'Age/Birth Year', 
            'Experience Level', 'Learning Goals', 'Found Us Via',
            'Membership Plan', 'Completion Date'
        ]
        
        # Add table header for regular sheet
        for col_idx, header in enumerate(regular_headers, 1):
            cell = ws_regular.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=12, name='Calibri')
            cell.fill = PatternFill(start_color=colors['header_bg'], end_color=colors['header_bg'], fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                bottom=Side(border_style="medium", color="FFFFFF")
            )
        
        # Make the header row taller
        ws_regular.row_dimensions[1].height = 30
            
        # ---- SUPREME MEMBERS SHEET ----
        ws_supreme = wb.create_sheet(title="Supreme Members")
        
        # Supreme membership headers
        supreme_headers = [
            'User ID', 'Username', 'Full Name', 'Birthday', 'Phone Number',
            'Time Zone', 'Expertise Level', 'Trading Time', 'Interest Reason',
            'Personal Goals', 'Call Preference', 'Challenges', 
            'Membership Plan', 'Completion Date'
        ]
        
        # Add table header for supreme sheet with the same styling
        for col_idx, header in enumerate(supreme_headers, 1):
            cell = ws_supreme.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=12, name='Calibri')
            cell.fill = PatternFill(start_color=colors['header_bg'], end_color=colors['header_bg'], fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                bottom=Side(border_style="medium", color="FFFFFF")
            )
        
        # Make the header row taller
        ws_supreme.row_dimensions[1].height = 30
        
        # Split data by membership type and add to appropriate sheet
        regular_row = 2
        supreme_row = 2
        
        # Update processing message
        bot.edit_message_text("ðŸ“Š Populating data with modern formatting...", 
                             chat_id=message.chat.id, 
                             message_id=processing_msg.message_id)
        
        for record in form_records:
            if record.get('mentorship_type') == 'supreme':
                # Add to supreme sheet
                ws_supreme.cell(row=supreme_row, column=1, value=record.get('user_id'))
                ws_supreme.cell(row=supreme_row, column=2, value=record.get('username'))
                ws_supreme.cell(row=supreme_row, column=3, value=record.get('full_name'))
                ws_supreme.cell(row=supreme_row, column=4, value=record.get('birthday'))
                ws_supreme.cell(row=supreme_row, column=5, value=record.get('phone_number'))
                ws_supreme.cell(row=supreme_row, column=6, value=record.get('time_zone'))
                ws_supreme.cell(row=supreme_row, column=7, value=record.get('expertise_level'))
                ws_supreme.cell(row=supreme_row, column=8, value=record.get('trading_time'))
                ws_supreme.cell(row=supreme_row, column=9, value=record.get('interest_reason'))
                ws_supreme.cell(row=supreme_row, column=10, value=record.get('personal_goals'))
                ws_supreme.cell(row=supreme_row, column=11, value=record.get('call_preference'))
                ws_supreme.cell(row=supreme_row, column=12, value=record.get('challenges'))
                ws_supreme.cell(row=supreme_row, column=13, value=record.get('plan'))
                ws_supreme.cell(row=supreme_row, column=14, value=record.get('form_completion_date'))
                
                # Apply alternating row colors with modern styling
                if supreme_row % 2 == 0:
                    for col in range(1, 15):
                        cell = ws_supreme.cell(row=supreme_row, column=col)
                        cell.fill = PatternFill(start_color=colors['alt_row'], end_color=colors['alt_row'], fill_type="solid")
                        # Add light border
                        cell.border = Border(
                            bottom=Side(border_style="thin", color="D3D3D3")
                        )
                else:
                    # Add just the bottom border to odd rows
                    for col in range(1, 15):
                        cell = ws_supreme.cell(row=supreme_row, column=col)
                        cell.border = Border(
                            bottom=Side(border_style="thin", color="D3D3D3")
                        )
                
                supreme_row += 1
            else:
                # Add to regular sheet
                ws_regular.cell(row=regular_row, column=1, value=record.get('user_id'))
                ws_regular.cell(row=regular_row, column=2, value=record.get('username'))
                ws_regular.cell(row=regular_row, column=3, value=record.get('full_name'))
                ws_regular.cell(row=regular_row, column=4, value=record.get('age_birth_year'))
                ws_regular.cell(row=regular_row, column=5, value=record.get('experience_level'))
                ws_regular.cell(row=regular_row, column=6, value=record.get('learning_goals'))
                ws_regular.cell(row=regular_row, column=7, value=record.get('source'))
                ws_regular.cell(row=regular_row, column=8, value=record.get('plan'))
                ws_regular.cell(row=regular_row, column=9, value=record.get('form_completion_date'))
                
                # Apply alternating row colors with modern styling
                if regular_row % 2 == 0:
                    for col in range(1, 10):
                        cell = ws_regular.cell(row=regular_row, column=col)
                        cell.fill = PatternFill(start_color=colors['alt_row'], end_color=colors['alt_row'], fill_type="solid")
                        # Add light border
                        cell.border = Border(
                            bottom=Side(border_style="thin", color="D3D3D3")
                        )
                else:
                    # Add just the bottom border to odd rows
                    for col in range(1, 10):
                        cell = ws_regular.cell(row=regular_row, column=col)
                        cell.border = Border(
                            bottom=Side(border_style="thin", color="D3D3D3")
                        )
                
                regular_row += 1
        
        # Auto-adjust column widths for regular sheet
        for col in ws_regular.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 3) if max_length < 50 else 50
            ws_regular.column_dimensions[column].width = adjusted_width
            
        # Auto-adjust column widths for supreme sheet
        for col in ws_supreme.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 3) if max_length < 50 else 50
            ws_supreme.column_dimensions[column].width = adjusted_width
        
        # Add table totals row for each sheet
        # Regular sheet
        total_row = regular_row
        ws_regular.cell(row=total_row, column=1, value="TOTAL")
        ws_regular.cell(row=total_row, column=2, value=f"{regular_row - 2} members")
        ws_regular.merge_cells(f'A{total_row}:B{total_row}')
        
        for col in range(1, 10):
            cell = ws_regular.cell(row=total_row, column=col)
            cell.font = Font(bold=True, color=colors['dark_text'])
            cell.fill = PatternFill(start_color=colors['secondary'], end_color=colors['secondary'], fill_type="solid")
        
        # Supreme sheet
        total_row = supreme_row
        ws_supreme.cell(row=total_row, column=1, value="TOTAL")
        ws_supreme.cell(row=total_row, column=2, value=f"{supreme_row - 2} members")
        ws_supreme.merge_cells(f'A{total_row}:B{total_row}')
        
        for col in range(1, 15):
            cell = ws_supreme.cell(row=total_row, column=col)
            cell.font = Font(bold=True, color=colors['dark_text'])
            cell.fill = PatternFill(start_color=colors['secondary'], end_color=colors['secondary'], fill_type="solid")
        
        # Update processing message
        bot.edit_message_text("ðŸ“Š Finalizing export with professional formatting...", 
                             chat_id=message.chat.id, 
                             message_id=processing_msg.message_id)
        
        # Freeze panes to keep headers visible on scroll
        ws_regular.freeze_panes = "A2"
        ws_supreme.freeze_panes = "A2"
        
        # Save to BytesIO
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        # Prepare for sending with more descriptive filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        excel_bytes.name = f"PTA_Member_Forms_{timestamp}.xlsx"
        
        # Edit processing message to show success
        bot.edit_message_text("âœ… Modern export generated successfully!", 
                             chat_id=message.chat.id, 
                             message_id=processing_msg.message_id)
        
        # Generate stats for the caption
        regular_count = regular_row - 2
        supreme_count = supreme_row - 2
        total_count = regular_count + supreme_count
        
        # Send Excel file as document with more professional caption
        bot.send_document(
            message.chat.id, 
            excel_bytes, 
            caption=f"ðŸ“Š *Member Form Responses Export*\n\nâ€¢ *Total Members:* {total_count}\nâ€¢ *Regular Members:* {regular_count}\nâ€¢ *Supreme Members:* {supreme_count}\n\nExported on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
            parse_mode="Markdown"
        )
        
        # Log the export activity
        admin_username = message.from_user.username or f"Admin {message.from_user.id}"
        logging.info(f"Form responses exported by {admin_username} ({message.from_user.id})")
        
        # Also notify all other admins for transparency
        for admin_id in ADMIN_IDS:
            if admin_id != message.from_user.id:  # Don't send to the admin who requested it
                try:
                    bot.send_message(
                        admin_id, 
                        f"ðŸ“Š *Form Data Export*\n\n"
                        f"@{admin_username} has exported member form responses containing data for {total_count} members.",
                        parse_mode="Markdown"
                    )
                except Exception as e:
                    logging.error(f"Failed to notify admin {admin_id} about form data export: {e}")
        
    except Exception as e:
        bot.edit_message_text(f"âŒ Error generating export: {str(e)}", 
                             chat_id=message.chat.id, 
                             message_id=processing_msg.message_id)
        logging.error(f"Error in export_form_responses: {e}")

@bot.message_handler(commands=['enrollment'])
def handle_enrollment_command(message):
    """Manage enrollment status for Regular and Supreme memberships (admin only)"""
    # Check if user is admin or creator
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    # Get current status
    regular_enrollment_open = BOT_SETTINGS.get('regular_enrollment_open', True)  # Default to open
    supreme_enrollment_open = BOT_SETTINGS.get('supreme_enrollment_open', True)  # Default to open
    
    regular_status = "ðŸŸ¢ OPEN" if regular_enrollment_open else "ðŸ”´ CLOSED"
    supreme_status = "ðŸŸ¢ OPEN" if supreme_enrollment_open else "ðŸ”´ CLOSED"
    
    # Create a keyboard with enrollment options
    markup = InlineKeyboardMarkup(row_width=2)
    markup.add(
        InlineKeyboardButton("ðŸŸ¢ Open Regular", callback_data="enrollment_regular_open"),
        InlineKeyboardButton("ðŸ”´ Close Regular", callback_data="enrollment_regular_close")
    )
    markup.add(
        InlineKeyboardButton("ðŸŸ¢ Open Supreme", callback_data="enrollment_supreme_open"),
        InlineKeyboardButton("ðŸ”´ Close Supreme", callback_data="enrollment_supreme_close")
    )
    
    bot.reply_to(
        message,
        f"ðŸ”„ *ENROLLMENT STATUS CONTROL*\n\n"
        f"*Regular Membership:* {regular_status}\n"
        f"*Supreme Membership:* {supreme_status}\n\n"
        f"Select an option to change enrollment status:",
        parse_mode="Markdown",
        reply_markup=markup
    )

# Update this in the handle_enrollment_callback function
@bot.callback_query_handler(func=lambda call: call.data.startswith("enrollment_"))
def handle_enrollment_callback(call):
    """Handle enrollment status change callbacks"""
    global BOT_SETTINGS
    
    # Check if user is admin or creator
    if call.from_user.id not in ADMIN_IDS and call.from_user.id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    # Parse callback data to determine enrollment type and action
    parts = call.data.split("_")
    enrollment_type = parts[1]  # "regular", "supreme", or "both"
    action = parts[2]           # "open" or "close"
    
    # Handle the different options
    if enrollment_type == "regular" and action == "open":
        # Open regular enrollment
        BOT_SETTINGS['regular_enrollment_open'] = True
        save_settings(BOT_SETTINGS)
        status_message = "Regular Membership enrollment is now *ðŸŸ¢ OPEN*"
        admin_message = "has *OPENED* Regular Membership enrollment"
        action_text = "Regular Membership enrollment is now OPEN"
        is_open = True
        
    elif enrollment_type == "regular" and action == "close":
        # Close regular enrollment
        BOT_SETTINGS['regular_enrollment_open'] = False
        save_settings(BOT_SETTINGS)
        status_message = "Regular Membership enrollment is now *ðŸ”´ CLOSED*"
        admin_message = "has *CLOSED* Regular Membership enrollment"
        action_text = "Regular Membership enrollment is now CLOSED"
        is_open = False
        
    elif enrollment_type == "supreme" and action == "open":
        # Open supreme enrollment
        BOT_SETTINGS['supreme_enrollment_open'] = True
        save_settings(BOT_SETTINGS)
        status_message = "Supreme Membership enrollment is now *ðŸŸ¢ OPEN*"
        admin_message = "has *OPENED* Supreme Membership enrollment"
        action_text = "Supreme Membership enrollment is now OPEN"
        is_open = True
        
    elif enrollment_type == "supreme" and action == "close":
        # Close supreme enrollment
        BOT_SETTINGS['supreme_enrollment_open'] = False
        save_settings(BOT_SETTINGS)
        status_message = "Supreme Membership enrollment is now *ðŸ”´ CLOSED*"
        admin_message = "has *CLOSED* Supreme Membership enrollment"
        action_text = "Supreme Membership enrollment is now CLOSED"
        is_open = False
        
    elif enrollment_type == "both" and action == "open":
        # Open both enrollment types
        BOT_SETTINGS['regular_enrollment_open'] = True
        BOT_SETTINGS['supreme_enrollment_open'] = True
        save_settings(BOT_SETTINGS)
        status_message = "Regular and Supreme Membership enrollments are now *ðŸŸ¢ OPEN*"
        admin_message = "has *OPENED* both Regular and Supreme Membership enrollments"
        action_text = "Both enrollment types are now OPEN"
        is_open = True
        
    elif enrollment_type == "both" and action == "close":
        # Close both enrollment types
        BOT_SETTINGS['regular_enrollment_open'] = False
        BOT_SETTINGS['supreme_enrollment_open'] = False
        save_settings(BOT_SETTINGS)
        status_message = "Regular and Supreme Membership enrollments are now *ðŸ”´ CLOSED*"
        admin_message = "has *CLOSED* both Regular and Supreme Membership enrollments"
        action_text = "Both enrollment types are now CLOSED"
        is_open = False
    
    # Update the message
    bot.edit_message_text(
        f"âœ… *Enrollment Status Updated*\n\n"
        f"{status_message}\n\n"
        f"{'New users can now purchase this membership type.' if is_open else 'Only existing members can renew this membership type.'}",
        call.message.chat.id,
        call.message.message_id,
        parse_mode="Markdown"
    )
    
    bot.answer_callback_query(call.id, action_text)
    
    # Log the action
    admin_username = call.from_user.username or f"Admin {call.from_user.id}"
    logging.info(f"{enrollment_type.capitalize()} enrollment {action}ed by {admin_username}")
    
    # Notify other admins
    for admin_id in ADMIN_IDS:
        if admin_id != call.from_user.id:  # Don't send to the admin who performed the action
            bot.send_message(
                admin_id,
                f"â„¹ï¸ *Enrollment Status Changed*\n\n"
                f"@{safe_markdown_escape(admin_username)} {admin_message}.\n\n"
                f"{'New users can now purchase this membership type.' if is_open else 'Only existing members can renew this membership type.'}",
                parse_mode="Markdown"
            )
    
    # Notify subscribers about the enrollment change
    if enrollment_type == "both":
        # Notify about both enrollment types changing
        notify_enrollment_change_specific("regular", is_open)
        notify_enrollment_change_specific("supreme", is_open)
    else:
        # Notify about the specific enrollment type changing
        notify_enrollment_change_specific(enrollment_type, is_open)

@bot.message_handler(commands=['export_payments'])
def export_payment_data(message):
    """Export payment data to a professionally formatted Excel file"""
    # Check if user is admin or creator
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    try:
        # Send processing message
        processing_msg = bot.reply_to(message, "ðŸ“Š *Generating Payment Data Export*\n\nCollecting payment records and creating report...", parse_mode="Markdown")
        
        # ADDED: Refresh payment data from MongoDB before checking
        global PAYMENT_DATA
        PAYMENT_DATA = load_payment_data()
        
        # Check if there's any payment data to export
        if not PAYMENT_DATA:
            bot.edit_message_text("âŒ No payment records found to export.", 
                                 chat_id=message.chat.id, 
                                 message_id=processing_msg.message_id)
            return
        
        # Collect payment records into a list of dictionaries
        payment_records = []
        
        for user_id_str, data in PAYMENT_DATA.items():
            # Get user info for display
            username = data.get('username', 'No Username')
            
            # Calculate payment status
            has_paid = data.get('haspayed', False)
            cancelled = data.get('cancelled', False)
            
            if has_paid and cancelled:
                status = "Cancelled (Active)"
            elif has_paid:
                status = "Active"
            elif cancelled:
                status = "Cancelled"
            else:
                status = "Inactive"
            
            # Calculate days remaining in subscription
            days_remaining = None
            expiry_date = None
            
            if 'due_date' in data:
                try:
                    due_date = datetime.strptime(data['due_date'], '%Y-%m-%d %H:%M:%S')
                    current_date = datetime.now()
                    days_remaining = (due_date - current_date).days
                    expiry_date = due_date.strftime('%Y-%m-%d')
                except Exception:
                    expiry_date = data['due_date']
            
            # Calculate lifetime value (never expires)
            is_lifetime = False
            if data.get('payment_plan') == 'Legacy' and data.get('mentorship_type', '').lower() == 'supreme':
                is_lifetime = True
            
            # Create the record
            record = {
                'user_id': user_id_str,
                'username': username,
                'status': status,
                'plan': data.get('payment_plan', 'Unknown'),
                'mentorship_type': data.get('mentorship_type', 'Unknown'),
                'payment_method': data.get('payment_mode', 'Unknown'),
                'expiry_date': "Lifetime" if is_lifetime else expiry_date,
                'days_remaining': "âˆž" if is_lifetime else days_remaining,
                'terms_accepted': data.get('terms_accepted', False),
                'privacy_accepted': data.get('privacy_accepted', False),
                'signup_date': data.get('signup_date', 'Unknown'),
                'last_renewal': data.get('last_renewal_date', 'Unknown'),
                'cancelled_date': data.get('cancellation_date', 'N/A') if cancelled else 'N/A',
                'has_form_data': 'form_answers' in data
            }
            
            payment_records.append(record)
        
        # Sort records by status and expiry date
        sorted_records = sorted(payment_records, key=lambda x: (0 if x['status'] == 'Active' else 1, 
                                                             x['expiry_date'] if x['expiry_date'] != 'Lifetime' else '9999-12-31'))
        
        # Create DataFrame for processing
        df = pd.DataFrame(sorted_records)
        
        # Update processing message
        bot.edit_message_text("ðŸ“Š Creating beautifully formatted Excel workbook...", 
                             chat_id=message.chat.id, 
                             message_id=processing_msg.message_id)
        
        # Create workbook
        wb = Workbook()
        
        # Create cover sheet (Dashboard)
        dashboard = wb.active
        dashboard.title = "Dashboard"
        
        # Define modern color scheme
        colors = {
            'primary': '4472C4',      # Modern blue
            'secondary': '5B9BD5',     # Lighter blue
            'accent1': '70AD47',      # Green
            'accent2': 'ED7D31',      # Orange
            'accent3': 'FFC000',      # Yellow
            'red': 'FF0000',          # Red for expired
            'light_bg': 'F2F2F2',     # Light gray
            'dark_text': '262626',    # Dark gray
            'header_bg': '203864',    # Dark blue
            'alt_row': 'E6F0FF'       # Very light blue
        }
        
        # ---- DASHBOARD SHEET ----
        
        # Add title
        dashboard.merge_cells('A1:H1')
        title_cell = dashboard['A1']
        title_cell.value = "PRODIGY TRADING ACADEMY"
        title_cell.font = Font(size=28, bold=True, color=colors['dark_text'], name='Calibri')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        dashboard.row_dimensions[1].height = 45
        
        # Add subtitle
        dashboard.merge_cells('A2:H2')
        subtitle_cell = dashboard['A2']
        subtitle_cell.value = "Member Payment Records"
        subtitle_cell.font = Font(size=16, italic=True, color=colors['dark_text'], name='Calibri')
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add export date
        dashboard.merge_cells('A3:H3')
        date_cell = dashboard['A3']
        date_cell.value = f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        date_cell.font = Font(size=12, color=colors['dark_text'], name='Calibri')
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add some space
        dashboard.row_dimensions[4].height = 20
        
        # Calculate key metrics
        active_members = len([r for r in payment_records if r['status'] == 'Active'])
        cancelled_members = len([r for r in payment_records if 'Cancelled' in r['status']])
        inactive_members = len([r for r in payment_records if r['status'] == 'Inactive'])
        regular_members = len([r for r in payment_records if r['mentorship_type'].lower() == 'regular' and r['status'] == 'Active'])
        supreme_members = len([r for r in payment_records if r['mentorship_type'].lower() == 'supreme' and r['status'] == 'Active'])
        lifetime_members = len([r for r in payment_records if r['expiry_date'] == 'Lifetime' and r['status'] == 'Active'])
        
        # Add expiring soon count (within 7 days but still active)
        expiring_soon = len([r for r in payment_records if r['status'] == 'Active' and 
                             isinstance(r['days_remaining'], int) and
                             0 <= r['days_remaining'] <= 7])
        
        # Add expired count (negative days remaining)
        expired = len([r for r in payment_records if isinstance(r['days_remaining'], int) and r['days_remaining'] < 0])
        
        # Add summary section header
        dashboard.merge_cells('B5:G5')
        summary_header = dashboard['B5']
        summary_header.value = "MEMBERSHIP SUMMARY"
        summary_header.font = Font(size=14, bold=True, color="FFFFFF", name='Calibri')
        summary_header.alignment = Alignment(horizontal='center', vertical='center')
        summary_header.fill = PatternFill(start_color=colors['header_bg'], end_color=colors['header_bg'], fill_type="solid")
        
        # Add key statistics in a modern card layout - first row
        stats_row1 = [
            ["Total Members", len(payment_records), colors['primary']],
            ["Active Members", active_members, colors['accent1']],
            ["Inactive Members", inactive_members, colors['secondary']]
        ]
        
        # Create stat cards - first row
        for i, (label, value, color) in enumerate(stats_row1):
            col = chr(ord('B') + i*2)  # B, D, F
            
            # Merge cells for each stat card
            dashboard.merge_cells(f'{col}6:{chr(ord(col)+1)}7')
            dashboard.merge_cells(f'{col}8:{chr(ord(col)+1)}9')
            
            # Add label
            label_cell = dashboard[f'{col}6']
            label_cell.value = label
            label_cell.font = Font(size=12, color="FFFFFF", name='Calibri')
            label_cell.alignment = Alignment(horizontal='center', vertical='center')
            label_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            # Add value
            value_cell = dashboard[f'{col}8']
            value_cell.value = value
            value_cell.font = Font(size=24, bold=True, color=colors['dark_text'], name='Calibri')
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add border to value cell
            thin_border = Border(
                left=Side(style='thin', color=color),
                right=Side(style='thin', color=color),
                top=Side(style='thin', color=color),
                bottom=Side(style='thin', color=color)
            )
            value_cell.border = thin_border
        
        # Second row of stats
        stats_row2 = [
            ["Lifetime Members", lifetime_members, colors['accent3']],
            ["Expiring Soon", expiring_soon, colors['accent2']],
            ["Expired", expired, colors['red']]
        ]
        
        # Create stat cards - second row
        for i, (label, value, color) in enumerate(stats_row2):
            col = chr(ord('B') + i*2)  # B, D, F
            row_offset = 5  # Move down 5 rows from first set of cards
            
            # Merge cells for each stat card
            dashboard.merge_cells(f'{col}{6+row_offset}:{chr(ord(col)+1)}{7+row_offset}')
            dashboard.merge_cells(f'{col}{8+row_offset}:{chr(ord(col)+1)}{9+row_offset}')
            
            # Add label
            label_cell = dashboard[f'{col}{6+row_offset}']
            label_cell.value = label
            label_cell.font = Font(size=12, color="FFFFFF", name='Calibri')
            label_cell.alignment = Alignment(horizontal='center', vertical='center')
            label_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            # Add value
            value_cell = dashboard[f'{col}{8+row_offset}']
            value_cell.value = value
            value_cell.font = Font(size=24, bold=True, color=colors['dark_text'], name='Calibri')
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add border to value cell
            thin_border = Border(
                left=Side(style='thin', color=color),
                right=Side(style='thin', color=color),
                top=Side(style='thin', color=color),
                bottom=Side(style='thin', color=color)
            )
            value_cell.border = thin_border
        
        # Add third row with additional stats (membership types)
        membership_stats = [
            ["Regular Members", regular_members, colors['secondary']],
            ["Supreme Members", supreme_members, colors['primary']],
            ["Cancelled", cancelled_members, colors['accent2']]
        ]
        
        # Create stat cards - third row
        for i, (label, value, color) in enumerate(membership_stats):
            col = chr(ord('B') + i*2)  # B, D, F
            row_offset = 10  # Move down 10 rows from first set of cards
            
            # Merge cells for each stat card
            dashboard.merge_cells(f'{col}{6+row_offset}:{chr(ord(col)+1)}{7+row_offset}')
            dashboard.merge_cells(f'{col}{8+row_offset}:{chr(ord(col)+1)}{9+row_offset}')
            
            # Add label
            label_cell = dashboard[f'{col}{6+row_offset}']
            label_cell.value = label
            label_cell.font = Font(size=12, color="FFFFFF", name='Calibri')
            label_cell.alignment = Alignment(horizontal='center', vertical='center')
            label_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            # Add value
            value_cell = dashboard[f'{col}{8+row_offset}']
            value_cell.value = value
            value_cell.font = Font(size=24, bold=True, color=colors['dark_text'], name='Calibri')
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add border to value cell
            thin_border = Border(
                left=Side(style='thin', color=color),
                right=Side(style='thin', color=color),
                top=Side(style='thin', color=color),
                bottom=Side(style='thin', color=color)
            )
            value_cell.border = thin_border
        
        # Add navigation section
        row_pos = 27  # Position after the third row of stat cards
        dashboard.merge_cells(f'B{row_pos}:G{row_pos}')
        nav_header = dashboard[f'B{row_pos}']
        nav_header.value = "PAYMENT RECORDS"
        nav_header.font = Font(size=14, bold=True, color="FFFFFF", name='Calibri')
        nav_header.alignment = Alignment(horizontal='center', vertical='center')
        nav_header.fill = PatternFill(start_color=colors['header_bg'], end_color=colors['header_bg'], fill_type="solid")
        
        # Add footer
        dashboard.merge_cells(f'B{row_pos+4}:G{row_pos+4}')
        footer = dashboard[f'B{row_pos+4}']
        footer.value = "Confidential - For Administrative Use Only"
        footer.font = Font(size=10, italic=True, color=colors['dark_text'], name='Calibri')
        footer.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths for dashboard
        for col in ['A', 'H']:
            dashboard.column_dimensions[col].width = 2  # Margin columns
        
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            dashboard.column_dimensions[col].width = 15
        
        # Update processing message
        bot.edit_message_text("ðŸ“Š Creating payment records sheet...", 
                             chat_id=message.chat.id, 
                             message_id=processing_msg.message_id)
        
        # ---- PAYMENT RECORDS SHEET ----
        payment_sheet = wb.create_sheet(title="Payment Records")
        
        # Payment records headers
        headers = [
            'User ID', 'Username', 'Status', 'Type', 'Plan', 'Payment Method',
            'Expiry Date', 'Days Remaining', 'Signup Date', 'Last Renewal', 'Cancellation Date'
        ]
        
        # Add table header with stylish formatting
        for col_idx, header in enumerate(headers, 1):
            cell = payment_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=12, name='Calibri')
            cell.fill = PatternFill(start_color=colors['header_bg'], end_color=colors['header_bg'], fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                bottom=Side(border_style="medium", color="FFFFFF")
            )
        
        # Make the header row taller
        payment_sheet.row_dimensions[1].height = 30
        
        # Populate the payment records with data
        for row_idx, record in enumerate(sorted_records, 2):
            payment_sheet.cell(row=row_idx, column=1, value=record['user_id'])
            payment_sheet.cell(row=row_idx, column=2, value=record['username'])
            payment_sheet.cell(row=row_idx, column=3, value=record['status'])
            payment_sheet.cell(row=row_idx, column=4, value=record['mentorship_type'])
            payment_sheet.cell(row=row_idx, column=5, value=record['plan'])
            payment_sheet.cell(row=row_idx, column=6, value=record['payment_method'])
            payment_sheet.cell(row=row_idx, column=7, value=record['expiry_date'])
            
            # Format days remaining column
            days_cell = payment_sheet.cell(row=row_idx, column=8)
            if record['days_remaining'] == "âˆž":
                days_cell.value = "âˆž"
            else:
                days_cell.value = record['days_remaining']
                # Color code based on days remaining
                if isinstance(record['days_remaining'], int):
                    if record['days_remaining'] < 0:
                        days_cell.font = Font(color="FF0000", bold=True)  # Red for expired
                    elif record['days_remaining'] <= 7:
                        days_cell.font = Font(color="FF9900", bold=True)  # Orange for expiring soon
            
            payment_sheet.cell(row=row_idx, column=9, value=record['signup_date'])
            payment_sheet.cell(row=row_idx, column=10, value=record['last_renewal'])
            payment_sheet.cell(row=row_idx, column=11, value=record['cancelled_date'])
            
            # Apply alternating row colors
            if row_idx % 2 == 0:
                for col in range(1, len(headers) + 1):
                    cell = payment_sheet.cell(row=row_idx, column=col)
                    cell.fill = PatternFill(start_color=colors['alt_row'], end_color=colors['alt_row'], fill_type="solid")
                    cell.border = Border(bottom=Side(border_style="thin", color="D3D3D3"))
            else:
                for col in range(1, len(headers) + 1):
                    cell = payment_sheet.cell(row=row_idx, column=col)
                    cell.border = Border(bottom=Side(border_style="thin", color="D3D3D3"))
        
        # Add table totals row
        totals_row = len(sorted_records) + 2
        payment_sheet.cell(row=totals_row, column=1, value="TOTAL")
        payment_sheet.cell(row=totals_row, column=2, value=f"{len(sorted_records)} members")
        payment_sheet.merge_cells(f'A{totals_row}:B{totals_row}')
        
        for col in range(1, len(headers) + 1):
            cell = payment_sheet.cell(row=totals_row, column=col)
            cell.font = Font(bold=True, color=colors['dark_text'])
            cell.fill = PatternFill(start_color=colors['secondary'], end_color=colors['secondary'], fill_type="solid")
        
        # Auto-adjust column widths
        for col in payment_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 3) if max_length < 50 else 50
            payment_sheet.column_dimensions[column].width = adjusted_width
        
        # Freeze panes to keep headers visible on scroll
        payment_sheet.freeze_panes = "A2"
        
        # Update processing message
        bot.edit_message_text("ðŸ“Š Finalizing export with professional formatting...", 
                             chat_id=message.chat.id, 
                             message_id=processing_msg.message_id)
        
        # Save to BytesIO
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        # Prepare for sending with descriptive filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        excel_bytes.name = f"PTA_Payment_Records_{timestamp}.xlsx"
        
        # Edit processing message to show success
        bot.edit_message_text("âœ… Payment records export generated successfully!", 
                             chat_id=message.chat.id, 
                             message_id=processing_msg.message_id)
        
        # Send Excel file as document with professional caption
        bot.send_document(
            message.chat.id, 
            excel_bytes, 
            caption=f"ðŸ“Š *Payment Records Export*\n\nâ€¢ *Total Records:* {len(payment_records)}\nâ€¢ *Active Members:* {active_members}\nâ€¢ *Expiring Soon:* {expiring_soon}\n\nExported on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
            parse_mode="Markdown"
        )
        
        # Log the export activity
        admin_username = message.from_user.username or f"Admin {message.from_user.id}"
        logging.info(f"Payment records exported by {admin_username} ({message.from_user.id})")
        
        # Also notify all other admins for transparency
        for admin_id in ADMIN_IDS:
            if admin_id != message.from_user.id:  # Don't send to the admin who requested it
                try:
                    bot.send_message(
                        admin_id, 
                        f"ðŸ“Š *Payment Records Export*\n\n"
                        f"@{admin_username} has exported payment records containing data for {len(payment_records)} members.",
                        parse_mode="Markdown"
                    )
                except Exception as e:
                    logging.error(f"Failed to notify admin {admin_id} about payment export: {e}")
        
    except Exception as e:
        bot.edit_message_text(f"âŒ Error generating payment export: {str(e)}", 
                             chat_id=message.chat.id, 
                             message_id=processing_msg.message_id)
        logging.error(f"Error in export_payment_data: {e}")

def cleanup_inactive_pending_users():
    """Clean up inactive pending users every 30 minutes, except those waiting for payment approval"""
    logging.info("Pending users cleanup thread started")
    
    while True:
        try:
            # Sleep first to avoid immediate cleanup after startup
            time.sleep(1800)  # 30 minutes
            
            # Get the current time for logging
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Track statistics for logging
            users_before = len(PENDING_USERS)
            removed_count = 0
            preserved_count = 0
            
            # Create a copy of keys to avoid modifying dictionary during iteration
            user_ids = list(PENDING_USERS.keys())
            
            for user_id in user_ids:
                # Get user status
                status = PENDING_USERS[user_id].get('status', '')
                
                # Preserve users waiting for payment approval
                if status == 'waiting_approval':
                    preserved_count += 1
                    continue
                
                # Remove all other pending users
                PENDING_USERS.pop(user_id, None)
                delete_pending_user(user_id)  # Remove from MongoDB
                removed_count += 1
                
            # Save changes
            save_pending_users()
            
            # Log the cleanup results
            users_after = len(PENDING_USERS)
            logging.info(f"Pending users cleanup completed at {current_time}: "
                        f"Removed {removed_count} inactive users, "
                        f"preserved {preserved_count} users waiting for approval. "
                        f"Users before: {users_before}, users after: {users_after}")
            
        except Exception as e:
            logging.error(f"Error in pending users cleanup: {e}")
            time.sleep(300)  # Wait 5 minutes on error before trying again

@bot.message_handler(commands=['remove_all'])
def remove_all_pending_users(message):
    """Manually remove all pending users except those waiting for payment approval"""
    # Check if user is admin or creator
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    try:
        # Send processing message
        processing_msg = bot.reply_to(message, "ðŸ”„ Removing pending users... Please wait.")
        
        # Track statistics
        users_before = len(PENDING_USERS)
        removed_count = 0
        preserved_count = 0
        
        # Create a copy of keys to avoid modifying dictionary during iteration
        user_ids = list(PENDING_USERS.keys())
        
        for user_id in user_ids:
            # Get user status
            status = PENDING_USERS[user_id].get('status', '')
            
            # Preserve users waiting for payment approval
            if status == 'waiting_approval':
                preserved_count += 1
                continue
            
            # Remove all other pending users
            PENDING_USERS.pop(user_id, None)
            delete_pending_user(user_id)  # Remove from MongoDB
            removed_count += 1
        
        # Save changes
        save_pending_users()
        
        # Send completion message with statistics
        users_after = len(PENDING_USERS)
        bot.edit_message_text(
            f"âœ… *Pending Users Cleanup Complete*\n\n"
            f"â€¢ *Removed:* {removed_count} inactive users\n"
            f"â€¢ *Preserved:* {preserved_count} users waiting for payment approval\n"
            f"â€¢ *Before:* {users_before} total users\n"
            f"â€¢ *After:* {users_after} total users",
            chat_id=message.chat.id,
            message_id=processing_msg.message_id,
            parse_mode="Markdown"
        )
        
        # Log the action
        admin_username = message.from_user.username or f"Admin {message.from_user.id}"
        logging.info(f"Manual cleanup of pending users by {admin_username}: removed {removed_count}, preserved {preserved_count}")
        
    except Exception as e:
        bot.reply_to(message, f"âŒ Error removing pending users: {str(e)}")
        logging.error(f"Error in manual pending users cleanup: {e}")

@bot.message_handler(commands=['april8'])
def handle_april8_command(message):
    """Send an April 8 meme GIF to the group chat with global usage limits"""
    if message.chat.type not in ['group', 'supergroup']:
        bot.reply_to(message, "âŒ This command can only be used in group chats.")
        return
        
    current_time = datetime.now()
    
    try:
        # Get global usage data
        global_data = jarvis_usage_collection.find_one({"_id": "april8_counter"})
        
        if not global_data:
            # First time the command is being used after implementation
            global_data = {
                "_id": "april8_counter",
                "count": 0,
                "last_reset": current_time.strftime('%Y-%m-%d %H:%M:%S')
            }
        
        # Check if 6 hours have passed since last reset
        last_reset = datetime.strptime(global_data["last_reset"], '%Y-%m-%d %H:%M:%S')
        hours_passed = (current_time - last_reset).total_seconds() / 3600
        
        # Reset counter if 6 hours have passed
        if hours_passed >= 6:
            global_data = {
                "_id": "april8_counter",
                "count": 0,
                "last_reset": current_time.strftime('%Y-%m-%d %H:%M:%S')
            }
        
        # Check if global limit has been reached
        if global_data["count"] >= 4:
            next_reset_time = last_reset + timedelta(hours=6)
            time_until_reset = next_reset_time - current_time
            hours, remainder = divmod(time_until_reset.seconds, 3600)
            minutes, _ = divmod(remainder, 60)
            
            # Send message and self-destruct after 5 seconds
            limit_msg = bot.reply_to(
                message, 
                f"â³ The April 8 command has reached its global limit (4 uses per 6 hours).\nTry again in {hours} hours and {minutes} minutes."
            )
            
            # Create a thread to delete BOTH messages after 5 seconds
            def delete_after_delay(chat_id, message_ids):
                time.sleep(3)
                for msg_id in message_ids:
                    try:
                        bot.delete_message(chat_id, msg_id)
                        logging.info(f"Auto-deleted message ID {msg_id} in chat {chat_id}")
                    except Exception as e:
                        logging.error(f"Failed to auto-delete message ID {msg_id}: {e}")
            
            # Delete both the original command message and the rate limit message
            threading.Thread(target=delete_after_delay, 
                         args=(message.chat.id, [message.message_id, limit_msg.message_id])).start()
            return
        
        # Increment the global usage counter
        global_data["count"] += 1
        jarvis_usage_collection.replace_one({"_id": "april8_counter"}, global_data, upsert=True)
        
        remaining = 4 - global_data["count"]
        
        # Path to the April 8 GIF
        april8_gif = "gifs/april8.gif"  # Using existing GIFs directory
        
        # Send the GIF
        with open(april8_gif, 'rb') as animation:
            bot.send_animation(
                message.chat.id, 
                animation,
                timeout=60
            )
            logging.info(f"Sent April 8 GIF in chat {message.chat.id} (requested by {message.from_user.id}, {remaining} global uses remaining)")
            
            # Also delete the original command after sending the GIF for cleaner chat
            try:
                bot.delete_message(message.chat.id, message.message_id)
            except Exception as e:
                logging.error(f"Failed to delete original command message: {e}")
                
    except FileNotFoundError:
        bot.reply_to(message, "âŒ GIF not found.")
        logging.warning(f"April 8 GIF not found (requested by {message.from_user.id})")
    except Exception as e:
        bot.reply_to(message, "âŒ Error sending GIF.")
        logging.error(f"Error in April 8 command: {e}")

@bot.message_handler(commands=['announce'])
def start_announcement(message):
    """Start the announcement creation process for admins"""
    # Check if user is admin or creator
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    user_id = message.from_user.id
    
    # Initialize announcement state
    ADMIN_ANNOUNCING[user_id] = {
        'status': 'waiting_for_announcement',
        'sent_to': set()  # Track where the announcement has been sent
    }
    
    bot.reply_to(
        message,
        "ðŸ“£ *ANNOUNCEMENT CREATION*\n\n"
        "Please send your announcement message below.\n"
        "You can include text formatting, links, images, GIFs or videos.\n\n"
        "Type /cancel to abort the announcement.",
        parse_mode="Markdown"
    )

@bot.message_handler(func=lambda message: message.from_user.id in ADMIN_ANNOUNCING 
                    and ADMIN_ANNOUNCING[message.from_user.id]['status'] == 'waiting_for_announcement')
def handle_announcement_message(message):
    """Process the announcement message from the admin"""
    user_id = message.from_user.id
    
    # Store the announcement message
    ADMIN_ANNOUNCING[user_id]['message'] = message
    ADMIN_ANNOUNCING[user_id]['status'] = 'selecting_destinations'
    
    # Create keyboard with destination options
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(InlineKeyboardButton("ðŸ“£ Send to All Destinations", callback_data="announce_all"))
    markup.add(InlineKeyboardButton("ðŸ” Choose Specific Destinations", callback_data="announce_choose"))
    markup.add(InlineKeyboardButton("âŒ Cancel Announcement", callback_data="announce_cancel"))
    
    # Send confirmation with the announcement preview
    preview_msg = get_announcement_preview(message)
    
    bot.send_message(
        user_id,
        f"ðŸ“ *Announcement Preview*\n\n{preview_msg}\n\n"
        f"Where would you like to send this announcement?",
        parse_mode="Markdown",
        reply_markup=markup
    )

def get_announcement_preview(message):
    """Generate a preview of the announcement message"""
    # For text messages, return the content
    if message.content_type == 'text':
        return message.text[:200] + ('...' if len(message.text) > 200 else '')
    # For media messages, return a description
    elif message.content_type == 'photo':
        caption = message.caption or ""
        return f"[IMAGE] {caption[:200]}" + ('...' if len(caption) > 200 else '')
    elif message.content_type == 'video':
        caption = message.caption or ""
        return f"[VIDEO] {caption[:200]}" + ('...' if len(caption) > 200 else '')
    elif message.content_type == 'animation':
        caption = message.caption or ""
        return f"[GIF] {caption[:200]}" + ('...' if len(caption) > 200 else '')
    elif message.content_type == 'document':
        caption = message.caption or ""
        return f"[DOCUMENT] {caption[:200]}" + ('...' if len(caption) > 200 else '')
    # For other types
    else:
        return f"[{message.content_type.upper()}]"

@bot.message_handler(commands=['cancel'])
def cancel_announcement(message):
    """Cancel an in-progress announcement"""
    user_id = message.from_user.id
    
    # Check if user is in announcing mode
    if user_id in ADMIN_ANNOUNCING:
        ADMIN_ANNOUNCING.pop(user_id)
        bot.send_message(user_id, "âœ… Announcement process canceled.")
    else:
        # This might be for another cancellation action or invalid command
        pass

@bot.callback_query_handler(func=lambda call: call.data.startswith("announce_"))
def handle_announcement_callback(call):
    """Handle announcement destination selection callbacks"""
    user_id = call.from_user.id
    
    # Check if user is admin and is in announcing mode
    if user_id not in ADMIN_ANNOUNCING or user_id not in ADMIN_IDS and user_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    # Add logging to debug callback data
    logging.info(f"Announcement callback received: {call.data}")
    
    if call.data == "announce_all":
        # Send to all destinations
        handle_send_to_all(call)
    
    elif call.data == "announce_choose":
        # Show list of destinations to choose from
        show_destination_choices(call)
    
    elif call.data == "announce_cancel":
        # Cancel the announcement
        ADMIN_ANNOUNCING.pop(user_id, None)
        
        bot.edit_message_text(
            "âŒ Announcement cancelled.",
            call.message.chat.id,
            call.message.message_id
        )
        
        bot.answer_callback_query(call.id, "Announcement cancelled")
    
    elif call.data == "announce_done":
        # Finish the announcement process
        sent_count = len(ADMIN_ANNOUNCING[user_id]['sent_to'])
        ADMIN_ANNOUNCING.pop(user_id, None)
        
        bot.edit_message_text(
            f"âœ… Announcement completed! Sent to {sent_count} destination(s).",
            call.message.chat.id,
            call.message.message_id
        )
        
        bot.answer_callback_query(call.id, "Announcement process completed")
    
    elif call.data.startswith("announce_dest_"):
        # Improved error handling for destination extraction
        try:
            parts = call.data.split("_")
            if len(parts) >= 3:
                # Get everything after "announce_dest_" as the destination key
                destination_key = "_".join(parts[2:])
                logging.info(f"Extracted destination key: '{destination_key}'")
                
                # Log available destinations for debugging
                logging.info(f"Available destinations: {list(ANNOUNCEMENT_DESTINATIONS.keys())}")
                
                handle_send_to_destination(call, destination_key)
            else:
                logging.error(f"Invalid callback data format: {call.data}")
                bot.answer_callback_query(call.id, "âŒ Invalid callback data format", show_alert=True)
        except Exception as e:
            logging.error(f"Error processing announcement destination: {e}")
            bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}", show_alert=True)

def handle_send_to_all(call):
    """Send announcement to all configured destinations"""
    user_id = call.from_user.id
    announcement_data = ADMIN_ANNOUNCING[user_id]
    message = announcement_data['message']
    
    # Update the message to show progress
    bot.edit_message_text(
        "ðŸ”„ Sending announcement to all destinations...",
        call.message.chat.id,
        call.message.message_id
    )
    
    # Counter for successful sends
    success_count = 0
    
    # Send to each destination
    for key, destination in ANNOUNCEMENT_DESTINATIONS.items():
        try:
            # Get destination details
            dest_id = destination['id']
            topic_id = destination['topic_id']
            
            # Send the announcement based on message type
            send_success = send_announcement_to_destination(message, dest_id, topic_id)
            
            if send_success:
                # Mark destination as sent
                announcement_data['sent_to'].add(key)
                success_count += 1
                
                # Log the successful send
                logging.info(f"Announcement sent to {destination['name']} by admin {user_id}")
            else:
                logging.error(f"Failed to send announcement to {destination['name']}")
            
        except Exception as e:
            logging.error(f"Error sending announcement to {destination['name']}: {e}")
    
    # Update message to show completion
    bot.edit_message_text(
        f"âœ… Announcement sent successfully to {success_count}/{len(ANNOUNCEMENT_DESTINATIONS)} destination(s)!",
        call.message.chat.id,
        call.message.message_id
    )
    
    # Clean up
    ADMIN_ANNOUNCING.pop(user_id, None)
    
    bot.answer_callback_query(call.id, f"Sent to {success_count} destination(s)")

def show_destination_choices(call):
    """Show buttons for each destination the admin can send to"""
    user_id = call.from_user.id
    announcement_data = ADMIN_ANNOUNCING[user_id]
    
    # Create keyboard with all destinations
    markup = InlineKeyboardMarkup(row_width=1)
    
    for key, destination in ANNOUNCEMENT_DESTINATIONS.items():
        # Skip destinations that have already received the announcement
        if key in announcement_data['sent_to']:
            continue
            
        markup.add(InlineKeyboardButton(
            f"ðŸ“ {destination['name']}", 
            callback_data=f"announce_dest_{key}"
        ))
    
    # Add Done button
    markup.add(InlineKeyboardButton("âœ… Done", callback_data="announce_done"))
    
    # Update message with destination choices
    bot.edit_message_text(
        "ðŸ“ *SELECT DESTINATIONS*\n\n"
        "Choose where to send your announcement:\n"
        "(Buttons will be removed after sending to that destination)",
        call.message.chat.id,
        call.message.message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    bot.answer_callback_query(call.id, "Select destinations")

def handle_send_to_destination(call, destination_key):
    """Send announcement to a specific destination"""
    user_id = call.from_user.id
    announcement_data = ADMIN_ANNOUNCING[user_id]
    message = announcement_data['message']
    
    # Check if the destination exists
    if destination_key not in ANNOUNCEMENT_DESTINATIONS:
        bot.answer_callback_query(call.id, "âŒ Invalid destination", show_alert=True)
        return
    
    destination = ANNOUNCEMENT_DESTINATIONS[destination_key]
    dest_id = destination['id']
    topic_id = destination['topic_id']
    
    try:
        # Send the announcement
        send_success = send_announcement_to_destination(message, dest_id, topic_id)
        
        if send_success:
            # Mark destination as sent
            announcement_data['sent_to'].add(destination_key)
            
            # Log the successful send
            logging.info(f"Announcement sent to {destination['name']} by admin {user_id}")
            
            # Show updated destination choices with this destination removed
            show_destination_choices(call)
            
            bot.answer_callback_query(call.id, f"Sent to {destination['name']}")
        else:
            bot.answer_callback_query(call.id, f"âŒ Failed to send to {destination['name']}", show_alert=True)
        
    except Exception as e:
        logging.error(f"Error sending announcement to {destination['name']}: {e}")
        bot.answer_callback_query(call.id, f"âŒ Error: {str(e)}", show_alert=True)

def send_announcement_to_destination(message, dest_id, topic_id=None):
    """Send a message to a destination with proper support for media types"""
    try:
        # Handle different message types
        if message.content_type == 'text':
            # Text message
            kwargs = {
                'chat_id': dest_id,
                'text': message.text,
                'parse_mode': 'Markdown'
            }
            
            # Add topic_id if applicable
            if topic_id:
                kwargs['message_thread_id'] = topic_id
                
            bot.send_message(**kwargs)
            
        elif message.content_type == 'photo':
            # Photo message
            photo = message.photo[-1].file_id  # Get highest resolution
            caption = message.caption or ""
            
            kwargs = {
                'chat_id': dest_id,
                'photo': photo,
                'caption': caption,
                'parse_mode': 'Markdown'
            }
            
            if topic_id:
                kwargs['message_thread_id'] = topic_id
                
            bot.send_photo(**kwargs)
            
        elif message.content_type == 'video':
            # Video message
            video = message.video.file_id
            caption = message.caption or ""
            
            kwargs = {
                'chat_id': dest_id,
                'video': video,
                'caption': caption,
                'parse_mode': 'Markdown'
            }
            
            if topic_id:
                kwargs['message_thread_id'] = topic_id
                
            bot.send_video(**kwargs)
            
        elif message.content_type == 'animation':
            # GIF/Animation message
            animation = message.animation.file_id
            caption = message.caption or ""
            
            kwargs = {
                'chat_id': dest_id,
                'animation': animation,
                'caption': caption,
                'parse_mode': 'Markdown'
            }
            
            if topic_id:
                kwargs['message_thread_id'] = topic_id
                
            bot.send_animation(**kwargs)
            
        elif message.content_type == 'document':
            # Document message
            document = message.document.file_id
            caption = message.caption or ""
            
            kwargs = {
                'chat_id': dest_id,
                'document': document,
                'caption': caption,
                'parse_mode': 'Markdown'
            }
            
            if topic_id:
                kwargs['message_thread_id'] = topic_id
                
            bot.send_document(**kwargs)
            
        else:
            # Unsupported message type
            logging.warning(f"Unsupported message type for announcement: {message.content_type}")
            return False
            
        return True
        
    except Exception as e:
        logging.error(f"Error in send_announcement_to_destination: {e}")
        return False

# Command to add a new announcement destination
@bot.message_handler(commands=['add_destination'])
def add_destination_command(message):
    """Command to add a new announcement destination"""
    # Check if user is admin or creator
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    # Start the add destination wizard
    bot.reply_to(message, 
                "ðŸ” *Add Announcement Destination*\n\n"
                "Let's set up a new destination where announcements can be sent.\n\n"
                "First, please provide a name for this destination (e.g., 'Main Group', 'Announcements Channel'):",
                parse_mode="Markdown")
    
    # Set user state
    PENDING_USERS[message.from_user.id] = {
        'status': 'add_destination_name'
    }
    save_pending_users()

# Handle destination name
@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'add_destination_name')
def handle_destination_name(message):
    user_id = message.from_user.id
    dest_name = message.text.strip()
    
    # Validate name
    if len(dest_name) < 3 or len(dest_name) > 30:
        bot.send_message(user_id, "âŒ Destination name must be between 3-30 characters. Please try again:")
        return
    
    # Store name and ask for destination type
    PENDING_USERS[user_id]['dest_name'] = dest_name
    PENDING_USERS[user_id]['status'] = 'add_destination_type'
    save_pending_users()
    
    # Create keyboard for destination type
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row("Group", "Channel")
    
    bot.send_message(
        user_id,
        "ðŸ“ Is this destination a Group or a Channel?",
        reply_markup=markup
    )

# Handle destination type
@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'add_destination_type')
def handle_destination_type(message):
    user_id = message.from_user.id
    dest_type = message.text.strip().lower()
    
    # Validate type
    if dest_type not in ["group", "channel"]:
        bot.send_message(user_id, "âŒ Please select either 'Group' or 'Channel'.")
        return
    
    # Store type and ask for destination ID
    PENDING_USERS[user_id]['dest_type'] = dest_type
    PENDING_USERS[user_id]['status'] = 'add_destination_id'
    save_pending_users()
    
    bot.send_message(
        user_id,
        f"ðŸ†” Please enter the {dest_type} ID for this destination.\n\n"
        f"You can use the /gettopic command in the {dest_type} to get its ID.",
        reply_markup=ReplyKeyboardRemove()
    )

# Handle destination ID
@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'add_destination_id')
def handle_destination_id(message):
    user_id = message.from_user.id
    
    try:
        # Try to parse as integer
        dest_id = int(message.text.strip())
        
        # Store ID and ask if it has a topic
        PENDING_USERS[user_id]['dest_id'] = dest_id
        PENDING_USERS[user_id]['status'] = 'add_destination_has_topic'
        save_pending_users()
        
        # Create keyboard for yes/no
        markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        markup.row("Yes", "No")
        
        bot.send_message(
            user_id,
            "â“ Does this destination have a specific topic ID where announcements should be posted?",
            reply_markup=markup
        )
    except ValueError:
        bot.send_message(user_id, "âŒ Invalid ID. Please enter a valid numeric ID.")

# Handle has topic response
@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'add_destination_has_topic')
def handle_destination_has_topic(message):
    user_id = message.from_user.id
    has_topic = message.text.strip().lower()
    
    if has_topic == "yes":
        # Ask for topic ID
        PENDING_USERS[user_id]['status'] = 'add_destination_topic_id'
        save_pending_users()
        
        bot.send_message(
            user_id,
            "ðŸ†” Please enter the topic ID.\n\n"
            "You can use the /gettopic command in the specific topic to get its ID.",
            reply_markup=ReplyKeyboardRemove()
        )
    elif has_topic == "no":
        # No topic, set topic_id to None
        PENDING_USERS[user_id]['dest_topic_id'] = None
        show_destination_confirmation(user_id)
    else:
        bot.send_message(user_id, "âŒ Please select either 'Yes' or 'No'.")

# Handle topic ID
@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'add_destination_topic_id')
def handle_destination_topic_id(message):
    user_id = message.from_user.id
    
    try:
        # Try to parse as integer
        topic_id = int(message.text.strip())
        
        # Store topic ID
        PENDING_USERS[user_id]['dest_topic_id'] = topic_id
        show_destination_confirmation(user_id)
    except ValueError:
        bot.send_message(user_id, "âŒ Invalid topic ID. Please enter a valid numeric ID.")

def show_destination_confirmation(user_id):
    """Show confirmation message for the new destination"""
    # Change status
    PENDING_USERS[user_id]['status'] = 'add_destination_confirm'
    save_pending_users()
    
    # Get destination details
    dest_name = PENDING_USERS[user_id]['dest_name']
    dest_type = PENDING_USERS[user_id]['dest_type']
    dest_id = PENDING_USERS[user_id]['dest_id']
    dest_topic_id = PENDING_USERS[user_id].get('dest_topic_id')
    
    # Create destination ID (slug) from name
    dest_slug = dest_name.lower().replace(' ', '_')
    PENDING_USERS[user_id]['dest_slug'] = dest_slug
    
    # Create confirmation message
    confirmation = (
        f"âœ… *Destination Summary*\n\n"
        f"â€¢ *Name:* {dest_name}\n"
        f"â€¢ *Type:* {dest_type.capitalize()}\n"
        f"â€¢ *ID:* {dest_id}\n"
    )
    
    if dest_topic_id:
        confirmation += f"â€¢ *Topic ID:* {dest_topic_id}\n"
    else:
        confirmation += "â€¢ *Topic ID:* None (Main chat)\n"
        
    confirmation += f"\nDestination ID will be: `{dest_slug}`"
    
    # Create confirm/cancel keyboard
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row("âœ… Confirm", "âŒ Cancel")
    
    bot.send_message(
        user_id,
        f"{confirmation}\n\nIs this correct?",
        parse_mode="Markdown",
        reply_markup=markup
    )

# Handle confirmation
@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'add_destination_confirm')
def handle_destination_confirmation(message):
    user_id = message.from_user.id
    response = message.text.strip()
    
    if response == "âœ… Confirm":
        # Get destination details
        dest_slug = PENDING_USERS[user_id]['dest_slug']
        dest_name = PENDING_USERS[user_id]['dest_name']
        dest_type = PENDING_USERS[user_id]['dest_type']
        dest_id = PENDING_USERS[user_id]['dest_id']
        dest_topic_id = PENDING_USERS[user_id].get('dest_topic_id')
        
        # Create destination object
        destination_data = {
            'name': dest_name,
            'type': dest_type,
            'id': dest_id,
            'topic_id': dest_topic_id
        }
        
        # Save to database
        success = save_announcement_destination(dest_slug, destination_data)
        
        # Update global destinations dictionary
        global ANNOUNCEMENT_DESTINATIONS
        ANNOUNCEMENT_DESTINATIONS = load_announcement_destinations()
        
        if success:
            bot.send_message(
                user_id,
                f"âœ… *Success!*\n\nThe destination '{dest_name}' has been added successfully.\n\n"
                f"You can now use it when sending announcements.",
                parse_mode="Markdown",
                reply_markup=ReplyKeyboardRemove()
            )
        else:
            bot.send_message(
                user_id,
                "âŒ There was an error saving the destination. Please try again later.",
                reply_markup=ReplyKeyboardRemove()
            )
    elif response == "âŒ Cancel":
        bot.send_message(
            user_id,
            "âŒ Destination creation cancelled.",
            reply_markup=ReplyKeyboardRemove()
        )
    else:
        bot.send_message(user_id, "âŒ Please select either 'âœ… Confirm' or 'âŒ Cancel'.")
        return
    
    # Clean up
    PENDING_USERS.pop(user_id, None)
    delete_pending_user(user_id)

# Command to list all announcement destinations
@bot.message_handler(commands=['list_destinations'])
def list_destinations_command(message):
    """Command to list all configured announcement destinations"""
    # Check if user is admin or creator
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    # Check if there are any destinations
    if not ANNOUNCEMENT_DESTINATIONS:
        bot.reply_to(message, "â„¹ï¸ No announcement destinations configured.")
        return
    
    # Build the list
    destination_list = "ðŸ“‹ *Configured Announcement Destinations*\n\n"
    
    for dest_id, dest in ANNOUNCEMENT_DESTINATIONS.items():
        destination_list += f"*{dest.get('name', 'Unknown')}* (`{dest_id}`)\n"
        destination_list += f"â€¢ Type: {dest.get('type', 'Unknown').capitalize()}\n"
        destination_list += f"â€¢ ID: `{dest.get('id', 'Unknown')}`\n"
        
        if dest.get('topic_id'):
            destination_list += f"â€¢ Topic ID: `{dest.get('topic_id')}`\n"
        else:
            destination_list += f"â€¢ Topic ID: None (Main chat)\n"
            
        destination_list += "\n"
    
    # Add instructions
    destination_list += "Use `/add_destination` to add a new destination or `/remove_destination` to remove one."
    
    bot.reply_to(message, destination_list, parse_mode="Markdown")

# Command to remove an announcement destination
@bot.message_handler(commands=['remove_destination'])
def remove_destination_command(message):
    """Command to remove an announcement destination"""
    # Check if user is admin or creator
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    args = message.text.split()
    
    # Show list of destinations if no ID provided
    if len(args) < 2:
        # Create inline buttons for each destination
        markup = InlineKeyboardMarkup(row_width=1)
        
        for dest_id, dest in ANNOUNCEMENT_DESTINATIONS.items():
            markup.add(InlineKeyboardButton(
                f"âŒ {dest.get('name', 'Unknown')} ({dest_id})",
                callback_data=f"remove_dest_{dest_id}"
            ))
        
        if not markup.keyboard:
            bot.reply_to(message, "â„¹ï¸ No announcement destinations configured.")
            return
            
        bot.reply_to(
            message,
            "ðŸ—‘ï¸ *Remove Announcement Destination*\n\n"
            "Please select a destination to remove:",
            parse_mode="Markdown",
            reply_markup=markup
        )
        return
    
    # If a destination ID is provided
    dest_id = args[1]
    remove_destination(message.chat.id, dest_id)

# Handle destination removal via callback
@bot.callback_query_handler(func=lambda call: call.data.startswith("remove_dest_"))
def handle_remove_destination_callback(call):
    """Handle removal of announcement destination via inline button"""
    # Check if user is admin or creator
    if call.from_user.id not in ADMIN_IDS and call.from_user.id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    # Extract destination ID
    dest_id = call.data[len("remove_dest_"):]
    
    # Create confirmation markup
    markup = InlineKeyboardMarkup()
    markup.add(
        InlineKeyboardButton("âœ… Yes, delete it", callback_data=f"confirm_remove_dest_{dest_id}"),
        InlineKeyboardButton("âŒ Cancel", callback_data="cancel_remove_dest")
    )
    
    # Get destination name
    dest_name = ANNOUNCEMENT_DESTINATIONS.get(dest_id, {}).get('name', 'Unknown')
    
    # Update message with confirmation
    bot.edit_message_text(
        f"ðŸ—‘ï¸ *Confirm Removal*\n\n"
        f"Are you sure you want to remove the destination:\n"
        f"*{dest_name}* (`{dest_id}`)?\n\n"
        f"This action cannot be undone.",
        call.message.chat.id,
        call.message.message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    bot.answer_callback_query(call.id)

# Handle removal confirmation
@bot.callback_query_handler(func=lambda call: call.data.startswith("confirm_remove_dest_"))
def handle_confirm_remove_destination(call):
    """Handle confirmation of destination removal"""
    global ANNOUNCEMENT_DESTINATIONS  # Add global declaration at the top
    
    # Check if user is admin or creator
    if call.from_user.id not in ADMIN_IDS and call.from_user.id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    # Extract destination ID
    dest_id = call.data[len("confirm_remove_dest_"):]
    
    # Delete from database
    success = delete_announcement_destination(dest_id)
    
    # Update global destinations dictionary
    ANNOUNCEMENT_DESTINATIONS = load_announcement_destinations()
    
    if success:
        bot.edit_message_text(
            f"âœ… Destination `{dest_id}` has been removed successfully.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
    else:
        bot.edit_message_text(
            f"âŒ Error removing destination `{dest_id}`. It may not exist or there was a database error.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
    
    bot.answer_callback_query(call.id, "Destination removal processed")

# Handle cancellation of removal
@bot.callback_query_handler(func=lambda call: call.data == "cancel_remove_dest")
def handle_cancel_remove_destination(call):
    """Handle cancellation of destination removal"""
    bot.edit_message_text(
        "âŒ Destination removal cancelled.",
        call.message.chat.id,
        call.message.message_id
    )
    
    bot.answer_callback_query(call.id, "Removal cancelled")

def remove_destination(chat_id, dest_id):
    """Remove a destination by ID"""
    global ANNOUNCEMENT_DESTINATIONS  # Move global declaration to the top
    
    # Check if destination exists
    if dest_id not in ANNOUNCEMENT_DESTINATIONS:
        bot.send_message(chat_id, f"âŒ Destination `{dest_id}` not found.", parse_mode="Markdown")
        return
    
    # Get destination name for confirmation
    dest_name = ANNOUNCEMENT_DESTINATIONS[dest_id]['name']
    
    # Delete from database
    success = delete_announcement_destination(dest_id)
    
    # Update global destinations dictionary
    ANNOUNCEMENT_DESTINATIONS = load_announcement_destinations()
    
    if success:
        bot.send_message(
            chat_id,
            f"âœ… Destination *{dest_name}* (`{dest_id}`) has been removed successfully.",
            parse_mode="Markdown"
        )
    else:
        bot.send_message(
            chat_id,
            f"âŒ Error removing destination `{dest_id}`. Please try again later.",
            parse_mode="Markdown"
        )

@bot.message_handler(commands=['config'])
def handle_config_command(message):
    """Central configuration menu for admin commands"""
    # Check if user is admin or creator
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    # Create inline keyboard with configuration options
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(
        InlineKeyboardButton("ðŸ”“ Enrollment Status", callback_data="config_enrollment"),
        InlineKeyboardButton("ðŸ’° Manage Discounts", callback_data="config_discount"),
        InlineKeyboardButton("ðŸŽŸï¸ Serial Number Management", callback_data="config_serial"),
        InlineKeyboardButton("ðŸŒ Destination Management", callback_data="config_destinations"),
        InlineKeyboardButton("ðŸ“¢ Send Announcement", callback_data="config_announce"),
        InlineKeyboardButton("ðŸ“Š Data Export", callback_data="config_export")
    )
    
    bot.send_message(
        message.chat.id,
        "âš™ï¸ *ADMIN CONFIGURATION PANEL*\n\n"
        "Select an option to configure:",
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.callback_query_handler(func=lambda call: call.data.startswith("config_"))
def handle_config_callbacks(call):
    """Handle config menu callbacks"""
    # Check if user is admin or creator
    if call.from_user.id not in ADMIN_IDS and call.from_user.id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
        
    option = call.data.split("_")[1]
    
    if option == "enrollment":
        # Show enrollment status menu
        show_enrollment_options(call)
    elif option == "discount":
        # Show discount options
        show_discount_options(call)
    elif option == "serial":
        # Show serial number options
        show_serial_options(call)
    elif option == "announce":
        # Start the announcement process
        handle_start_announcement(call)
    elif option == "export":
        # Show data export options
        show_export_options(call)
    elif option == "destinations":
        # Show destination management options
        show_destination_management(call)
    elif option == "back":
        # Return to main config menu
        show_config_menu(call)

def show_enrollment_options(call):
    """Show enrollment status options"""
    # Get current status for both membership types
    regular_enrollment_open = BOT_SETTINGS.get('regular_enrollment_open', True)  # Default to open
    supreme_enrollment_open = BOT_SETTINGS.get('supreme_enrollment_open', True)  # Default to open
    
    regular_status = "ðŸŸ¢ OPEN" if regular_enrollment_open else "ðŸ”´ CLOSED"
    supreme_status = "ðŸŸ¢ OPEN" if supreme_enrollment_open else "ðŸ”´ CLOSED"
    
    # Create keyboard with options for both membership types
    markup = InlineKeyboardMarkup(row_width=2)
    markup.add(
        InlineKeyboardButton("ðŸŸ¢ Open XM Partnership", callback_data="enrollment_regular_open"),
        InlineKeyboardButton("ðŸ”´ Close XM Partnership", callback_data="enrollment_regular_close")
    )
    markup.add(
        InlineKeyboardButton("ðŸŸ¢ Open Supreme", callback_data="enrollment_supreme_open"),
        InlineKeyboardButton("ðŸ”´ Close Supreme", callback_data="enrollment_supreme_close")
    )
    # Add buttons to control both enrollment types at once
    markup.add(
        InlineKeyboardButton("ðŸŸ¢ Open Both", callback_data="enrollment_both_open"),
        InlineKeyboardButton("ðŸ”´ Close Both", callback_data="enrollment_both_close")
    )
    markup.add(InlineKeyboardButton("Â« Back to Config Menu", callback_data="config_back"))
    
    # Edit the message with enrollment options for both types
    bot.edit_message_text(
        f"ðŸ”“ *ENROLLMENT STATUS*\n\n"
        f"*XM Partnership (Free):* {regular_status}\n"
        f"*Supreme Membership (Paid):* {supreme_status}\n\n"
        f"Select an option to change enrollment status:",
        call.message.chat.id,
        call.message.message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    bot.answer_callback_query(call.id)

def show_discount_options(call):
    """Show discount management options"""
    # Get current discount status
    reg_status = "ðŸŸ¢ Active" if DISCOUNTS.get('regular') else "ðŸ”´ Not active"
    sup_status = "ðŸŸ¢ Active" if DISCOUNTS.get('supreme') else "ðŸ”´ Not active"
    
    # Create keyboard with options
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(
        InlineKeyboardButton("âž• Create New Discount", callback_data="start_discount"),
        InlineKeyboardButton("âŒ Remove Discount", callback_data="remove_discount")
    )
    
    # Add discount-specific removal buttons if discounts are active
    if DISCOUNTS.get('regular'):
        markup.add(InlineKeyboardButton("âŒ Remove Regular Discount", callback_data="remove_discount_regular"))
    if DISCOUNTS.get('supreme'):
        markup.add(InlineKeyboardButton("âŒ Remove Supreme Discount", callback_data="remove_discount_supreme"))
    
    markup.add(InlineKeyboardButton("Â« Back to Config Menu", callback_data="config_back"))
    
    # Edit the message with discount options
    bot.edit_message_text(
        f"ðŸ’° *DISCOUNT MANAGEMENT*\n\n"
        f"Current Status:\n"
        f"â€¢ Regular Discount: {reg_status}\n"
        f"â€¢ Supreme Discount: {sup_status}\n\n"
        f"Select an option to manage discounts:",
        call.message.chat.id,
        call.message.message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    bot.answer_callback_query(call.id)

def show_config_menu(call):
    """Show main config menu (for back button)"""
    # Create inline keyboard with configuration options
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(
        InlineKeyboardButton("ðŸ”“ Enrollment Status", callback_data="config_enrollment"),
        InlineKeyboardButton("ðŸ’° Manage Discounts", callback_data="config_discount"),
        InlineKeyboardButton("ðŸŽŸï¸ Serial Number Management", callback_data="config_serial"),
        InlineKeyboardButton("ðŸŒ Destination Management", callback_data="config_destinations"),
        InlineKeyboardButton("ðŸ“¢ Send Announcement", callback_data="config_announce"),
        InlineKeyboardButton("ðŸ“Š Data Export", callback_data="config_export")
    )
    
    # Edit the message with main config menu
    bot.edit_message_text(
        "âš™ï¸ *ADMIN CONFIGURATION PANEL*\n\n"
        "Select an option to configure:",
        call.message.chat.id,
        call.message.message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda call: call.data == "start_discount")
def handle_start_discount_callback(call):
    """Start the discount creation process"""
    # Check if user is admin or creator
    if call.from_user.id not in ADMIN_IDS and call.from_user.id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    user_id = call.from_user.id
    
    # Answer the callback first
    bot.answer_callback_query(call.id)
    
    # Edit the message to acknowledge the action
    bot.edit_message_text(
        "ðŸ’° *DISCOUNT CREATION*\n\n"
        "Starting discount creation process...\n"
        "Please check your messages for the next steps.",
        call.message.chat.id,
        call.message.message_id,
        parse_mode="Markdown"
    )
    
    # Simulate the /discount command
    fake_message = types.Message(
        message_id=0,
        from_user=types.User(
            id=user_id,
            is_bot=False,
            first_name="Admin",
            username=call.from_user.username,
        ),
        date=0,
        chat=types.Chat(
            id=user_id,  # Send to the user's private chat
            type="private",
        ),
        content_type="text",
        options={},
        json_string="{}",
    )
    fake_message.text = "/discount"
    
    # Call the discount command handler
    start_discount_setup(fake_message)

@bot.callback_query_handler(func=lambda call: call.data == "remove_discount")
def handle_remove_discount_callback(call):
    """Start the discount removal process"""
    # Check if user is admin or creator
    if call.from_user.id not in ADMIN_IDS and call.from_user.id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    user_id = call.from_user.id
    
    # Answer the callback first
    bot.answer_callback_query(call.id)
    
    # Simulate the /remove_discount command
    fake_message = types.Message(
        message_id=0,
        from_user=types.User(
            id=user_id,
            is_bot=False,
            first_name="Admin",
            username=call.from_user.username,
        ),
        date=0,
        chat=types.Chat(
            id=call.message.chat.id,
            type="private",
        ),
        content_type="text",
        options={},
        json_string="{}",
    )
    fake_message.text = "/remove_discount"
    
    # Call the remove_discount command handler
    remove_discount(fake_message)

# Add this new function
def show_export_options(call):
    """Show data export options"""
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(
        InlineKeyboardButton("ðŸ“‹ Export Form Responses", callback_data="export_forms"),
        InlineKeyboardButton("ðŸ’µ Export Payment Records", callback_data="export_payments"),
        InlineKeyboardButton("Â« Back to Config Menu", callback_data="config_back")
    )
    
    bot.edit_message_text(
        "ðŸ“Š *DATA EXPORT OPTIONS*\n\n"
        "Select what data you want to export:",
        call.message.chat.id,
        call.message.message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda call: call.data in ["export_forms", "export_payments"])
def handle_export_callbacks(call):
    """Handle export button callbacks"""
    # Check if user is admin or creator
    if call.from_user.id not in ADMIN_IDS and call.from_user.id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    # Update message to show processing
    bot.edit_message_text(
        "ðŸ”„ Starting export process...\n\nPlease check your messages for the exported file.",
        call.message.chat.id,
        call.message.message_id
    )
    
    # Simulate the appropriate export command
    fake_message = types.Message(
        message_id=0,
        from_user=types.User(
            id=call.from_user.id,
            is_bot=False,
            first_name="Admin",
            username=call.from_user.username,
        ),
        date=0,
        chat=types.Chat(
            id=call.from_user.id,
            type="private",
        ),
        content_type="text",
        options={},
        json_string="{}",
    )
    
    if call.data == "export_forms":
        fake_message.text = "/export_forms"
        export_form_responses(fake_message)
    else:  # export_payments
        fake_message.text = "/export_payments"
        export_payment_data(fake_message)
    
    bot.answer_callback_query(call.id)

def handle_start_announcement(call):
    """Start the announcement creation process from config menu"""
    user_id = call.from_user.id
    
    # Answer the callback first
    bot.answer_callback_query(call.id)
    
    # Edit the message to acknowledge the action
    bot.edit_message_text(
        "ðŸ“¢ *ANNOUNCEMENT CREATION*\n\n"
        "Starting announcement creation process...\n"
        "Please check your messages for the next steps.",
        call.message.chat.id,
        call.message.message_id,
        parse_mode="Markdown"
    )
    
    # Simulate the /announce command
    fake_message = types.Message(
        message_id=0,
        from_user=types.User(
            id=user_id,
            is_bot=False,
            first_name="Admin",
            username=call.from_user.username,
        ),
        date=0,
        chat=types.Chat(
            id=user_id,  # Send to the user's private chat
            type="private",
        ),
        content_type="text",
        options={},
        json_string="{}",
    )
    fake_message.text = "/announce"
    
    # Call the announcement command handler
    start_announcement(fake_message)

def show_destination_management(call):
    """Show destination management options"""
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(
        InlineKeyboardButton("ðŸ“‹ List All Destinations", callback_data="list_destinations"),
        InlineKeyboardButton("âž• Add New Destination", callback_data="add_destination"),
        InlineKeyboardButton("ðŸ—‘ï¸ Remove Destination", callback_data="remove_dest")
    )
    markup.add(InlineKeyboardButton("Â« Back to Config Menu", callback_data="config_back"))
    
    bot.edit_message_text(
        "ðŸŒ *DESTINATION MANAGEMENT*\n\n"
        "Configure channels and groups where announcements can be sent:",
        call.message.chat.id,
        call.message.message_id,
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda call: call.data in ["list_destinations", "add_destination", "remove_dest"])
def handle_destination_callbacks(call):
    """Handle destination management callbacks"""
    # Check if user is admin or creator
    if call.from_user.id not in ADMIN_IDS and call.from_user.id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    user_id = call.from_user.id
    
    # Answer the callback first
    bot.answer_callback_query(call.id)
    
    if call.data == "list_destinations":
        # Use existing list_destinations functionality
        if not ANNOUNCEMENT_DESTINATIONS:
            bot.edit_message_text(
                "ðŸŒ *DESTINATION MANAGEMENT*\n\n"
                "â„¹ï¸ No announcement destinations configured yet.\n\n"
                "Use the 'Add New Destination' option to create one.",
                call.message.chat.id,
                call.message.message_id,
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup().add(
                    InlineKeyboardButton("Â« Back", callback_data="config_destinations")
                )
            )
            return
        
        # Build the list
        destination_list = "ðŸŒ *CONFIGURED DESTINATIONS*\n\n"
        
        for dest_id, dest in ANNOUNCEMENT_DESTINATIONS.items():
            destination_list += f"*{dest.get('name', 'Unknown')}* (`{dest_id}`)\n"
            destination_list += f"â€¢ Type: {dest.get('type', 'Unknown').capitalize()}\n"
            destination_list += f"â€¢ ID: `{dest.get('id', 'Unknown')}`\n"
            
            if dest.get('topic_id'):
                destination_list += f"â€¢ Topic ID: `{dest.get('topic_id')}`\n"
            else:
                destination_list += f"â€¢ Topic ID: None (Main chat)\n"
                
            destination_list += "\n"
        
        # Add back button
        markup = InlineKeyboardMarkup()
        markup.add(InlineKeyboardButton("Â« Back", callback_data="config_destinations"))
        
        bot.edit_message_text(
            destination_list,
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
        
    elif call.data == "add_destination":
        # Simulate /add_destination command
        fake_message = types.Message(
            message_id=0,
            from_user=types.User(
                id=user_id,
                is_bot=False,
                first_name="Admin",
                username=call.from_user.username,
            ),
            date=0,
            chat=types.Chat(
                id=user_id,  # Send to user's private chat
                type="private",
            ),
            content_type="text",
            options={},
            json_string="{}",
        )
        fake_message.text = "/add_destination"
        
        # Call the add_destination command handler
        add_destination_command(fake_message)
        
        # Update the message to indicate the action
        bot.edit_message_text(
            "ðŸŒ *DESTINATION MANAGEMENT*\n\n"
            "âž• Add destination process started!\n\n"
            "Please check your direct messages to continue setting up the new destination.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup().add(
                InlineKeyboardButton("Â« Back to Config Menu", callback_data="config_back")
            )
        )
        
    elif call.data == "remove_dest":
        # Show list of destinations to remove
        markup = InlineKeyboardMarkup(row_width=1)
        
        for dest_id, dest in ANNOUNCEMENT_DESTINATIONS.items():
            markup.add(InlineKeyboardButton(
                f"âŒ {dest.get('name', 'Unknown')} ({dest_id})",
                callback_data=f"remove_dest_{dest_id}"
            ))
        
        # Add back button
        markup.add(InlineKeyboardButton("Â« Back", callback_data="config_destinations"))
        
        if not markup.keyboard or len(markup.keyboard) == 1:  # Only back button exists
            bot.edit_message_text(
                "ðŸŒ *DESTINATION MANAGEMENT*\n\n"
                "â„¹ï¸ No destinations available to remove.",
                call.message.chat.id,
                call.message.message_id,
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup().add(
                    InlineKeyboardButton("Â« Back", callback_data="config_destinations")
                )
            )
        else:
            bot.edit_message_text(
                "ðŸ—‘ï¸ *REMOVE DESTINATION*\n\n"
                "Select a destination to remove:",
                call.message.chat.id,
                call.message.message_id,
                parse_mode="Markdown",
                reply_markup=markup
            )

def show_serial_options(call):
    """Show serial number management options"""
    try:
        # Count available and used serials
        total_serials = len(SERIAL_NUMBERS) if SERIAL_NUMBERS else 0
        used_serials = sum(1 for data in SERIAL_NUMBERS.values() if data.get('used', False)) if SERIAL_NUMBERS else 0
        available_serials = total_serials - used_serials
        
        # Create keyboard with options
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(
            InlineKeyboardButton("âž• Generate New Serial", callback_data="serial_generate"),
            InlineKeyboardButton("âŒ Delete Serial", callback_data="serial_delete"),
            InlineKeyboardButton("Â« Back to Config Menu", callback_data="config_back")
        )
        
        # Edit the message with serial options
        bot.edit_message_text(
            f"ðŸŽŸï¸ *SERIAL NUMBER MANAGEMENT*\n\n"
            f"Current Status:\n"
            f"â€¢ Total Serials: {total_serials}\n"
            f"â€¢ Available Serials: {available_serials}\n"
            f"â€¢ Used Serials: {used_serials}\n\n"
            f"Select an option to manage serial numbers:",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
        
        bot.answer_callback_query(call.id)
        
    except Exception as e:
        # Log the error
        logging.error(f"Error in show_serial_options: {e}")
        
        # Try to notify user
        try:
            bot.answer_callback_query(call.id, "Error showing options", show_alert=True)
        except:
            pass

@bot.callback_query_handler(func=lambda call: call.data.startswith("serial_"))
def handle_serial_callbacks(call):
    """Handle serial number management callbacks"""
    # Check if user is admin or creator
    if call.from_user.id not in ADMIN_IDS and call.from_user.id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    try:
        option = call.data.split("_")[1]
        user_id = call.from_user.id
        
        # Log the callback for debugging
        logging.info(f"Serial callback received: {call.data} from user {user_id}")
        
        if option == "generate":
            # Start the serial generation process
            bot.answer_callback_query(call.id)
            bot.edit_message_text(
                "ðŸŽŸï¸ *SERIAL NUMBER GENERATION*\n\n"
                "Starting serial generation process...\n"
                "Please check your messages for the next steps.",
                call.message.chat.id,
                call.message.message_id,
                parse_mode="Markdown"
            )
            
            # Simulate the /generate command
            fake_message = types.Message(
                message_id=0,
                from_user=types.User(
                    id=user_id,
                    is_bot=False,
                    first_name="Admin",
                    username=call.from_user.username,
                ),
                date=0,
                chat=types.Chat(
                    id=user_id,  # Send to the user's private chat
                    type="private",
                ),
                content_type="text",
                options={},
                json_string="{}",
            )
            fake_message.text = "/generate"
            
            # Call the generate command handler
            generate_serial(fake_message)
            
        elif option == "list_all" or option == "list_available" or option == "list_used":
            # Determine filter status
            filter_status = None
            if option == "list_available":
                filter_status = False
            elif option == "list_used":
                filter_status = True
            
            # Log filter info
            logging.info(f"Listing serials with filter_status: {filter_status}")
            
            # Check if SERIAL_NUMBERS is properly loaded
            if not SERIAL_NUMBERS:
                markup = InlineKeyboardMarkup()
                markup.add(InlineKeyboardButton("Â« Back to Serial Menu", callback_data="config_serial"))
                
                bot.edit_message_text(
                    "ðŸ”‘ *SERIAL NUMBERS*\n\n"
                    "No serial numbers found in database.",
                    call.message.chat.id,
                    call.message.message_id,
                    parse_mode="Markdown",
                    reply_markup=markup
                )
                bot.answer_callback_query(call.id, "No serial numbers found")
                return
            
            # Build the list
            serial_list = "ðŸ”‘ *SERIAL NUMBERS*\n\n"
            count = 0
            
            # Create a safe copy for iteration
            serials_to_process = dict(SERIAL_NUMBERS)
            
            # Process each serial based on filter
            for serial, data in serials_to_process.items():
                is_used = data.get('used', False)
                
                # Apply filter if specified
                if filter_status is not None and is_used != filter_status:
                    continue
                
                count += 1
                if count > 20:  # Limit to 20 serials to avoid message length issues
                    serial_list += f"\n... and {len(serials_to_process) - 20} more serials."
                    break
                
                # Format serial info
                status = "âœ… Used" if is_used else "â³ Available"
                mentorship_type = data.get('mentorship_type', 'regular').capitalize()
                plan = data.get('plan', 'Unknown')
                
                serial_list += f"*{serial}*\n"
                serial_list += f"â€¢ Status: {status}\n"
                serial_list += f"â€¢ Type: {mentorship_type}\n"
                serial_list += f"â€¢ Plan: {plan}\n"
                
                if is_used:
                    used_by = data.get('used_by', 'Unknown')
                    used_at = data.get('used_at', 'Unknown date')
                    serial_list += f"â€¢ Used by: User {used_by}\n"
                    serial_list += f"â€¢ Used on: {used_at}\n"
                
                serial_list += "\n"
            
            # Create back button
            markup = InlineKeyboardMarkup()
            markup.add(InlineKeyboardButton("Â« Back to Serial Menu", callback_data="config_serial"))
            
            # Add filter info
            filter_note = ""
            if filter_status is not None:
                filter_type = "used" if filter_status else "available"
                filter_note = f" (filtered: {filter_type} only)"
            
            # Handle no matching serials case
            if count == 0:
                if filter_status is not None:
                    msg = f"No {filter_type} serials found."
                else:
                    msg = "No serials found."
                
                bot.edit_message_text(
                    f"ðŸ”‘ *SERIAL NUMBERS*\n\n{msg}",
                    call.message.chat.id,
                    call.message.message_id,
                    parse_mode="Markdown",
                    reply_markup=markup
                )
            else:
                # Complete the message with count note
                serial_list += f"Total: {count} serials{filter_note}"
                
                # Edit the message with the serial list
                try:
                    bot.edit_message_text(
                        serial_list,
                        call.message.chat.id,
                        call.message.message_id,
                        parse_mode="Markdown",
                        reply_markup=markup
                    )
                except Exception as e:
                    logging.error(f"Error sending serial list: {e}")
                    # Try sending without markdown if that might be the issue
                    bot.edit_message_text(
                        f"Error with markdown formatting. {count} serials found.",
                        call.message.chat.id,
                        call.message.message_id,
                        reply_markup=markup
                    )
            
            bot.answer_callback_query(call.id, f"Showing {count} serials")
            logging.info(f"Listed {count} serials with filter {filter_status}")
            
        elif option == "delete":
            # Show serial deletion interface
            show_serial_deletion(call)
        
    except Exception as e:
        # Log the error
        logging.error(f"Error in serial_callbacks: {e}")
        
        # Try to send an error message
        try:
            markup = InlineKeyboardMarkup()
            markup.add(InlineKeyboardButton("Â« Back to Serial Menu", callback_data="config_serial"))
            
            bot.edit_message_text(
                "âŒ *An error occurred*\n\nThe operation could not be completed.",
                call.message.chat.id,
                call.message.message_id,
                parse_mode="Markdown",
                reply_markup=markup
            )
            bot.answer_callback_query(call.id, "Error occurred", show_alert=True)
        except:
            # Last resort attempt
            try:
                bot.answer_callback_query(call.id, "Error handling request", show_alert=True)
            except:
                pass

def show_serial_deletion(call):
    """Show interface to delete a serial number"""
    # Check if there are any serials
    if not SERIAL_NUMBERS:
        markup = InlineKeyboardMarkup()
        markup.add(InlineKeyboardButton("Â« Back to Serial Menu", callback_data="config_serial"))
        
        bot.edit_message_text(
            "ðŸ”‘ *DELETE SERIAL*\n\n"
            "No serials found to delete.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown",
            reply_markup=markup
        )
        bot.answer_callback_query(call.id)
        return
    
    # Edit message to guide user
    bot.edit_message_text(
        "ðŸ”‘ *DELETE SERIAL*\n\n"
        "Please enter the serial number you want to delete in the chat.\n\n"
        "Or use the inline keyboard to go back.",
        call.message.chat.id,
        call.message.message_id,
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup().add(
            InlineKeyboardButton("Â« Back to Serial Menu", callback_data="config_serial")
        )
    )
    
    # Set user state to handle the delete in the next message
    PENDING_USERS[call.from_user.id] = {
        'status': 'deleting_serial',
        'config_msg_id': call.message.message_id,
        'config_chat_id': call.message.chat.id
    }
    save_pending_users()
    
    bot.answer_callback_query(call.id)

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'deleting_serial')
def handle_delete_serial_input(message):
    """Process admin's input of serial to delete"""
    user_id = message.from_user.id
    
    # Check if user is admin or creator
    if user_id not in ADMIN_IDS and user_id != CREATOR_ID:
        return
    
    # Get the serial number
    serial = message.text.strip()
    
    # Get stored message info
    config_chat_id = PENDING_USERS[user_id].get('config_chat_id')
    config_msg_id = PENDING_USERS[user_id].get('config_msg_id')
    
    # Check if serial exists
    if serial not in SERIAL_NUMBERS:
        bot.reply_to(message, f"âŒ Serial `{serial}` not found.", parse_mode="Markdown")
        return
    
    # Get serial info for confirmation
    serial_data = SERIAL_NUMBERS[serial]
    status = "Used" if serial_data.get('used', False) else "Available"
    mentorship_type = serial_data.get('mentorship_type', 'regular').capitalize()
    plan = serial_data.get('plan', 'Unknown')
    
    # Create a confirmation keyboard
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(
        InlineKeyboardButton("âœ… Yes, delete this serial", callback_data=f"confirm_delete_serial_{serial}"),
        InlineKeyboardButton("âŒ No, cancel", callback_data="config_serial")
    )
    
    # Send confirmation message
    bot.send_message(
        user_id,
        f"ðŸ—‘ï¸ *CONFIRM SERIAL DELETION*\n\n"
        f"Are you sure you want to delete this serial?\n\n"
        f"*Serial:* `{serial}`\n"
        f"*Status:* {status}\n"
        f"*Type:* {mentorship_type}\n"
        f"*Plan:* {plan}\n\n"
        f"âš ï¸ This action cannot be undone.",
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    # Update the config menu message to avoid confusion
    if config_chat_id and config_msg_id:
        try:
            bot.edit_message_text(
                "ðŸ”‘ *DELETE SERIAL*\n\n"
                "Please check your most recent message for the deletion confirmation.",
                config_chat_id,
                config_msg_id,
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup().add(
                    InlineKeyboardButton("Â« Back to Serial Menu", callback_data="config_serial")
                )
            )
        except:
            pass  # Ignore errors if message can't be edited
    
    # Clean up user state
    PENDING_USERS.pop(user_id, None)
    delete_pending_user(user_id)

@bot.callback_query_handler(func=lambda call: call.data.startswith("confirm_delete_serial_"))
def handle_confirm_delete_serial(call):
    """Handle confirmation of serial deletion"""
    user_id = call.from_user.id
    
    # Check if user is admin or creator
    if user_id not in ADMIN_IDS and user_id != CREATOR_ID:
        bot.answer_callback_query(call.id, "âŒ You are not authorized to perform this action.", show_alert=True)
        return
    
    # Extract the serial number
    serial = call.data[len("confirm_delete_serial_"):]
    
    # Check if serial exists
    if serial not in SERIAL_NUMBERS:
        bot.edit_message_text(
            "âŒ Error: Serial not found.",
            call.message.chat.id,
            call.message.message_id
        )
        bot.answer_callback_query(call.id, "Serial not found")
        return
    
    # Store serial info before deleting
    serial_data = SERIAL_NUMBERS[serial]
    status = "Used" if serial_data.get('used', False) else "Available"
    mentorship_type = serial_data.get('mentorship_type', 'regular').capitalize()
    plan = serial_data.get('plan', 'Unknown')
    
    # Delete from database
    try:
        serial_numbers_collection.delete_one({"serial": serial})
        del SERIAL_NUMBERS[serial]
        
        # Update message to show success
        bot.edit_message_text(
            f"âœ… *SERIAL DELETED SUCCESSFULLY*\n\n"
            f"*Serial:* `{serial}`\n"
            f"*Status:* {status}\n"
            f"*Type:* {mentorship_type}\n"
            f"*Plan:* {plan}\n\n"
            f"This serial has been permanently deleted.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup().add(
                InlineKeyboardButton("Â« Back to Serial Menu", callback_data="config_serial")
            )
        )
        
        # Log the action
        admin_username = call.from_user.username or f"Admin {user_id}"
        logging.info(f"Serial {serial} deleted by {admin_username}")
        
        # Notify other admins for transparency
        for admin_id in ADMIN_IDS:
            if admin_id != user_id:  # Don't send to the admin who deleted it
                try:
                    bot.send_message(
                        admin_id,
                        f"ðŸ—‘ï¸ *SERIAL DELETED*\n\n"
                        f"Admin @{admin_username} has deleted serial:\n"
                        f"`{serial}`\n\n"
                        f"*Type:* {mentorship_type}\n"
                        f"*Plan:* {plan}\n"
                        f"*Status:* {status}",
                        parse_mode="Markdown"
                    )
                except Exception as e:
                    logging.error(f"Failed to notify admin {admin_id} about serial deletion: {e}")
        
        bot.answer_callback_query(call.id, "Serial deleted successfully")
        
    except Exception as e:
        # Handle error
        bot.edit_message_text(
            f"âŒ Error deleting serial: {str(e)}",
            call.message.chat.id,
            call.message.message_id,
            reply_markup=InlineKeyboardMarkup().add(
                InlineKeyboardButton("Â« Back to Serial Menu", callback_data="config_serial")
            )
        )
        logging.error(f"Error deleting serial {serial}: {e}")
        bot.answer_callback_query(call.id, "Error deleting serial")

def send_birthday_greetings():
    """Check for birthdays and send greetings to users"""
    logging.info("Running birthday greeting check")
    
    # Get current date in Manila timezone for consistency
    manila_tz = pytz.timezone('Asia/Manila')
    now = datetime.now(manila_tz)
    today_day = now.day
    today_month = now.month
    
    # Count for logging
    greetings_sent = 0
    
    # Check each user in PAYMENT_DATA
    for user_id_str, data in PAYMENT_DATA.items():
        # Skip users who aren't active members
        if not data.get('haspayed', False) or data.get('cancelled', False):
            continue
            
        try:
            # Get birthday from form answers if available
            birthday = None
            if 'form_answers' in data and 'birthday' in data['form_answers']:
                birthday = data['form_answers']['birthday']
                
            # Skip if no birthday info
            if not birthday:
                continue
                
            # Try to parse birthday in DD/MM/YYYY format
            try:
                day, month, year = map(int, birthday.split('/'))
                
                # Check if today is their birthday (day and month match)
                if day == today_day and month == today_month:
                    # Get username or full name for personalization
                    user_id = int(user_id_str)
                    username = data.get('username', '')
                    full_name = data.get('form_answers', {}).get('full_name', '')
                    display_name = full_name or f"@{username}" or f"User {user_id}"
                    
                    # Calculate age
                    age = now.year - year
                    
                    # Send birthday greeting with personalized touches
                    greeting_templates = [
                        f"ðŸŽ‚ *Happy Birthday, {display_name}!* ðŸŽ‰\n\nWishing you a fantastic day filled with joy and success! May your {age}th year be your best trading year yet! ðŸ“ˆ",
                        
                        f"ðŸŽˆ *Birthday Alert!* ðŸŽˆ\n\nHappy Birthday to {display_name}! The entire Prodigy Trading Academy team wishes you an amazing {age}th birthday. May your candles be many and your losses few! ðŸ’°",
                        
                        f"ðŸ¥³ *Special Day Alert!* ðŸŽ\n\n{display_name}, happy {age}th birthday! May your charts be green, your strategies profitable, and your year ahead absolutely spectacular!"
                    ]
                    
                    greeting = random.choice(greeting_templates)
                    
                    # Send the greeting
                    bot.send_message(
                        user_id,
                        greeting,
                        parse_mode="Markdown"
                    )
                    
                    # Log successful greeting
                    logging.info(f"Birthday greeting sent to user {user_id} ({display_name})")
                    greetings_sent += 1
                    
                    # Optional: If you want to announce birthdays in the group
                    if ANNOUNCEMENT_TOPIC_ID:
                        # Only share first name for privacy in public
                        first_name = full_name.split()[0] if full_name else (username or f"one of our members")
                        
                        # Send more generic announcement to the group
                        bot.send_message(
                            PAID_GROUP_ID,
                            f"ðŸŽ‚ *Happy Birthday!* ðŸŽ‰\n\nPlease join us in wishing {first_name} a wonderful birthday today! ðŸ¥³",
                            parse_mode="Markdown",
                            message_thread_id=ANNOUNCEMENT_TOPIC_ID
                        )
            except (ValueError, IndexError):
                # Skip users with incorrectly formatted birthdays
                logging.warning(f"Invalid birthday format for user {user_id_str}: {birthday}")
                continue
                
        except Exception as e:
            logging.error(f"Error processing birthday for user {user_id_str}: {e}")
            continue
    
    logging.info(f"Birthday check complete. Sent {greetings_sent} birthday greetings")

def birthday_check_thread():
    """Thread to check for user birthdays daily"""
    logging.info("Birthday check thread started")
    
    # Track the last date we checked birthdays
    last_check_date = None
    
    while True:
        try:
            # Get current time in Manila timezone
            manila_tz = pytz.timezone('Asia/Manila')
            now = datetime.now(manila_tz)
            current_time = now.strftime('%H:%M')
            current_date = now.strftime('%Y-%m-%d')
            
            # Run birthday checks at 9:00 AM Manila time, once per day
            if current_time == '09:00' and current_date != last_check_date:
                send_birthday_greetings()
                
                # Update the last check date
                last_check_date = current_date
            
            # Calculate the time to sleep until the start of the next minute
            sleep_time = 60 - now.second - now.microsecond / 1_000_000
            time.sleep(sleep_time)
            
        except Exception as e:
            logging.error(f"Error in birthday check thread: {e}")
            time.sleep(60)  # Wait a minute on error before trying again

@bot.message_handler(commands=['test_birthday'])
def test_birthday_message(message):
    """Admin command to test birthday messages"""
    # Check if user is admin or creator
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    args = message.text.split()
    
    if len(args) < 2:
        # Send to the admin who requested it
        user_id = message.from_user.id
        display_name = message.from_user.first_name
        
        # Send a test greeting
        greeting = f"ðŸŽ‚ *Happy Birthday, {display_name}!* ðŸŽ‰\n\nThis is a test birthday greeting. Your real one will look similar to this!"
        
        bot.send_message(
            user_id,
            greeting,
            parse_mode="Markdown"
        )
        
        bot.reply_to(message, "âœ… Test birthday message sent to you. To send to another user, use /test_birthday [user_id]")
    else:
        try:
            # Send to specified user
            target_user_id = int(args[1])
            
            # Check if user exists in PAYMENT_DATA
            if str(target_user_id) in PAYMENT_DATA:
                data = PAYMENT_DATA[str(target_user_id)]
                username = data.get('username', '')
                full_name = data.get('form_answers', {}).get('full_name', '')
                display_name = full_name or f"@{username}" or f"User {target_user_id}"
                
                # Send a test greeting
                greeting = f"ðŸŽ‚ *Happy Birthday, {display_name}!* ðŸŽ‰\n\n(This is a test birthday greeting sent by an admin)"
                
                bot.send_message(
                    target_user_id,
                    greeting,
                    parse_mode="Markdown"
                )
                
                bot.reply_to(message, f"âœ… Test birthday message sent to user {target_user_id}")
            else:
                bot.reply_to(message, f"âŒ User {target_user_id} not found in payment data")
        except ValueError:
            bot.reply_to(message, "âŒ Invalid user ID. Please provide a numeric ID.")
        except Exception as e:
            bot.reply_to(message, f"âŒ Error sending test birthday message: {str(e)}")

@bot.message_handler(commands=['generate'])
def generate_serial(message):
    """Generate a serial number for giveaways (admin only)"""
    # Check if user is admin or creator
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
        
    # Ask for mentorship type
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row("Regular Mentorship", "Supreme Mentorship")
    
    msg = bot.reply_to(
        message,
        "ðŸŽ *SERIAL NUMBER GENERATOR*\n\n"
        "What type of mentorship is this serial for?",
        parse_mode="Markdown",
        reply_markup=markup
    )
    
    # Set user state
    PENDING_USERS[message.from_user.id] = {
        'status': 'generating_serial_type',
        'message_id': msg.message_id
    }
    save_pending_users()

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'generating_serial_type')
def handle_serial_mentorship_type(message):
    """Handle mentorship type selection for serial generation"""
    user_id = message.from_user.id
    mentorship_type = message.text.strip().lower()
    
    # Validate mentorship type
    valid_types = ["regular mentorship", "supreme mentorship"]
    if mentorship_type not in valid_types:
        bot.send_message(user_id, "âŒ Invalid selection. Please choose Regular or Supreme Mentorship.")
        return
    
    # Extract type
    mentorship_type = mentorship_type.split()[0].lower()  # "regular" or "supreme"
    
    # Store the selection
    PENDING_USERS[user_id]['mentorship_type'] = mentorship_type
    PENDING_USERS[user_id]['status'] = 'generating_serial_plan'
    save_pending_users()
    
    # Create plan selection keyboard
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    
    if mentorship_type == "regular":
        markup.row("Trial Plan (1 Month)")
        markup.row("Momentum Plan (3 Months)")
        markup.row("Legacy Plan (Year)")
    else:  # Supreme
        markup.row("Apprentice Plan (3 Month)")
        markup.row("Disciple Plan (6 Months)")
        markup.row("Lifetime Plan (Permanent)")
    
    # Ask for plan selection
    bot.send_message(
        user_id,
        f"ðŸ“ *SELECT PLAN*\n\n"
        f"Which {mentorship_type.capitalize()} Mentorship plan is this serial for?",
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'generating_serial_plan')
def handle_serial_plan(message):
    """Handle plan selection for serial generation"""
    user_id = message.from_user.id
    plan_text = message.text.strip()
    
    # Extract plan ID from text
    if "trial" in plan_text.lower():
        plan_id = "Trial"
    elif "momentum" in plan_text.lower():
        plan_id = "Momentum"
    elif "legacy" in plan_text.lower():
        plan_id = "Legacy"
    elif "apprentice" in plan_text.lower():
        plan_id = "Apprentice"
    elif "disciple" in plan_text.lower():
        plan_id = "Disciple"
    elif "lifetime" in plan_text.lower():
        plan_id = "Legacy"  # Using Legacy for Lifetime plan
    else:
        bot.send_message(user_id, "âŒ Invalid plan selection. Please select from the provided options.")
        return
    
    # Store the plan
    PENDING_USERS[user_id]['plan'] = plan_id
    PENDING_USERS[user_id]['mentorship_type_display'] = PENDING_USERS[user_id]['mentorship_type'].capitalize()
    PENDING_USERS[user_id]['status'] = 'confirming_serial_generation'
    save_pending_users()
    
    # Create confirmation keyboard
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row("âœ… Yes, Generate Serial", "âŒ No, Cancel")
    
    # Ask for confirmation
    bot.send_message(
        user_id,
        f"ðŸ” *CONFIRM SERIAL GENERATION*\n\n"
        f"You are about to generate a serial for:\n\n"
        f"â€¢ *Mentorship Type:* {PENDING_USERS[user_id]['mentorship_type_display']}\n"
        f"â€¢ *Plan:* {plan_id}\n\n"
        f"Is this correct?",
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'confirming_serial_generation')
def handle_serial_confirmation(message):
    """Handle confirmation for serial generation"""
    user_id = message.from_user.id
    response = message.text.strip()
    
    # Create a keyboard remover
    markup = ReplyKeyboardRemove()
    
    if response == "âœ… Yes, Generate Serial":
        # Proceed with generation
        # Call the function without parameters and process the result
        serial = generate_unique_serial()
        process_generated_serial(user_id, message, serial)
    elif response == "âŒ No, Cancel":
        bot.send_message(
            user_id,
            "âŒ Serial generation cancelled.",
            reply_markup=markup
        )
        PENDING_USERS.pop(user_id, None)
        delete_pending_user(user_id)
    else:
        bot.send_message(user_id, "âŒ Invalid response. Please select one of the provided options.")


def generate_unique_serial():
    """Generate a unique serial number in Synapse X style format"""
    while True:
        # Generate UUID-like components using hexadecimal characters
        hex_chars = '0123456789abcdef'
        
        # Generate the 5 parts with appropriate lengths
        part1 = ''.join(random.choice(hex_chars) for _ in range(8))
        part2 = ''.join(random.choice(hex_chars) for _ in range(4))
        part3 = ''.join(random.choice(hex_chars) for _ in range(4))
        part4 = ''.join(random.choice(hex_chars) for _ in range(4))
        part5 = ''.join(random.choice(hex_chars) for _ in range(12))
        
        # Combine with hyphens in Synapse X format
        serial = f"{part1}-{part2}-{part3}-{part4}-{part5}"
        
        # Check if serial already exists
        if serial not in SERIAL_NUMBERS:
            return serial


def process_generated_serial(user_id, message, serial):
    """Process a newly generated serial number"""
    # Create a keyboard remover
    markup = ReplyKeyboardRemove()
    
    # Get the user data from the pending users dict
    user_data = PENDING_USERS.get(user_id, {})
    mentorship_type = user_data.get('mentorship_type', 'regular')
    plan = user_data.get('plan', 'Trial')
    
    # Create a document for the serial
    serial_doc = {
        "serial": serial,
        "mentorship_type": mentorship_type,
        "plan": plan,
        "created_by": user_id,
        "created_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "used": False
    }
    
    # Save to database
    try:
        serial_numbers_collection.insert_one(serial_doc)
        
        # Update the global dictionary
        SERIAL_NUMBERS[serial] = serial_doc
        
        # Format mentorship type for display
        mentorship_display = mentorship_type.capitalize()
        
        # Send confirmation with the serial
        bot.send_message(
            user_id,
            f"âœ… *SERIAL GENERATED SUCCESSFULLY*\n\n"
            f"â€¢ *Serial:* `{serial}`\n"
            f"â€¢ *Mentorship Type:* {mentorship_display}\n"
            f"â€¢ *Plan:* {plan}\n\n"
            f"This serial can be redeemed using the `/redeem` command.",
            parse_mode="Markdown",
            reply_markup=markup
        )
        
        # Log the action
        admin_username = message.from_user.username or f"Admin {user_id}"
        logging.info(f"Serial {serial} generated by {admin_username} for {mentorship_type} {plan}")
        
    except Exception as e:
        # Send error message
        bot.send_message(
            user_id,
            f"âŒ Error generating serial: {str(e)}",
            reply_markup=markup
        )
        logging.error(f"Error saving serial: {e}")
    
    # Clean up
    PENDING_USERS.pop(user_id, None)
    delete_pending_user(user_id)
        
@bot.message_handler(commands=['redeem'])
def redeem_serial(message):
    """Redeem a serial number for a free membership"""
    user_id = message.from_user.id
    chat_id = message.chat.id
    
    # Only allow in private chat
    if message.chat.type != 'private':
        bot.reply_to(message, "ðŸ”’ Please use this command in a private message with the bot for security.")
        return
    
    # Check if user is already in the middle of something
    if user_id in PENDING_USERS:
        current_status = PENDING_USERS[user_id].get('status', '')
        if current_status.startswith('redeeming'):
            bot.reply_to(message, "â³ You already have a redemption in progress. Please complete it or send /cancel to stop.")
            return
    
    # Start the redemption process
    PENDING_USERS[user_id] = {
        'status': 'redeeming_serial',
    }
    save_pending_users()
    
    # Ask for serial number
    bot.reply_to(
        message,
        "ðŸŽ *SERIAL NUMBER REDEMPTION*\n\n"
        "Please enter your serial number in the exact format it was provided (e.g., 629326d5-a191-cefb-e7e8-e18bcd774b1a).\n\n"
        "Type /cancelRedeem to cancel the redemption process.",
        parse_mode="Markdown"
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'redeeming_serial')
def handle_serial_input(message):
    """Process the serial number input from the user"""
    user_id = message.from_user.id
    serial_input = message.text.strip()  # Don't convert to uppercase here
    
    # Check if it's a cancel command
    if serial_input.strip() == '/cancelRedeem':
        PENDING_USERS.pop(user_id, None)
        delete_pending_user(user_id)
        bot.reply_to(message, "âŒ Serial redemption cancelled.")
        return
    
    # Validate serial number format (optional)
    if not re.match(r'^[a-f0-9]{8}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{12}$', serial_input.lower()):
        bot.reply_to(
            message, 
            "âŒ Invalid serial format. Please enter the serial exactly as provided (e.g., 629326d5-a191-cefb-e7e8-e18bcd774b1a).\n\n"
            "Try again or send /cancelRedeem to stop."
        )
        return
    
    # Case-insensitive check for the serial key
    serial = None
    for key in SERIAL_NUMBERS:
        if key.lower() == serial_input.lower():
            serial = key  # Use the original key from the dictionary
            break
    
    # Check if serial exists in database
    if serial is None:
        bot.reply_to(
            message,
            "âŒ *Invalid Serial Number*\n\n"
            "The serial number you entered is not recognized. Please check and try again, or contact an administrator for assistance.\n\n"
            "Try again or send /cancelRedeem to stop.",
            parse_mode="Markdown"
        )
        return
    
    # Rest of your code remains the same...
    # Check if serial has already been used
    serial_data = SERIAL_NUMBERS[serial]
    if serial_data.get('used', False):
        used_at = serial_data.get('used_at', 'Unknown date')
        bot.reply_to(
            message,
            f"âŒ *Serial Already Used*\n\n"
            f"This serial number has already been redeemed on {used_at}.\n\n"
            f"Please contact an administrator if you believe this is an error.",
            parse_mode="Markdown"
        )
        return
    
    # Store serial and proceed with confirmation
    PENDING_USERS[user_id]['serial'] = serial
    PENDING_USERS[user_id]['serial_data'] = serial_data
    PENDING_USERS[user_id]['status'] = 'confirming_redemption'
    save_pending_users()
    
    # Get plan details
    mentorship_type = serial_data.get('mentorship_type', 'regular').capitalize()
    plan = serial_data.get('plan', 'Unknown')
    
    # Create confirmation keyboard
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row("âœ… Yes, Redeem Now", "âŒ Cancel")
    
    # Send confirmation message
    bot.reply_to(
        message,
        f"ðŸ” *SERIAL VERIFICATION SUCCESSFUL*\n\n"
        f"You are about to redeem a serial for:\n"
        f"â€¢ *Mentorship Type:* {mentorship_type}\n"
        f"â€¢ *Plan:* {plan}\n\n"
        f"Would you like to redeem this serial now?",
        parse_mode="Markdown",
        reply_markup=markup
    )

@bot.message_handler(func=lambda message: PENDING_USERS.get(message.from_user.id, {}).get('status') == 'confirming_redemption')
def handle_redemption_confirmation(message):
    """Process the user's confirmation for serial redemption"""
    user_id = message.from_user.id
    response = message.text.strip()
    
    # Create a keyboard remover
    markup = ReplyKeyboardRemove()
    
    if response == "âœ… Yes, Redeem Now":
        # Get serial data
        serial = PENDING_USERS[user_id]['serial']
        serial_data = PENDING_USERS[user_id]['serial_data']
        
        # Mark serial as used
        serial_data['used'] = True
        serial_data['used_by'] = user_id
        serial_data['used_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Update database
        if save_serial_number(serial, serial_data):
            # Update global variable
            SERIAL_NUMBERS[serial] = serial_data
            
            # Get user info
            username = message.from_user.username or f"User {user_id}"
            
            # Set up payment data for this user
            user_id_str = str(user_id)
            mentorship_type = serial_data.get('mentorship_type', 'regular')
            plan = serial_data.get('plan', 'Trial')
            
            # Calculate due date based on plan - UPDATED TO MATCH UI
            due_date = None
            if plan == "Trial":  # Regular 1 month
                due_date = datetime.now() + timedelta(days=30)
            elif plan == "Momentum":  # Regular 3 months
                due_date = datetime.now() + timedelta(days=90)
            elif plan == "Legacy" and mentorship_type == "regular":  # Regular year
                due_date = datetime.now() + timedelta(days=365)  # Fixed: Was 180, now 365 days
            elif plan == "Apprentice":  # Supreme 3 months
                due_date = datetime.now() + timedelta(days=90)  # Fixed: Was 30, now 90 days
            elif plan == "Disciple":  # Supreme 6 months
                due_date = datetime.now() + timedelta(days=180)  # Fixed: Was 90, now 180 days
            elif plan == "Legacy" and mentorship_type == "supreme":  # Supreme lifetime
                due_date = datetime.now() + timedelta(days=3650)  # ~10 years (effectively permanent)
                
            # Create or update payment data
            if user_id_str in PAYMENT_DATA:
                # Update existing data
                PAYMENT_DATA[user_id_str].update({
                    'username': username,
                    'payment_plan': plan,
                    'mentorship_type': mentorship_type.capitalize(),
                    'payment_mode': "Serial Redemption",
                    'due_date': due_date.strftime('%Y-%m-%d %H:%M:%S'),
                    'haspayed': True,
                    'terms_accepted': True,
                    'privacy_accepted': True,
                    'enrollment_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
            else:
                # Create new entry
                PAYMENT_DATA[user_id_str] = {
                    'username': username,
                    'payment_plan': plan,
                    'mentorship_type': mentorship_type.capitalize(),
                    'payment_mode': "Serial Redemption",
                    'due_date': due_date.strftime('%Y-%m-%d %H:%M:%S'),
                    'haspayed': True,
                    'terms_accepted': True,
                    'privacy_accepted': True,
                    'enrollment_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'form_answers': {}  # Empty form answers for now
                }
            
            # Save payment data
            save_payment_data()
            
            # Reset PENDING_USERS status
            PENDING_USERS[user_id] = {
                'status': 'completed_onboarding',
                'membership_type': mentorship_type
            }
            save_pending_users()
            
            # Send success message
            bot.send_message(
                user_id,
                f"ðŸŽ‰ *SERIAL REDEEMED SUCCESSFULLY!*\n\n"
                f"Congratulations! You've successfully redeemed your serial number for a free membership:\n\n"
                f"â€¢ *Membership Type:* {mentorship_type.capitalize()}\n"
                f"â€¢ *Plan:* {plan}\n"
                f"â€¢ *Expiry Date:* {due_date.strftime('%Y-%m-%d')}\n\n"
                f"Your membership is now active! You'll now need to complete the onboarding form.",
                parse_mode="Markdown",
                reply_markup=markup
            )
            
            # Log the redemption
            logging.info(f"Serial {serial} redeemed by user {username} ({user_id}) for {mentorship_type} {plan}")
            
            # Notify admins
            for admin_id in ADMIN_IDS:
                try:
                    bot.send_message(
                        admin_id,
                        f"ðŸŽŸï¸ *Serial Redeemed*\n\n"
                        f"User @{username} has redeemed serial `{serial}`\n"
                        f"â€¢ *Membership Type:* {mentorship_type.capitalize()}\n"
                        f"â€¢ *Plan:* {plan}\n"
                        f"â€¢ *Expiry Date:* {due_date.strftime('%Y-%m-%d')}\n\n"
                        f"The user now has free access to this membership.",
                        parse_mode="Markdown"
                    )
                except Exception as e:
                    logging.error(f"Failed to notify admin {admin_id} about serial redemption: {e}")
            
            # Start the onboarding form process
            send_onboarding_form(user_id)
            
        else:
            bot.send_message(
                user_id,
                "âŒ There was an error processing your serial redemption. Please contact an administrator.",
                reply_markup=markup
            )
            PENDING_USERS.pop(user_id, None)
            delete_pending_user(user_id)
            
    elif response == "âŒ Cancel":
        bot.send_message(
            user_id,
            "âŒ Serial redemption cancelled.",
            reply_markup=markup
        )
        PENDING_USERS.pop(user_id, None)
        delete_pending_user(user_id)
        
    else:
        bot.send_message(user_id, "âŒ Invalid response. Please select one of the provided options.")

@bot.message_handler(commands=['list_serials'])
def list_serials(message):
    """List all serial numbers (admin only)"""
    # Check if user is admin or creator
    if message.from_user.id not in ADMIN_IDS and message.from_user.id != CREATOR_ID:
        bot.reply_to(message, "âŒ This command is only available to administrators.")
        return
    
    # Check if there are any serials
    if not SERIAL_NUMBERS:
        bot.reply_to(message, "â„¹ï¸ No serial numbers have been generated yet.")
        return
    
    # Parse command arguments
    args = message.text.split()
    filter_status = None
    
    if len(args) > 1:
        if args[1].lower() == "used":
            filter_status = True
        elif args[1].lower() == "unused":
            filter_status = False
    
    # Build the list
    serial_list = "ðŸ”‘ *SERIAL NUMBERS*\n\n"
    count = 0
    
    for serial, data in SERIAL_NUMBERS.items():
        # Apply filter if specified
        if filter_status is not None and data.get('used', False) != filter_status:
            continue
            
        count += 1
        if count > 20:  # Limit to 20 serials to avoid message length issues
            serial_list += f"\n... and {len(SERIAL_NUMBERS) - 20} more serials."
            break
            
        # Format serial info
        status = "âœ… Used" if data.get('used', False) else "â³ Available"
        mentorship_type = data.get('mentorship_type', 'regular').capitalize()
        plan = data.get('plan', 'Unknown')
        
        serial_list += f"*{serial}*\n"
        serial_list += f"â€¢ Status: {status}\n"
        serial_list += f"â€¢ Type: {mentorship_type}\n"
        serial_list += f"â€¢ Plan: {plan}\n"
        
        if data.get('used', False):
            used_by = data.get('used_by', 'Unknown')
            used_at = data.get('used_at', 'Unknown date')
            serial_list += f"â€¢ Used by: User {used_by}\n"
            serial_list += f"â€¢ Used on: {used_at}\n"
        
        serial_list += "\n"
    
    # Add usage instructions
    filter_note = ""
    if filter_status is not None:
        filter_type = "used" if filter_status else "unused"
        filter_note = f" (filtered: {filter_type} only)"
    
    serial_list += f"Total: {count} serials{filter_note}\n"
    serial_list += "Use `/list_serials used` or `/list_serials unused` to filter."
    
    bot.reply_to(message, serial_list, parse_mode="Markdown")

# Handler for /update command
@bot.message_handler(commands=['update'])
def handle_update_command(message):
    """Handle the /update command to opt in for notifications"""
    user_id = message.from_user.id
    
    # Check if user is already subscribed
    is_subscribed = user_id in UPDATE_SUBSCRIBERS
    
    # Create inline keyboard with Yes/No options
    markup = InlineKeyboardMarkup(row_width=2)
    
    if is_subscribed:
        markup.add(
            InlineKeyboardButton("âœ… Keep Me Updated", callback_data="update_yes"),
            InlineKeyboardButton("âŒ Unsubscribe", callback_data="update_no")
        )
        
        bot.reply_to(
            message,
            "ðŸ”” *Stay In The Loop!*\n\n"
            "You're currently receiving notifications about:\n"
            "â€¢ ðŸ”“ Enrollment openings & closings\n"
            "â€¢ ðŸ’° Special discount offers\n\n"
            "Would you like to continue receiving these updates?",
            parse_mode="Markdown",
            reply_markup=markup
        )
    else:
        markup.add(
            InlineKeyboardButton("âœ… Yes, Keep Me Updated", callback_data="update_yes"),
            InlineKeyboardButton("âŒ No Thanks", callback_data="update_no")
        )
        
        bot.reply_to(
            message,
            "ðŸ”” *Never Miss An Opportunity!*\n\n"
            "Would you like to receive instant notifications about:\n\n"
            "â€¢ ðŸ”“ When enrollment opens or closes\n"
            "â€¢ ðŸ’° Special discount promotions & offers\n"
            "â€¢ ðŸŽ Limited-time opportunities\n\n"
            "Stay informed about Prodigy Trading Academy opportunities!",
            parse_mode="Markdown",
            reply_markup=markup
        )

# Callback handler for update subscription
@bot.callback_query_handler(func=lambda call: call.data.startswith("update_"))
def handle_update_callback(call):
    """Handle the user's choice to receive updates"""
    user_id = call.from_user.id
    choice = call.data.split("_")[1]  # "yes" or "no"
    
    if choice == "yes":
        # Add user to subscribers list if not already there
        if user_id not in UPDATE_SUBSCRIBERS:
            UPDATE_SUBSCRIBERS.add(user_id)
            save_update_subscriber(user_id)
        
        bot.edit_message_text(
            "âœ… *You're All Set!*\n\n"
            "You'll now receive important notifications about:\n"
            "â€¢ When enrollment opens or closes\n"
            "â€¢ Special discount offers\n\n"
            "You can unsubscribe anytime by using the `/update` command again.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        
        bot.answer_callback_query(call.id, "You'll receive enrollment & discount updates! âœ…")
        logging.info(f"User {user_id} subscribed to updates")
    else:
        # Remove user from subscribers list if present
        if user_id in UPDATE_SUBSCRIBERS:
            UPDATE_SUBSCRIBERS.remove(user_id)
            remove_update_subscriber(user_id)
        
        bot.edit_message_text(
            "âŒ *Notifications Disabled*\n\n"
            "You won't receive updates about enrollment changes or discount offers.\n\n"
            "You can subscribe anytime by using the `/update` command.",
            call.message.chat.id,
            call.message.message_id,
            parse_mode="Markdown"
        )
        
        bot.answer_callback_query(call.id, "Update notifications disabled")
        logging.info(f"User {user_id} unsubscribed from updates")

def notify_enrollment_change_specific(enrollment_type, is_open):
    """Notify subscribers about enrollment status changes for a specific membership type"""
    status = "ðŸŸ¢ OPEN" if is_open else "ðŸ”´ CLOSED"
    action = "opened" if is_open else "closed"
    membership_name = "Regular Membership" if enrollment_type == "regular" else "Supreme Membership"
    
    message_text = (
        f"ðŸ“£ *ENROLLMENT STATUS UPDATE*\n\n"
        f"{membership_name} enrollment is now {status}!\n\n"
    )
    
    if is_open:
        message_text += (
            f"âœ… You can now purchase {membership_name} plans through the bot.\n\n"
            f"Use /start to begin the enrollment process!"
        )
    else:
        message_text += (
            f"âš ï¸ New {membership_name} purchases are temporarily unavailable.\n\n"
            f"Existing members can still renew their memberships."
        )
    
    # Send to all subscribers
    success_count = 0
    fail_count = 0
    
    for user_id in UPDATE_SUBSCRIBERS:
        try:
            bot.send_message(user_id, message_text, parse_mode="Markdown")
            success_count += 1
        except Exception as e:
            logging.error(f"Failed to notify user {user_id} about enrollment change: {e}")
            fail_count += 1
    
    logging.info(f"{enrollment_type.capitalize()} enrollment {action}: Notified {success_count} subscribers ({fail_count} failed)")

def notify_discount_created(discount_name, reg_discount, sup_discount):
    """Notify subscribers about new discount offers"""
    # Format discount details
    reg_percentage = reg_discount['percentage']
    sup_percentage = sup_discount['percentage']
    
    reg_end_date = datetime.strptime(reg_discount['end_date'], '%Y-%m-%d %H:%M:%S')
    sup_end_date = datetime.strptime(sup_discount['end_date'], '%Y-%m-%d %H:%M:%S')
    
    # Get user limits if available
    reg_user_limit = reg_discount.get('user_limit')
    sup_user_limit = sup_discount.get('user_limit')
    
    # Safely escape discount name for markdown
    safe_discount_name = safe_markdown_escape(discount_name)
    
    # Build the message
    message_text = (
        f"ðŸ”¥ *SPECIAL DISCOUNT ALERT!*\n\n"
        f"ðŸ“¢ **{safe_discount_name}**\n\n"
        f"*REGULAR MEMBERSHIP PLANS*\n"
        f"ðŸ’° **{reg_percentage}% OFF** on all regular membership plans!\n"
        f"â° Valid until: {reg_end_date.strftime('%B %d, %Y')}"
    )
    
    # Add user limit info for regular plans if available
    if reg_user_limit:
        message_text += f"\nðŸ‘¥ Limited to first {reg_user_limit} users only!"
    
    message_text += (
        f"\n\n*SUPREME MEMBERSHIP PLANS*\n"
        f"ðŸ’° **{sup_percentage}% OFF** on all supreme membership plans!\n"
        f"â° Valid until: {sup_end_date.strftime('%B %d, %Y')}"
    )
    
    # Add user limit info for supreme plans if available
    if sup_user_limit:
        message_text += f"\nðŸ‘¥ Limited to first {sup_user_limit} users only!"
    
    message_text += (
        f"\n\nðŸ”¸ Regular plans include: Trial, Momentum & Legacy\n"
        f"ðŸ”¸ Supreme plans include: Apprentice, Disciple & Lifetime\n\n"
        f"Act fast! Use the /start command to take advantage of this limited-time offer!"
    )
    
    # Create a plain text version for fallback
    plain_message_text = message_text.replace('*', '')
    
    # Send to all subscribers
    success_count = 0
    fail_count = 0
    
    for user_id in UPDATE_SUBSCRIBERS:
        try:
            bot.send_chat_action(user_id, 'typing')  # First check if user can receive messages
            
            # Try with Markdown first
            try:
                bot.send_message(
                    user_id,
                    message_text,
                    parse_mode="Markdown"
                )
            except ApiException as e:
                # If Markdown fails, try with plain text as fallback
                if "can't parse entities" in str(e):
                    bot.send_message(
                        user_id,
                        plain_message_text,
                        parse_mode=None
                    )
                    logging.warning(f"Sent discount update to user {user_id} without markdown formatting")
                else:
                    # Re-raise if it's not a markdown parsing issue
                    raise
                    
            success_count += 1
            # Add small delay to prevent hitting rate limits
            time.sleep(0.05)
        except Exception as e:
            logging.error(f"Failed to send discount update to user {user_id}: {e}")
            fail_count += 1
    
    logging.info(f"Discount '{discount_name}' created: Notified {success_count} subscribers ({fail_count} failed)")

def check_and_reset_rate_limits():
    """Background thread to periodically check and reset expired rate limits"""
    logging.info("Rate limit checker thread started")
    
    while True:
        try:
            now = time.time()
            users_reset = 0
            
            # Check each user's rate limit
            for user_id, data in list(QWEN_USAGE.items()):
                last_used = data.get('last_used', 0)
                
                # If more than an hour has passed since last use
                if now - last_used >= 3600:
                    # Reset count to 0
                    QWEN_USAGE[user_id]['count'] = 0
                    
                    # If there was a 'reset_scheduled' flag, clear it
                    if 'reset_scheduled' in QWEN_USAGE[user_id]:
                        QWEN_USAGE[user_id]['reset_scheduled'] = False
                    
                    # Update in MongoDB too
                    jarvis_usage_collection.update_one(
                        {"user_id": user_id},
                        {"$set": {"count": 0}},
                        upsert=True
                    )
                    users_reset += 1
            
            if users_reset > 0:
                logging.info(f"Reset rate limits for {users_reset} users")
            
            # Run this check every minute
            time.sleep(60)
            
        except Exception as e:
            logging.error(f"Error in rate limit checker: {e}")
            time.sleep(60)  # Wait a minute before retrying on error

keep_alive()

# Start the rate limit checker thread
# threading.Thread(target=check_and_reset_rate_limits, daemon=True).start()

# Start reminder thread
reminder_thread = threading.Thread(target=send_payment_reminder)
reminder_thread.daemon = True  # This ensures the thread will exit when the main program exits
reminder_thread.start()

# Start the scheduled GIF thread
scheduled_gif_thread = threading.Thread(target=send_scheduled_gifs, daemon=True)
scheduled_gif_thread.start()

# Start the daily challenge thread
daily_challenge_thread = threading.Thread(target=send_daily_challenges, daemon=True)
daily_challenge_thread.start()

# Start MongoDB refresh thread
refresh_thread = threading.Thread(target=mongodb_refresh_thread)
refresh_thread.daemon = True
refresh_thread.start()

# Start reminder thread
pending_reminder_thread = threading.Thread(target=send_pending_request_reminders)
pending_reminder_thread.daemon = True
pending_reminder_thread.start()

# Start the leaderboard thread
leaderboard_thread = threading.Thread(target=send_daily_leaderboard, daemon=True)
leaderboard_thread.start()

# Add this to your bot startup code
midnight_thread = threading.Thread(target=midnight_cleanup_thread, daemon=True)
midnight_thread.start()

# Start the discount expiry thread
discount_thread = threading.Thread(target=discount_expiry_thread, daemon=True)
discount_thread.start()

# Start the pending users cleanup thread
pending_cleanup_thread = threading.Thread(target=cleanup_inactive_pending_users, daemon=True)
pending_cleanup_thread.start()

# Start the birthday greeting thread
birthday_thread = threading.Thread(target=birthday_check_thread, daemon=True)
birthday_thread.start()

# Start the trial reminder checker thread
threading.Thread(target=check_trial_reminders, daemon=True).start()

# Function to start the bot with auto-restart
@server.route('/telegram_webhook', methods=['POST'])
def webhook():
    try:
        json_string = request.get_data().decode('utf-8')
        update = telebot.types.Update.de_json(json_string)
        bot.process_new_updates([update])
        return '', 200
    except Exception as e:
        logging.error(f"Error in webhook handler: {e}", exc_info=True)
        return '', 500

# Add health check endpoint
@server.route('/health', methods=['GET'])
def health_check():
    return 'Bot is running', 200

# Add root endpoint
@server.route('/', methods=['GET'])
def index():
    return 'PTA Bot is running', 200

# Now modify your start_bot function to not define these routes again
def start_bot():
    """Start the bot using webhooks instead of polling"""
    # Set up webhook
    WEBHOOK_URL = os.environ.get('WEBHOOK_URL', 'https://ptabot.up.railway.app/')
    WEBHOOK_SECRET_PATH = os.environ.get('WEBHOOK_SECRET_PATH', 'telegram_webhook')
    PORT = int(os.environ.get('PORT', 5001))
    
    # Log webhook setup
    logging.info(f"Setting up webhook at {WEBHOOK_URL + WEBHOOK_SECRET_PATH}")
    
    try:
        # Remove any existing webhook first
        bot.remove_webhook()
        time.sleep(1)  # Small delay to ensure webhook is removed
        
        # Set up the new webhook
        bot.set_webhook(url=WEBHOOK_URL + WEBHOOK_SECRET_PATH)
        logging.info("Webhook set successfully")
        
        # Start the Flask server
        logging.info(f"Starting Flask server on port {PORT}")
        server.run(host='0.0.0.0', port=PORT)
        
    except Exception as e:
        logging.critical(f"Critical error setting up webhook: {e}", exc_info=True)
        
        # If webhook setup fails, fall back to polling as a backup
        logging.info("Falling back to polling method")
        use_polling_fallback()

def use_polling_fallback():
    """Fallback to polling if webhook fails"""
    consecutive_errors = 0
    max_consecutive_errors = 5
    backoff_time = 5  # Initial backoff time in seconds
    max_backoff_time = 300  # Maximum backoff time (5 minutes)
    
    logging.warning("Using polling as fallback method")
    
    while True:
        try:
            # Make sure webhook is removed before polling
            bot.remove_webhook()
            time.sleep(2)
            
            logging.info("Starting polling mode...")
            connection_time = datetime.now()
            logging.info(f"Bot connecting at {connection_time.strftime('%Y-%m-%d %H:%M:%S')}")
            
            # Start polling with better parameters
            bot.polling(none_stop=True, interval=1, timeout=60)
            
            # Reset error counter on successful connection
            consecutive_errors = 0
            backoff_time = 5
            
            logging.info("Bot is online and processing messages via polling")
            
        except requests.exceptions.ReadTimeout:
            # Handle timeout separately - this is generally not critical
            logging.warning("Connection timeout, reconnecting...")
            time.sleep(3)
            
        except requests.exceptions.ConnectionError:
            # Network connection issues
            consecutive_errors += 1
            wait_time = min(backoff_time * consecutive_errors, max_backoff_time)
            logging.error(f"Network connection error. Attempt {consecutive_errors}. Waiting {wait_time}s before retry.")
            time.sleep(wait_time)
            
        except telebot.apihelper.ApiException as api_error:
            # Handle Telegram API errors
            consecutive_errors += 1
            
            if "429" in str(api_error):  # Rate limiting
                # Extract retry_after if available
                retry_after = 30  # Default
                if hasattr(api_error, 'result') and 'retry_after' in api_error.result.json:
                    retry_after = api_error.result.json['retry_after']
                    
                logging.warning(f"Rate limited by Telegram. Waiting {retry_after}s before retry.")
                time.sleep(retry_after + 1)  # Add 1 second buffer
            elif "409" in str(api_error):  # Conflict error
                logging.error("Conflict error (409). Another instance might be running.")
                time.sleep(15)  # Wait longer for conflict errors
            else:
                wait_time = min(backoff_time * consecutive_errors, max_backoff_time)
                logging.error(f"Telegram API error: {api_error}. Waiting {wait_time}s before retry.")
                time.sleep(wait_time)
        
        except Exception as e:
            # General exception handling with exponential backoff
            consecutive_errors += 1
            wait_time = min(backoff_time * consecutive_errors, max_backoff_time)
            
            # Log more detailed error information
            logging.error(f"Error occurred: {e}", exc_info=True)
            logging.error(f"Consecutive errors: {consecutive_errors}. Waiting {wait_time}s before retry.")
            
            # Check if we've hit too many consecutive errors
            if consecutive_errors >= max_consecutive_errors:
                logging.critical(f"Hit {max_consecutive_errors} consecutive errors. Attempting bot reset.")
                # Could add additional recovery logic here
                consecutive_errors = 0  # Reset counter after recovery attempt
            
            time.sleep(wait_time)
            logging.info("Restarting the bot...")

if __name__ == "__main__":
    # Make sure all background threads are started before setting up webhooks
    start_bot()